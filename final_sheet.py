"""
final_sheet.py
--------------
Confluence module. Unlike every other module in this pipeline, this one
does NOT take a raw OHLC data_dict -- it reads the '<Indicator> Recomm'
row that each indicator module already wrote into its own sheet of the
SAME workbook, and combines them. It must therefore run strictly AFTER
every indicator's write_matrix() has finished (i.e. as its own pipeline
step following Step 7 in 01_Master_Code.py), not in the parallel
build_matrix() phase.

Confluence rule -- category-based majority, not strict unanimous vote
(see the algo-trading conversation this was requested from: 7-way literal
unanimity was found to be over-counting correlated indicators rather than
genuinely diversifying the signal -- EMA 20, VWAP, and TW ALL are all
trend-direction indicators computed differently, so requiring all three to
agree was really "one trend opinion, asked three times", not three
independent confirmations):

    Indicators are grouped by what they actually measure, and confluence
    requires agreement WITHIN each category, not literal unanimity across
    every sheet:

        TREND     (EMA 20, VWAP, TW ALL)  -- >= 2 of 3 must agree
        MOMENTUM  (RSI, SQZMOM)           -- both must agree (kept strict:
                                              only 2 indicators here, and
                                              momentum is the entry trigger)
        VOLUME    (OBV CMF)               -- must agree (single vote,
                                              validates real participation
                                              behind the move)

    Final Recomm is 'BUY CE' only if TREND, MOMENTUM, and VOLUME all
    resolve to 'BUY CE' by the rules above; symmetrically for 'BUY PE';
    'WAIT' otherwise -- including when a category's own indicator sheet
    is entirely missing, same as before.

    [CHANGED -- 13-Jul-26] The third category's voter is now OBV CMF, not
    BRKPRO. A prior single-day audit (see obv_cmf.py's module docstring)
    measured BRKPRO's Recomm at ~49.95% forward hit-rate with roughly
    7,500 direction flips in one day -- statistically indistinguishable
    from noise on that sample, yet BRKPRO held a mandatory, single-vote
    veto over every trade in the system (BREAKOUT_MIN_AGREE = 1 in a
    category with only one member). obv_cmf.py was written specifically
    to replace that vote with a real, largely independent confirmation
    (OBV's cumulative running volume agreeing with CMF's bounded
    high/low/close-range oscillator) -- see that module's docstring --
    but the swap was never actually made in this file until now. BRKPRO's
    own sheet is still computed and shown in the 'Final' sheet's
    passthrough rows for visibility; it simply no longer gates
    Final Recomm. Rename the category back to BREAKOUT_SHEETS = ["BRKPRO"]
    below if you want to A/B the two.

    ADX is intentionally EXCLUDED from the vote (though its passthrough
    row is still shown in the 'Final' sheet for visibility). It's already
    consumed as a regime-STRENGTH gate in order_sheet.py (ADX_MIN
    threshold on the pre-entry bar) -- also requiring its own BUY CE/PE
    opinion here would double-count the same underlying DI/ADX
    computation as both a gate AND a vote.

    Set CONFLUENCE_MODE = "strict_unanimous" below to restore the exact
    original 7-way-unanimous behavior (BRKPRO included) if you want to
    A/B the two.

Which sheets vote:
    SQZMOM, RSI, ADX, TW ALL, EMA 20, VWAP, OBV CMF -- BRKPRO is still
    read and shown as a passthrough row but no longer contributes a vote
    (see CHANGED note above).
    'HTF Bias' is deliberately excluded: it never produces a Recomm of its
    own (see htf_bias.py's module docstring for why) -- it's a gate
    consumed inside ema20.py's own Recomm, not an independent vote here.

Output: a 'Final' sheet with one block of rows per symbol -- each source
indicator's Recomm row passed through unchanged (so you can see exactly
which indicator(s) broke confluence on a given cell), followed by a bolded
'Final Recomm' row.
"""

import os
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

CODES_DIR = os.path.dirname(os.path.abspath(__file__))
if CODES_DIR not in sys.path:
    sys.path.append(CODES_DIR)
import excel_utils

# (sheet name in the workbook, exact 'Metrics' label of its Recomm row)
# BRKPRO stays here so its own row is still shown as a passthrough in the
# 'Final' sheet -- see CHANGED note above for why it no longer votes.
INDICATOR_SHEETS = [
    ("SQZMOM",     "SQZMOM Recomm"),
    ("RSI",        "RSI Recomm"),
    ("BRKPRO",     "BRKPRO Recomm"),
    ("ADX",        "ADX Recomm"),
    ("TW ALL",     "TW ALL Recomm"),
    ("EMA 20",     "EMA 20 Recomm"),
    ("VWAP",       "VWAP Recomm"),
    ("OBV CMF",    "OBV CMF Recomm"),
    # [ADDED] Passthrough only, same as ADX/BRKPRO above -- see
    # zerolag.py's docstring for why this doesn't get its own vote here
    # (it's consumed as an order_sheet.py gate instead, avoiding
    # double-counting the same trend computation as both a vote and a
    # gate). Deliberately NOT listed in TREND_SHEETS below.
    ("ZLTREND",    "ZL Recomm"),
]

# "category" (default) = grouped majority per the docstring above.
# "strict_unanimous" = restore the original all-7-must-agree behavior
# (BRKPRO included, OBV CMF excluded, ADX still excluded).
CONFLUENCE_MODE = "category"

# ADX is deliberately absent from every category below -- see docstring.
TREND_SHEETS = ["EMA 20", "VWAP", "TW ALL"]
TREND_MIN_AGREE = 2          # >= 2 of 3
MOMENTUM_SHEETS = ["RSI", "SQZMOM"]
MOMENTUM_MIN_AGREE = 2       # both (2 of 2)
# [CHANGED] OBV CMF replaces BRKPRO as the third category's voter -- see
# module docstring. BRKPRO is intentionally NOT listed in any category
# below, so it contributes no vote while still being computed/shown.
VOLUME_SHEETS = ["OBV CMF"]
VOLUME_MIN_AGREE = 1         # the only vote in its category

# ---------------------------------------------------------------------------
# Excel styling (kept visually consistent with the other sheets)
# ---------------------------------------------------------------------------
FILL_LIME = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")   # BUY CE
FILL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # BUY PE
FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")   # WAIT
FONT_WHITE = Font(color="FFFFFF")
FONT_BLACK = Font(color="000000")
FONT_BOLD = Font(bold=True)


def _load_recomm_table(output_excel_path, sheet_name, recomm_label):
    """Returns {symbol: {time_str: value}} for one indicator sheet's Recomm
    row. Returns {} (rather than raising) if the sheet is missing or
    malformed -- a missing indicator simply contributes no votes, which
    correctly fails confluence for every cell it would have covered
    instead of crashing the whole Final sheet build."""
    try:
        df = pd.read_excel(output_excel_path, sheet_name=sheet_name)
    except Exception as e:
        print(f"[WARNING] Final Sheet: '{sheet_name}' sheet not found ({e}) -- its vote is WAIT everywhere.")
        return {}

    if 'Metrics' not in df.columns or 'Symbol' not in df.columns:
        print(f"[WARNING] Final Sheet: '{sheet_name}' sheet is missing 'Symbol'/'Metrics' columns -- skipped.")
        return {}

    recomm_df = df[df['Metrics'] == recomm_label]
    if recomm_df.empty:
        print(f"[WARNING] Final Sheet: no '{recomm_label}' row found in '{sheet_name}' -- its vote is WAIT everywhere.")
        return {}

    time_cols = [c for c in df.columns if c not in ('Symbol', 'Metrics')]

    table = {}
    for _, row in recomm_df.iterrows():
        sym = str(row['Symbol']).strip().upper()
        table[sym] = {
            str(t): row[t] for t in time_cols
            if pd.notna(row[t]) and str(row[t]).strip() != ""
        }
    return table


def _category_vote(tables, category_sheets, sym, t, min_agree):
    """'BUY CE' / 'BUY PE' / 'WAIT' for one category at one (sym, t)
    cell. A sheet missing from `tables` (never loaded) or with no vote
    for this cell counts as 'WAIT', same as the original per-cell
    default -- it can contribute to neither side reaching min_agree."""
    votes = [tables.get(s, {}).get(sym, {}).get(t, 'WAIT') for s in category_sheets]
    ce_votes = sum(1 for v in votes if v == 'BUY CE')
    pe_votes = sum(1 for v in votes if v == 'BUY PE')
    if ce_votes >= min_agree:
        return 'BUY CE'
    if pe_votes >= min_agree:
        return 'BUY PE'
    return 'WAIT'


def build_final_matrix(output_excel_path):
    """Reads every sheet in INDICATOR_SHEETS out of the already-written
    workbook and builds the pivoted Symbol x Time 'Final' matrix. No
    Excel writing here -- see write_matrix()."""
    tables = {}
    all_symbols = set()
    all_times = set()

    for sheet_name, recomm_label in INDICATOR_SHEETS:
        table = _load_recomm_table(output_excel_path, sheet_name, recomm_label)
        tables[sheet_name] = table
        for sym, time_map in table.items():
            all_symbols.add(sym)
            all_times.update(time_map.keys())

    if not all_symbols:
        raise RuntimeError(
            "Final Sheet: none of the indicator sheets produced any Recomm rows -- matrix build aborted."
        )

    sorted_times = sorted(all_times)
    matrix_rows = [['Symbol', 'Metrics'] + sorted_times]

    for sym in sorted(all_symbols):
        per_indicator_rows = {
            sheet_name: [sym, recomm_label] for sheet_name, recomm_label in INDICATOR_SHEETS
        }
        final_row = [sym, 'Final Recomm']

        for t in sorted_times:
            for sheet_name, recomm_label in INDICATOR_SHEETS:
                val = tables[sheet_name].get(sym, {}).get(t, 'WAIT')
                per_indicator_rows[sheet_name].append(val)

            if CONFLUENCE_MODE == "strict_unanimous":
                # Original behavior, preserved for A/B comparison: every
                # sheet in INDICATOR_SHEETS (including ADX and BRKPRO,
                # excluding OBV CMF) must agree.
                legacy_sheets = [s for s in INDICATOR_SHEETS if s[0] != "OBV CMF"]
                votes = [tables[s].get(sym, {}).get(t, 'WAIT') for s, _ in legacy_sheets]
                if all(v == 'BUY CE' for v in votes):
                    final_row.append('BUY CE')
                elif all(v == 'BUY PE' for v in votes):
                    final_row.append('BUY PE')
                else:
                    final_row.append('WAIT')
            else:
                trend = _category_vote(tables, TREND_SHEETS, sym, t, TREND_MIN_AGREE)
                momentum = _category_vote(tables, MOMENTUM_SHEETS, sym, t, MOMENTUM_MIN_AGREE)
                volume = _category_vote(tables, VOLUME_SHEETS, sym, t, VOLUME_MIN_AGREE)

                if trend == momentum == volume == 'BUY CE':
                    final_row.append('BUY CE')
                elif trend == momentum == volume == 'BUY PE':
                    final_row.append('BUY PE')
                else:
                    final_row.append('WAIT')

        for sheet_name, _ in INDICATOR_SHEETS:
            matrix_rows.append(per_indicator_rows[sheet_name])
        matrix_rows.append(final_row)

    return matrix_rows


def write_matrix(matrix_rows, output_excel_path):
    """Injects matrix_rows into the 'Final' sheet of an existing workbook
    (preserves every other sheet) and autofits + atomically saves."""
    wb = load_workbook(output_excel_path)
    sheet_name = "Final"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    for r in matrix_rows:
        ws.append(r)

    ws.freeze_panes = "C2"

    # Color applies ONLY to 'Final Recomm' -- applying it to every
    # per-indicator passthrough row too would bury the one row that
    # actually matters (the confluence result) under several other
    # identically-colored rows.
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1), start=2):
        metric_type = ws.cell(row=r_idx, column=2).value
        is_final_recomm_row = (metric_type == 'Final Recomm')
        if not is_final_recomm_row:
            continue

        for cell in row[2:]:
            val = cell.value
            if val == "":
                continue
            if val == 'BUY CE':
                cell.fill = FILL_LIME
                cell.font = FONT_BLACK
            elif val == 'BUY PE':
                cell.fill = FILL_RED
                cell.font = FONT_WHITE
            elif val == 'WAIT':
                cell.fill = FILL_GRAY
                cell.font = FONT_BLACK

        # Bold the label cells so the confluence row stands out from the
        # (now uncolored) per-indicator passthrough rows above it.
        row[0].font = FONT_BOLD
        row[1].font = FONT_BOLD

    excel_utils.autofit_columns(ws)
    excel_utils.atomic_save(wb, output_excel_path)


def run_final_sheet_step(output_excel_path):
    """Single entry point for 01_Master_Code.py. Must run AFTER every
    indicator's write_matrix() has already saved into output_excel_path."""
    print("[SYSTEM] Reading Recomm rows from all indicator sheets for confluence check...")
    matrix_rows = build_final_matrix(output_excel_path)
    write_matrix(matrix_rows, output_excel_path)
    print("[SUCCESS] Final Recomm matrix written to sheet 'Final'.")


# ---------------------------------------------------------------------------
# Disclaimer: 'Final Recomm' is a mechanical category-majority aggregation
# of eight heuristic indicators (seven voting, BRKPRO shown only), none of
# which has been backtested in this conversation, individually or in
# combination. It is not financial advice. Paper-trade the full pipeline
# and run it through scripts/backtester.py before risking real capital.
# ---------------------------------------------------------------------------
