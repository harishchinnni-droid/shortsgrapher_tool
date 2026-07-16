"""
SQZMOM.py
---------
Squeeze Momentum indicator, wired to run as Step 6 of the daily pipeline.
Output format is the pivoted Symbol x Time matrix shared by every
indicator module in this pipeline.

[FIX -- 13-Jul-26] Two bugs fixed in this pass, both contributing to the
same symptom (the workbook showing candle data before the market had
even opened):

1. process_symbol() never applied CUTOFF_TIME truncation or a timezone
   strip, unlike every other module in this pipeline (RSI.py, adx_di.py,
   etc.). Practically: SQZMOM's sheet carried a DIFFERENT set of time
   columns than the other seven voting sheets for the same symbol/day
   (through 15:25 instead of 15:15), so final_sheet.py's time-keyed
   lookup could silently miss SQZMOM's vote on the last two candles of
   the day. Fixed by adopting the same brute-force time/timezone/cutoff
   protocol every other module already uses.
2. The Symbol x Time pivot used to dedupe purely on time-of-day
   ('09:15', '09:20', ...) with no date component, keeping the LAST
   occurrence of each slot across up to 90 days of cached history. Before
   target_date's own candle for a given slot existed, "last occurrence"
   silently fell back to the most recent PRIOR trading day's value for
   that same clock time -- showing a "complete" day of numbers even
   pre-market. Fixed via excel_utils.restrict_to_target_date(), called
   right before process_symbol() returns -- see that function's docstring
   and 01_Master_Code.py's market-open scheduling fix, which the two
   fixes are designed to work together with.
"""
import pandas as pd
import numpy as np
import concurrent.futures
import os
import sys
from datetime import time as dtime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

CODES_DIR = os.path.dirname(os.path.abspath(__file__))
if CODES_DIR not in sys.path:
    sys.path.append(CODES_DIR)
import excel_utils

CUTOFF_TIME = dtime(15, 15)
INTERVAL = "5minute"

# Define exact color hexes
FILL_LIME = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
FILL_GREEN = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
FILL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
FILL_MAROON = PatternFill(start_color="800000", end_color="800000", fill_type="solid")
FILL_ORANGE = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
FILL_BLACK = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
FILL_GRAY = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
FONT_WHITE = Font(color="FFFFFF")
FONT_BLACK = Font(color="000000")

def linreg(x, length):
    """Vectorized linear regression for momentum calculation."""
    result = np.full_like(x, np.nan, dtype=float)
    x_mean = (length - 1) / 2.0
    x_dev = np.arange(length) - x_mean
    sum_sq_x_dev = np.sum(x_dev**2)
    
    for i in range(length - 1, len(x)):
        y_val = x[i - length + 1 : i + 1]
        y_mean = np.mean(y_val)
        y_dev = y_val - y_mean
        slope = np.sum(x_dev * y_dev) / sum_sq_x_dev
        intercept = y_mean - slope * x_mean
        result[i] = intercept + slope * (length - 1)
    return result

def apply_signal_hold(dot_vals, bar_vals, entry_ce_vals, entry_pe_vals):
    """Turns single-bar entry triggers into a held state that persists on
    every subsequent bar until an exit condition fires:
        BUY CE exits when dot reverts Black->Orange, or bar flips Red/Maroon
        BUY PE exits when dot reverts Black->Orange, or bar flips Lime/Green
    Inherently stateful (bar t depends on whether a call was open at t-1),
    so it's a single sequential pass rather than a vectorized window.
    Returns a numpy object array of 'BUY CE' / 'BUY PE' / 'WAIT'.
    """
    n = len(dot_vals)
    recomm = np.full(n, 'WAIT', dtype=object)
    state = None
    for i in range(n):
        if state == 'BUY CE' and (dot_vals[i] == 'Orange' or bar_vals[i] in ('Red', 'Maroon')):
            state = None
        elif state == 'BUY PE' and (dot_vals[i] == 'Orange' or bar_vals[i] in ('Lime', 'Green')):
            state = None

        if state is None:
            if entry_ce_vals[i]:
                state = 'BUY CE'
            elif entry_pe_vals[i]:
                state = 'BUY PE'

        recomm[i] = state if state is not None else 'WAIT'
    return recomm


def calculate_sqzmom(df, length=20, mult=2.0, lengthKC=20, multKC=1.5):
    """Core logic translation with user's exhausted-breakout parameters."""
    df = df.copy()
    
    # Bollinger Bands
    basis = df['close'].rolling(window=length).mean()
    dev = mult * df['close'].rolling(window=length).std(ddof=0) 
    upperBB = basis + dev
    lowerBB = basis - dev
    
    # Keltner Channels
    ma = df['close'].rolling(window=lengthKC).mean()
    tr0 = abs(df['high'] - df['low'])
    tr1 = abs(df['high'] - df['close'].shift())
    tr2 = abs(df['low'] - df['close'].shift())
    tr = pd.concat([tr0, tr1, tr2], axis=1).max(axis=1)
    rangema = tr.rolling(window=lengthKC).mean()
    upperKC = ma + rangema * multKC
    lowerKC = ma - rangema * multKC
    
    # Squeeze Conditions
    sqzOn = (lowerBB > lowerKC) & (upperBB < upperKC)
    sqzOff = (lowerBB < lowerKC) & (upperBB > upperKC)
    
    # Momentum Value
    highest_high = df['high'].rolling(window=lengthKC).max()
    lowest_low = df['low'].rolling(window=lengthKC).min()
    avg_hl = (highest_high + lowest_low) / 2
    avg_hl_sma = (avg_hl + ma) / 2
    source_diff = df['close'] - avg_hl_sma
    
    val = linreg(source_diff.values, lengthKC)
    
    df['SQZ_VAL'] = val
    # sqzOn  (BB inside KC  -> low volatility, squeeze building)   = Orange
    # sqzOff (BB outside KC -> volatility released, squeeze fired) = Black
    # This matches the chart (orange during quiet consolidation, black across
    # the expansion hump) and the strategy comment two lines below, which
    # explicitly watches for Orange-building-to-Black as the entry trigger.
    df['DOT_COLOR'] = np.where(sqzOn, 'Orange', np.where(sqzOff, 'Black', 'Gray'))
    
    # Histogram Color Logic
    diff = df['SQZ_VAL'].diff()
    df['BAR_COLOR'] = 'Gray'
    df.loc[(df['SQZ_VAL'] > 0) & (diff > 0), 'BAR_COLOR'] = 'Lime'  
    df.loc[(df['SQZ_VAL'] > 0) & (diff <= 0), 'BAR_COLOR'] = 'Green'
    df.loc[(df['SQZ_VAL'] < 0) & (diff < 0), 'BAR_COLOR'] = 'Red'   
    df.loc[(df['SQZ_VAL'] < 0) & (diff >= 0), 'BAR_COLOR'] = 'Maroon'
    
    # SIGNAL GENERATION: 6 consecutive Orange dots shifting to Black = ENTRY.
    orange_count = (df['DOT_COLOR'] == 'Orange').astype(int).groupby((df['DOT_COLOR'] != 'Orange').cumsum()).cumsum()
    prev_orange_count = orange_count.shift(1)
    
    buy_ce_entry = (prev_orange_count >= 6) & (df['DOT_COLOR'] == 'Black') & (df['BAR_COLOR'] == 'Lime')
    buy_pe_entry = (prev_orange_count >= 6) & (df['DOT_COLOR'] == 'Black') & (df['BAR_COLOR'] == 'Red')

    # HOLD: once triggered, the call is carried forward on every subsequent
    # bar (not just the entry bar) until an exit condition fires:
    #   BUY CE exits when the dot reverts Black->Orange (squeeze re-forming)
    #                   or the bar flips to Red/Maroon (momentum turned bearish)
    #   BUY PE exits when the dot reverts Black->Orange (squeeze re-forming)
    #                   or the bar flips to Lime/Green (momentum turned bullish)
    df['SQZMOM Recomm'] = apply_signal_hold(
        df['DOT_COLOR'].to_numpy(),
        df['BAR_COLOR'].to_numpy(),
        buy_ce_entry.to_numpy(),
        buy_pe_entry.to_numpy(),
    )
    
    return df


def process_symbol(symbol_data):
    """Normalizes columns, extracts a tz-naive 'HH:MM' time_str (same
    brute-force protocol as every other module in this pipeline), computes
    the indicator on the full multi-day history, then restricts the
    returned rows to target_date only. See module docstring for why both
    the cutoff/tz handling and the date restriction were added here."""
    symbol, df, target_date = symbol_data
    df = df.copy()

    df.columns = df.columns.astype(str).str.strip().str.lower()

    if not all(col in df.columns for col in ['open', 'high', 'low', 'close']):
        print(f"[WARNING] {symbol}: Missing required price columns. Discarding.")
        return symbol, None

    time_col = next((col for col in ['date', 'time', 'datetime', 'timestamp'] if col in df.columns), None)
    if time_col:
        parsed = pd.to_datetime(df[time_col], errors='coerce')
    elif pd.api.types.is_datetime64_any_dtype(df.index):
        parsed = pd.Series(df.index, index=df.index)
    else:
        unnamed_cols = [c for c in df.columns if 'unnamed' in c]
        if unnamed_cols:
            parsed = pd.to_datetime(df[unnamed_cols[0]], errors='coerce')
        else:
            print(f"[WARNING] {symbol}: Failed to locate any valid timestamp data. Discarding.")
            return symbol, None

    # [FIX] Strip tz label if present (Kite returns Asia/Kolkata tz-aware
    # timestamps) -- openpyxl can't write a tz-aware datetime cell, and
    # every other module already does this. SQZMOM previously skipped it,
    # which was harmless only because time_str is a plain string, but kept
    # SQZMOM inconsistent with the shared protocol.
    try:
        if parsed.dt.tz is not None:
            parsed = parsed.dt.tz_localize(None)
    except (AttributeError, TypeError):
        pass

    df['_sort_dt'] = parsed
    df['time_str'] = parsed.dt.strftime('%H:%M')
    df.dropna(subset=['time_str', '_sort_dt'], inplace=True)
    if df.empty:
        print(f"[WARNING] {symbol}: No rows with a valid timestamp. Discarding.")
        return symbol, None

    df.sort_values('_sort_dt', inplace=True)

    # [FIX] CUTOFF_TIME truncation, same as every other module -- SQZMOM
    # previously carried candles through the literal end of the CSV
    # (15:25+), producing a different time-column set than the other
    # seven voting sheets for the same symbol/day.
    df = df[df['_sort_dt'].dt.time <= CUTOFF_TIME]
    if df.empty:
        print(f"[WARNING] {symbol}: No candles at/before {CUTOFF_TIME.strftime('%H:%M')}. Discarding.")
        return symbol, None

    df = calculate_sqzmom(df)
    df.dropna(subset=['close'], inplace=True)
    if df.empty:
        return symbol, None

    # [FIX] Date-scope the OUTPUT to target_date only -- see module
    # docstring and excel_utils.restrict_to_target_date()'s docstring.
    df = excel_utils.restrict_to_target_date(df, target_date)
    if df is None:
        return symbol, None

    return symbol, df

def build_matrix(data_dict, target_date, max_workers=None):
    """Executes multi-core calculation and builds the pivoted Matrix
    (no Excel I/O -- safe to run concurrently with other indicators'
    build_matrix() calls since it never touches the shared workbook).

    [CHANGED] Now requires target_date -- threaded into every
    process_symbol() call so the output can be restricted to that day's
    own rows (see module docstring). Passing the wrong/no target_date is
    now a TypeError at the call site rather than a silent stale-data bug
    at runtime -- see 01_Master_Code.py's _compute_and_write_matrices()."""
    results = {}

    with concurrent.futures.ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(process_symbol, (sym, df, target_date)): sym for sym, df in data_dict.items()}
        for future in concurrent.futures.as_completed(futures):
            sym = futures[future]
            try:
                processed_sym, processed_df = future.result()
                if processed_df is not None and not processed_df.empty and sym.lower() != 'summary':
                    results[sym] = processed_df
            except Exception as e:
                print(f"[ERROR] Engine Failure on symbol {sym}: {str(e)}")

    if not results:
        # [NOTE] This now fires legitimately (not just on a real failure)
        # if called before target_date's first candle has closed -- e.g.
        # a manual/backtest run against a date with no data yet. That's
        # intentional: failing loudly here beats silently falling back to
        # a prior day's data, which was the original bug. Under normal
        # LIVE operation, 01_Master_Code.py's market-open scheduling fix
        # means this function is never called before 09:20:05 IST, so
        # this should not fire in practice.
        raise RuntimeError("SQZMOM: zero symbols processed successfully for target_date -- matrix build aborted.")

    all_times = set()
    for df in results.values():
        if 'time_str' in df.columns:
            all_times.update(df['time_str'].dropna().tolist())
    
    if not all_times:
        raise RuntimeError("SQZMOM: no valid timestamps extracted across all symbols -- matrix build aborted.")
        
    sorted_times = sorted(list(all_times))

    matrix_rows = []
    matrix_rows.append(['Symbol', 'Metrics'] + sorted_times) # Header Row
    
    for sym in sorted(results.keys()):
        df = results[sym]
        # De-duplicate times in case of overlapping data rows before pivoting
        df = df.drop_duplicates(subset=['time_str'], keep='last')
        df_indexed = df.set_index('time_str')
        
        def get_metric_row(metric_name, col_name):
            row = [sym, metric_name]
            for t in sorted_times:
                if t in df_indexed.index:
                    val = df_indexed.loc[t, col_name]
                    if isinstance(val, pd.Series): val = val.iloc[-1]
                    row.append(val)
                else:
                    row.append("")
            return row
        
        matrix_rows.append(get_metric_row('Open', 'open'))
        matrix_rows.append(get_metric_row('Close', 'close'))
        matrix_rows.append(get_metric_row('DOT_COLOR', 'DOT_COLOR'))
        matrix_rows.append(get_metric_row('BAR_COLOR', 'BAR_COLOR'))
        matrix_rows.append(get_metric_row('SQZMOM Recomm', 'SQZMOM Recomm'))

    return matrix_rows


def write_matrix_to_workbook(matrix_rows, wb):
    """Applies the 'SQZMOM' sheet into an ALREADY-OPEN workbook (wb) --
    no file load/save here. Use this from an orchestrator that batches
    several indicators into a single load -> write-all -> save cycle
    (see 01_Master_Code.py's _compute_and_write_matrices()) instead of
    calling write_matrix() once per indicator -- see
    excel_utils.replace_sheet_with_matrix()'s docstring for why that
    reload/resave-per-indicator pattern was the real cost behind a slow
    'Computing indicator matrices' step."""
    sheet_name = "SQZMOM"
    ws = excel_utils.replace_sheet_with_matrix(wb, sheet_name, matrix_rows)

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1), start=2):
        metric_type = ws.cell(row=r_idx, column=2).value

        for c_idx, cell in enumerate(row[2:], start=3):
            val = cell.value
            if val == "": continue

            if metric_type == 'DOT_COLOR':
                if val == 'Black':
                    cell.fill = FILL_BLACK
                    cell.font = FONT_WHITE
                elif val == 'Orange':
                    cell.fill = FILL_ORANGE
                    cell.font = FONT_BLACK
                else:
                    cell.fill = FILL_GRAY
                    cell.font = FONT_BLACK

            elif metric_type == 'BAR_COLOR':
                if val == 'Lime': cell.fill = FILL_LIME
                elif val == 'Green': cell.fill = FILL_GREEN
                elif val == 'Red': cell.fill = FILL_RED
                elif val == 'Maroon': cell.fill = FILL_MAROON

            elif metric_type == 'SQZMOM Recomm':
                if val == 'BUY CE':
                    cell.fill = FILL_LIME
                    cell.font = FONT_BLACK
                elif val == 'BUY PE':
                    cell.fill = FILL_RED
                    cell.font = FONT_WHITE
                elif val == 'WAIT':
                    cell.fill = FILL_GRAY
                    cell.font = FONT_BLACK

    excel_utils.autofit_columns(ws)


def write_matrix(matrix_rows, output_excel_path):
    """Standalone/back-compat entry point: load -> apply -> save. Prefer
    write_matrix_to_workbook() when writing several indicators in the
    same run (see 01_Master_Code.py)."""
    wb = load_workbook(output_excel_path)
    write_matrix_to_workbook(matrix_rows, wb)
    excel_utils.atomic_save(wb, output_excel_path)


def parallel_compute_and_export(data_dict, target_date, output_excel_path, max_workers=None):
    """Backward-compatible fused entry point (build + write in one call) --
    what run_pipeline.py used to call directly for Step 6. Standalone
    scripts can still call this exactly as before; the pipeline itself now
    calls build_matrix() and write_matrix() separately so SQZMOM's
    computation can run concurrently with the other indicators while the
    actual Excel write stays serialized (see run_pipeline.py).

    [CHANGED] Now requires target_date -- see build_matrix()."""
    matrix_rows = build_matrix(data_dict, target_date, max_workers=max_workers)
    write_matrix(matrix_rows, output_excel_path)


def run_sqzmom_step(df_ref, target_date, output_excel_path):
    """Single entry point for run_pipeline.py's Step 6 (standalone use)."""
    import data_ingestion
    print("[SYSTEM] Loading 5-minute historical data for SQZMOM...")
    data_dict = data_ingestion.load_interval_data(df_ref, target_date, interval=INTERVAL)
    if not data_dict:
        raise RuntimeError("SQZMOM: no 5-minute data files found for any symbol.")
    print(f"[PROCESS] Computing SQZMOM matrix for {len(data_dict)} symbol(s)...")
    parallel_compute_and_export(data_dict, target_date, output_excel_path)
    print("[SUCCESS] SQZMOM matrix written to sheet 'SQZMOM'.")


# ---------------------------------------------------------------------------
# Disclaimer: this module estimates a directional bias from historical/live
# price data only. It is not financial advice, no result here is a
# guarantee of future performance, and the 'SQZMOM Recomm' rule has not
# been backtested in this conversation -- paper-trade it and run it
# through scripts/backtester.py before risking real capital.
# ---------------------------------------------------------------------------
