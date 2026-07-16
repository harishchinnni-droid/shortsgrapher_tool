"""
sheets_sync.py
---------------
Optional VIEWER MIRROR only -- after order_sheet.py resolves the 'Orders'
sheet and rebuilds 'Dashboard' (dashboard.py) in the local/Drive xlsx
workbook, this pushes plain-value copies of those two sheets into a
separate Google Sheet, purely so Harish can check trades/P&L from a
phone or browser without opening the Drive-mounted xlsx.

Deliberately NOT a replacement for the xlsx workbook. Every indicator
matrix, the Reference sheet, and all actual trading computation still
read/write the fast local/Drive xlsx exactly as before -- see
file_mgmt.py's BASE_DIR docstring. The Google Sheets API is network-bound
and rate-limited (roughly 60 write requests/minute/user by default),
which is a bad fit for this pipeline's high-frequency small writes across
9 parallel indicators -- moving THAT to Sheets would make the pipeline
slower, not faster. Mirroring just the two sheets a human actually reads
avoids that problem entirely: at most one sync per LIVE cycle (~every 5
minutes) or once at the end of a BACKTEST, each sync being only 1-2 API
calls (a clear + one bulk update per sheet), nowhere near any quota.

One-time setup required before this does anything (ENABLE_SHEETS_SYNC
stays False until it's done):
    1. console.cloud.google.com -> create/select a project.
    2. Enable "Google Sheets API" and "Google Drive API" for that project
       (search bar at the top -> API name -> Enable).
    3. IAM & Admin -> Service Accounts -> Create Service Account (any
       name, e.g. "sheets-writer") -> Create and Continue -> Done (no
       project role needed -- sharing the specific Sheet below is what
       actually grants access).
    4. Click into that service account -> Keys tab -> Add Key -> Create
       New Key -> JSON. This downloads a .json credentials file.
    5. Note the service account's email (looks like
       "xxxx@xxxx.iam.gserviceaccount.com", shown on the Service Accounts
       list page).
    6. Create a new Google Sheet (or pick an existing one) to be the
       "viewer" copy -- e.g. inside the 02_Claude_Trading Drive folder.
       Share it with the service account's email as Editor, exactly like
       sharing with a person.
    7. Copy that Sheet's ID (the long string in its URL between /d/ and
       /edit) into TARGET_SHEET_ID below.
    8. Upload the JSON key file from step 4 into
       <BASE_DIR>/01_JSON_Files/google_service_account.json (same Drive
       folder as the other credential JSONs, so both desktop and Colab
       find it via file_mgmt.BASE_DIR).
    9. Flip ENABLE_SHEETS_SYNC to True below (or leave it as an env var
       override -- see the flag itself).

Every call site wraps sync_to_google_sheets() in try/except -- a sync
failure (network hiccup, quota, misconfigured key, step 6-9 not done
yet) must NEVER block or fail the actual trading pipeline. Worst case,
the phone/browser view is stale until the next successful sync.
"""

import os
import file_mgmt

# [FLAG] One-time setup (see docstring) is done -- service account JSON
# is in 01_JSON_Files, and 01_SourceFile has been shared with it as
# Editor (see TARGET_SHEET_ID below: this reuses that existing sheet
# rather than a separate dedicated one -- new 'Orders'/'Dashboard' tabs
# get added alongside 'Reference'/'Copy of Reference 2' without touching
# them, since sync_to_google_sheets() only ever writes to sheet names in
# SHEETS_TO_MIRROR). Can still be overridden off via env var if needed:
#   os.environ["ENABLE_SHEETS_SYNC"] = "0"
ENABLE_SHEETS_SYNC = os.environ.get("ENABLE_SHEETS_SYNC", "1") == "1"

JSON_DIR = os.path.join(file_mgmt.BASE_DIR, "01_JSON_Files")
SERVICE_ACCOUNT_FILE = os.path.join(JSON_DIR, "google_service_account.json")

# [SET] Harish's existing 01_SourceFile Google Sheet -- reused as the
# viewer mirror target rather than a separate dedicated sheet (deliberate
# choice, see sheets_sync.py conversation notes). Override via env var
# without editing this file if the target ever needs to change.
TARGET_SHEET_ID = os.environ.get(
    "SHEETS_SYNC_TARGET_ID", "1ObfXwDkTZTCkiJIVzYDP1NAl1n19TLqgucZp-mOGIdY"
)

SHEETS_TO_MIRROR = ("Orders", "Dashboard")

_client = None  # lazy singleton -- one authenticated gspread client per process


def _get_client():
    global _client
    if _client is not None:
        return _client

    # Lazy import -- gspread/google-auth are only required when this
    # feature is actually enabled. Importing them unconditionally at
    # module load time would break every desktop run that hasn't pip
    # installed these two new packages yet, even though ENABLE_SHEETS_SYNC
    # defaults to off.
    import gspread
    from google.oauth2.service_account import Credentials

    if not os.path.exists(SERVICE_ACCOUNT_FILE):
        raise FileNotFoundError(
            f"Google service account key not found at '{SERVICE_ACCOUNT_FILE}'. "
            "See sheets_sync.py's module docstring for one-time setup steps."
        )
    if not TARGET_SHEET_ID:
        raise ValueError(
            "sheets_sync.TARGET_SHEET_ID (or the SHEETS_SYNC_TARGET_ID env "
            "var) is empty -- set it to your target spreadsheet's ID "
            "(from its URL, between /d/ and /edit) before enabling sync."
        )

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=scopes)
    _client = gspread.authorize(creds)
    return _client


def _sheet_to_rows(output_excel_path, sheet_name):
    """Plain-value snapshot of one sheet as a list of lists. Uses openpyxl
    directly (not pandas.read_excel) because 'Dashboard' is a free-form
    report -- merged title cells, a KPI panel, an embedded chart -- not a
    clean rectangular table with one header row the way pandas.read_excel
    assumes. data_only=True reads the last-saved VALUE of any formula
    cell rather than the formula text itself (there aren't any formulas
    written by this pipeline, but this is the safe default regardless)."""
    from openpyxl import load_workbook
    wb = load_workbook(output_excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    rows = [
        ["" if cell is None else cell for cell in row]
        for row in ws.iter_rows(values_only=True)
    ]
    return rows


def sync_to_google_sheets(output_excel_path, sheet_names=SHEETS_TO_MIRROR):
    """Best-effort mirror of the given sheets into the target Google
    Sheet. Call this AFTER dashboard.build_dashboard_sheet() so both
    sheets being mirrored are current. No-op (returns immediately) unless
    ENABLE_SHEETS_SYNC is True."""
    if not ENABLE_SHEETS_SYNC:
        return

    import gspread  # safe here -- _get_client() below already requires it to exist

    client = _get_client()
    spreadsheet = client.open_by_key(TARGET_SHEET_ID)

    synced = []
    for sheet_name in sheet_names:
        rows = _sheet_to_rows(output_excel_path, sheet_name)
        if rows is None:
            continue

        try:
            worksheet = spreadsheet.worksheet(sheet_name)
            worksheet.clear()
        except gspread.exceptions.WorksheetNotFound:
            n_rows = max(len(rows), 100)
            n_cols = max((len(r) for r in rows), default=10)
            worksheet = spreadsheet.add_worksheet(title=sheet_name, rows=n_rows, cols=n_cols)

        if rows:
            worksheet.update(values=rows, range_name="A1")
        synced.append(sheet_name)

    if synced:
        print(f"[SYSTEM] Mirrored {synced} to Google Sheets (viewer copy only -- xlsx remains the source of truth).")


# ---------------------------------------------------------------------------
# Disclaimer: this module only copies already-computed values for viewing.
# It reads nothing back from the Google Sheet and makes no trading
# decision -- the xlsx workbook (Reference/Orders/indicator sheets) is,
# and remains, the only source of truth the rest of this pipeline acts on.
# ---------------------------------------------------------------------------
