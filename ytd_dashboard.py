"""
ytd_dashboard.py
-----------------
STANDALONE, INDEPENDENT SCRIPT -- imports NOTHING from the rest of this
project (no excel_utils.py, file_mgmt.py, dashboard.py, order_sheet.py,
etc.). Deliberately decoupled per Harish's explicit request: he is
actively iterating on multiple versions of the trading pipeline
(01_Master_Code.py, 02_Master_Code_3Indicator.py, and whatever comes
next) and does not want this reporting tool to break or need rework
every time that code changes underneath it.

WHAT THIS DOES
    Scans F:\05_Claude_Automation (top level only, no subfolders) for
    every "DD-Mon-YY FNO.xlsx"-style backtest output file, pulls each
    day's 'Orders' sheet (every closed trade -- BACKTEST mode never
    leaves a position open overnight, so every row already has an Exit
    Time/Reason), and builds/updates a single consolidated
    "YTD Dashboard.xlsx" in the same folder with:

        Trade Log     -- one row per individual trade, across every
                         file, raw Orders columns preserved as-is, with
                         a prepended Date column. The single source of
                         truth everything else below is computed from.
        Overall       -- YTD KPI card dashboard + a 'MONTHLY P&L' table,
                         one row per calendar month.
        <Mon-YY> ...  -- [CHANGED -- 18-Jul-26] one sheet PER CALENDAR
                         MONTH found in the data (e.g. 'Jul-26'), each
                         with its own KPI card dashboard scoped to that
                         month plus a 'DAILY P&L' table (one row per
                         trading date in that month). Replaces the old
                         single 'Monthly' sheet -- these are rebuilt
                         fresh every run based on whatever months are
                         actually present, so sheets appear/disappear as
                         data does (per Harish's "based on data
                         availability").
        _RunLog       -- hidden. {Date: source file's mtime last
                         processed}. Never shown to Harish, just
                         internal state.

INCREMENTAL REFRESH
    Comparing "have I loaded this date before" is not enough -- Harish
    frequently re-runs the SAME date against different/updated pipeline
    code while iterating, and a naive "skip once ever seen" rule would
    silently freeze the first result forever. Instead, each date's
    _RunLog entry stores the source file's last-modified time: unchanged
    mtime -> skip (fast path, satisfies "don't reload everything every
    run"); new date OR changed mtime (a re-run with fresh numbers) ->
    that date's old Trade Log rows are deleted and replaced with
    freshly-read ones. Dates already logged are never forgotten even if
    the source .xlsx is later deleted from disk (this is meant to be a
    persistent YTD ledger, not a live mirror of the folder's contents).

VISUAL STYLE
    [CHANGED -- 18-Jul-26, Harish's feedback + reference mockup] Redone
    as a card-based KPI dashboard (dark title banner, bordered metric
    cards with a colored top accent bar, standard green=good/red=bad/
    navy=neutral coding) matching the layout Harish provided, plus the
    same blue-header/zebra-striped table convention used elsewhere in
    this project for the data tables underneath. All styling is local to
    this file (not imported), per the "no dependency" requirement.

CAPITAL DEPLOYED
    [ADDED -- 18-Jul-26] Same formula dashboard.py uses for its own
    'Total Capital Deployed (Rs)' KPI: sum(Entry LTP x Quantity (Units)).
    Shown per day in each month sheet's Daily P&L table; the Overall
    sheet's Monthly P&L table shows the MIN and MAX of those daily
    figures within each month. Note: a zero-trade day contributes Rs 0,
    which will pull a month's Min down to 0 if any zero-trade day falls
    in it -- flagged to Harish, not silently hidden.

HOW TO RUN
    python ytd_dashboard.py
    (Or the equivalent `py ytd_dashboard.py` on Harish's Windows setup.)
    Prints a summary of which dates were reprocessed vs. skipped.

Disclaimer: this is purely a reporting/aggregation tool over historical
backtest data already produced elsewhere. It contains no trading/signal
logic of its own, and none of the underlying numbers are a guarantee of
future performance.
"""

import os
import re
import sys
import traceback
from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Config -- same ALGO_BASE_DIR portability convention the rest of the
# project uses (see file_mgmt.py), reproduced here inline rather than
# imported, per the "no dependency on other codes" requirement.
# ---------------------------------------------------------------------------
BASE_DIR = os.environ.get("ALGO_BASE_DIR", r"F:\05_Claude_Automation")
YTD_FILENAME = "YTD Dashboard.xlsx"
YTD_PATH = os.path.join(BASE_DIR, YTD_FILENAME)

# Matches "08-Jul-26 FNO.xlsx" and also tolerates a trailing suffix before
# the extension (e.g. upload-dedup artifacts like "08-Jul-26 FNO-450ae647.xlsx")
# so the same script behaves the same way against either naming.
FNO_FILE_PATTERN = re.compile(r'^(\d{2}-[A-Za-z]{3}-\d{2})\s+FNO.*\.xlsx$', re.IGNORECASE)

ORDERS_SHEET = "Orders"
TRADE_LOG_SHEET = "Trade Log"
OVERALL_SHEET = "Overall"
RUNLOG_SHEET = "_RunLog"
RESERVED_SHEETS = {TRADE_LOG_SHEET, OVERALL_SHEET, RUNLOG_SHEET}
MONTH_SHEET_PATTERN = re.compile(r'^[A-Za-z]{3}-\d{2}$')  # e.g. 'Jul-26' -- used to identify old month sheets to clear

PL_COL = "Net P/L (Rs)"

# ---------------------------------------------------------------------------
# Styling -- card-based KPI dashboard + blue-header zebra tables, matching
# the reference layout Harish provided (18-Jul-26). Local to this file,
# not imported, per the "no dependency" requirement.
# ---------------------------------------------------------------------------
TITLE_BG = "111827"
SUBTITLE_BG = "1F2937"
SECTION_BG = "374151"
TABLE_HEAD_BLUE = "2F5596"
GREEN = "1E9E4C"
RED = "D33B2C"
NAVY = "1F3864"
CARD_BORDER = "E5E7EB"
BAND_LIGHT = "FFFFFF"
BAND_DARK = "F2F2F2"
BORDER_COLOR = "D9D9D9"
LABEL_GRAY = "6B7280"
SUBTITLE_GRAY = "9CA3AF"

FILL_TITLE = PatternFill(start_color=TITLE_BG, end_color=TITLE_BG, fill_type="solid")
FILL_SUBTITLE = PatternFill(start_color=SUBTITLE_BG, end_color=SUBTITLE_BG, fill_type="solid")
FILL_SECTION = PatternFill(start_color=SECTION_BG, end_color=SECTION_BG, fill_type="solid")
FILL_TABLE_HEAD = PatternFill(start_color=TABLE_HEAD_BLUE, end_color=TABLE_HEAD_BLUE, fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

FONT_TITLE = Font(color="FFFFFF", bold=True, size=16)
FONT_SUBTITLE = Font(color=SUBTITLE_GRAY, italic=True, size=9)
FONT_SECTION = Font(color="FFFFFF", bold=True, size=11)
FONT_TABLE_HEAD = Font(color="FFFFFF", bold=True)
FONT_CARD_LABEL = Font(color=LABEL_GRAY, bold=True, size=9)
FONT_BOLD = Font(bold=True)

THIN = Side(style="thin", color=BORDER_COLOR)
CARD_SIDE = Side(style="thin", color=CARD_BORDER)
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LEFT = Alignment(horizontal="left", vertical="center", indent=1)
LEFT_NOINDENT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")

# Card grid geometry
CARD_WIDTH = 3       # columns per card
CARD_GAP = 1          # columns between cards
CARDS_PER_ROW = 4
GRID_START_COL = 2    # column B


def _classify_good_bad(good):
    """True -> GREEN, False -> RED, None -> NAVY (neutral)."""
    if good is True:
        return GREEN
    if good is False:
        return RED
    return NAVY


def _parse_numeric(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace("Rs.", "").replace("Rs", "").replace("₹", "").replace(",", "").replace("%", "").replace("x", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _fmt_rs(value):
    try:
        v = float(value)
    except (TypeError, ValueError):
        return "N/A"
    sign = "-" if v < 0 else ""
    return f"{sign}₹{abs(v):,.2f}"


def autofit_columns(ws, min_width=8, max_width=42, padding=2):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col = cell.column_letter
            length = len(str(cell.value))
            widths[col] = max(widths.get(col, min_width), min(length + padding, max_width))
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# Card-dashboard drawing helpers
# ---------------------------------------------------------------------------
def draw_title_banner(ws, title_text, subtitle_text, last_col):
    last_col_letter = get_column_letter(last_col)
    ws.merge_cells(f"B1:{last_col_letter}1")
    title_cell = ws.cell(row=1, column=2, value=title_text)
    title_cell.font = FONT_TITLE
    title_cell.alignment = LEFT
    for c in range(2, last_col + 1):
        ws.cell(row=1, column=c).fill = FILL_TITLE
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f"B2:{last_col_letter}2")
    subtitle_cell = ws.cell(row=2, column=2, value=subtitle_text)
    subtitle_cell.font = FONT_SUBTITLE
    subtitle_cell.alignment = LEFT
    for c in range(2, last_col + 1):
        ws.cell(row=2, column=c).fill = FILL_SUBTITLE
    ws.row_dimensions[2].height = 18


def draw_section_header(ws, row, start_col, end_col, text):
    cell = ws.cell(row=row, column=start_col, value=text)
    if end_col > start_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    for c in range(start_col, end_col + 1):
        ws.cell(row=row, column=c).fill = FILL_SECTION
    cell.font = FONT_SECTION
    cell.alignment = LEFT
    ws.row_dimensions[row].height = 20


def draw_kpi_card(ws, top_row, left_col, label, value, good):
    """One metric card: thin colored accent bar, gray uppercase label,
    large bold value, light gray border box. good=True/False/None picks
    the standard green/red/navy color coding."""
    accent_hex = _classify_good_bad(good)
    accent_fill = PatternFill(start_color=accent_hex, end_color=accent_hex, fill_type="solid")
    right_col = left_col + CARD_WIDTH - 1

    accent_row, label_row, value_row = top_row, top_row + 1, top_row + 2

    for c in range(left_col, right_col + 1):
        ws.cell(row=accent_row, column=c).fill = accent_fill
    ws.row_dimensions[accent_row].height = 4

    ws.merge_cells(start_row=label_row, start_column=left_col, end_row=label_row, end_column=right_col)
    label_cell = ws.cell(row=label_row, column=left_col, value=str(label).upper())
    label_cell.font = FONT_CARD_LABEL
    label_cell.alignment = LEFT
    ws.row_dimensions[label_row].height = 16

    ws.merge_cells(start_row=value_row, start_column=left_col, end_row=value_row, end_column=right_col)
    value_cell = ws.cell(row=value_row, column=left_col, value=value)
    value_cell.font = Font(color=accent_hex if good is not None else "1F2937", bold=True, size=15)
    value_cell.alignment = LEFT
    ws.row_dimensions[value_row].height = 26

    for r in (label_row, value_row):
        for c in range(left_col, right_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = FILL_WHITE
            top = CARD_SIDE if r == label_row else None
            bottom = CARD_SIDE if r == value_row else None
            leftb = CARD_SIDE if c == left_col else None
            rightb = CARD_SIDE if c == right_col else None
            cell.border = Border(top=top, bottom=bottom, left=leftb, right=rightb)


def draw_kpi_grid(ws, start_row, cards):
    """cards: list of (label, value, good). Lays out CARDS_PER_ROW per
    row, 3 rows tall each (accent/label/value), 1-row gap between card
    rows. Returns the first free row after the grid."""
    r = start_row
    for i, (label, value, good) in enumerate(cards):
        col_in_row = i % CARDS_PER_ROW
        if i > 0 and col_in_row == 0:
            r += 3 + 1  # 3 rows per card + 1 row gap
        left_col = GRID_START_COL + col_in_row * (CARD_WIDTH + CARD_GAP)
        draw_kpi_card(ws, r, left_col, label, value, good)
    return r + 3 + 1  # row after the last card row's value row, plus a blank line


def grid_last_col():
    return GRID_START_COL + (CARDS_PER_ROW * CARD_WIDTH + (CARDS_PER_ROW - 1) * CARD_GAP) - 1


def style_wide_table(ws, header_row, start_col, end_col, pl_col_names=(), pct_col_names=()):
    """Blue header row + zebra-striped body, green/red text on P&L
    columns, number formatting on Rs/percent columns."""
    pl_cols, pct_cols, rs_cols = set(), set(), set()
    for c in range(start_col, end_col + 1):
        header_text = str(ws.cell(row=header_row, column=c).value or "")
        cell = ws.cell(row=header_row, column=c)
        cell.fill = FILL_TABLE_HEAD
        cell.font = FONT_TABLE_HEAD
        cell.alignment = CENTER
        cell.border = CELL_BORDER
        if header_text.strip() in pl_col_names:
            pl_cols.add(c)
        if header_text.strip() in pct_col_names or "(%)" in header_text:
            pct_cols.add(c)
        elif "(Rs)" in header_text:
            rs_cols.add(c)
    ws.row_dimensions[header_row].height = 18

    band_idx = 0
    for r in range(header_row + 1, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(start_col, end_col + 1)]
        if all(v in (None, "") for v in row_vals):
            continue
        band = BAND_LIGHT if band_idx % 2 == 0 else BAND_DARK
        band_fill = PatternFill(start_color=band, end_color=band, fill_type="solid")
        band_idx += 1
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = band_fill
            cell.border = CELL_BORDER
            if c in rs_cols and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
            if c in pct_cols and isinstance(cell.value, (int, float)):
                cell.number_format = '0.0'
            if c in pl_cols:
                numeric_val = _parse_numeric(cell.value)
                cell.font = Font(color=GREEN, bold=True) if (numeric_val is None or numeric_val >= 0) else Font(color=RED, bold=True)
                cell.alignment = RIGHT
            elif isinstance(cell.value, (int, float)):
                cell.font = Font(color="000000")
                cell.alignment = RIGHT
            else:
                cell.font = Font(color="000000")
                cell.alignment = LEFT_NOINDENT if c == start_col else CENTER


# ---------------------------------------------------------------------------
# Step 1: Discover candidate files on disk
# ---------------------------------------------------------------------------
def discover_fno_files(base_dir):
    """Returns {date_str ('DD-Mon-YY'): (full_path, mtime)} for every
    top-level file matching the FNO naming pattern. If more than one file
    matches the same date (shouldn't happen in normal use -- Harish
    deletes before re-running -- but guarded anyway), the most recently
    modified one wins."""
    found = {}
    if not os.path.isdir(base_dir):
        raise RuntimeError(f"Base directory not found: {base_dir}")
    for name in os.listdir(base_dir):
        full_path = os.path.join(base_dir, name)
        if not os.path.isfile(full_path):
            continue
        m = FNO_FILE_PATTERN.match(name)
        if not m:
            continue
        date_str = m.group(1)
        mtime = os.path.getmtime(full_path)
        if date_str not in found or mtime > found[date_str][1]:
            found[date_str] = (full_path, mtime)
    return found


def _parse_date(date_str):
    """'08-Jul-26' -> datetime.date(2026, 7, 8)."""
    return datetime.strptime(date_str, "%d-%b-%y").date()


def _month_sheet_name(d):
    """datetime.date(2026, 7, 8) -> 'Jul-26'."""
    return d.strftime("%b-%y")


# ---------------------------------------------------------------------------
# Step 2: Read one day's Orders sheet
# ---------------------------------------------------------------------------
def read_orders_for_date(file_path, date_str):
    """Returns a DataFrame of that date's trades (raw Orders columns, plus
    a prepended 'Date' column), or an empty DataFrame if the sheet is
    missing/empty (zero-trade day, e.g. no signals fired) -- never raises
    for that reason alone."""
    try:
        df = pd.read_excel(file_path, sheet_name=ORDERS_SHEET)
    except Exception as e:
        print(f"    [INFO] {date_str}: no '{ORDERS_SHEET}' sheet found ({e}) -- zero-trade day.")
        return pd.DataFrame()

    if df.empty:
        return pd.DataFrame()

    df.insert(0, 'Date', _parse_date(date_str))
    return df


# ---------------------------------------------------------------------------
# Step 3: Load existing YTD workbook state (Trade Log + RunLog), if any
# ---------------------------------------------------------------------------
def load_existing_state(ytd_path):
    """Returns (trade_log_df, runlog_dict). Both empty/blank if the YTD
    file doesn't exist yet (first-ever run)."""
    if not os.path.exists(ytd_path):
        return pd.DataFrame(), {}

    try:
        trade_log_df = pd.read_excel(ytd_path, sheet_name=TRADE_LOG_SHEET)
        if 'Date' in trade_log_df.columns:
            trade_log_df['Date'] = pd.to_datetime(trade_log_df['Date']).dt.date
    except Exception:
        trade_log_df = pd.DataFrame()

    runlog = {}
    try:
        runlog_df = pd.read_excel(ytd_path, sheet_name=RUNLOG_SHEET)
        for _, row in runlog_df.iterrows():
            runlog[str(row['Date'])] = float(row['Last Processed Mtime'])
    except Exception:
        runlog = {}

    return trade_log_df, runlog


# ---------------------------------------------------------------------------
# Step 4: Incrementally refresh the Trade Log
# ---------------------------------------------------------------------------
def refresh_trade_log(base_dir, ytd_path):
    """Core incremental logic. Returns (trade_log_df, runlog, processed,
    skipped) -- runlog accumulates every date ever processed (old entries
    are carried forward untouched), so history persists even if a source
    .xlsx is later deleted from disk."""
    found_files = discover_fno_files(base_dir)
    trade_log_df, runlog = load_existing_state(ytd_path)

    processed, skipped = [], []
    new_frames = []

    for date_str, (file_path, mtime) in sorted(found_files.items(), key=lambda kv: _parse_date(kv[0])):
        last_mtime = runlog.get(date_str)
        if last_mtime is not None and abs(last_mtime - mtime) < 1e-6:
            skipped.append(date_str)
            continue

        print(f"  [PROCESS] {date_str}: reading '{ORDERS_SHEET}' from {os.path.basename(file_path)}...")
        day_df = read_orders_for_date(file_path, date_str)

        # Drop any existing rows for this date before re-inserting -- handles
        # both first-time load and a re-run with different results.
        if not trade_log_df.empty and 'Date' in trade_log_df.columns:
            trade_log_df = trade_log_df[trade_log_df['Date'] != _parse_date(date_str)]

        if not day_df.empty:
            new_frames.append(day_df)

        runlog[date_str] = mtime
        processed.append(date_str)

    if new_frames:
        trade_log_df = pd.concat([trade_log_df] + new_frames, ignore_index=True, sort=False)

    return trade_log_df, runlog, processed, skipped


# ---------------------------------------------------------------------------
# Step 5: Aggregation
# ---------------------------------------------------------------------------
def _pf(gross_win, gross_loss):
    if gross_loss > 0:
        return round(gross_win / gross_loss, 2)
    return 'Inf' if gross_win > 0 else ''


def _capital_deployed(trades_df):
    """Same formula as dashboard.py's 'Total Capital Deployed (Rs)':
    sum(Entry LTP x Quantity (Units)) across the given trades."""
    if trades_df.empty:
        return 0.0
    entry_ltp = pd.to_numeric(trades_df.get('Entry LTP', pd.Series(dtype=float)), errors='coerce').fillna(0)
    qty = pd.to_numeric(trades_df.get('Quantity (Units)', pd.Series(dtype=float)), errors='coerce').fillna(0)
    return round((entry_ltp * qty).sum(), 2)


def compute_kpi_stats(trades_df):
    """Reusable KPI block for both the Overall (YTD) sheet and each
    per-month sheet -- same formulas, different scope of trades_df."""
    n = len(trades_df)
    if n == 0:
        return {'total_trades': 0, 'wins': 0, 'losses': 0, 'win_rate': 0, 'net_pl': 0.0,
                'pf': '', 'avg_win': 0.0, 'avg_loss': 0.0, 'max_dd': 0.0}
    pl = pd.to_numeric(trades_df[PL_COL], errors='coerce').fillna(0)
    wins = int((pl > 0).sum())
    losses = int((pl < 0).sum())
    gross_win = pl[pl > 0].sum()
    gross_loss = abs(pl[pl < 0].sum())
    return {
        'total_trades': n, 'wins': wins, 'losses': losses,
        'win_rate': round(wins / n * 100, 1),
        'net_pl': round(pl.sum(), 2),
        'pf': _pf(gross_win, gross_loss),
        'avg_win': round(pl[pl > 0].mean(), 2) if wins else 0.0,
        'avg_loss': round(pl[pl < 0].mean(), 2) if losses else 0.0,
        'max_dd': compute_max_drawdown(trades_df),
    }


def build_card_list(stats, extra_cards=None):
    """Standard 8-card KPI set (Total Trades, Win Rate, Net P/L, Profit
    Factor, Wins/Losses, Avg Profit, Avg Loss, Max Drawdown) with
    standard green=good/red=bad/navy=neutral coding. extra_cards (e.g.
    Best/Worst Month, Overall sheet only) appended after."""
    total_trades = stats['total_trades']
    win_rate = stats['win_rate']
    net_pl = stats['net_pl']
    pf = stats['pf']
    pf_num = _parse_numeric(pf)
    wins, losses = stats['wins'], stats['losses']
    avg_win, avg_loss, max_dd = stats['avg_win'], stats['avg_loss'], stats['max_dd']

    cards = [
        ("Total Trades", total_trades, None),
        ("Overall Win Rate", f"{win_rate}%" if total_trades else "N/A", (win_rate >= 50) if total_trades else None),
        ("Total Net P/L (Rs)", _fmt_rs(net_pl) if total_trades else "N/A", (net_pl >= 0) if total_trades else None),
        ("Profit Factor", f"{pf}x" if pf not in ('', None) else "N/A", (pf_num >= 1) if pf_num is not None else None),
        ("Wins / Losses", f"{wins} / {losses}", None),
        ("Avg Profit per Win (Rs)", _fmt_rs(avg_win), True if wins else None),
        ("Avg Loss per Loss (Rs)", _fmt_rs(avg_loss), False if losses else None),
        ("Max Drawdown (Rs)", _fmt_rs(-abs(max_dd)) if max_dd else _fmt_rs(0), False if max_dd else None),
    ]
    if extra_cards:
        cards.extend(extra_cards)
    return cards


def build_daily_rollup(trade_log_df, dates):
    """One row per trading date in `dates` (used both for a single
    month's Daily P&L table and internally for Min/Max Capital Deployed)."""
    rows = []
    for d in sorted(dates):
        day_trades = trade_log_df[trade_log_df['Date'] == d] if not trade_log_df.empty else pd.DataFrame()
        n = len(day_trades)
        capital_deployed = _capital_deployed(day_trades)
        if n == 0:
            rows.append({
                'Date': d, 'Trades': 0, 'Wins': 0, 'Losses': 0,
                'Win Rate (%)': '', 'Net P/L (Rs)': 0.0,
                'Profit Factor': '', 'Capital Deployed (Rs)': capital_deployed,
                'Best Symbol': '', 'Worst Symbol': '',
            })
            continue
        pl = pd.to_numeric(day_trades[PL_COL], errors='coerce').fillna(0)
        wins = int((pl > 0).sum())
        losses = int((pl < 0).sum())
        gross_win = pl[pl > 0].sum()
        gross_loss = abs(pl[pl < 0].sum())
        best_idx = pl.idxmax()
        worst_idx = pl.idxmin()
        rows.append({
            'Date': d, 'Trades': n, 'Wins': wins, 'Losses': losses,
            'Win Rate (%)': round(wins / n * 100, 1),
            'Net P/L (Rs)': round(pl.sum(), 2),
            'Profit Factor': _pf(gross_win, gross_loss),
            'Capital Deployed (Rs)': capital_deployed,
            'Best Symbol': day_trades.loc[best_idx, 'Symbol'],
            'Worst Symbol': day_trades.loc[worst_idx, 'Symbol'],
        })
    return pd.DataFrame(rows)


def build_monthly_rollup(trade_log_df, all_dates):
    """One row per calendar month -- P&L/win-rate figures recomputed from
    the underlying trades in that month (never by averaging daily
    ratios); Min/Max Capital Deployed computed from that month's own
    daily rollup. Embedded in the 'Overall' sheet."""
    if not all_dates:
        return pd.DataFrame()
    month_of = {d: _month_sheet_name(d) for d in all_dates}
    months_sorted = sorted(set(month_of.values()), key=lambda m: datetime.strptime(m, '%b-%y'))

    rows = []
    for month_label in months_sorted:
        month_dates = [d for d, m in month_of.items() if m == month_label]
        month_trades = trade_log_df[trade_log_df['Date'].isin(month_dates)] if not trade_log_df.empty else pd.DataFrame()
        n = len(month_trades)
        trading_days = len(month_dates)

        month_daily = build_daily_rollup(trade_log_df, month_dates)
        min_cap = round(month_daily['Capital Deployed (Rs)'].min(), 2) if not month_daily.empty else 0.0
        max_cap = round(month_daily['Capital Deployed (Rs)'].max(), 2) if not month_daily.empty else 0.0

        if n == 0:
            rows.append({
                'Month': month_label, 'Trading Days': trading_days, 'Trades': 0,
                'Wins': 0, 'Losses': 0, 'Win Rate (%)': '',
                'Net P/L (Rs)': 0.0, 'Profit Factor': '',
                'Min Capital Deployed (Rs)': min_cap, 'Max Capital Deployed (Rs)': max_cap,
            })
            continue
        pl = pd.to_numeric(month_trades[PL_COL], errors='coerce').fillna(0)
        wins = int((pl > 0).sum())
        losses = int((pl < 0).sum())
        gross_win = pl[pl > 0].sum()
        gross_loss = abs(pl[pl < 0].sum())
        rows.append({
            'Month': month_label, 'Trading Days': trading_days, 'Trades': n,
            'Wins': wins, 'Losses': losses,
            'Win Rate (%)': round(wins / n * 100, 1),
            'Net P/L (Rs)': round(pl.sum(), 2),
            'Profit Factor': _pf(gross_win, gross_loss),
            'Min Capital Deployed (Rs)': min_cap, 'Max Capital Deployed (Rs)': max_cap,
        })
    return pd.DataFrame(rows)


def compute_max_drawdown(trade_log_df):
    """Peak-to-trough decline on the chronological (Date + Exit Time)
    cumulative Net P/L curve across the given trades."""
    if trade_log_df.empty or PL_COL not in trade_log_df.columns:
        return 0.0
    df = trade_log_df.copy()
    exit_time = df['Exit Time'].astype(str) if 'Exit Time' in df.columns else ''
    df['_sort_key'] = pd.to_datetime(df['Date'].astype(str) + ' ' + exit_time, errors='coerce')
    df = df.sort_values('_sort_key')
    cum = pd.to_numeric(df[PL_COL], errors='coerce').fillna(0).cumsum()
    running_max = cum.cummax()
    drawdown = running_max - cum
    return round(drawdown.max(), 2) if not drawdown.empty else 0.0


# ---------------------------------------------------------------------------
# Step 6: Write sheets
# ---------------------------------------------------------------------------
def write_trade_log_sheet(wb, trade_log_df):
    if TRADE_LOG_SHEET in wb.sheetnames:
        del wb[TRADE_LOG_SHEET]
    ws = wb.create_sheet(TRADE_LOG_SHEET)

    if trade_log_df.empty:
        ws['A1'] = "No trades recorded yet."
        ws['A1'].font = FONT_BOLD
        return

    df = trade_log_df.sort_values(['Date'], kind='stable') if 'Date' in trade_log_df.columns else trade_log_df
    cols = list(df.columns)
    ws.append(cols)
    for _, row in df.iterrows():
        ws.append([row[c] for c in cols])

    style_wide_table(ws, header_row=1, start_col=1, end_col=len(cols), pl_col_names={PL_COL})
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    autofit_columns(ws)


def write_month_sheet(wb, month_label, month_trades_df, month_daily_df):
    """One sheet per calendar month (e.g. 'Jul-26') -- KPI card dashboard
    for that month + a 'DAILY P&L' table below."""
    if month_label in wb.sheetnames:
        del wb[month_label]
    ws = wb.create_sheet(month_label)

    stats = compute_kpi_stats(month_trades_df)
    cards = build_card_list(stats)
    last_col = grid_last_col()

    month_full = datetime.strptime(month_label, "%b-%y").strftime("%B %Y")
    draw_title_banner(
        ws,
        "F&O TRADING DASHBOARD",
        f"{month_full}  |  Source: Trade Log tab (filtered to this month)  |  Rebuilt each time you run ytd_dashboard.py",
        last_col,
    )

    row = 4
    next_row = draw_kpi_grid(ws, row, cards)

    draw_section_header(ws, next_row, 2, last_col, "DAILY P&L")
    next_row += 1

    if month_daily_df is None or month_daily_df.empty:
        ws.cell(row=next_row, column=2, value="No trading days recorded yet.")
    else:
        header_row = next_row
        cols = list(month_daily_df.columns)
        for c_idx, col_name in enumerate(cols, start=2):
            ws.cell(row=header_row, column=c_idx, value=col_name)
        r = header_row + 1
        for _, d_row in month_daily_df.iterrows():
            for c_idx, col_name in enumerate(cols, start=2):
                ws.cell(row=r, column=c_idx, value=d_row[col_name])
            r += 1
        style_wide_table(ws, header_row=header_row, start_col=2, end_col=1 + len(cols),
                          pl_col_names={'Net P/L (Rs)'})

    ws.column_dimensions['A'].width = 3
    ws.sheet_view.showGridLines = False
    autofit_columns(ws)


def write_overall_sheet(wb, trade_log_df, monthly_df):
    if OVERALL_SHEET in wb.sheetnames:
        del wb[OVERALL_SHEET]
    ws = wb.create_sheet(OVERALL_SHEET, 0)  # first tab

    ytd_stats = compute_kpi_stats(trade_log_df)

    best_month = worst_month = None
    if monthly_df is not None and not monthly_df.empty:
        traded_months = monthly_df[monthly_df['Trades'] > 0]
        if not traded_months.empty:
            best_row = traded_months.loc[traded_months['Net P/L (Rs)'].idxmax()]
            worst_row = traded_months.loc[traded_months['Net P/L (Rs)'].idxmin()]
            best_month = (best_row['Month'], best_row['Net P/L (Rs)'])
            worst_month = (worst_row['Month'], worst_row['Net P/L (Rs)'])

    extra_cards = []
    if best_month:
        extra_cards.append(("Best Month", f"{best_month[0]} ({_fmt_rs(best_month[1])})", True))
        extra_cards.append(("Worst Month", f"{worst_month[0]} ({_fmt_rs(worst_month[1])})", False))

    cards = build_card_list(ytd_stats, extra_cards=extra_cards)
    last_col = max(grid_last_col(), 1 + (len(monthly_df.columns) if monthly_df is not None and not monthly_df.empty else 0))

    draw_title_banner(
        ws,
        "F&O TRADING DASHBOARD  --  YTD OVERVIEW",
        f"Generated {datetime.now().strftime('%d %b %Y, %H:%M')} IST  |  Source: Trade Log tab (all dates)  |  Rebuilt each time you run ytd_dashboard.py",
        last_col,
    )

    row = 4
    next_row = draw_kpi_grid(ws, row, cards)

    draw_section_header(ws, next_row, 2, last_col, "MONTHLY P&L")
    next_row += 1

    if monthly_df is None or monthly_df.empty:
        ws.cell(row=next_row, column=2, value="No trading months recorded yet.")
    else:
        header_row = next_row
        cols = list(monthly_df.columns)
        for c_idx, col_name in enumerate(cols, start=2):
            ws.cell(row=header_row, column=c_idx, value=col_name)
        r = header_row + 1
        for _, m_row in monthly_df.iterrows():
            for c_idx, col_name in enumerate(cols, start=2):
                ws.cell(row=r, column=c_idx, value=m_row[col_name])
            r += 1
        style_wide_table(ws, header_row=header_row, start_col=2, end_col=1 + len(cols),
                          pl_col_names={'Net P/L (Rs)'})

    ws.column_dimensions['A'].width = 3
    ws.sheet_view.showGridLines = False
    autofit_columns(ws)


def write_runlog_sheet(wb, runlog):
    if RUNLOG_SHEET in wb.sheetnames:
        del wb[RUNLOG_SHEET]
    ws = wb.create_sheet(RUNLOG_SHEET)
    ws.sheet_state = 'hidden'
    ws.append(['Date', 'Last Processed Mtime'])
    for date_str, mtime in sorted(runlog.items(), key=lambda kv: _parse_date(kv[0])):
        ws.append([date_str, mtime])


def clear_old_month_sheets(wb):
    """Removes every sheet that looks like a month sheet (e.g. 'Jul-26')
    so stale months from a prior run's data don't linger, and also
    migrates away from the old single 'Monthly' sheet name if present."""
    for name in list(wb.sheetnames):
        if name in RESERVED_SHEETS:
            continue
        if name == "Monthly" or MONTH_SHEET_PATTERN.match(name):
            del wb[name]


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("=" * 60)
    print("   YTD DASHBOARD -- standalone consolidator")
    print("=" * 60)
    print(f"[SYSTEM] Base directory: {BASE_DIR}")
    print(f"[SYSTEM] Output file:    {YTD_PATH}")

    try:
        trade_log_df, runlog, processed, skipped = refresh_trade_log(BASE_DIR, YTD_PATH)
    except Exception as e:
        print(f"\n[FATAL] Failed to refresh Trade Log: {e}")
        print(traceback.format_exc())
        sys.exit(1)

    print("-" * 60)
    print(f"[SUMMARY] {len(processed)} date(s) (re)processed, {len(skipped)} date(s) unchanged/skipped.")
    if processed:
        print(f"          Processed: {', '.join(processed)}")
    if skipped:
        print(f"          Skipped:   {', '.join(skipped)}")

    all_dates = sorted({_parse_date(ds) for ds in runlog.keys()})
    if not all_dates:
        print("\n[WARNING] No FNO.xlsx files found in the base directory -- nothing to consolidate yet.")

    monthly_df = build_monthly_rollup(trade_log_df, all_dates)

    months_present = sorted(
        {_month_sheet_name(d) for d in all_dates},
        key=lambda m: datetime.strptime(m, '%b-%y'),
    )

    try:
        if os.path.exists(YTD_PATH):
            wb = load_workbook(YTD_PATH)
        else:
            wb = Workbook()
            wb.remove(wb.active)

        clear_old_month_sheets(wb)

        write_trade_log_sheet(wb, trade_log_df)

        for month_label in months_present:
            month_dates = [d for d in all_dates if _month_sheet_name(d) == month_label]
            month_trades_df = trade_log_df[trade_log_df['Date'].isin(month_dates)] if not trade_log_df.empty else pd.DataFrame()
            month_daily_df = build_daily_rollup(trade_log_df, month_dates)
            print(f"  [WRITE] Sheet '{month_label}': {len(month_trades_df)} trade(s) across {len(month_dates)} day(s).")
            write_month_sheet(wb, month_label, month_trades_df, month_daily_df)

        write_overall_sheet(wb, trade_log_df, monthly_df)
        write_runlog_sheet(wb, runlog)

        # Order: Overall first, most-recent month next, ... , Trade Log, hidden _RunLog last.
        desired_order = [OVERALL_SHEET] + list(reversed(months_present)) + [TRADE_LOG_SHEET, RUNLOG_SHEET]
        wb._sheets.sort(key=lambda ws: desired_order.index(ws.title) if ws.title in desired_order else len(desired_order))

        wb.save(YTD_PATH)
        print(f"\n[SUCCESS] {YTD_PATH} updated. Sheets: {', '.join(wb.sheetnames)}")
    except Exception as e:
        print(f"\n[FATAL] Failed to write {YTD_PATH}: {e}")
        print(traceback.format_exc())
        sys.exit(1)

    print("=" * 60)


if __name__ == "__main__":
    main()


# ---------------------------------------------------------------------------
# Disclaimer: this is purely a reporting/aggregation tool over historical
# backtest data already produced elsewhere. It contains no trading/signal
# logic of its own, and none of the underlying numbers are a guarantee of
# future performance.
# ---------------------------------------------------------------------------
