"""
order_sheet.py
---------------
Reads the 'Final Recomm' row that final_sheet.py already wrote to the
'Final' sheet, and turns it into order records. Must run AFTER
final_sheet.run_final_sheet_step() has already saved the 'Final' sheet
into the same workbook.

[REWRITTEN] -- fixes applied, and why
--------------------------------------
Backtests were showing a DIFFERENT result every time they ran for the
SAME historical date, and every trade showed Entry LTP = Current LTP =
Max LTP with P/L = 0. Root causes, all fixed here:

  1. Every gate (spot price, option LTP, VIX, PCR, OI-buildup) called
     kite_api.quote() -- a LIVE market snapshot -- regardless of which
     date was being backtested. A backtest for 06-Jul-26 run on the
     evening of 10-Jul-26 was silently evaluated against 10-Jul-26's
     live market, not against 06-Jul-26 at all. Fixed: mode-aware
     sourcing -- BACKTEST reads point-in-time HISTORICAL data via the
     new historical_lookup.py (cached per token/date, so a rerun of the
     same backtest date never hits the API twice and always resolves to
     the same number); LIVE is unchanged.
  2. No exit-price tracking existed at all. 'Current LTP'/'Max LTP'/'P/L'
     were hardcoded equal to entry values at creation and never updated.
     Fixed: position_manager.simulate_backtest_exit() now walks the
     option's own historical 5-min candles forward from entry to a real
     stop-loss / target / trailing-stop / signal-reversal / max-hold /
     EOD exit, and a real net P/L is computed via
     position_manager.estimate_round_trip_costs().
  3. The expiry-day gate and reversal-exit timestamp used
     datetime.now(IST) even in BACKTEST -- graded against the real
     wall-clock date/time, not the historical date being tested. Fixed:
     both now use target_date / the signal's own historical bar time.
  4. The OI-buildup gate compared against a single cross-run JSON
     snapshot file with no date in its key, so replaying several
     backtest dates for the same recurring contract compared one date's
     OI against a different date's. Fixed: BACKTEST now compares two
     HISTORICAL bars on the same day (historical_lookup.get_historical_
     oi_buildup); that file is LIVE-only now.
  5. Orders/Rejected were read-and-appended-onto on every run, so
     rerunning the same backtest date piled a new (differently-live)
     pass on top of the previous run's leftover rows instead of
     reproducing it. Fixed: BACKTEST always starts from a clean slate.
  6. Position sizing and risk management didn't exist at all -- no
     stop-loss, no target, a fixed lot size regardless of account size
     or volatility. Fixed: position_manager.py adds ATR-based stop/
     target sizing and equity-risk-based position sizing (see gate #9
     below), and a real cost model so Net P/L reflects what you'd
     actually keep, not gross signal accuracy.
  7. The PCR gate rejected on a single absolute snapshot, which is noisy
     for a single stock (one large trade can skew it). Fixed:
     PCRTrendTracker now requires PCR to be both low/high AND trending
     the wrong way before rejecting -- the first few readings of the day
     always pass since there's no trend yet.

Trigger rule (per your instruction -- "check for BUY CE or BUY PE signals
in 03 subsequent columns"):
    Slide a 3-column window across the sorted time columns of a symbol's
    Final Recomm row. The moment three CONSECUTIVE time columns all read
    the same 'BUY CE' (or all 'BUY PE'), that's a confirmed entry signal:
        t1 = Pre-Entry Trigger Time / Status
        t2 = Entry Trigger Time / Status
        t3 = Support Entry Time / Status
    A position already open for that (symbol, t1) is never re-created.
    While a streak is active, the position is closed the moment a later
    column's value stops matching the streak ("5M Reversal").

Risk / liquidity rejection filters, applied in this order before an order
record is created (each rejection is logged to the 'Rejected' sheet with
a reason and resets the streak so a fresh 3-bar confirmation is required
again):
    1. India VIX gate       -- skip if VIX > VIX_MAX (IV too rich for a
                                long-premium entry).
    2. PCR trend gate       -- [CHANGED] skip only if PCR is BOTH beyond
                                its band AND trending the wrong way (see
                                PCRTrendTracker), not a single snapshot.
    3. Entry LTP floor      -- skip if the option's LTP < MIN_ENTRY_LTP
                                (spread too wide / dead premium).
    4. Liquidity gate       -- skip if the option's traded volume is 0.
    5. Low-momentum gate    -- skip if the underlying's ADX (read from the
                                'HTF Bias' sheet's 'ADX Value' row, at the
                                pre-entry bar) is below ADX_MIN.
    6. Expiry-day gate      -- skip if it's the contract's own expiry day
                                and the signal fired within the last 45
                                minutes of the session (theta/gamma risk).
                                [CHANGED for BACKTEST] uses target_date and
                                the signal's own bar time, not real "now".
    7. OI buildup gate      -- skip if the option's OI+price delta since
                                the comparison point contradicts the trade
                                direction. First sighting always passes.
                                [CHANGED for BACKTEST]: sourced from the
                                option's own two historical candles
                                instead of a same-day-only JSON snapshot
                                cache, which would otherwise compare one
                                backtest date's OI against a DIFFERENT
                                date's OI for a recurring contract symbol.
    8. Sector cap gate      -- skip if this would push the symbol's sector
                                beyond MAX_POSITIONS_PER_SECTOR concurrent
                                open positions.
    9. [ADDED] Risk-sizing gate -- skip if the ATR-derived stop distance
                                means even ONE lot's risk exceeds this
                                trade's risk budget (account_equity x
                                RISK_PCT_PER_TRADE). Taking 1 lot anyway
                                would silently blow past the stated
                                per-trade risk limit -- see
                                position_manager.compute_position_size().

Writes two sheets: 'Orders' (one row per tracked position, open or
closed) and 'Rejected' (an audit trail of signals that did NOT become an
order, and why). run_order_sheet_step() also rebuilds the 'Dashboard'
sheet from the real Net P/L this step now produces (see dashboard.py).
"""

import os
import sys
import json
import time
from datetime import datetime
from datetime import time as dtime  # [ADDED -- Task 49] time-of-day constants; 'time' above is the stdlib module

import pandas as pd
import pytz
from openpyxl import load_workbook

CODES_DIR = os.path.dirname(os.path.abspath(__file__))
if CODES_DIR not in sys.path:
    sys.path.append(CODES_DIR)
import excel_utils
import calendar_mgmt
import historical_lookup
import position_manager
import dashboard
import file_mgmt
import sheets_sync
import zerolag
from ist_clock import today_ist

IST = pytz.timezone('Asia/Kolkata')

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
# [CHANGED -- cloud/Colab portability] derives from file_mgmt.BASE_DIR --
# see file_mgmt.py's BASE_DIR docstring.
JSON_DIR = os.path.join(file_mgmt.BASE_DIR, "01_JSON_Files")
OI_SNAPSHOT_CACHE = os.path.join(JSON_DIR, "oi_snapshot_cache.json")   # LIVE mode only, see fix #4 above
# [ADDED -- 13-Jul-26] PCR trend history now persists to disk across
# LIVE cycles for the same reason the OI cache does -- see
# PCRTrendTracker below.
PCR_TREND_CACHE = os.path.join(JSON_DIR, "pcr_trend_cache.json")      # LIVE mode only
# [ADDED -- risk_and_signal_patches audit] Daily drawdown guard state,
# LIVE mode only -- same persistence pattern as PCR_TREND_CACHE, for the
# same reason (build_order_sheet() is re-instantiated once per cycle).
DAILY_DRAWDOWN_CACHE = os.path.join(JSON_DIR, "daily_drawdown_cache.json")

VIX_MAX = 18.0
MIN_ENTRY_LTP = 6.0
ADX_MIN = 18.0
DEFAULT_STRIKE_STEP = 50
MAX_POSITIONS_PER_SECTOR = 2

# [ADDED -- 18-Jul-26, Task 49, Harish's training material] NSE intraday
# volume/liquidity thins out around midday -- 12:00-1:00pm is the
# commonly-cited low-liquidity chop window (wide spreads, low
# conviction, more prone to whipsaw). Off by default -- an untested
# hypothesis like every other experimental gate here; the WINDOW itself
# is a well-known intraday pattern, but whether blocking new entries in
# it actually improves THIS pipeline's results (vs. just cutting trade
# count) still needs its own A/B backtest.
ENABLE_LOW_LIQUIDITY_WINDOW_GATE = False
LOW_LIQUIDITY_WINDOW_START = dtime(12, 0)
LOW_LIQUIDITY_WINDOW_END = dtime(13, 0)

# [ADDED -- Harish's Pine script idea, 17-Jul-26, see zerolag.py] An
# otherwise-confirmed signal must ALSO have the Zero-Lag trend cloud
# agreeing with its direction (price on the correct side of the ZLEMA/ATR
# band right now) AND real volume behind it (RVOL >= ZEROLAG_RVOL_MIN) at
# the pre-entry bar, or it's rejected. This is meant to make the pipeline
# pickier, not looser -- fewer, better-confirmed entries.
#
# [CHANGED -- 18-Jul-26, Task 41 A/B] Temporarily flipped ON to run the
# actual multi-day backtest this flag has needed since it was added. A
# single-day replay against 17-Jul-26 (6 trades, gate would have kept 1)
# showed the gate rejecting purely on RVOL, not direction -- promising on
# that one day but not proof; needs the full 01-15 Jul range to mean
# anything. OFF-baseline for that range already captured before this run:
# 38 trades, Net P/L -Rs 4,719.77, 44.7% win rate (see task 41 notes).
# Flip back to False after this A/B once the ON-range results are in and
# compared, unless the delta clearly favors leaving it on.
ENABLE_ZEROLAG_GATE = True
ZEROLAG_RVOL_MIN = zerolag.RVOL_MIN  # 0.8 as of 18-Jul-26 recalibration -- see zerolag.py

# [ADDED -- 18-Jul-26, Task 41 Q1] The chart's own 'X' cross marker for
# this indicator fires only on the exact bar a trend flips (edge-
# triggered), but zerolag.py's Trend Dir is a PERSISTENT state -- a flip
# from 40 bars ago reads identically to one from the last bar. This adds
# a genuine freshness check on top of the existing direction+RVOL checks:
# the flip backing this signal must be within ZEROLAG_MAX_FLIP_AGE bars
# of the pre-entry bar, i.e. actually mimic what the X marker means
# instead of "cloud agrees, whenever that started being true." Separate
# flag from ENABLE_ZEROLAG_GATE itself so it can be A/B'd independently.
ENABLE_ZEROLAG_FRESHNESS = True
ZEROLAG_MAX_FLIP_AGE = 3  # bars (5min candles) -- 15 minutes since the flip

# [ADDED -- 18-Jul-26, Task 48] Same pattern as ENABLE_ZEROLAG_GATE: an
# otherwise-confirmed signal must also have LuxAlgo SuperTrend AI's
# adaptive trend (see supertrend_ai.py) agreeing with its direction at
# the pre-entry bar, or it's rejected. Off by default -- untested
# hypothesis, needs its own A/B backtest, same as every other
# experimental gate in this file. Deliberately left OFF while the
# still-pending Zero-Lag/PCR re-test (Task 46) is outstanding, so this
# doesn't confound that comparison.
ENABLE_SUPERTREND_GATE = False

# [ADDED -- 18-Jul-26, Task 50] An otherwise-confirmed signal must also
# have a RECENT, volume-confirmed Support/Resistance break agreeing with
# its direction (see support_resistance.py) -- a genuinely different
# dimension from the oscillator-based gates above (price actually
# breaking a real structural level with volume, not just an indicator
# threshold). Off by default -- untested hypothesis, needs its own A/B
# backtest, same as every other experimental gate here.
ENABLE_SR_GATE = False
SR_MAX_BREAK_AGE = 3  # bars (5min candles) -- 15 minutes since the break, same window as ZeroLag's freshness check

# [ADDED -- 18-Jul-26, Task 51] An otherwise-confirmed signal must also
# have a RECENT Smart Money Concepts order-block retest agreeing with
# its direction (see smart_money_concepts.py) -- structural/order-flow
# confirmation, a different dimension again from the gates above. Off by
# default -- first version of a genuinely new signal paradigm for this
# codebase, needs its own A/B backtest before being trusted, same as
# every other experimental gate here.
ENABLE_SMC_GATE = False
SMC_MAX_ZONE_AGE = 3  # bars (5min candles) -- 15 minutes since the retest

# [CHANGED] PCR is now a TREND gate, not a single-value cutoff -- see
# PCRTrendTracker. These bands are deliberately WIDER than the old
# PCR_CE_MIN=0.6 / PCR_PE_MAX=1.3 hard cutoffs, because the level alone no
# longer rejects a trade; level AND trend direction both have to agree.
#
# [CHANGED -- 18-Jul-26, Task 41 Q4] Widened further (CE 0.75->0.65, PE
# 1.15->1.25): PCR was the #2 rejector in the 01-17 Jul rejection-log
# audit (254 of 864 total rejections), largely "Overbought/Oversold Trap"
# band rejections. This is a rough ~13% widening, NOT derived from this
# data the way the RVOL_MIN change was -- treat it as a starting point to
# be judged by this same backtest pass, not a settled number.
# PCR_REQUIRE_SUFFICIENT_DATA / PCR_TREND_MIN_READINGS below are
# DELIBERATELY untouched -- that guard exists because the "insufficient
# data" bucket was the single worst-performing PCR state across two
# earlier independent samples (see that flag's own docstring); loosening
# band width is not the same as removing that guard, and doing the
# latter would knowingly reintroduce an already-fixed loss pattern.
PCR_TREND_BAND_CE = 0.65
PCR_TREND_BAND_PE = 1.25
PCR_TREND_MIN_READINGS = 3

# [ADDED -- 16-Jul-26 audit, 01-15 Jul 26 sample: 42 trades, -Rs 9,742 net,
# 33.3% win rate, PF 0.42] PCR Trend = INSUFFICIENT_DATA was the worst-
# performing bucket by per-trade average in BOTH this sample (9 trades,
# -Rs 6,174, avg -Rs 686) and the earlier 06-13 Jul sample (32/49 trades,
# 93% of that period's total loss) -- the ONE finding that has now held up
# across two independent samples rather than flipping (see the OI gate
# note below for a signal that DID flip). Previously, insufficient PCR
# history made pcr_tracker.evaluate() return passes=True ("gate skipped")
# -- i.e. the gate that exists specifically to screen out crowded/trapped
# setups had no opinion at all on 21% of this sample's trades, and those
# trades were the worst performers. This flag blocks new entries entirely
# until PCR_TREND_MIN_READINGS readings exist for that symbol, rather than
# waving them through. ENABLE_EAGER_PCR_RECORDING (below) already makes
# those readings accrue within the first ~15-20 min a symbol is watched,
# so this mostly delays a symbol's FIRST possible entry each session
# rather than removing it outright. Set False to restore the old
# gate-skipped behavior for an A/B re-run.
PCR_REQUIRE_SUFFICIENT_DATA = True

# [ADDED -- risk_and_signal_patches audit, 13-Jul-26] The 06-13 Jul 26
# sample showed 65% of executed trades (32/49) carried PCR Trend =
# INSUFFICIENT_DATA, and that bucket alone accounted for 93% of total
# losses (-Rs 11,231 of -Rs 12,026) at a 31.2% win rate. Root cause:
# pcr_tracker.record() only fires when a NEW 3-bar signal streak starts
# for a symbol -- most symbols never see 3 separate streak events before
# their entry fires. This flag makes recording ALSO fire on any single-
# bar BUY CE/PE reading (not requiring all 3 bars to agree), which
# happens far more often, so 3 readings accrue earlier in a symbol's
# session. Trade-off: more PCR lookups per symbol per day -- cheap in
# BACKTEST (reads cached historical data), a real rate-limit cost in
# LIVE. Consider disabling in LIVE if Kite rate limits become an issue;
# the persistence fix (PCR_TREND_CACHE, already applied) still helps on
# its own across cycles even with this off.
ENABLE_EAGER_PCR_RECORDING = True

# [REVERTED -- 16-Jul-26, second audit] Was briefly set to informational-
# only (see prior comment, kept below for the full history) on the theory
# that the "good" OI quadrant was too unstable across samples to gate on.
# That change was tested by actually re-running the 01-15 Jul range with
# it live: trade count went 42 -> 61 (+19 new trades let through), win
# rate went 33.3% -> 27.9%, and Net P/L went Rs -9,742 -> Rs -17,132 --
# WORSE, not better. Isolating just the newly-admitted SHORT_COVERING
# trades (16 of them, Rs -6,840, 25% win rate) and removing them still
# only gets back to Rs -10,291 -- still worse than the Rs -9,742 baseline.
# SHORT_COVERING has now looked bad in every sample it's been allowed to
# enter in (it's WHY ALLOW_SHORT_COVERING_CONFIRM was set False on
# historical_lookup.py in the first place) -- unlike LONG_BUILDUP/
# LONG_UNWINDING/SHORT_BUILDUP, which is the part that's genuinely
# flip-flopped between samples. Reverted to hard-reject rather than try to
# hand-tune a SHORT_COVERING-only carve-out from this same one sample --
# that risks overfitting to n=61 the same way a full quadrant rule would
# have. oi_quadrant is still computed and written to 'OI Signal' either
# way, so the data keeps accumulating either way.
ENABLE_OI_BUILDUP_GATE = True

# [ADDED -- 16-Jul-26, Harish's idea] When the PCR trend gate rejects a
# momentum-confluence signal as an "Overbought Trap" (BUY CE, PCR low AND
# still falling) or "Oversold Trap" (BUY PE, PCR high AND still rising),
# that rejection itself IS the classic contrarian/mean-reversion setup:
# the crowd is positioned one way and PCR shows it fading, which argues
# for fading the crowd rather than just standing aside. When this flag is
# True, a signal rejected by the PCR trap gate is not simply dropped --
# the OPPOSITE direction (CE<->PE) is tried instead, re-resolving the
# option contract and re-running every downstream gate (LTP floor,
# volume, ADX, expiry-day, OI-buildup, sector cap, risk-sizing) against
# THAT contract's own data (see _try_contrarian_flip()). This is a brand
# new, UNTESTED hypothesis -- n=0 backtest evidence at the time this flag
# was added. Default OFF. Must be A/B backtested (same date range, flag
# on vs off) before ever being trusted, same discipline as every other
# gate change in this file's audit trail. Rows created this way are
# tagged 'Entry Type' = 'Contrarian Flip' so they're separately
# measurable and never silently blended into ordinary confluence-trade
# stats.
ENABLE_PCR_CONTRARIAN_FLIP = False

ORDER_HEADERS = [
    'Symbol', 'PCR', 'PCR Trend', 'OI Signal', 'Entry Type',
    'Pre-Entry Trigger Time', 'Pre-Entry Trigger Status',
    'Entry Trigger Time', 'Entry Trigger Status',
    'Support Entry Time', 'Support Trigger Status',
    'Exit Trigger Time', 'Exit Trigger Status',
    'Spot Price', 'ATM Strike', 'Option Symbol', 'Option Token', 'Lot Size',
    'ATR (Underlying)', 'Entry LTP', 'Stop Loss LTP', 'Target LTP', 'Risk/Unit (Rs)',
    'Quantity (Lots)', 'Quantity (Units)', 'Risk Amount (Rs)',
    'Current LTP', 'Max LTP', 'Min LTP',
    # [ADDED -- ENABLE_TSL_CONFIRMATION_HOLD] Persists position_manager.
    # check_live_exit()'s confirmation counter across LIVE polling cycles
    # for this position -- see that function's docstring. Always "" (read
    # back as 0) when the flag is off; BACKTEST doesn't need this column
    # at all since simulate_backtest_exit() tracks its own streak
    # internally within one call.
    'TSL Breach Streak',
    'Gross P/L (Rs)', 'Costs (Rs)', 'Net P/L (Rs)',
    'Order ID', 'Exit Time', 'Exit Reason',
]
REJECTED_HEADERS = ['Symbol', 'Trigger Time', 'Signal', 'Reason', 'Timestamp']


# ---------------------------------------------------------------------------
# PCR trend tracking -- see module docstring, gate #2
# ---------------------------------------------------------------------------
class PCRTrendTracker:
    """Accumulates PCR readings per symbol so the PCR gate acts on a TREND
    rather than a single absolute snapshot. Call .record() for EVERY
    pcr_val computed (even ones that end up rejected by an earlier gate),
    then .evaluate() when the PCR gate itself runs.

    [FIX -- 13-Jul-26] PERSISTENCE BUG. build_order_sheet() previously did
    `pcr_tracker = PCRTrendTracker()` fresh on every call -- and in LIVE
    mode it's called once per candle-close cycle (see 01_Master_Code.
    run_cycle()). Since `.history` was a plain in-memory dict, it was
    thrown away and recreated empty on every single cycle. A symbol only
    gets `.record()` called when a NEW 3-bar signal streak starts for it
    (see build_order_sheet's main loop), which is usually at most once
    per cycle per symbol -- so `len(hist)` almost never reached
    PCR_TREND_MIN_READINGS (3) within one call, and the trend gate fell
    back to "insufficient PCR history -- gate skipped" -> PCR Trend =
    INSUFFICIENT_DATA essentially FOREVER in LIVE, not just at session
    start. This is exactly what the screenshot showed. The gate wasn't
    broken logically, it just never got the chance to accumulate history
    across cycles the way its own docstring says it should.

    Fix: when `cache_path` is given (LIVE only -- BACKTEST still starts
    from a clean slate every run, per fix #5 in the module docstring),
    history is loaded from and saved back to a small JSON file, the same
    pattern get_oi_buildup_signal() already uses for the OI cache. The
    cache is keyed by trading DATE (today_ist()) so a new session always
    starts cold instead of picking up yesterday's PCR readings for a
    symbol -- the same date-scoping fix applied to the OI cache below.
    """

    def __init__(self, min_readings=PCR_TREND_MIN_READINGS,
                 band_ce=PCR_TREND_BAND_CE, band_pe=PCR_TREND_BAND_PE,
                 cache_path=None):
        self.min_readings = min_readings
        self.band_ce = band_ce
        self.band_pe = band_pe
        self.cache_path = cache_path
        self._date_key = today_ist().strftime('%Y-%m-%d')
        self.history = {}   # symbol -> [pcr, pcr, ...] in time order
        if cache_path:
            cache_all = _load_json(cache_path, {})
            # Only adopt today's entries -- anything under a different
            # date key is a previous session's leftover and must NOT be
            # blended into today's trend.
            self.history = dict(cache_all.get(self._date_key, {}))

    def _persist(self):
        if not self.cache_path:
            return
        # Save ONLY today's date key -- this both resets the cache for a
        # new day automatically (no separate purge step to remember) and
        # keeps the file from growing across sessions indefinitely.
        _save_json(self.cache_path, {self._date_key: self.history})

    def record(self, symbol, pcr):
        if pcr is not None:
            hist = self.history.setdefault(symbol, [])
            hist.append(pcr)
            # Cap so a stuck/misbehaving cycle loop can't grow this
            # unboundedly across a session; only the most recent
            # min_readings-worth is ever actually used by evaluate().
            del hist[:-50]
            self._persist()

    def evaluate(self, symbol, signal, current_pcr):
        """Returns (passes: bool, reason_str, trend_label)."""
        hist = self.history.get(symbol, [])
        if len(hist) < self.min_readings or current_pcr is None:
            return True, "insufficient PCR history -- gate skipped", "INSUFFICIENT_DATA"

        recent = hist[-self.min_readings:]
        if recent[-1] > recent[0]:
            trend = "RISING"
        elif recent[-1] < recent[0]:
            trend = "FALLING"
        else:
            trend = "FLAT"

        if signal == "BUY CE":
            # Reject only when PCR is BOTH low AND still falling -- a
            # low-but-RISING PCR is recovering FROM crowded-call
            # territory, which argues FOR a CE, not against it.
            if current_pcr < self.band_ce and trend == "FALLING":
                return False, f"Overbought Trap (PCR {current_pcr:.2f} < {self.band_ce}, still FALLING).", trend
            return True, "PCR trend does not confirm an overbought trap", trend

        if signal == "BUY PE":
            # Symmetric: reject only when PCR is BOTH high AND still
            # rising -- a high-but-FALLING PCR is unwinding FROM an
            # oversold trap, arguing FOR a PE, not against it.
            if current_pcr > self.band_pe and trend == "RISING":
                return False, f"Oversold Trap (PCR {current_pcr:.2f} > {self.band_pe}, still RISING).", trend
            return True, "PCR trend does not confirm an oversold trap", trend

        return True, "signal is not a directional entry", trend


# ---------------------------------------------------------------------------
# Small local helpers
# ---------------------------------------------------------------------------
def _load_json(path, default):
    try:
        with open(path, 'r') as f:
            return json.load(f)
    except Exception:
        return default


def _save_json(path, data):
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, 'w') as f:
            json.dump(data, f)
    except Exception as e:
        print(f"[WARNING] Failed to persist {path}: {e}")


def _match_symbol(df_ref, base_symbol):
    """Case/whitespace-insensitive lookup row for base_symbol in
    df_ref['Symbol / StrikePrice'], or None if not found."""
    normalized = df_ref['Symbol / StrikePrice'].astype(str).str.strip().str.upper()
    match = df_ref[normalized == base_symbol]
    return match if not match.empty else None


def _build_kite_master(kite_api):
    """NSE+NFO instrument master. Kept local (rather than importing
    token_mgmt.py) since this file doesn't need the Angel One master that
    token_mgmt.py also builds.

    [ADDED -- 14-Jul-26 resilience patch] This call runs ONCE, before any
    symbol is processed, and was previously unguarded -- a Kite access
    token expiring mid-session, a rate-limit hit, or a network blip here
    used to crash build_order_sheet() immediately, before a single symbol
    (including one with a fully-confirmed 3-bar streak) was even looked
    at. Wrapping it doesn't prevent the failure, but it turns a bare,
    hard-to-diagnose traceback into a clear, actionable message the next
    time it happens."""
    try:
        kite_nse = pd.DataFrame(kite_api.instruments(exchange=kite_api.EXCHANGE_NSE))
        kite_nfo = pd.DataFrame(kite_api.instruments(exchange=kite_api.EXCHANGE_NFO))
        return pd.concat([kite_nse, kite_nfo], ignore_index=True)
    except Exception as e:
        raise RuntimeError(
            f"Kite instrument master fetch failed -- check access token/session validity "
            f"and Kite API rate limits before retrying: {e}"
        ) from e


def _load_final_recomm_table(output_excel_path):
    """{symbol: {time_str: value}} from the 'Final' sheet's 'Final Recomm'
    row, plus the sorted list of time columns."""
    try:
        df = pd.read_excel(output_excel_path, sheet_name='Final')
    except Exception as e:
        raise RuntimeError(
            f"Order Sheet: could not read the 'Final' sheet ({e}) -- "
            "run final_sheet.run_final_sheet_step() first."
        )

    recomm_df = df[df['Metrics'] == 'Final Recomm']
    time_cols = [c for c in df.columns if c not in ('Symbol', 'Metrics')]

    table = {}
    for _, row in recomm_df.iterrows():
        sym = str(row['Symbol']).strip().upper()
        table[sym] = {
            str(t): row[t] for t in time_cols
            if pd.notna(row[t]) and str(row[t]).strip() != ""
        }
    return table, sorted(str(t) for t in time_cols)


def _load_metric_lookup(output_excel_path, sheet_name, metric_label):
    """{symbol: {time_str: value}} from any '<sheet>'/'Metrics'==metric_label
    row -- shared by the ADX lookup (gate #5) and the ATR lookup (position
    sizing). Returns {} (gate/sizing skipped) if the sheet or row isn't
    present."""
    try:
        df = pd.read_excel(output_excel_path, sheet_name=sheet_name)
    except Exception:
        print(f"[WARNING] Order Sheet: '{sheet_name}' sheet not found -- '{metric_label}' lookup will be skipped.")
        return {}

    metric_df = df[df['Metrics'] == metric_label]
    time_cols = [c for c in df.columns if c not in ('Symbol', 'Metrics')]
    table = {}
    for _, row in metric_df.iterrows():
        sym = str(row['Symbol']).strip().upper()
        table[sym] = {str(t): row[t] for t in time_cols if pd.notna(row[t])}
    return table


# ---------------------------------------------------------------------------
# LIVE-ONLY: option chain / risk-context lookups via live broker quotes.
# ---------------------------------------------------------------------------
def resolve_option_chain(base_symbol, spot_price, signal, df_ref, kite_master):
    """Returns (tradingsymbol, instrument_token, lot_size, strike), or
    (None, None, 1, strike) if no matching contract is found. Pure
    lookup against the instrument master -- no live quote here, so this
    is safe to call in both modes."""
    diff_val = DEFAULT_STRIKE_STEP
    if 'Option Price Difference' in df_ref.columns:
        match = _match_symbol(df_ref, base_symbol)
        if match is not None:
            val = match['Option Price Difference'].values[0]
            if pd.notna(val) and val != 0:
                diff_val = val

    strike = round(spot_price / diff_val) * diff_val
    opt_type = "CE" if "CE" in signal else "PE"

    matches = kite_master[
        (kite_master['name'] == base_symbol)
        & (kite_master['instrument_type'] == opt_type)
        & (kite_master['strike'] == strike)
        & (kite_master['segment'] == 'NFO-OPT')
    ]
    if matches.empty:
        return None, None, 1, strike

    target = matches.sort_values('expiry').iloc[0]
    return target['tradingsymbol'], target['instrument_token'], target['lot_size'], strike


def calculate_local_pcr(base_symbol, spot_price, df_ref, kite_master, kite_api):
    """LIVE ONLY. Put-Call OI ratio across the ATM +/-5 strikes of the
    nearest expiry, from a live quote() snapshot."""
    diff_val = DEFAULT_STRIKE_STEP
    if 'Option Price Difference' in df_ref.columns:
        match = _match_symbol(df_ref, base_symbol)
        if match is not None:
            val = match['Option Price Difference'].values[0]
            if pd.notna(val) and val != 0:
                diff_val = val

    atm_strike = round(spot_price / diff_val) * diff_val
    matches = kite_master[(kite_master['name'] == base_symbol) & (kite_master['segment'] == 'NFO-OPT')]
    if matches.empty:
        return None

    current_expiry = matches.sort_values('expiry').iloc[0]['expiry']
    strikes = [atm_strike + (i * diff_val) for i in range(-5, 6)]
    chain = matches[(matches['expiry'] == current_expiry) & (matches['strike'].isin(strikes))]
    if chain.empty:
        return None

    queries = [f"NFO:{ts}" for ts in chain['tradingsymbol']]
    try:
        quotes = kite_api.quote(queries)
        put_oi = sum(d.get('oi', 0) for ts, d in quotes.items() if ts.endswith('PE'))
        call_oi = sum(d.get('oi', 0) for ts, d in quotes.items() if ts.endswith('CE'))
        return 2.0 if call_oi == 0 else put_oi / call_oi
    except Exception as e:
        print(f"[WARNING] PCR quote fetch failed for {base_symbol}: {e}")
        return None


def get_india_vix(kite_api):
    """LIVE ONLY."""
    try:
        quote_data = kite_api.quote(["NSE:INDIA VIX"])
        return quote_data.get("NSE:INDIA VIX", {}).get('last_price', None)
    except Exception as e:
        print(f"[WARNING] India VIX fetch failed: {e}")
        return None


def get_expiry_context(base_symbol, df_ref, as_of_date):
    """Returns (is_expiry_day, is_expiry_week, days_to_expiry) using the
    'Expiry Date' column in df_ref, or (False, False, None) if that
    column is absent / unparseable for this symbol. as_of_date is an
    explicit argument (target_date.date() for both LIVE and BACKTEST
    callers) instead of datetime.now(IST).date() -- previously this
    always used the real wall-clock date, which meant a backtest for a
    past date was silently graded against TODAY's days-to-expiry."""
    if 'Expiry Date' not in df_ref.columns:
        return False, False, None
    match = _match_symbol(df_ref, base_symbol)
    if match is None:
        return False, False, None
    try:
        expiry_dt = pd.to_datetime(match['Expiry Date'].values[0]).date()
    except Exception:
        return False, False, None

    days_to_expiry = (expiry_dt - as_of_date).days
    return (days_to_expiry == 0), (0 <= days_to_expiry <= 4), days_to_expiry


def get_oi_buildup_signal(opt_symbol, current_oi, current_price, signal, cache_path=OI_SNAPSHOT_CACHE):
    """LIVE ONLY. Compares this poll's OI+price for opt_symbol against
    the last cached snapshot to classify the OI buildup quadrant. First
    sighting of a contract always confirms. This intentionally persists
    across runs -- correct for LIVE's poll-to-poll comparison during a
    single trading session. BACKTEST never calls this; see
    historical_lookup.get_historical_oi_buildup(), which compares two
    historical bars instead and never touches this file.

    [FIX -- 13-Jul-26] DATE-SCOPING. The cache used to be a flat
    {opt_symbol: {...}} dict with no date in it at all, so it was never
    cleared between sessions. That's the opposite-direction bug from the
    PCR one above: instead of always showing INSUFFICIENT_DATA, a
    contract polled on a PREVIOUS day (or an earlier, stale run earlier
    today) would silently compare today's first real poll against
    yesterday's leftover OI/price -- producing a confident-looking
    LONG_BUILDUP / SHORT_COVERING / etc. quadrant that has nothing to do
    with today's session. The fix mirrors PCRTrendTracker: nest the cache
    under today's date key and only ever save that key back out, so a
    new trading day (or a fresh weekly-expiry contract, as in the
    AXISBANK example) always starts from a true INSUFFICIENT_DATA first
    reading, and old dates are dropped rather than accumulated forever.
    """
    date_key = today_ist().strftime('%Y-%m-%d')
    cache_all = _load_json(cache_path, {})
    cache_today = dict(cache_all.get(date_key, {}))
    prev = cache_today.get(opt_symbol)
    cache_today[opt_symbol] = {'oi': current_oi, 'price': current_price, 'ts': time.time()}
    _save_json(cache_path, {date_key: cache_today})

    if prev is None:
        return "INSUFFICIENT_DATA", True

    oi_delta = current_oi - prev.get('oi', current_oi)
    price_delta = current_price - prev.get('price', current_price)

    if price_delta >= 0 and oi_delta > 0:
        quadrant = "LONG_BUILDUP"
    elif price_delta >= 0 and oi_delta <= 0:
        quadrant = "SHORT_COVERING"
    elif price_delta < 0 and oi_delta > 0:
        quadrant = "SHORT_BUILDUP"
    else:
        quadrant = "LONG_UNWINDING"

    # [FIX -- risk_and_signal_patches audit] see historical_lookup.
    # get_historical_oi_buildup()'s matching fix for the full rationale
    # and backtest evidence; kept identical here for LIVE/BACKTEST
    # consistency. ALLOW_SHORT_COVERING_CONFIRM lives on historical_lookup
    # (imported below) so both code paths share one toggle.
    bullish_quadrants = {"LONG_BUILDUP"}
    if historical_lookup.ALLOW_SHORT_COVERING_CONFIRM:
        bullish_quadrants.add("SHORT_COVERING")
    bearish_quadrants = {"SHORT_BUILDUP", "LONG_UNWINDING"}

    if signal == "BUY CE":
        confirms = quadrant in bullish_quadrants
    elif signal == "BUY PE":
        confirms = quadrant in bearish_quadrants
    else:
        confirms = True
    return quadrant, confirms


# ---------------------------------------------------------------------------
# Core: 3-column streak detection -> order creation / 5M-reversal exit
# ---------------------------------------------------------------------------
def build_order_sheet(output_excel_path, kite_api, df_ref, mode=calendar_mgmt.LIVE, target_date=None,
                       account_equity=position_manager.ACCOUNT_EQUITY_DEFAULT,
                       max_positions_per_sector=MAX_POSITIONS_PER_SECTOR,
                       oi_cache_path=OI_SNAPSHOT_CACHE):
    is_backtest = (mode == calendar_mgmt.BACKTEST)
    if is_backtest and target_date is None:
        raise ValueError("build_order_sheet: target_date is required when mode=calendar_mgmt.BACKTEST")

    final_table, sorted_times = _load_final_recomm_table(output_excel_path)
    adx_lookup = _load_metric_lookup(output_excel_path, 'HTF Bias', 'ADX Value')
    atr_lookup = _load_metric_lookup(output_excel_path, 'HTF Bias', 'ATR Value')
    # [ADDED -- ENABLE_ZEROLAG_GATE] {symbol: {time_str: value}} straight
    # off the 'ZLTREND' sheet zerolag.py already wrote -- same
    # _load_metric_lookup() helper as the ADX/ATR lookups above, just a
    # different source sheet/row.
    zl_trend_lookup = _load_metric_lookup(output_excel_path, 'ZLTREND', 'Trend Dir')
    zl_rvol_lookup = _load_metric_lookup(output_excel_path, 'ZLTREND', 'RVOL')
    # [ADDED -- ENABLE_ZEROLAG_FRESHNESS] {symbol: {time_str: bars_since_flip}}
    # -- see zerolag.py's 'Flip Age' row and this flag's own docstring above.
    zl_flip_age_lookup = _load_metric_lookup(output_excel_path, 'ZLTREND', 'Flip Age')
    # [ADDED -- ENABLE_SUPERTREND_GATE] {symbol: {time_str: 'BUY CE'/'BUY PE'/'WAIT'}}
    # straight off the 'Supertrend' sheet supertrend_ai.py writes.
    supertrend_lookup = _load_metric_lookup(output_excel_path, 'Supertrend', 'Supertrend Recomm')
    # [ADDED -- ENABLE_SR_GATE] {symbol: {time_str: value}} lookups off
    # the 'SUPRES' sheet support_resistance.py writes. 'Last Break Dir'
    # (persistent, like zerolag's Trend Dir) is used rather than the
    # edge-triggered 'SR Recomm', since a pre-entry bar rarely lands
    # exactly on the break bar itself -- see that column's own docstring.
    sr_break_dir_lookup = _load_metric_lookup(output_excel_path, 'SUPRES', 'Last Break Dir')
    sr_break_age_lookup = _load_metric_lookup(output_excel_path, 'SUPRES', 'SR Break Age')
    # [ADDED -- ENABLE_SMC_GATE] {symbol: {time_str: value}} lookups off
    # the 'SMC' sheet smart_money_concepts.py writes.
    smc_dir_lookup = _load_metric_lookup(output_excel_path, 'SMC', 'Last OB Dir')
    smc_age_lookup = _load_metric_lookup(output_excel_path, 'SMC', 'SMC Zone Age')
    kite_master = _build_kite_master(kite_api)
    # [FIX -- 13-Jul-26] cache_path=None for BACKTEST keeps the existing,
    # correct "clean slate every run" behavior (fix #5 in the module
    # docstring); LIVE now persists PCR history across cycles via
    # PCR_TREND_CACHE -- see PCRTrendTracker's docstring for why this was
    # the actual cause of PCR Trend showing INSUFFICIENT_DATA constantly.
    pcr_tracker = PCRTrendTracker(cache_path=None if is_backtest else PCR_TREND_CACHE)
    hist_cache = historical_lookup.HistoricalCache(kite_api) if is_backtest else None

    # [ADDED -- risk_and_signal_patches audit] Daily drawdown circuit
    # breaker. LIVE: persists across cycles via DAILY_DRAWDOWN_CACHE, so
    # .breached() genuinely blocks new entries mid-session once crossed
    # -- same pattern as pcr_tracker above. BACKTEST: Net P/L doesn't
    # exist yet during this loop (Phase 1 decides entries before Phase 2
    # resolves any exit), so .breached() will never trip here regardless
    # of how the day resolves -- see DailyDrawdownGuard's docstring in
    # position_manager.py. The Phase 2.5 block after Phase 2 reports what
    # the guard WOULD have done, retroactively, without pretending it
    # blocked anything it didn't.
    dd_guard = position_manager.DailyDrawdownGuard(
        max_daily_loss_rs=position_manager.DAILY_MAX_LOSS_RS,
        # [FIXED -- 15-Jul-26] DAILY_DRAWDOWN_CACHE is defined locally in this
        # module (see line ~149), NOT on position_manager -- position_manager
        # only exports DAILY_MAX_LOSS_RS. The old `position_manager.DAILY_
        # DRAWDOWN_CACHE` reference raised AttributeError and crashed
        # build_order_sheet() entirely, every single cycle.
        cache_path=None if is_backtest else DAILY_DRAWDOWN_CACHE,
        date_key=None if is_backtest else target_date.strftime('%Y-%m-%d') if target_date else None,
    )

    # BACKTEST: start from a clean slate every run -- previously this
    # always read the existing Orders/Rejected sheets and merged onto
    # them, which is correct for LIVE (resuming an intraday session) but
    # meant a second run of the same backtest date piled its results on
    # top of the first run's instead of reproducing them.
    if is_backtest:
        df_orders = pd.DataFrame(columns=ORDER_HEADERS)
        rejected_rows = []
    else:
        try:
            df_orders = pd.read_excel(output_excel_path, sheet_name='Orders')
        except Exception:
            df_orders = pd.DataFrame(columns=ORDER_HEADERS)
        for col in ORDER_HEADERS:
            if col not in df_orders.columns:
                df_orders[col] = ""
        try:
            df_rejected_existing = pd.read_excel(output_excel_path, sheet_name='Rejected')
            rejected_rows = df_rejected_existing.to_dict('records')
        except Exception:
            rejected_rows = []

    def _timestamp_str(bar_time_str=None):
        if is_backtest and bar_time_str:
            return f"{target_date.strftime('%Y-%m-%d')} {bar_time_str}:00"
        return datetime.now(IST).strftime('%H:%M:%S')

    def _log_rejection(sym, t1, s1, reason):
        rejected_rows.append({
            'Symbol': sym, 'Trigger Time': t1, 'Signal': s1, 'Reason': reason,
            'Timestamp': _timestamp_str(t1),
        })

    existing_orders = {}
    for _, r in df_orders.iterrows():
        sym_val = r.get('Symbol')
        t1_val = r.get('Pre-Entry Trigger Time')
        if pd.notna(sym_val) and str(sym_val).strip() != "" and pd.notna(t1_val):
            existing_orders[f"{sym_val}_{t1_val}"] = r.to_dict()

    sector_map = {}
    if 'Sector' in df_ref.columns:
        for _, ref_row in df_ref.iterrows():
            sector_map[str(ref_row['Symbol / StrikePrice']).strip().upper()] = str(ref_row.get('Sector', 'Unknown')).strip()

    def count_open_positions_in_sector(sector_name):
        count = 0
        for o in existing_orders.values():
            o_sym = str(o.get('Symbol', '')).strip().upper()
            o_sector = sector_map.get(o_sym, 'Unknown')
            exit_val = o.get('Exit Time', "")
            is_open = pd.isna(exit_val) or str(exit_val).strip() == ""
            if is_open and o_sector == sector_name:
                count += 1
        return count

    def has_open_position_in_symbol(sym_name):
        """[ADDED] The existing_orders dedup key is (symbol, t1), so it
        only blocks re-creating the SAME streak's order twice -- it does
        NOT stop a second, independent position opening in the same
        symbol once current_signal_streak resets and a fresh t1 starts a
        new streak. That gap was low-risk while the 5M Reversal exit
        closed positions almost immediately (this symbol was usually flat
        again within a bar or two). With ENABLE_SIGNAL_REVERSAL_EXIT=False
        a position can now stay open for a while (until stop/target/
        trailing/max-hold/EOD resolves it), so this explicit same-symbol
        check is now required to prevent stacking a duplicate position on
        top of one that's already open."""
        for o in existing_orders.values():
            o_sym = str(o.get('Symbol', '')).strip().upper()
            exit_val = o.get('Exit Time', "")
            is_open = pd.isna(exit_val) or str(exit_val).strip() == ""
            if is_open and o_sym == sym_name:
                return True
        return False

    def _zerolag_gate_check(sym, signal, t1):
        """[ADDED -- ENABLE_ZEROLAG_GATE] (ok, detail) for one (sym, t1,
        signal) -- shared by both the main Phase-1 loop and
        _try_contrarian_flip() so the two paths can't drift apart. No-op
        (always ok) when the flag is off, matching every other gate's
        convention in this file.

        [CHANGED -- ENABLE_ZEROLAG_FRESHNESS] Added a third check on top
        of direction+RVOL: the flip backing this trend agreement must be
        recent (see zerolag.py's 'Flip Age' / bars_since_flip). Without
        this, 'cloud trend agrees' can be true because of a flip from an
        hour ago -- not the same thing as the chart's own 'X' cross
        marker, which only fires on the actual flip bar. Independently
        toggleable from ENABLE_ZEROLAG_GATE itself."""
        if not ENABLE_ZEROLAG_GATE:
            return True, ""

        zl_dir = zl_trend_lookup.get(sym, {}).get(t1)
        zl_rvol = zl_rvol_lookup.get(sym, {}).get(t1)
        zl_flip_age = zl_flip_age_lookup.get(sym, {}).get(t1)
        try:
            zl_dir = int(float(zl_dir)) if zl_dir is not None else None
        except (TypeError, ValueError):
            zl_dir = None
        try:
            zl_rvol = float(zl_rvol) if zl_rvol is not None else None
        except (TypeError, ValueError):
            zl_rvol = None
        try:
            zl_flip_age = int(float(zl_flip_age)) if zl_flip_age is not None else None
        except (TypeError, ValueError):
            zl_flip_age = None

        expected_dir = 1 if signal == 'BUY CE' else -1 if signal == 'BUY PE' else None
        dir_ok = expected_dir is not None and zl_dir == expected_dir
        rvol_ok = zl_rvol is not None and zl_rvol >= ZEROLAG_RVOL_MIN
        fresh_ok = (not ENABLE_ZEROLAG_FRESHNESS) or (
            zl_flip_age is not None and zl_flip_age <= ZEROLAG_MAX_FLIP_AGE
        )

        if dir_ok and rvol_ok and fresh_ok:
            return True, ""

        zl_dir_label = {1: 'BUY CE', -1: 'BUY PE', 0: 'WAIT', None: 'N/A'}.get(zl_dir, 'N/A')
        rvol_label = f"{zl_rvol:.2f}" if zl_rvol is not None else "N/A"
        age_label = str(zl_flip_age) if zl_flip_age is not None else "N/A"
        return False, (f"Zero-Lag gate failed (cloud trend {zl_dir_label}, RVOL {rvol_label} "
                        f"< {ZEROLAG_RVOL_MIN}, flip age {age_label} bars "
                        f"{'> ' + str(ZEROLAG_MAX_FLIP_AGE) if ENABLE_ZEROLAG_FRESHNESS else '(freshness off)'}). "
                        f"Cloud/volume/freshness doesn't confirm this move.")

    def _supertrend_gate_check(sym, signal, t1):
        """[ADDED -- ENABLE_SUPERTREND_GATE, Task 48] (ok, detail) for one
        (sym, t1, signal) -- same shared-helper pattern as
        _zerolag_gate_check(). No-op (always ok) when the flag is off."""
        if not ENABLE_SUPERTREND_GATE:
            return True, ""

        st_recomm = supertrend_lookup.get(sym, {}).get(t1)
        if st_recomm == signal:
            return True, ""

        st_label = st_recomm if st_recomm else "N/A"
        return False, (f"SuperTrend AI gate failed (adaptive trend reads {st_label}, "
                        f"signal wants {signal}). Trend disagreement.")

    def _sr_gate_check(sym, signal, t1):
        """[ADDED -- ENABLE_SR_GATE, Task 50] (ok, detail) for one
        (sym, t1, signal) -- same shared-helper pattern as
        _zerolag_gate_check(). No-op (always ok) when the flag is off."""
        if not ENABLE_SR_GATE:
            return True, ""

        sr_dir = sr_break_dir_lookup.get(sym, {}).get(t1)
        sr_age = sr_break_age_lookup.get(sym, {}).get(t1)
        try:
            sr_dir = int(float(sr_dir)) if sr_dir is not None else None
        except (TypeError, ValueError):
            sr_dir = None
        try:
            sr_age = int(float(sr_age)) if sr_age is not None else None
        except (TypeError, ValueError):
            sr_age = None

        expected_dir = 1 if signal == 'BUY CE' else -1 if signal == 'BUY PE' else None
        dir_ok = expected_dir is not None and sr_dir == expected_dir
        fresh_ok = sr_age is not None and 0 <= sr_age <= SR_MAX_BREAK_AGE

        if dir_ok and fresh_ok:
            return True, ""

        sr_dir_label = {1: 'Resistance Break (bullish)', -1: 'Support Break (bearish)', 0: 'None yet', None: 'N/A'}.get(sr_dir, 'N/A')
        age_label = str(sr_age) if sr_age is not None else "N/A"
        return False, (f"Support/Resistance gate failed (last break: {sr_dir_label}, "
                        f"{age_label} bars ago, needs <= {SR_MAX_BREAK_AGE}). "
                        f"No recent, direction-agreeing structural break.")

    def _smc_gate_check(sym, signal, t1):
        """[ADDED -- ENABLE_SMC_GATE, Task 51] (ok, detail) for one
        (sym, t1, signal) -- same shared-helper pattern as
        _zerolag_gate_check(). No-op (always ok) when the flag is off."""
        if not ENABLE_SMC_GATE:
            return True, ""

        smc_dir = smc_dir_lookup.get(sym, {}).get(t1)
        smc_age = smc_age_lookup.get(sym, {}).get(t1)
        try:
            smc_dir = int(float(smc_dir)) if smc_dir is not None else None
        except (TypeError, ValueError):
            smc_dir = None
        try:
            smc_age = int(float(smc_age)) if smc_age is not None else None
        except (TypeError, ValueError):
            smc_age = None

        expected_dir = 1 if signal == 'BUY CE' else -1 if signal == 'BUY PE' else None
        dir_ok = expected_dir is not None and smc_dir == expected_dir
        fresh_ok = smc_age is not None and 0 <= smc_age <= SMC_MAX_ZONE_AGE

        if dir_ok and fresh_ok:
            return True, ""

        smc_dir_label = {1: 'Bullish OB retest', -1: 'Bearish OB retest', 0: 'None yet', None: 'N/A'}.get(smc_dir, 'N/A')
        age_label = str(smc_age) if smc_age is not None else "N/A"
        return False, (f"Smart Money Concepts gate failed (last order-block retest: {smc_dir_label}, "
                        f"{age_label} bars ago, needs <= {SMC_MAX_ZONE_AGE}). "
                        f"No recent, direction-agreeing order-block retest.")

    def _try_contrarian_flip(sym, original_signal, t1, spot_price):
        """[ADDED -- ENABLE_PCR_CONTRARIAN_FLIP, Harish's idea, 16-Jul-26]
        Called only when the ORIGINAL momentum-confluence signal
        (original_signal) was about to be rejected by the PCR trap gate
        (pcr_tracker.evaluate() returned passes=False). Rather than reject
        outright, tries the OPPOSITE direction -- e.g. a rejected BUY CE
        (Overbought Trap: PCR low + still falling) becomes a contrarian BUY
        PE betting the crowded/fading call side reverses down; symmetric for
        a rejected BUY PE (Oversold Trap) becoming a contrarian BUY CE.

        Re-runs every downstream gate against the FLIPPED contract's own
        data -- it does NOT reuse the original contract's entry_ltp/volume/
        OI, since flipping CE<->PE means a different tradingsymbol with its
        own price/liquidity. Does NOT re-check VIX (direction-independent,
        already passed above) or PCR itself (that's what triggered this in
        the first place). This is a NEW, UNTESTED hypothesis -- must be A/B
        backtested (flag on vs off) before ever being trusted live.

        Returns (order_fields_dict, flipped_signal) on success, or
        (None, reason_str) on rejection.
        """
        flipped_signal = "BUY PE" if original_signal == "BUY CE" else "BUY CE"

        opt_symbol, opt_token, lot_size, atm_strike = resolve_option_chain(
            sym, spot_price, flipped_signal, df_ref, kite_master
        )

        entry_ltp, opt_vol, opt_oi = 0, 0, 0
        if opt_symbol:
            if is_backtest:
                snap = historical_lookup.get_option_snapshot(kite_api, opt_token, target_date, t1, hist_cache)
                if snap:
                    entry_ltp, opt_vol, opt_oi = snap['close'], snap['volume'], snap['oi']
            else:
                try:
                    opt_quote = kite_api.quote([f"NFO:{opt_symbol}"])
                    opt_data = opt_quote.get(f"NFO:{opt_symbol}", {})
                    entry_ltp = opt_data.get('last_price', 0)
                    opt_vol = opt_data.get('volume', 0)
                    opt_oi = opt_data.get('oi', 0)
                except Exception:
                    pass

        if entry_ltp < MIN_ENTRY_LTP:
            return None, f"{opt_symbol}: [Contrarian Flip] Entry LTP (Rs.{entry_ltp:.2f}) < Rs.{MIN_ENTRY_LTP}. Spread too wide / dead premium."

        if opt_vol == 0:
            return None, f"{opt_symbol}: [Contrarian Flip] Option volume is zero. Dead candle/no liquidity."

        pre_entry_adx = adx_lookup.get(sym, {}).get(t1)
        try:
            pre_entry_adx = float(pre_entry_adx) if pre_entry_adx is not None else None
        except (TypeError, ValueError):
            pre_entry_adx = None
        if pre_entry_adx is not None and pre_entry_adx < ADX_MIN:
            return None, f"{opt_symbol}: [Contrarian Flip] Low momentum / choppy regime (ADX {pre_entry_adx:.1f} < {ADX_MIN}). Skipping."

        zl_ok, zl_detail = _zerolag_gate_check(sym, flipped_signal, t1)
        if not zl_ok:
            return None, f"{opt_symbol}: [Contrarian Flip] {zl_detail}"

        as_of_date = target_date.date() if is_backtest else datetime.now(IST).date()
        is_expiry_day, is_expiry_week, days_to_expiry = get_expiry_context(sym, df_ref, as_of_date)
        t1_time = datetime.strptime(t1, "%H:%M").time()
        if is_expiry_day and (t1_time.hour > 14 or (t1_time.hour == 14 and t1_time.minute >= 30)):
            return None, f"{opt_symbol}: [Contrarian Flip] Expiry Day, within last 45min of session -- theta/gamma risk too high for fresh entry."

        if is_backtest:
            oi_quadrant, oi_confirms = historical_lookup.get_historical_oi_buildup(
                kite_api, opt_token, target_date, t1, hist_cache, flipped_signal
            )
        else:
            oi_quadrant, oi_confirms = get_oi_buildup_signal(opt_symbol, opt_oi, entry_ltp, flipped_signal, oi_cache_path)
        if ENABLE_OI_BUILDUP_GATE and not oi_confirms:
            return None, f"{opt_symbol}: [Contrarian Flip] OI buildup contradicts signal ({oi_quadrant} on {flipped_signal}). Writers positioned against this move."

        sym_sector = sector_map.get(sym, 'Unknown')
        open_in_sector = count_open_positions_in_sector(sym_sector)
        if sym_sector != 'Unknown' and open_in_sector >= max_positions_per_sector:
            return None, f"{opt_symbol}: [Contrarian Flip] Sector cap reached ({sym_sector}: {open_in_sector}/{max_positions_per_sector} open)."

        underlying_atr = atr_lookup.get(sym, {}).get(t1)
        try:
            underlying_atr = float(underlying_atr) if underlying_atr is not None else 0.0
        except (TypeError, ValueError):
            underlying_atr = 0.0
        stop_ltp, target_ltp, risk_per_unit = position_manager.compute_stop_and_target(entry_ltp, underlying_atr)
        num_lots, quantity, risk_amount = position_manager.compute_position_size(
            account_equity, risk_per_unit, lot_size
        )
        if num_lots <= 0:
            return None, f"{opt_symbol}: [Contrarian Flip] even 1 lot's risk (Rs.{risk_per_unit * lot_size:.2f}) exceeds this trade's risk budget (Rs.{risk_amount:.2f})."

        order_id = (
            f"BT_{target_date.strftime('%Y%m%d')}_{sym}_{t1.replace(':', '')}_CFLIP"
            if is_backtest else f"SIM_{int(time.time())}_CFLIP"
        )

        fields = {
            'OI Signal': oi_quadrant,
            'ATM Strike': atm_strike if opt_symbol else "N/A",
            'Option Symbol': opt_symbol or "NOT_FOUND",
            'Option Token': opt_token or "",
            'Lot Size': lot_size,
            'ATR (Underlying)': round(underlying_atr, 2),
            'Entry LTP': entry_ltp, 'Stop Loss LTP': stop_ltp, 'Target LTP': target_ltp,
            'Risk/Unit (Rs)': risk_per_unit,
            'Quantity (Lots)': num_lots, 'Quantity (Units)': quantity, 'Risk Amount (Rs)': risk_amount,
            'Current LTP': entry_ltp, 'Max LTP': entry_ltp, 'Min LTP': entry_ltp,
            'Order ID': order_id,
        }
        return fields, flipped_signal

    def _process_symbol(sym, time_map):
        """[ADDED -- 14-Jul-26 resilience patch] The full 3-bar-streak
        scan for ONE symbol, pulled out of the main loop below so it can
        be called inside a try/except. Previously this body ran directly
        inside `for sym, time_map in final_table.items():` with no fault
        isolation -- an exception raised while processing ANY symbol
        (not necessarily the one with the interesting signal) aborted
        build_order_sheet() entirely, discarding every order/rejection
        already computed for every other symbol in this cycle, including
        ones with a fully-confirmed streak. See the wrapper loop after
        this function for how each symbol is now isolated."""
        current_signal_streak = None
        active_trade_key = None
        # [ADDED -- Task 46, Harish's 18-Jul-26 audit] Tracks the raw,
        # single-bar underlying Final Recomm value of whatever continuous
        # run is currently in progress, and whether THAT run has already
        # been through a full gate-chain audit once. Without this, a
        # rejected 3-bar streak (t1=10:10) re-qualifies as a "new" 3-bar
        # streak on the very next bar (t1=10:15) for as long as the raw
        # signal simply keeps holding -- one continuous BUY CE run was
        # generating a rejection row on EVERY bar instead of one audit for
        # the whole run, exactly the "rejecting at every candle" behavior
        # Harish flagged. This makes it ONE audit per continuous run,
        # matching how a discretionary trader reads a persisting signal:
        # you don't re-ask "is this a buy?" every 5 minutes the same trend
        # just continues to hold -- you asked once, at the point it first
        # looked like 3 bars of confluence, and you live with that answer
        # until the signal itself actually changes.
        run_signal = None
        run_audited = False

        for i in range(len(sorted_times) - 2):
            t1, t2, t3 = sorted_times[i], sorted_times[i + 1], sorted_times[i + 2]
            s1 = time_map.get(t1, "WAIT")
            s2 = time_map.get(t2, "WAIT")
            s3 = time_map.get(t3, "WAIT")

            if s1 != run_signal:
                run_signal = s1
                run_audited = False

            # [ADDED -- risk_and_signal_patches audit] Eager PCR recording,
            # see ENABLE_EAGER_PCR_RECORDING docstring above. Fires on any
            # single-bar BUY CE/PE reading, independent of whether a full
            # 3-bar streak has formed, so pcr_tracker accrues its
            # PCR_TREND_MIN_READINGS readings earlier in the session
            # instead of only when a streak happens to complete. Wrapped
            # defensively -- a failure here must never abort entry
            # detection for this candle.
            if ENABLE_EAGER_PCR_RECORDING and s1 in ("BUY CE", "BUY PE"):
                try:
                    eager_base_match = _match_symbol(df_ref, sym)
                    if eager_base_match is not None:
                        if is_backtest:
                            eager_spot = historical_lookup.get_spot_snapshot(sym, target_date, t1) or 0
                        else:
                            eager_quote = kite_api.quote([f"NSE:{sym}"])
                            eager_spot = eager_quote.get(f"NSE:{sym}", {}).get('last_price', 0)
                        if eager_spot:
                            if is_backtest:
                                eager_pcr = historical_lookup.get_historical_pcr(
                                    sym, eager_spot, target_date, t1, df_ref, kite_master, kite_api, hist_cache
                                )
                            else:
                                eager_pcr = calculate_local_pcr(sym, eager_spot, df_ref, kite_master, kite_api)
                            pcr_tracker.record(sym, eager_pcr)
                except Exception as e:
                    print(f"[WARNING] Eager PCR recording failed for {sym} at {t1}: {e}")

            if current_signal_streak is None:
                if not (s1 == s2 == s3 and s1 in ("BUY CE", "BUY PE")):
                    continue

                # [ADDED -- Task 46] This exact run already went through
                # the gate chain once (accepted or rejected) -- see
                # run_signal/run_audited setup above. Skip re-auditing
                # every subsequent bar of the same continuous signal.
                if run_audited:
                    continue
                run_audited = True

                current_signal_streak = s1
                key = f"{sym}_{t1}"

                if key not in existing_orders:
                    # [ADDED -- Task 49, Harish's training material] Cheap,
                    # pure time-of-day check -- deliberately first in this
                    # chain so it short-circuits before any of the more
                    # expensive lookups/API calls below run at all.
                    if ENABLE_LOW_LIQUIDITY_WINDOW_GATE:
                        t1_clock = datetime.strptime(t1, "%H:%M").time()
                        if LOW_LIQUIDITY_WINDOW_START <= t1_clock < LOW_LIQUIDITY_WINDOW_END:
                            reason = (f"{sym}: Pre-entry bar ({t1}) falls in the low-liquidity "
                                      f"window ({LOW_LIQUIDITY_WINDOW_START.strftime('%H:%M')}-"
                                      f"{LOW_LIQUIDITY_WINDOW_END.strftime('%H:%M')}). Skipping.")
                            print(f"[REJECTED] {sym} -> {reason}")
                            _log_rejection(sym, t1, s1, reason)
                            current_signal_streak = None
                            continue

                    if not position_manager.ENABLE_SIGNAL_REVERSAL_EXIT and has_open_position_in_symbol(sym):
                        reason = f"{sym}: position already open in this symbol -- skipping new entry until it closes."
                        print(f"[REJECTED] {sym} -> {reason}")
                        _log_rejection(sym, t1, s1, reason)
                        current_signal_streak = None
                        continue

                    # [ADDED -- risk_and_signal_patches audit] Daily
                    # drawdown circuit breaker. In LIVE mode (cache_path
                    # set) this genuinely blocks new entries once crossed.
                    # In BACKTEST mode this will never trip here -- see
                    # DailyDrawdownGuard's docstring -- it's still checked
                    # for LIVE/BACKTEST code-path symmetry and to make the
                    # limitation visible rather than silent.
                    if dd_guard.breached():
                        reason = (f"{sym}: Daily drawdown cap breached (realized Rs "
                                  f"{dd_guard.status()['realized_pl']:.2f}). No new entries for the rest of the session.")
                        print(f"[REJECTED] {sym} -> {reason}")
                        _log_rejection(sym, t1, s1, reason)
                        current_signal_streak = None
                        continue

                    base_match = _match_symbol(df_ref, sym)
                    spot_price = 0
                    if base_match is not None:
                        if is_backtest:
                            spot_price = historical_lookup.get_spot_snapshot(sym, target_date, t1) or 0
                        else:
                            try:
                                quote_data = kite_api.quote([f"NSE:{sym}"])
                                spot_price = quote_data.get(f"NSE:{sym}", {}).get('last_price', 0)
                            except Exception:
                                pass

                    opt_symbol, opt_token, lot_size, atm_strike = resolve_option_chain(
                        sym, spot_price, s1, df_ref, kite_master
                    )

                    if is_backtest:
                        pcr_val = historical_lookup.get_historical_pcr(
                            sym, spot_price, target_date, t1, df_ref, kite_master, kite_api, hist_cache
                        )
                    else:
                        pcr_val = calculate_local_pcr(sym, spot_price, df_ref, kite_master, kite_api)
                    pcr_tracker.record(sym, pcr_val)

                    entry_ltp, opt_vol, opt_oi = 0, 0, 0
                    if opt_symbol:
                        if is_backtest:
                            snap = historical_lookup.get_option_snapshot(kite_api, opt_token, target_date, t1, hist_cache)
                            if snap:
                                entry_ltp, opt_vol, opt_oi = snap['close'], snap['volume'], snap['oi']
                        else:
                            try:
                                opt_quote = kite_api.quote([f"NFO:{opt_symbol}"])
                                opt_data = opt_quote.get(f"NFO:{opt_symbol}", {})
                                entry_ltp = opt_data.get('last_price', 0)
                                opt_vol = opt_data.get('volume', 0)
                                opt_oi = opt_data.get('oi', 0)
                            except Exception:
                                pass

                    # --- Rejection filters (see module docstring) ---
                    if is_backtest:
                        vix_val = historical_lookup.get_vix_snapshot(kite_api, target_date, t1, kite_master, hist_cache)
                    else:
                        vix_val = get_india_vix(kite_api)
                    if vix_val is not None and vix_val > VIX_MAX:
                        reason = f"{opt_symbol}: India VIX ({vix_val:.2f}) > {VIX_MAX}. IV too rich for long-premium entry."
                        print(f"[REJECTED] {sym} -> {reason}")
                        _log_rejection(sym, t1, s1, reason)
                        current_signal_streak = None
                        continue

                    pcr_ok, pcr_reason, pcr_trend = pcr_tracker.evaluate(sym, s1, pcr_val)
                    if not pcr_ok:
                        # [ADDED -- ENABLE_PCR_CONTRARIAN_FLIP, Harish's idea,
                        # 16-Jul-26] See flag docstring above and
                        # _try_contrarian_flip()'s own docstring. Default OFF
                        # -- untested hypothesis, must be A/B backtested
                        # before ever being trusted.
                        flip_fields, flip_result = (None, None)
                        if ENABLE_PCR_CONTRARIAN_FLIP:
                            flip_fields, flip_result = _try_contrarian_flip(sym, s1, t1, spot_price)

                        if flip_fields is None:
                            reason = f"{opt_symbol}: {pcr_reason}"
                            if ENABLE_PCR_CONTRARIAN_FLIP:
                                reason += f" | Contrarian flip also rejected -- {flip_result}"
                            print(f"[REJECTED] {sym} -> {reason}")
                            _log_rejection(sym, t1, s1, reason)
                            current_signal_streak = None
                            continue

                        # Contrarian flip succeeded -- build the order record
                        # directly from the flipped contract's own data. Does
                        # NOT fall through to the normal (same-direction)
                        # gate chain below, since flip_fields already passed
                        # its own equivalent checks against the flipped
                        # contract -- re-running the checks below would
                        # wrongly apply the ORIGINAL (unflipped) opt_symbol's
                        # entry_ltp/opt_vol/opt_oi to a trade that isn't that
                        # contract.
                        flipped_signal = flip_result
                        print(f"[EXECUTION] {sym}: PCR trap on {s1} -- contrarian flip to {flipped_signal} instead.")
                        new_rec = {h: "" for h in ORDER_HEADERS}
                        new_rec.update({
                            'Symbol': sym,
                            'PCR': round(pcr_val, 2) if pcr_val else "N/A",
                            'PCR Trend': pcr_trend,
                            'Entry Type': f"Contrarian Flip ({s1} -> {flipped_signal})",
                            'Pre-Entry Trigger Time': t1, 'Pre-Entry Trigger Status': s1,
                            'Entry Trigger Time': t2, 'Entry Trigger Status': s2,
                            'Support Entry Time': t3, 'Support Trigger Status': s3,
                            'Spot Price': spot_price,
                            'Gross P/L (Rs)': "", 'Costs (Rs)': "", 'Net P/L (Rs)': "",
                            'Exit Time': "",
                        })
                        new_rec.update(flip_fields)
                        existing_orders[key] = new_rec

                        if key in existing_orders:
                            exit_val = existing_orders[key].get('Exit Time', "")
                            if pd.isna(exit_val) or str(exit_val).strip() == "":
                                active_trade_key = key
                        continue

                    # [ADDED -- 16-Jul-26 audit] See PCR_REQUIRE_SUFFICIENT_DATA
                    # docstring above -- pcr_tracker.evaluate() itself always
                    # returns passes=True on insufficient history ("gate
                    # skipped"), so this is checked separately, AFTER the
                    # gate above, specifically to catch and reject that case.
                    if PCR_REQUIRE_SUFFICIENT_DATA and pcr_trend == "INSUFFICIENT_DATA":
                        reason = (f"{opt_symbol}: PCR trend history insufficient "
                                  f"(< {PCR_TREND_MIN_READINGS} readings yet) -- this bucket was the "
                                  f"single worst-performing PCR state across two independent backtest "
                                  f"samples. Waiting for more readings before trusting this entry.")
                        print(f"[REJECTED] {sym} -> {reason}")
                        _log_rejection(sym, t1, s1, reason)
                        current_signal_streak = None
                        continue

                    if entry_ltp < MIN_ENTRY_LTP:
                        reason = f"{opt_symbol}: Entry LTP (Rs.{entry_ltp:.2f}) < Rs.{MIN_ENTRY_LTP}. Spread too wide / dead premium."
                        print(f"[REJECTED] {sym} -> {reason}")
                        _log_rejection(sym, t1, s1, reason)
                        current_signal_streak = None
                        continue

                    if opt_vol == 0:
                        reason = f"{opt_symbol}: Option volume is zero. Dead candle/no liquidity."
                        print(f"[REJECTED] {sym} -> {reason}")
                        _log_rejection(sym, t1, s1, reason)
                        current_signal_streak = None
                        continue

                    pre_entry_adx = adx_lookup.get(sym, {}).get(t1)
                    try:
                        pre_entry_adx = float(pre_entry_adx) if pre_entry_adx is not None else None
                    except (TypeError, ValueError):
                        pre_entry_adx = None
                    if pre_entry_adx is not None and pre_entry_adx < ADX_MIN:
                        reason = f"{opt_symbol}: Low momentum / choppy regime (ADX {pre_entry_adx:.1f} < {ADX_MIN}). Skipping."
                        print(f"[REJECTED] {sym} -> {reason}")
                        _log_rejection(sym, t1, s1, reason)
                        current_signal_streak = None
                        continue

                    zl_ok, zl_detail = _zerolag_gate_check(sym, s1, t1)
                    if not zl_ok:
                        reason = f"{opt_symbol}: {zl_detail}"
                        print(f"[REJECTED] {sym} -> {reason}")
                        _log_rejection(sym, t1, s1, reason)
                        current_signal_streak = None
                        continue

                    st_ok, st_detail = _supertrend_gate_check(sym, s1, t1)
                    if not st_ok:
                        reason = f"{opt_symbol}: {st_detail}"
                        print(f"[REJECTED] {sym} -> {reason}")
                        _log_rejection(sym, t1, s1, reason)
                        current_signal_streak = None
                        continue

                    sr_ok, sr_detail = _sr_gate_check(sym, s1, t1)
                    if not sr_ok:
                        reason = f"{opt_symbol}: {sr_detail}"
                        print(f"[REJECTED] {sym} -> {reason}")
                        _log_rejection(sym, t1, s1, reason)
                        current_signal_streak = None
                        continue

                    smc_ok, smc_detail = _smc_gate_check(sym, s1, t1)
                    if not smc_ok:
                        reason = f"{opt_symbol}: {smc_detail}"
                        print(f"[REJECTED] {sym} -> {reason}")
                        _log_rejection(sym, t1, s1, reason)
                        current_signal_streak = None
                        continue

                    as_of_date = target_date.date() if is_backtest else datetime.now(IST).date()
                    is_expiry_day, is_expiry_week, days_to_expiry = get_expiry_context(sym, df_ref, as_of_date)
                    # Evaluated against the SIGNAL'S OWN bar time (t1), not
                    # the real wall clock -- correct for BACKTEST (there is
                    # no "now" for a past date) and at least as accurate for
                    # LIVE (t1 is already ~now during live polling).
                    t1_time = datetime.strptime(t1, "%H:%M").time()
                    if is_expiry_day and (t1_time.hour > 14 or (t1_time.hour == 14 and t1_time.minute >= 30)):
                        reason = f"{opt_symbol}: Expiry Day, within last 45min of session -- theta/gamma risk too high for fresh entry."
                        print(f"[REJECTED] {sym} -> {reason}")
                        _log_rejection(sym, t1, s1, reason)
                        current_signal_streak = None
                        continue

                    if is_backtest:
                        oi_quadrant, oi_confirms = historical_lookup.get_historical_oi_buildup(
                            kite_api, opt_token, target_date, t1, hist_cache, s1
                        )
                    else:
                        oi_quadrant, oi_confirms = get_oi_buildup_signal(opt_symbol, opt_oi, entry_ltp, s1, oi_cache_path)
                    # [CHANGED -- 16-Jul-26 audit] See ENABLE_OI_BUILDUP_GATE
                    # docstring above -- oi_quadrant is still resolved and
                    # still lands in the 'OI Signal' column below either way;
                    # only the hard reject is now conditional.
                    if ENABLE_OI_BUILDUP_GATE and not oi_confirms:
                        reason = f"{opt_symbol}: OI buildup contradicts signal ({oi_quadrant} on {s1}). Writers positioned against this move."
                        print(f"[REJECTED] {sym} -> {reason}")
                        _log_rejection(sym, t1, s1, reason)
                        current_signal_streak = None
                        continue

                    sym_sector = sector_map.get(sym, 'Unknown')
                    open_in_sector = count_open_positions_in_sector(sym_sector)
                    if sym_sector != 'Unknown' and open_in_sector >= max_positions_per_sector:
                        reason = f"{opt_symbol}: Sector cap reached ({sym_sector}: {open_in_sector}/{max_positions_per_sector} open). Avoiding correlated concentration."
                        print(f"[REJECTED] {sym} -> {reason}")
                        _log_rejection(sym, t1, s1, reason)
                        current_signal_streak = None
                        continue

                    # --- Position sizing (see position_manager.py) ---
                    underlying_atr = atr_lookup.get(sym, {}).get(t1)
                    try:
                        underlying_atr = float(underlying_atr) if underlying_atr is not None else 0.0
                    except (TypeError, ValueError):
                        underlying_atr = 0.0
                    stop_ltp, target_ltp, risk_per_unit = position_manager.compute_stop_and_target(entry_ltp, underlying_atr)
                    num_lots, quantity, risk_amount = position_manager.compute_position_size(
                        account_equity, risk_per_unit, lot_size
                    )
                    if num_lots <= 0:
                        reason = (f"{opt_symbol}: even 1 lot's risk (Rs.{risk_per_unit * lot_size:.2f}) exceeds "
                                  f"this trade's risk budget (Rs.{risk_amount:.2f}).")
                        print(f"[REJECTED] {sym} -> {reason}")
                        _log_rejection(sym, t1, s1, reason)
                        current_signal_streak = None
                        continue

                    order_id = (
                        f"BT_{target_date.strftime('%Y%m%d')}_{sym}_{t1.replace(':', '')}"
                        if is_backtest else f"SIM_{int(time.time())}"
                    )

                    new_rec = {h: "" for h in ORDER_HEADERS}
                    new_rec.update({
                        'Symbol': sym,
                        'PCR': round(pcr_val, 2) if pcr_val else "N/A",
                        'PCR Trend': pcr_trend,
                        'OI Signal': oi_quadrant,
                        'Entry Type': 'Confluence',
                        'Pre-Entry Trigger Time': t1, 'Pre-Entry Trigger Status': s1,
                        'Entry Trigger Time': t2, 'Entry Trigger Status': s2,
                        'Support Entry Time': t3, 'Support Trigger Status': s3,
                        'Spot Price': spot_price,
                        'ATM Strike': atm_strike if opt_symbol else "N/A",
                        'Option Symbol': opt_symbol or "NOT_FOUND",
                        'Option Token': opt_token or "",
                        'Lot Size': lot_size,
                        'ATR (Underlying)': round(underlying_atr, 2),
                        'Entry LTP': entry_ltp, 'Stop Loss LTP': stop_ltp, 'Target LTP': target_ltp,
                        'Risk/Unit (Rs)': risk_per_unit,
                        'Quantity (Lots)': num_lots, 'Quantity (Units)': quantity, 'Risk Amount (Rs)': risk_amount,
                        'Current LTP': entry_ltp, 'Max LTP': entry_ltp, 'Min LTP': entry_ltp,
                        'Gross P/L (Rs)': "", 'Costs (Rs)': "", 'Net P/L (Rs)': "",
                        'Order ID': order_id, 'Exit Time': "",
                    })
                    existing_orders[key] = new_rec

                if key in existing_orders:
                    exit_val = existing_orders[key].get('Exit Time', "")
                    if pd.isna(exit_val) or str(exit_val).strip() == "":
                        active_trade_key = key
            else:
                if s3 != current_signal_streak:
                    if active_trade_key:
                        exit_val = existing_orders[active_trade_key].get('Exit Time', "")
                        if pd.isna(exit_val) or str(exit_val).strip() == "":
                            # [CHANGED] 'Exit Trigger Time'/'Status' are
                            # always recorded -- useful audit trail of when
                            # the confirmed signal itself broke, regardless
                            # of whether we act on it. Actually CLOSING the
                            # position here only happens if
                            # ENABLE_SIGNAL_REVERSAL_EXIT is True; otherwise
                            # the position stays open and is left for
                            # stop/target/trailing/max-hold/EOD to resolve
                            # (BACKTEST: position_manager.simulate_backtest_
                            # exit() in Phase 2 below; LIVE:
                            # update_open_positions_live()'s per-cycle
                            # price-based check). See position_manager.py's
                            # ENABLE_SIGNAL_REVERSAL_EXIT docstring for the
                            # backtest data that motivated this.
                            update_fields = {'Exit Trigger Time': t3, 'Exit Trigger Status': s3}
                            if position_manager.ENABLE_SIGNAL_REVERSAL_EXIT:
                                # Stamps the historical bar time in BACKTEST,
                                # real live time in LIVE -- never the wall
                                # clock the script happened to run under.
                                update_fields.update({
                                    'Exit Time': _timestamp_str(t3),
                                    'Exit Reason': f"5M Reversal ({s3})",
                                })
                                print(f"[EXECUTION] Position closed via 5M Logic for {sym}. Reason: Reversal to {s3}")
                            existing_orders[active_trade_key].update(update_fields)
                    current_signal_streak = None
                    active_trade_key = None

    # [ADDED -- 14-Jul-26 resilience patch] Actually run the per-symbol
    # scan now, with each symbol isolated: a crash processing one symbol
    # (bad option-chain lookup, missing historical data, an unexpected
    # None somewhere, etc.) is logged and recorded as a SYSTEM_ERROR
    # rejection row instead of aborting every other symbol's
    # already-computed orders/rejections for this cycle.
    for sym, time_map in final_table.items():
        try:
            _process_symbol(sym, time_map)
        except Exception as e:
            print(f"[ERROR] Symbol {sym} processing crashed mid-scan -- "
                  f"skipping rest of this symbol's signal history: {e}")
            _log_rejection(sym, "N/A", "SYSTEM_ERROR",
                            f"Symbol processing crashed and was skipped: {e}")

    # --- Phase 2, BACKTEST only: resolve the REAL exit for every order
    # that doesn't have one yet, via position_manager's historical candle
    # walk-forward. This is what actually produces a trustworthy Net P/L
    # instead of a hardcoded zero. ---
    if is_backtest:
        for key, order in existing_orders.items():
            if str(order.get('Net P/L (Rs)', "")).strip() not in ("", "nan"):
                continue  # already resolved on a prior pass within this same run
            opt_token = order.get('Option Token')
            if not opt_token:
                continue

            entry_ltp = float(order.get('Entry LTP') or 0)
            stop_ltp = float(order.get('Stop Loss LTP') or 0)
            target_ltp = float(order.get('Target LTP') or 0)
            risk_per_unit = float(order.get('Risk/Unit (Rs)') or 0)
            entry_time_str = order.get('Pre-Entry Trigger Time')
            reversal_time_str = order.get('Exit Trigger Time') or None
            quantity = int(order.get('Quantity (Units)') or 0)  # [MOVED] up so it can be passed in below

            result = position_manager.simulate_backtest_exit(
                kite_api, opt_token, target_date, entry_time_str, entry_ltp,
                stop_ltp, target_ltp, risk_per_unit, hist_cache,
                signal_reversal_time_str=reversal_time_str,
                quantity=quantity,  # [ADDED] enables the Rs 2,000 hard-cap check
            )

            gross_pl = (result['exit_ltp'] - entry_ltp) * quantity
            costs = position_manager.estimate_round_trip_costs(entry_ltp, result['exit_ltp'], quantity)
            net_pl = round(gross_pl - costs['total'], 2)

            order.update({
                'Current LTP': result['exit_ltp'], 'Max LTP': result['max_ltp_seen'], 'Min LTP': result['min_ltp_seen'],
                'Gross P/L (Rs)': round(gross_pl, 2), 'Costs (Rs)': costs['total'], 'Net P/L (Rs)': net_pl,
                'Exit Time': result['exit_time'], 'Exit Reason': result['exit_reason'],
            })

    # --- Phase 2.5, BACKTEST only: POST-HOC daily drawdown replay. ---
    # [ADDED -- risk_and_signal_patches audit] Net P/L only exists after
    # Phase 2 above, by which point every entry for the day was already
    # taken in Phase 1 -- so dd_guard.breached() was never True during
    # entry decisions this run (see DailyDrawdownGuard's docstring). This
    # block answers a different, still-useful question after the fact:
    # "walking through today's trades in the order they were ENTERED, at
    # what point would the daily cap have been crossed, and how much of
    # the day's loss came from trades opened after that point?" It's a
    # diagnostic on top of the backtest, not a claim that any trade below
    # was actually prevented.
    if is_backtest:
        entered_orders = [o for o in existing_orders.values() if str(o.get('Net P/L (Rs)', "")).strip() not in ("", "nan")]
        entered_orders.sort(key=lambda o: str(o.get('Pre-Entry Trigger Time', "")))
        replay_guard = position_manager.DailyDrawdownGuard(max_daily_loss_rs=position_manager.DAILY_MAX_LOSS_RS)
        pl_after_breach = 0.0
        trades_after_breach = 0
        for o in entered_orders:
            net_pl = float(o.get('Net P/L (Rs)') or 0)
            was_breached_before = replay_guard.breached()
            replay_guard.update(net_pl, at_time_str=o.get('Pre-Entry Trigger Time'))
            if was_breached_before:
                pl_after_breach += net_pl
                trades_after_breach += 1
        if replay_guard.breached():
            print(f"[RISK] Post-hoc replay: a Rs {position_manager.DAILY_MAX_LOSS_RS:.0f} daily cap would have "
                  f"triggered at {replay_guard.status()['breach_time']} this session. {trades_after_breach} "
                  f"trade(s) were entered after that point, contributing Rs {pl_after_breach:.2f} to the day's "
                  f"total -- these would NOT have been taken live with the guard active.")

    df_orders_out = pd.DataFrame(list(existing_orders.values()), columns=ORDER_HEADERS)
    df_rejected_out = pd.DataFrame(rejected_rows, columns=REJECTED_HEADERS)
    return df_orders_out, df_rejected_out


def write_order_sheet(df_orders, df_rejected, output_excel_path):
    """Writes/replaces the 'Orders' and 'Rejected' sheets, then autofits
    just those two and atomically re-saves."""
    # [CHANGED -- 17-Jul-26, Harish's request] Both sheets used to land in
    # whatever order build_order_sheet()'s per-symbol loop produced them
    # (effectively Symbol-ascending, since that loop walks df_ref's
    # Reference rows in symbol order) -- sorted here instead by the
    # trade's own Pre-Entry Time / Trigger Time so the sheet reads
    # chronologically (what happened first in the session), tie-broken by
    # Symbol ascending for entries sharing the same bar.
    if not df_orders.empty and 'Pre-Entry Trigger Time' in df_orders.columns:
        df_orders = df_orders.sort_values(
            by=['Pre-Entry Trigger Time', 'Symbol'], na_position='last', ignore_index=True
        )
    if not df_rejected.empty and 'Trigger Time' in df_rejected.columns:
        df_rejected = df_rejected.sort_values(
            by=['Trigger Time', 'Symbol'], na_position='last', ignore_index=True
        )

    with pd.ExcelWriter(output_excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_orders.to_excel(writer, sheet_name='Orders', index=False)
        df_rejected.to_excel(writer, sheet_name='Rejected', index=False)

    wb = load_workbook(output_excel_path)
    for sheet_name in ('Orders', 'Rejected'):
        if sheet_name in wb.sheetnames:
            excel_utils.autofit_columns(wb[sheet_name])
    excel_utils.atomic_save(wb, output_excel_path)


def run_order_sheet_step(output_excel_path, kite_api, df_ref,
                          mode=calendar_mgmt.LIVE, target_date=None,
                          account_equity=position_manager.ACCOUNT_EQUITY_DEFAULT,
                          max_positions_per_sector=MAX_POSITIONS_PER_SECTOR):
    """Single entry point for 01_Master_Code.py. Must run AFTER
    final_sheet.run_final_sheet_step() has already saved the 'Final'
    sheet into output_excel_path.

    mode/target_date/account_equity let BACKTEST runs source historical
    (not live) data -- see module docstring -- and this now rebuilds the
    'Dashboard' sheet from the real Net P/L this step produces (see
    dashboard.py)."""
    print(f"[SYSTEM] Scanning Final Recomm for 3-consecutive-column confluence signals ({mode})...")
    try:
        df_orders, df_rejected = build_order_sheet(
            output_excel_path, kite_api, df_ref, mode=mode, target_date=target_date,
            account_equity=account_equity, max_positions_per_sector=max_positions_per_sector,
        )
    except Exception as e:
        # [ADDED -- 14-Jul-26 resilience patch] A total build_order_sheet()
        # failure (e.g. _build_kite_master() above hitting an expired
        # token) used to leave the workbook completely untouched -- no
        # Orders, no Rejected, no Dashboard -- with the only trace being a
        # console print in 01_Master_Code.py's outer try/except. That made
        # a real signal indistinguishable from "nothing happened this
        # cycle." Now it still writes an empty Orders sheet plus one
        # Rejected row documenting the crash, so the failure is visible in
        # the workbook itself, not only the console.
        print(f"[ERROR] build_order_sheet() failed entirely: {e}")
        df_orders = pd.DataFrame(columns=ORDER_HEADERS)
        df_rejected = pd.DataFrame([{
            'Symbol': 'SYSTEM', 'Trigger Time': 'N/A', 'Signal': 'N/A',
            'Reason': f"Order Sheet step crashed before processing any symbol: {e}",
            'Timestamp': datetime.now(IST).strftime('%H:%M:%S'),
        }], columns=REJECTED_HEADERS)

    write_order_sheet(df_orders, df_rejected, output_excel_path)
    print(f"[SUCCESS] Order sheet updated: {len(df_orders)} order(s) tracked, {len(df_rejected)} rejection(s) logged.")

    try:
        dashboard.build_dashboard_sheet(output_excel_path, mode=mode, target_date=target_date)
        print("[SUCCESS] Dashboard sheet rebuilt from real Net P/L.")
    except Exception as e:
        print(f"[WARNING] Dashboard rebuild failed (non-fatal): {e}")

    # [ADDED -- Phase B, viewer mirror] No-op unless ENABLE_SHEETS_SYNC is
    # on (see sheets_sync.py docstring for one-time setup). Wrapped here
    # too, on top of sheets_sync's own internal no-op check, so a genuine
    # sync failure (bad key, quota, network) can never take down the
    # actual trading pipeline -- worst case the phone/browser view is
    # simply stale until the next successful cycle.
    try:
        sheets_sync.sync_to_google_sheets(output_excel_path)
    except Exception as e:
        print(f"[WARNING] Google Sheets viewer sync failed (non-fatal): {e}")


# ---------------------------------------------------------------------------
# [ADDED] LIVE-mode open-position polling -- call once per candle-close
# cycle (see 01_Master_Code.py's run_cycle() patch notes). Reads every
# still-open row in 'Orders', checks it against position_manager's
# stop/target/trailing/max-hold/EOD logic via a single live quote() per
# position, and writes back Current/Max/Min LTP and, if triggered, the
# exit fields + real net P/L. Lighter-touch than the BACKTEST path --
# paper-trade thoroughly before trusting it with real order placement.
# ---------------------------------------------------------------------------
def update_open_positions_live(kite_api, output_excel_path):
    # [ADDED -- 13-Jul-26] This runs BEFORE order_sheet.run_order_sheet_step()
    # in 01_Master_Code.run_cycle() (intentionally -- see that function's
    # comment), so on the FIRST cycle of the day the 'Orders' sheet simply
    # doesn't exist yet: nothing has been created to poll. That used to
    # print a [WARNING] every single day at 09:20, for a condition that
    # resolves itself one call later in the same cycle. Checking sheet
    # existence first turns that into a silent, expected no-op and keeps
    # [WARNING] reserved for an actually-unreadable/corrupt workbook.
    try:
        sheet_names = load_workbook(output_excel_path, read_only=True).sheetnames
    except Exception as e:
        print(f"[WARNING] update_open_positions_live: could not open workbook ({e}).")
        return
    if 'Orders' not in sheet_names:
        return  # expected pre-first-order-sheet-run state -- not an error

    # [ADDED -- risk_and_signal_patches audit] Same persisted guard as
    # build_order_sheet() -- LIVE mode's sequential, real-time polling is
    # exactly where this circuit breaker is designed to work: closing a
    # position here that crosses the daily cap will cause the NEXT call
    # to build_order_sheet() (which re-reads DAILY_DRAWDOWN_CACHE) to
    # reject all new entries for the rest of the session.
    dd_guard = position_manager.DailyDrawdownGuard(
        max_daily_loss_rs=position_manager.DAILY_MAX_LOSS_RS,
        # [FIXED -- 15-Jul-26] same AttributeError as build_order_sheet() above
        # -- DAILY_DRAWDOWN_CACHE lives in this module, not on position_manager.
        cache_path=DAILY_DRAWDOWN_CACHE,
        date_key=today_ist().strftime('%Y-%m-%d'),
    )

    try:
        df_orders = pd.read_excel(output_excel_path, sheet_name='Orders')
    except Exception as e:
        print(f"[WARNING] update_open_positions_live: could not read 'Orders' sheet ({e}).")
        return

    if df_orders.empty:
        return

    now = datetime.now(IST)
    changed = False

    for idx, order in df_orders.iterrows():
        exit_val = order.get('Exit Time', "")
        is_open = pd.isna(exit_val) or str(exit_val).strip() == ""
        opt_symbol = order.get('Option Symbol')
        if not is_open or not opt_symbol or opt_symbol == "NOT_FOUND":
            continue

        try:
            quote_data = kite_api.quote([f"NFO:{opt_symbol}"])
            current_ltp = quote_data.get(f"NFO:{opt_symbol}", {}).get('last_price')
        except Exception as e:
            print(f"[WARNING] Live LTP poll failed for {opt_symbol}: {e}")
            continue
        if current_ltp is None:
            continue

        entry_ltp = float(order.get('Entry LTP') or 0)
        stop_ltp = float(order.get('Stop Loss LTP') or 0)
        target_ltp = float(order.get('Target LTP') or 0)
        risk_per_unit = float(order.get('Risk/Unit (Rs)') or 0)
        max_ltp_seen = float(order.get('Max LTP') or entry_ltp)
        min_ltp_seen = float(order.get('Min LTP') or entry_ltp)
        tsl_breach_streak = int(order.get('TSL Breach Streak') or 0)  # [ADDED] see position_manager.check_live_exit()
        quantity = int(order.get('Quantity (Units)') or 0)  # [MOVED] up so it can be passed in below
        pre_entry_time = order.get('Pre-Entry Trigger Time')
        try:
            entry_dt = now.replace(
                hour=int(str(pre_entry_time).split(':')[0]), minute=int(str(pre_entry_time).split(':')[1]),
                second=0, microsecond=0,
            )
        except Exception:
            entry_dt = now

        should_exit, exit_reason, exit_ltp, new_max, new_trailing_stop, new_tsl_streak = position_manager.check_live_exit(
            entry_ltp, stop_ltp, target_ltp, risk_per_unit, current_ltp, max_ltp_seen, entry_dt, now,
            quantity=quantity,  # [ADDED] enables the Rs 2,000 hard-cap check
            tsl_breach_streak=tsl_breach_streak,  # [ADDED] ENABLE_TSL_CONFIRMATION_HOLD -- carried across polls
        )

        df_orders.at[idx, 'Current LTP'] = current_ltp
        df_orders.at[idx, 'Max LTP'] = max(new_max, max_ltp_seen)
        df_orders.at[idx, 'Min LTP'] = min(min_ltp_seen, current_ltp)
        df_orders.at[idx, 'TSL Breach Streak'] = new_tsl_streak
        changed = True

        if should_exit:
            gross_pl = (exit_ltp - entry_ltp) * quantity
            costs = position_manager.estimate_round_trip_costs(entry_ltp, exit_ltp, quantity)
            net_pl = round(gross_pl - costs['total'], 2)
            df_orders.at[idx, 'Exit Time'] = now.strftime('%H:%M:%S')
            df_orders.at[idx, 'Exit Reason'] = exit_reason
            df_orders.at[idx, 'Gross P/L (Rs)'] = round(gross_pl, 2)
            df_orders.at[idx, 'Costs (Rs)'] = costs['total']
            df_orders.at[idx, 'Net P/L (Rs)'] = net_pl
            print(f"[EXECUTION] {order.get('Symbol')} closed live: {exit_reason}, Net P/L Rs.{net_pl:.2f}")
            dd_guard.update(net_pl, at_time_str=now.strftime('%H:%M:%S'))

    if changed:
        # [CHANGED -- 17-Jul-26] Same Pre-Entry-Time sort as
        # write_order_sheet() -- keeps LIVE's per-cycle re-write from
        # drifting back to Symbol-ascending order after every position
        # update.
        if not df_orders.empty and 'Pre-Entry Trigger Time' in df_orders.columns:
            df_orders = df_orders.sort_values(
                by=['Pre-Entry Trigger Time', 'Symbol'], na_position='last', ignore_index=True
            )
        with pd.ExcelWriter(output_excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_orders.to_excel(writer, sheet_name='Orders', index=False)
        wb = load_workbook(output_excel_path)
        excel_utils.autofit_columns(wb['Orders'])
        excel_utils.atomic_save(wb, output_excel_path)
        try:
            dashboard.build_dashboard_sheet(output_excel_path, mode=calendar_mgmt.LIVE, target_date=None)
        except Exception as e:
            print(f"[WARNING] Dashboard rebuild failed (non-fatal): {e}")

        # [ADDED -- Phase B, viewer mirror] Same no-op-unless-enabled
        # mirror as run_order_sheet_step() above -- see sheets_sync.py.
        try:
            sheets_sync.sync_to_google_sheets(output_excel_path)
        except Exception as e:
            print(f"[WARNING] Google Sheets viewer sync failed (non-fatal): {e}")


# ---------------------------------------------------------------------------
# Disclaimer: this module creates SIMULATED order records in BACKTEST mode
# (Order ID is a local tag) and, in LIVE mode, still only SIMULATES/tracks
# positions -- no order is actually placed with the broker anywhere in
# this file. It is not financial advice. Every threshold, the ATM-delta
# approximation, and the broker fee schedule in position_manager.py
# should be paper-traded and reviewed before this is ever wired to real
# order placement or real capital.
# ---------------------------------------------------------------------------
