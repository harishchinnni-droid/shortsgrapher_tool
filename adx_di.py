"""
adx_di.py
---------
Python translation of '04_ProjectFiles/ADX & DI.txt' (Pine v4, "ADX and DI
for v4" by BeikabuOyaji), wired into the pipeline the same way as
RSI.py/SQZMOM.py so it can be plugged straight back into run_pipeline.py's
INDICATORS list. Rebuilt from the Pine source only -- the original Python
version of this module was lost (source file gone, only a stale
__pycache__/adx_di.cpython-312.pyc leftover), so this is a fresh port, not
a recovered file. Flag anything that looks off against whatever the prior
version actually did.

Indicator logic (exact translation of the Pine source, len=14, th=20):
    TrueRange = max(high-low, abs(high-close_prev), abs(low-close_prev))
    DM+ = max(high-high_prev, 0) if (high-high_prev) > (low_prev-low) else 0
    DM- = max(low_prev-low, 0)   if (low_prev-low) > (high-high_prev) else 0

    Wilder's running-sum smoothing (NOT a plain EMA -- the Pine recursion
    adds the full new term each bar rather than an alpha-weighted one:
    S[t] = S[t-1] - S[t-1]/len + x[t], with S seeded at 0 like Pine's nz()
    on an undefined S[t-1]). Applied to TrueRange, DM+, DM- independently.

    DI+ = Smoothed(DM+) / Smoothed(TR) * 100
    DI- = Smoothed(DM-) / Smoothed(TR) * 100
    DX  = abs(DI+ - DI-) / (DI+ + DI-) * 100
    ADX = simple moving average of DX over `len` bars

Recommendation ('ADX Recomm') -- NOT part of the original Pine script,
which only plots DI+/DI-/ADX with no buy/sell logic. Added here using the
standard textbook ADX/DI trend-confirmation rule (DI crossover + ADX
above the script's own th=20 "trending" threshold), for consistency with
every other sheet in this workbook having a Recomm column. Flag to Harish
if the lost original used different Recomm logic than this:
    BUY CE if DI+ > DI- AND ADX > 20
    BUY PE if DI- > DI+ AND ADX > 20
    WAIT   otherwise (no clear trend, or DI+/DI- too close to call)

Per-bar, not hold-based -- same convention as RSI.py/ema20.py: recomputed
fresh every bar from that bar's own DI+/DI-/ADX values, no memory of a
previous "armed" state.

Timeframe: 5-minute candles, truncated per day to CUTOFF_TIME (15:15),
same as every other module in this pipeline.
"""

import os
import sys
import concurrent.futures
from datetime import time as dtime

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

CODES_DIR = os.path.dirname(os.path.abspath(__file__))
if CODES_DIR not in sys.path:
    sys.path.append(CODES_DIR)
import data_ingestion
import excel_utils

# ---------------------------------------------------------------------------
# Config -- mirrors the Pine script's `len`/`th` inputs
# ---------------------------------------------------------------------------
ADX_LEN = 14                  # `len` in the Pine source
ADX_TREND_THRESHOLD = 20      # `th` in the Pine source (its hline level)
CUTOFF_TIME = dtime(15, 15)
INTERVAL = "5minute"

# ---------------------------------------------------------------------------
# Excel styling (kept visually consistent with the rest of the workbook)
# ---------------------------------------------------------------------------
FILL_LIME = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")   # BUY CE
FILL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # BUY PE
FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")   # WAIT
FONT_WHITE = Font(color="FFFFFF")
FONT_BLACK = Font(color="000000")


# ---------------------------------------------------------------------------
# Indicator -- literal translation of the Pine source's recursion
# ---------------------------------------------------------------------------
def _wilder_running_sum(values: np.ndarray, period: int) -> np.ndarray:
    """S[t] = S[t-1] - S[t-1]/period + values[t], S seeded at 0 (Pine's
    nz() on an undefined S[t-1] at bar 1). A plain loop on purpose: this
    recursion has a coefficient of 1 (not alpha=1/period) on the new term,
    so it is NOT the same series as pandas .ewm() and can't be swapped
    for one without changing the warm-up behavior. Cheap either way at
    intraday row counts."""
    out = np.empty(len(values), dtype=float)
    prev = 0.0
    for i, x in enumerate(values):
        prev = prev - (prev / period) + x
        out[i] = prev
    return out


def calculate_adx_di(df, period=ADX_LEN, threshold=ADX_TREND_THRESHOLD):
    df = df.copy()

    high = df['high']
    low = df['low']
    close = df['close']
    prev_high = high.shift(1)
    prev_low = low.shift(1)
    prev_close = close.shift(1)

    true_range = pd.concat([
        high - low,
        (high - prev_close).abs(),
        (low - prev_close).abs(),
    ], axis=1).max(axis=1).fillna(0.0)

    up_move = high - prev_high
    down_move = prev_low - low
    dm_plus = np.where((up_move > down_move) & (up_move > 0), up_move, 0.0)
    dm_minus = np.where((down_move > up_move) & (down_move > 0), down_move, 0.0)
    # First bar has no prev_high/prev_low -- treat like Pine's nz(...) = 0.
    dm_plus = np.nan_to_num(dm_plus, nan=0.0)
    dm_minus = np.nan_to_num(dm_minus, nan=0.0)

    smoothed_tr = _wilder_running_sum(true_range.to_numpy(), period)
    smoothed_dm_plus = _wilder_running_sum(dm_plus, period)
    smoothed_dm_minus = _wilder_running_sum(dm_minus, period)

    with np.errstate(divide='ignore', invalid='ignore'):
        di_plus = np.where(smoothed_tr != 0, smoothed_dm_plus / smoothed_tr * 100, 0.0)
        di_minus = np.where(smoothed_tr != 0, smoothed_dm_minus / smoothed_tr * 100, 0.0)
        di_sum = di_plus + di_minus
        dx = np.where(di_sum != 0, np.abs(di_plus - di_minus) / di_sum * 100, 0.0)

    df['DI_PLUS'] = di_plus
    df['DI_MINUS'] = di_minus
    df['ADX'] = pd.Series(dx, index=df.index).rolling(window=period, min_periods=period).mean()

    buy_ce = (df['DI_PLUS'] > df['DI_MINUS']) & (df['ADX'] > threshold)
    buy_pe = (df['DI_MINUS'] > df['DI_PLUS']) & (df['ADX'] > threshold)
    df['ADX Recomm'] = np.where(buy_ce, 'BUY CE', np.where(buy_pe, 'BUY PE', 'WAIT'))

    return df


# ---------------------------------------------------------------------------
# Per-symbol processing (same brute-force time/timezone protocol as RSI.py)
# ---------------------------------------------------------------------------
def process_symbol(symbol_data):
    symbol, df, target_date = symbol_data
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip().str.lower()

    if not all(col in df.columns for col in ['open', 'high', 'low', 'close']):
        print(f"[WARNING] {symbol}: Missing required price columns. Discarding.")
        return symbol, None

    time_col = next((col for col in ['date', 'time', 'datetime', 'timestamp'] if col in df.columns), None)
    if time_col:
        parsed = pd.to_datetime(df[time_col], errors='coerce')
    elif pd.api.types.is_datetime64_any_dtype(df.index):
        parsed = pd.Series(df.index, index=df.index)
    else:
        unnamed_cols = [c for c in df.columns if 'unnamed' in c]
        if unnamed_cols:
            parsed = pd.to_datetime(df[unnamed_cols[0]], errors='coerce')
        else:
            print(f"[WARNING] {symbol}: Failed to locate any valid timestamp data. Discarding.")
            return symbol, None

    try:
        if parsed.dt.tz is not None:
            parsed = parsed.dt.tz_localize(None)
    except (AttributeError, TypeError):
        pass

    df['_sort_dt'] = parsed
    df['time_str'] = parsed.dt.strftime('%H:%M')
    df.dropna(subset=['time_str', '_sort_dt'], inplace=True)
    if df.empty:
        print(f"[WARNING] {symbol}: No rows with a valid timestamp. Discarding.")
        return symbol, None

    df.sort_values('_sort_dt', inplace=True)
    df = df[df['_sort_dt'].dt.time <= CUTOFF_TIME]
    if df.empty:
        print(f"[WARNING] {symbol}: No candles at/before {CUTOFF_TIME.strftime('%H:%M')}. Discarding.")
        return symbol, None

    df = calculate_adx_di(df)

    df = excel_utils.restrict_to_target_date(df, target_date)
    if df is None:
        return symbol, None

    return symbol, df


# ---------------------------------------------------------------------------
# Excel export -- pivoted Symbol x Time matrix (same pattern as RSI.py)
# ---------------------------------------------------------------------------
def build_matrix(data_dict, target_date, max_workers=None):
    results = {}
    with concurrent.futures.ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(process_symbol, (sym, df, target_date)): sym for sym, df in data_dict.items()}
        for future in concurrent.futures.as_completed(futures):
            sym = futures[future]
            try:
                processed_sym, processed_df = future.result()
                if processed_df is not None and 'ADX' in processed_df.columns:
                    processed_df = processed_df.dropna(subset=['ADX'])
                    if not processed_df.empty and sym.lower() != 'summary':
                        results[sym] = processed_df
            except Exception as e:
                print(f"[ERROR] ADX & DI Engine Failure on symbol {sym}: {str(e)}")

    if not results:
        raise RuntimeError("ADX & DI: zero symbols processed successfully for target_date -- matrix build aborted.")

    all_times = set()
    for df in results.values():
        if 'time_str' in df.columns:
            all_times.update(df['time_str'].dropna().tolist())

    if not all_times:
        raise RuntimeError("ADX & DI: no valid timestamps extracted across all symbols -- matrix build aborted.")

    sorted_times = sorted(all_times)
    matrix_rows = [['Symbol', 'Metrics'] + sorted_times]

    for sym in sorted(results.keys()):
        df = results[sym].drop_duplicates(subset=['time_str'], keep='last')
        df_indexed = df.set_index('time_str')

        def get_metric_row(metric_name, col_name):
            row = [sym, metric_name]
            for t in sorted_times:
                if t in df_indexed.index:
                    val = df_indexed.loc[t, col_name]
                    if isinstance(val, pd.Series):
                        val = val.iloc[-1]
                    row.append(val)
                else:
                    row.append("")
            return row

        matrix_rows.append(get_metric_row('DI+', 'DI_PLUS'))
        matrix_rows.append(get_metric_row('DI-', 'DI_MINUS'))
        matrix_rows.append(get_metric_row('ADX', 'ADX'))
        matrix_rows.append(get_metric_row('ADX Recomm', 'ADX Recomm'))

    return matrix_rows


def write_matrix_to_workbook(matrix_rows, wb):
    sheet_name = "ADX"
    ws = excel_utils.replace_sheet_with_matrix(wb, sheet_name, matrix_rows)

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1), start=2):
        metric_type = ws.cell(row=r_idx, column=2).value
        for cell in row[2:]:
            val = cell.value
            if val == "":
                continue
            if metric_type == 'ADX Recomm':
                if val == 'BUY CE':
                    cell.fill = FILL_LIME
                    cell.font = FONT_BLACK
                elif val == 'BUY PE':
                    cell.fill = FILL_RED
                    cell.font = FONT_WHITE
                elif val == 'WAIT':
                    cell.fill = FILL_GRAY
                    cell.font = FONT_BLACK

    excel_utils.autofit_columns(ws)


def write_matrix(matrix_rows, output_excel_path):
    wb = load_workbook(output_excel_path)
    write_matrix_to_workbook(matrix_rows, wb)
    excel_utils.atomic_save(wb, output_excel_path)


def parallel_compute_and_export(data_dict, target_date, output_excel_path, max_workers=None):
    matrix_rows = build_matrix(data_dict, target_date, max_workers=max_workers)
    write_matrix(matrix_rows, output_excel_path)


def run_adx_di_step(df_ref, target_date, output_excel_path):
    print("[SYSTEM] Loading 5-minute historical data for ADX & DI...")
    data_dict = data_ingestion.load_interval_data(df_ref, target_date, interval=INTERVAL)
    if not data_dict:
        raise RuntimeError("ADX & DI: no 5-minute data files found for any symbol.")
    print(f"[PROCESS] Computing ADX & DI matrix for {len(data_dict)} symbol(s)...")
    parallel_compute_and_export(data_dict, target_date, output_excel_path)
    print("[SUCCESS] ADX & DI matrix written to sheet 'ADX'.")


# ---------------------------------------------------------------------------
# Disclaimer: this module derives a signal from historical/live price data
# only. It is not financial advice, and no result here is a guarantee of
# future performance -- paper-trade it and run it through
# scripts/backtester.py before risking real capital.
# ---------------------------------------------------------------------------
