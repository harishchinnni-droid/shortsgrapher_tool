"""
lite_final_sheet.py
--------------------
NEW MODULE -- part of the 3-indicator standalone pipeline (Task 52,
18-Jul-26). Confluence rule per Harish's explicit answer: "confluence
value of all 3 rows... should give same signal either BUY CE or BUY PE
for entry in order sheet. If any disagreement in these 3 indicators, then
it will be WAIT." -- i.e. strict unanimous vote across exactly 3 sheets,
not the category-majority scheme final_sheet.py uses for the full
13-indicator pipeline.

Does NOT modify or replace final_sheet.py -- that module stays exactly
as-is for the existing full pipeline. This writes to the SAME 'Final'
sheet name and the SAME Symbol x Metrics x time-column matrix shape
final_sheet.py already produces, so order_sheet.py (which reads the
'Final' sheet's 'Final Recomm' row generically via pandas) works
completely unmodified against either pipeline's output -- see
order_sheet.py's _load_final_recomm_lookup()-equivalent read, which only
cares about the sheet/row existing, not which module wrote it.

Which sheets vote (all 3, strict unanimous):
    TW ALL  (lite_tw_all.py's 'TW ALL Recomm')
    RSI     (lite_rsi.py's 'RSI Recomm')
    ADX     (adx_di.py's own, UNMODIFIED 'ADX Recomm' -- reused as-is,
             see adx_di.py's docstring: DI+ > DI- and ADX > 20 -> BUY CE,
             mirrored for BUY PE, else WAIT)

Final Recomm = 'BUY CE' only if all 3 read 'BUY CE'; 'BUY PE' only if all
3 read 'BUY PE'; 'WAIT' otherwise -- including when any one of the three
is WAIT, or when a sheet/row is missing entirely (counts as WAIT for that
cell, never crashes the build).
"""

import os
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

CODES_DIR = os.path.dirname(os.path.abspath(__file__))
if CODES_DIR not in sys.path:
    sys.path.append(CODES_DIR)
import excel_utils

# (sheet name in the workbook, exact 'Metrics' label of its Recomm row)
INDICATOR_SHEETS = [
    ("TW ALL", "TW ALL Recomm"),
    ("RSI",    "RSI Recomm"),
    ("ADX",    "ADX Recomm"),
]

# ---------------------------------------------------------------------------
# Excel styling (kept visually consistent with the other sheets)
# ---------------------------------------------------------------------------
FILL_LIME = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")   # BUY CE
FILL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # BUY PE
FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")   # WAIT
FONT_WHITE = Font(color="FFFFFF")
FONT_BLACK = Font(color="000000")
FONT_BOLD = Font(bold=True)


def _load_recomm_table(output_excel_path, sheet_name, recomm_label):
    """Returns {symbol: {time_str: value}} for one indicator sheet's Recomm
    row. Returns {} (rather than raising) if the sheet is missing or
    malformed -- a missing indicator simply contributes no votes, which
    correctly fails confluence for every cell it would have covered
    instead of crashing the whole Final sheet build."""
    try:
        df = pd.read_excel(output_excel_path, sheet_name=sheet_name)
    except Exception as e:
        print(f"[WARNING] Final Sheet (lite): '{sheet_name}' sheet not found ({e}) -- its vote is WAIT everywhere.")
        return {}

    if 'Metrics' not in df.columns or 'Symbol' not in df.columns:
        print(f"[WARNING] Final Sheet (lite): '{sheet_name}' sheet is missing 'Symbol'/'Metrics' columns -- skipped.")
        return {}

    recomm_df = df[df['Metrics'] == recomm_label]
    if recomm_df.empty:
        print(f"[WARNING] Final Sheet (lite): no '{recomm_label}' row found in '{sheet_name}' -- its vote is WAIT everywhere.")
        return {}

    time_cols = [c for c in df.columns if c not in ('Symbol', 'Metrics')]

    table = {}
    for _, row in recomm_df.iterrows():
        sym = str(row['Symbol']).strip().upper()
        table[sym] = {
            str(t): row[t] for t in time_cols
            if pd.notna(row[t]) and str(row[t]).strip() != ""
        }
    return table


def build_final_matrix(output_excel_path):
    """Reads all 3 indicator sheets and builds the pivoted Symbol x Time
    'Final' matrix via strict unanimous vote. No Excel writing here -- see
    write_matrix()."""
    tables = {}
    all_symbols = set()
    all_times = set()

    for sheet_name, recomm_label in INDICATOR_SHEETS:
        table = _load_recomm_table(output_excel_path, sheet_name, recomm_label)
        tables[sheet_name] = table
        for sym, time_map in table.items():
            all_symbols.add(sym)
            all_times.update(time_map.keys())

    if not all_symbols:
        raise RuntimeError(
            "Final Sheet (lite): none of the 3 indicator sheets produced any Recomm rows -- matrix build aborted."
        )

    sorted_times = sorted(all_times)
    matrix_rows = [['Symbol', 'Metrics'] + sorted_times]

    for sym in sorted(all_symbols):
        per_indicator_rows = {
            sheet_name: [sym, recomm_label] for sheet_name, recomm_label in INDICATOR_SHEETS
        }
        final_row = [sym, 'Final Recomm']

        for t in sorted_times:
            votes = []
            for sheet_name, recomm_label in INDICATOR_SHEETS:
                val = tables[sheet_name].get(sym, {}).get(t, 'WAIT')
                per_indicator_rows[sheet_name].append(val)
                votes.append(val)

            if all(v == 'BUY CE' for v in votes):
                final_row.append('BUY CE')
            elif all(v == 'BUY PE' for v in votes):
                final_row.append('BUY PE')
            else:
                final_row.append('WAIT')

        for sheet_name, _ in INDICATOR_SHEETS:
            matrix_rows.append(per_indicator_rows[sheet_name])
        matrix_rows.append(final_row)

    return matrix_rows


def write_matrix(matrix_rows, output_excel_path):
    """Injects matrix_rows into the 'Final' sheet of an existing workbook
    (preserves every other sheet) and autofits + atomically saves. Same
    sheet name / row shape as final_sheet.py's write_matrix() so
    order_sheet.py needs zero changes to read it."""
    wb = load_workbook(output_excel_path)
    sheet_name = "Final"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    for r in matrix_rows:
        ws.append(r)

    ws.freeze_panes = "C2"

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1), start=2):
        metric_type = ws.cell(row=r_idx, column=2).value
        is_final_recomm_row = (metric_type == 'Final Recomm')
        if not is_final_recomm_row:
            continue

        for cell in row[2:]:
            val = cell.value
            if val == "":
                continue
            if val == 'BUY CE':
                cell.fill = FILL_LIME
                cell.font = FONT_BLACK
            elif val == 'BUY PE':
                cell.fill = FILL_RED
                cell.font = FONT_WHITE
            elif val == 'WAIT':
                cell.fill = FILL_GRAY
                cell.font = FONT_BLACK

        row[0].font = FONT_BOLD
        row[1].font = FONT_BOLD

    excel_utils.autofit_columns(ws)
    excel_utils.atomic_save(wb, output_excel_path)


def run_final_sheet_lite_step(output_excel_path):
    """Single entry point for run_pipeline_lite.py / 02_Master_Code_3Indicator.py.
    Must run AFTER TW ALL / RSI / ADX have already been written to
    output_excel_path."""
    print("[SYSTEM] Reading Recomm rows from TW ALL / RSI / ADX for 3-way unanimous confluence check...")
    matrix_rows = build_final_matrix(output_excel_path)
    write_matrix(matrix_rows, output_excel_path)
    print("[SUCCESS] Final Recomm matrix (3-indicator confluence) written to sheet 'Final'.")


# ---------------------------------------------------------------------------
# Disclaimer: 'Final Recomm' here is a mechanical unanimous-vote aggregation
# of 3 heuristic indicators, none of which has been backtested in this
# conversation, individually or in combination. It is not financial advice.
# Paper-trade the full pipeline and run it through scripts/backtester.py
# before risking real capital.
# ---------------------------------------------------------------------------
