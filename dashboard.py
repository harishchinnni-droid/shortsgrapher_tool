"""
dashboard.py
------------
Reads the 'Orders' sheet (after order_sheet.run_order_sheet_step() has
already written/resolved it -- real Net P/L, not a hardcoded zero, see
order_sheet.py's Phase 2 BACKTEST resolution pass and
position_manager.py) and writes a 'Dashboard' sheet summarizing
performance: trade counts, win rate, profit factor, P&L overview,
per-symbol breakdown, exit-reason breakdown, an equity curve with a
chart, and an hourly P&L breakdown.

This is a read-only reporting sheet -- it never feeds back into any
trading decision.

Two layers, kept deliberately separate:
    1. DATA  -- compute_dashboard_stats() and friends compute every
       number purely off the 'Orders' dataframe, no Excel I/O, so they
       can be unit-tested independently.
    2. STYLE -- style_dashboard_sheet() colors the sheet build_dashboard_
       sheet() just wrote (navy title bar, section headers, green/red
       fills on metrics that are meaningfully good/bad, banded tables) --
       pure formatting, it never touches a cell's VALUE. Call it again on
       its own (e.g. after a manual edit) and it re-derives color from
       whatever's currently in the cells; it never accumulates formatting.

Definitions (all computed off the 'Orders' sheet's 'Net P/L (Rs)' column):
    Trades Closed        -- rows with a non-empty 'Exit Time'.
    Trades Open          -- rows with an empty 'Exit Time' (still tracking,
                             LIVE mode only -- a BACKTEST run resolves
                             every position it creates, see order_sheet.py).
    Profitable / Loss / Breakeven -- Net P/L > 0 / < 0 / == 0, among
                             CLOSED trades with a numeric (resolved) P/L.
                             A trade order_sheet.py couldn't resolve (e.g.
                             'NO_HISTORICAL_DATA' with no cost applied) is
                             excluded from win/loss but still counted in
                             Trades Closed -- see 'unresolved_count'.
    Win Rate              -- Profitable / (resolved trades). 0 if none.
    Profit Factor         -- sum(P/L of wins) / abs(sum(P/L of losses)).
                             None ("N/A") if there are no losses to divide by.
    Expectancy per Trade  -- (win_rate x avg_win) - (loss_rate x avg_loss).
    Max Drawdown           -- largest peak-to-trough dip in the chronological
                             cumulative P/L curve (see compute_equity_curve).
"""

import os
import sys
from datetime import datetime

import pandas as pd
import pytz
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CODES_DIR = os.path.dirname(os.path.abspath(__file__))
if CODES_DIR not in sys.path:
    sys.path.append(CODES_DIR)
import excel_utils
import calendar_mgmt

IST = pytz.timezone('Asia/Kolkata')
MARKET_CLOSE_TIME = datetime.strptime("15:30", "%H:%M").time()

FONT_BOLD = Font(bold=True)


# ---------------------------------------------------------------------------
# DATA layer -- pure computation over the Orders dataframe, no Excel I/O.
# ---------------------------------------------------------------------------
def _num(val, default=0.0):
    """Coerce a value that may be '', NaN, None, or already numeric to a
    float, without raising."""
    try:
        if val is None or str(val).strip() == "":
            return default
        if isinstance(val, float) and pd.isna(val):
            return default
        return float(val)
    except (TypeError, ValueError):
        return default


def _closed_mask(df):
    return df['Exit Time'].notna() & (df['Exit Time'].astype(str).str.strip() != "")


def session_status(mode, target_date):
    """Returns (status_label, is_final). BACKTEST is always final (a
    backtest resolves every position it creates in one pass). LIVE is
    final only past market close -- otherwise the numbers are still
    accumulating and should be read as provisional. mode/target_date may
    be None (older callers) -- falls back to a neutral snapshot label
    rather than guessing finality."""
    now = datetime.now(IST)
    if mode == calendar_mgmt.BACKTEST:
        date_str = target_date.strftime('%d-%b-%Y') if target_date else now.strftime('%d-%b-%Y')
        return f"Status: FINAL -- Backtest complete for {date_str}", True

    if mode == calendar_mgmt.LIVE:
        date_str = now.strftime('%d-%b-%Y')
        if now.time() >= MARKET_CLOSE_TIME:
            return f"Status: FINAL -- End of Day ({date_str})", True
        return (f"Status: IN PROGRESS -- as of {now.strftime('%H:%M:%S')} IST "
                f"(rebuilds every cycle; treat ratios below as provisional until EOD)"), False

    return f"Status: Snapshot as of {now.strftime('%d-%b-%Y %H:%M:%S')} IST", False


def compute_equity_curve(valid_pl):
    """valid_pl: the 'closed, resolvable trades' dataframe
    compute_dashboard_stats() builds. Returns (curve_rows, max_drawdown)
    where curve_rows is a chronological (by Exit Time) list of {time,
    symbol, trade_pl, cumulative_pl}, and max_drawdown is the largest
    peak-to-trough dip in cumulative P/L seen along that curve (0 if
    never underwater)."""
    if valid_pl.empty:
        return [], 0.0

    ordered = valid_pl.sort_values('Exit Time')
    cumulative = 0.0
    peak = 0.0
    max_dd = 0.0
    curve = []
    for _, row in ordered.iterrows():
        trade_pl = row['Net P/L (Rs)']
        cumulative += trade_pl
        peak = max(peak, cumulative)
        max_dd = max(max_dd, peak - cumulative)
        curve.append({
            'time': row['Exit Time'], 'symbol': row['Symbol'],
            'trade_pl': round(trade_pl, 2), 'cumulative_pl': round(cumulative, 2),
        })
    return curve, round(max_dd, 2)


def compute_hourly_breakdown(valid_pl):
    """Groups resolved trades by the HOUR their exit fired (Exit Time is
    'HH:MM:SS'), so a session can be read back afterwards for which part
    of the day actually produced the edge. Returns a list sorted by hour."""
    if valid_pl.empty:
        return []

    hours = valid_pl['Exit Time'].astype(str).str.slice(0, 2)
    out = []
    for hour, grp in valid_pl.groupby(hours):
        out.append({
            'hour': f"{hour}:00", 'trades': len(grp),
            'wins': int((grp['Net P/L (Rs)'] > 0).sum()),
            'total_pl': round(grp['Net P/L (Rs)'].sum(), 2),
        })
    out.sort(key=lambda r: r['hour'])
    return out


def compute_dashboard_stats(df_orders):
    """Pure computation over the Orders dataframe -- no Excel I/O -- so
    it can be unit-tested / reused independently of write_dashboard().
    Returns None if there are no orders at all."""
    if df_orders is None or df_orders.empty:
        return None

    df = df_orders.copy()
    total_trades = len(df)
    closed = df[_closed_mask(df)].copy()
    trades_closed = len(closed)
    trades_open = total_trades - trades_closed

    closed['Net P/L (Rs)'] = closed.get('Net P/L (Rs)', pd.Series(dtype=float)).apply(lambda v: _num(v, default=None))
    valid_pl = closed[closed['Net P/L (Rs)'].notna()].copy()
    n_valid = len(valid_pl)

    wins = valid_pl[valid_pl['Net P/L (Rs)'] > 0]
    losses = valid_pl[valid_pl['Net P/L (Rs)'] < 0]
    n_wins, n_losses = len(wins), len(losses)
    n_be = n_valid - n_wins - n_losses

    win_rate = (n_wins / n_valid * 100) if n_valid > 0 else 0.0
    gross_win_sum = wins['Net P/L (Rs)'].sum() if n_wins else 0.0
    gross_loss_sum = abs(losses['Net P/L (Rs)'].sum()) if n_losses else 0.0
    profit_factor = round(gross_win_sum / gross_loss_sum, 2) if gross_loss_sum > 0 else None

    total_capital_deployed = (
        df.get('Entry LTP', pd.Series(dtype=float)).apply(_num)
        * df.get('Quantity (Units)', pd.Series(dtype=float)).apply(_num)
    ).sum()
    total_pl = valid_pl['Net P/L (Rs)'].sum() if n_valid else 0.0

    max_single_profit = wins['Net P/L (Rs)'].max() if n_wins else 0.0
    max_single_loss = losses['Net P/L (Rs)'].min() if n_losses else 0.0
    best_symbol = wins.loc[wins['Net P/L (Rs)'].idxmax(), 'Symbol'] if n_wins else "N/A"
    worst_symbol = losses.loc[losses['Net P/L (Rs)'].idxmin(), 'Symbol'] if n_losses else "N/A"

    avg_win = wins['Net P/L (Rs)'].mean() if n_wins else 0.0
    avg_loss = abs(losses['Net P/L (Rs)'].mean()) if n_losses else 0.0
    loss_rate = (n_losses / n_valid) if n_valid > 0 else 0.0
    expectancy = (win_rate / 100 * avg_win) - (loss_rate * avg_loss)

    per_symbol = []
    for sym, grp in valid_pl.groupby('Symbol'):
        per_symbol.append({
            'Symbol': sym, 'Trades': len(grp),
            'Wins': int((grp['Net P/L (Rs)'] > 0).sum()), 'Losses': int((grp['Net P/L (Rs)'] < 0).sum()),
            'Total P/L (Rs)': round(grp['Net P/L (Rs)'].sum(), 2),
            'Avg P/L (Rs)': round(grp['Net P/L (Rs)'].mean(), 2),
        })
    per_symbol.sort(key=lambda r: r['Symbol'])

    exit_dist = []
    if 'Exit Reason' in closed.columns:
        for reason, grp in closed.groupby('Exit Reason'):
            exit_dist.append({
                'Exit Reason': reason, 'Count': len(grp),
                'Total P/L (Rs)': round(grp['Net P/L (Rs)'].fillna(0).sum(), 2),
            })
        exit_dist.sort(key=lambda r: -r['Count'])

    equity_curve, max_drawdown = compute_equity_curve(valid_pl)
    hourly_breakdown = compute_hourly_breakdown(valid_pl)

    return {
        'total_trades': total_trades, 'trades_closed': trades_closed, 'trades_open': trades_open,
        'wins': n_wins, 'losses': n_losses, 'breakeven': n_be,
        'win_rate': win_rate, 'profit_factor': profit_factor,
        'total_capital_deployed': round(total_capital_deployed, 2), 'total_pl': round(total_pl, 2),
        'max_single_profit': round(max_single_profit, 2), 'best_symbol': best_symbol,
        'max_single_loss': round(max_single_loss, 2), 'worst_symbol': worst_symbol,
        'avg_win': round(avg_win, 2), 'avg_loss': round(avg_loss, 2), 'expectancy': round(expectancy, 2),
        'per_symbol': per_symbol, 'exit_dist': exit_dist,
        'unresolved_count': trades_closed - n_valid,
        'equity_curve': equity_curve, 'max_drawdown': max_drawdown,
        'hourly_breakdown': hourly_breakdown,
    }


# ---------------------------------------------------------------------------
# WRITE -- data layer: plain values, one row per metric/data point. Kept
# free of any fill/font decisions -- style_dashboard_sheet() below is the
# only place color gets decided, so the two layers can't drift apart.
# ---------------------------------------------------------------------------
def build_dashboard_sheet(output_excel_path, mode=None, target_date=None):
    try:
        df_orders = pd.read_excel(output_excel_path, sheet_name='Orders')
    except Exception as e:
        print(f"[WARNING] dashboard: could not read 'Orders' sheet ({e}) -- Dashboard not rebuilt.")
        return

    stats = compute_dashboard_stats(df_orders)
    status_label, is_final = session_status(mode, target_date)

    wb = load_workbook(output_excel_path)
    sheet_name = 'Dashboard'
    if sheet_name in wb.sheetnames:
        idx = wb.sheetnames.index(sheet_name)
        del wb[sheet_name]
        ws = wb.create_sheet(sheet_name, idx)
    else:
        ws = wb.create_sheet(sheet_name, 0)

    date_label = datetime.now(IST).strftime('%d %b %Y')
    ws.append([None, f"F&O TRADING DASHBOARD  --  {date_label}"])
    ws['B1'].font = Font(bold=True, size=14)
    ws.append([None, status_label])
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True, color=("2E7D32" if is_final else "B36B00"))

    if stats is None:
        ws.append([None, "No orders recorded for this run."])
        excel_utils.autofit_columns(ws)
        excel_utils.atomic_save(wb, output_excel_path)
        return

    pf_display = f"{stats['profit_factor']:.2f}" if stats['profit_factor'] is not None else "N/A"

    ws.append([])
    ws.append([None, "TRADE SUMMARY", None, None, "P/L OVERVIEW"])
    ws.append([None, "Total Trades Taken", stats['total_trades'], None, "Total Capital Deployed (Rs)", stats['total_capital_deployed']])
    ws.append([None, "Trades Closed", stats['trades_closed'], None, "Total Net P/L (Rs)", stats['total_pl']])
    ws.append([None, "Trades Open", stats['trades_open'], None, "Max Single Profit (Rs)", stats['max_single_profit']])
    ws.append([None, "Profitable Trades", stats['wins'], None, "Best Trade Symbol", stats['best_symbol']])
    ws.append([None, "Loss Trades", stats['losses'], None, "Max Single Loss (Rs)", stats['max_single_loss']])
    ws.append([None, "Breakeven Trades", stats['breakeven'], None, "Worst Trade Symbol", stats['worst_symbol']])
    ws.append([None, "Win Rate", f"{stats['win_rate']:.1f}%", None, "Avg Profit per Win (Rs)", stats['avg_win']])
    ws.append([None, "Profit Factor", pf_display, None, "Avg Loss per Loss (Rs)", stats['avg_loss']])
    ws.append([None, None, None, None, "Expectancy per Trade (Rs)", stats['expectancy']])
    ws.append([None, None, None, None, "Max Drawdown (Rs)", stats['max_drawdown']])

    for r in range(3, 13):
        ws.cell(row=r, column=2).font = FONT_BOLD
        ws.cell(row=r, column=5).font = FONT_BOLD

    ws.append([])
    n_valid = stats['wins'] + stats['losses'] + stats['breakeven']
    note = (f"Historical backtest result only -- not a guarantee of future performance. "
            f"Sample size: {n_valid} resolved trade(s).")
    if n_valid < 30:
        note += " Below the ~30-trade threshold generally treated as statistically meaningful -- read every ratio above as provisional."
    if stats['unresolved_count']:
        note += f" {stats['unresolved_count']} closed trade(s) could not be resolved (no historical data) and are excluded from win/loss."
    ws.append([None, note])
    ws.cell(row=ws.max_row, column=2).font = Font(italic=True, color="808080")

    ws.append([])
    ws.append([None, "PER-SYMBOL PERFORMANCE"])
    ws.cell(row=ws.max_row, column=2).font = FONT_BOLD
    ws.append([None, "Symbol", "Trades", "Wins", "Losses", "Total P/L (Rs)", "Avg P/L (Rs)"])
    for c in range(2, 8):
        ws.cell(row=ws.max_row, column=c).font = FONT_BOLD
    for row in stats['per_symbol']:
        ws.append([None, row['Symbol'], row['Trades'], row['Wins'], row['Losses'],
                   row['Total P/L (Rs)'], row['Avg P/L (Rs)']])

    ws.append([])
    ws.append([None, "EXIT REASON DISTRIBUTION"])
    ws.cell(row=ws.max_row, column=2).font = FONT_BOLD
    ws.append([None, "Exit Reason", "Count", "Total P/L (Rs)"])
    for c in range(2, 5):
        ws.cell(row=ws.max_row, column=c).font = FONT_BOLD
    for row in stats['exit_dist']:
        ws.append([None, row['Exit Reason'], row['Count'], row['Total P/L (Rs)']])

    # --- Equity curve: chronological cumulative P/L, for reading back HOW
    # the session's total was reached, not just what it totaled. Table
    # always written; a line chart is added whenever there are at least
    # 2 points to draw a line between.
    ws.append([])
    ws.append([None, "EQUITY CURVE (chronological, resolved trades only)"])
    ws.cell(row=ws.max_row, column=2).font = FONT_BOLD
    equity_header_row = ws.max_row + 1
    ws.append([None, "#", "Exit Time", "Symbol", "Trade P/L (Rs)", "Cumulative P/L (Rs)"])
    for c in range(2, 7):
        ws.cell(row=ws.max_row, column=c).font = FONT_BOLD
    equity_first_data_row = ws.max_row + 1
    for i, point in enumerate(stats['equity_curve'], start=1):
        ws.append([None, i, point['time'], point['symbol'], point['trade_pl'], point['cumulative_pl']])
    equity_last_data_row = ws.max_row

    if len(stats['equity_curve']) >= 2:
        chart = LineChart()
        chart.title = "Cumulative P/L"
        chart.y_axis.title = "Rs"
        chart.x_axis.title = "Trade #"
        chart.width, chart.height = 22, 9
        data_ref = Reference(ws, min_col=6, min_row=equity_header_row, max_row=equity_last_data_row)
        cats_ref = Reference(ws, min_col=2, min_row=equity_first_data_row, max_row=equity_last_data_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, f"I{equity_header_row}")

    ws.append([])
    ws.append([None, "HOURLY P&L BREAKDOWN"])
    ws.cell(row=ws.max_row, column=2).font = FONT_BOLD
    ws.append([None, "Hour", "Trades", "Wins", "Total P/L (Rs)"])
    for c in range(2, 6):
        ws.cell(row=ws.max_row, column=c).font = FONT_BOLD
    for row in stats['hourly_breakdown']:
        ws.append([None, row['hour'], row['trades'], row['wins'], row['total_pl']])

    style_dashboard_sheet(ws)
    excel_utils.autofit_columns(ws)
    excel_utils.atomic_save(wb, output_excel_path)


# ---------------------------------------------------------------------------
# STYLE layer -- pure formatting over an already-populated Dashboard
# worksheet. Never touches a cell's value, only fill/font/border/
# alignment, so it's safe to call again on its own (e.g. after a manual
# edit) -- it always re-derives color from whatever's currently in the
# cells rather than accumulating formatting.
# ---------------------------------------------------------------------------
NAVY = "1F3864"
SECTION_BLUE = "2F5596"
TABLE_HEAD_BLUE = "4472C4"
STATUS_BAND = "D9E2F3"
GREEN = "1E9E4C"
RED = "D33B2C"
BAND_LIGHT = "FFFFFF"
BAND_DARK = "F2F2F2"
BORDER_COLOR = "D9D9D9"

FONT_TITLE = Font(color="FFFFFF", bold=True, size=14)
FONT_SECTION = Font(color="FFFFFF", bold=True, size=11)
FONT_TABLE_HEAD = Font(color="FFFFFF", bold=True)
FONT_VALUE_ON_FILL = Font(color="FFFFFF", bold=True)
FONT_GREEN_TEXT = Font(color=GREEN, bold=True)
FONT_RED_TEXT = Font(color=RED, bold=True)

THIN = Side(style="thin", color=BORDER_COLOR)
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")

SECTION_TITLES = ("TRADE SUMMARY", "P/L OVERVIEW", "PER-SYMBOL PERFORMANCE",
                   "EXIT REASON DISTRIBUTION", "EQUITY CURVE", "HOURLY P&L BREAKDOWN")
THRESHOLD_RULES = (
    ("win rate", lambda v: v >= 50),
    ("profit factor", lambda v: v >= 1),
    ("net p/l", lambda v: v >= 0),
    ("total p/l", lambda v: v >= 0),
    ("expectancy", lambda v: v >= 0),
)
GOOD_LABEL_KEYWORDS = ("profitable trades", "max single profit", "best trade", "avg profit")
BAD_LABEL_KEYWORDS = ("loss trades", "max single loss", "worst trade", "avg loss", "max drawdown")


def _clean(text):
    return str(text).strip().upper() if text is not None else ""


def _parse_numeric(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace("Rs.", "").replace("Rs", "").replace(",", "").replace("%", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _classify_kpi_label(label_text):
    label = str(label_text).strip().lower()
    for keyword, is_good_fn in THRESHOLD_RULES:
        if keyword in label:
            return "__threshold__", is_good_fn
    if any(k in label for k in GOOD_LABEL_KEYWORDS):
        return "good", None
    if any(k in label for k in BAD_LABEL_KEYWORDS):
        return "bad", None
    return None, None


def _find_last_used_column(ws, min_col=2):
    last_col = min_col
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, "") and cell.column > last_col:
                last_col = cell.column
    return last_col


def _find_section_headers(ws, last_col):
    found = []
    for row in ws.iter_rows():
        for cell in row:
            text = _clean(cell.value)
            if not text:
                continue
            for title in SECTION_TITLES:
                if title in text:
                    found.append({"row": cell.row, "start_col": cell.column, "title": title})
                    break
    found.sort(key=lambda h: (h["row"], h["start_col"]))
    for h in found:
        same_row_next = [o for o in found if o["row"] == h["row"] and o["start_col"] > h["start_col"]]
        h["end_col"] = min(o["start_col"] for o in same_row_next) - 1 if same_row_next else last_col
    return found


def _section_row_extent(ws, header_row, start_col, end_col, max_row):
    last_row = header_row
    for r in range(header_row + 1, max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(start_col, end_col + 1)]
        if all(v in (None, "") for v in row_vals):
            break
        first_cell_text = _clean(ws.cell(r, start_col).value)
        if any(title in first_cell_text for title in SECTION_TITLES) and r != header_row:
            break
        last_row = r
    return last_row


def _style_title_and_status(ws, last_col):
    last_col_letter = get_column_letter(last_col)
    if ws.cell(1, 2).value not in (None, ""):
        ws.merge_cells(f"B1:{last_col_letter}1")
        ws.row_dimensions[1].height = 28
        title_cell = ws.cell(1, 2)
        title_cell.font = FONT_TITLE
        title_cell.alignment = CENTER
        for c in range(2, last_col + 1):
            ws.cell(1, c).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

    if ws.cell(2, 2).value not in (None, ""):
        ws.merge_cells(f"B2:{last_col_letter}2")
        status_cell = ws.cell(2, 2)
        status_cell.alignment = CENTER
        existing_font = status_cell.font
        status_cell.font = Font(bold=True, italic=True, color=(existing_font.color.rgb if existing_font.color else "44546A"))
        for c in range(2, last_col + 1):
            ws.cell(2, c).fill = PatternFill(start_color=STATUS_BAND, end_color=STATUS_BAND, fill_type="solid")


def _style_section_header(ws, header):
    row, start_col, end_col = header["row"], header["start_col"], header["end_col"]
    start_letter, end_letter = get_column_letter(start_col), get_column_letter(end_col)
    if end_col > start_col:
        ws.merge_cells(f"{start_letter}{row}:{end_letter}{row}")
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row, c)
        cell.fill = PatternFill(start_color=SECTION_BLUE, end_color=SECTION_BLUE, fill_type="solid")
        cell.font = FONT_SECTION
    ws.cell(row, start_col).alignment = LEFT
    ws.row_dimensions[row].height = 20


def _style_kpi_panel(ws, header):
    start_col, end_col = header["start_col"], header["end_col"]
    value_col = start_col + 1
    last_row = _section_row_extent(ws, header["row"], start_col, end_col, ws.max_row)

    band_idx = 0
    for r in range(header["row"] + 1, last_row + 1):
        label_cell = ws.cell(r, start_col)
        value_cell = ws.cell(r, value_col)
        if label_cell.value in (None, ""):
            continue

        band = BAND_LIGHT if band_idx % 2 == 0 else BAND_DARK
        band_idx += 1
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            cell.fill = PatternFill(start_color=band, end_color=band, fill_type="solid")
            cell.border = CELL_BORDER

        label_cell.alignment = LEFT

        classification, threshold_fn = _classify_kpi_label(label_cell.value)
        numeric_val = _parse_numeric(value_cell.value)

        fill_color = None
        if classification == "__threshold__" and numeric_val is not None:
            fill_color = GREEN if threshold_fn(numeric_val) else RED
        elif classification == "good":
            fill_color = GREEN
        elif classification == "bad":
            fill_color = RED

        if fill_color:
            value_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            value_cell.font = FONT_VALUE_ON_FILL
        else:
            value_cell.font = Font(color="000000", bold=True)
        value_cell.alignment = RIGHT


def _style_wide_table(ws, header):
    start_col, end_col = header["start_col"], header["end_col"]
    header_row = header["row"] + 1
    last_row = _section_row_extent(ws, header["row"], start_col, end_col, ws.max_row)
    if header_row > last_row:
        return

    pl_cols = set()
    for c in range(start_col, end_col + 1):
        cell = ws.cell(header_row, c)
        cell.fill = PatternFill(start_color=TABLE_HEAD_BLUE, end_color=TABLE_HEAD_BLUE, fill_type="solid")
        cell.font = FONT_TABLE_HEAD
        cell.alignment = CENTER
        cell.border = CELL_BORDER
        if "P/L" in _clean(cell.value):
            pl_cols.add(c)
    ws.row_dimensions[header_row].height = 18

    band_idx = 0
    for r in range(header_row + 1, last_row + 1):
        if all(ws.cell(r, c).value in (None, "") for c in range(start_col, end_col + 1)):
            continue
        band = BAND_LIGHT if band_idx % 2 == 0 else BAND_DARK
        band_idx += 1
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            cell.fill = PatternFill(start_color=band, end_color=band, fill_type="solid")
            cell.border = CELL_BORDER
            if c in pl_cols:
                numeric_val = _parse_numeric(cell.value)
                cell.font = FONT_GREEN_TEXT if (numeric_val is None or numeric_val >= 0) else FONT_RED_TEXT
                cell.alignment = RIGHT
            elif isinstance(cell.value, (int, float)):
                cell.font = Font(color="000000")
                cell.alignment = RIGHT
            else:
                cell.font = Font(color="000000")
                cell.alignment = LEFT if c == start_col else CENTER


def style_dashboard_sheet(ws):
    """Applies the reference color-coded look to an already-populated
    Dashboard worksheet, in place. Does not touch any cell VALUE."""
    ws.column_dimensions['A'].width = 3
    last_col = _find_last_used_column(ws)
    _style_title_and_status(ws, last_col)
    for header in _find_section_headers(ws, last_col):
        _style_section_header(ws, header)
        if header["title"] in ("TRADE SUMMARY", "P/L OVERVIEW"):
            _style_kpi_panel(ws, header)
        else:
            _style_wide_table(ws, header)
    ws.sheet_view.showGridLines = False


def run_dashboard_style_step(output_excel_path, sheet_name="Dashboard"):
    """Standalone entry point to (re)style an already-saved workbook's
    Dashboard sheet on disk, without recomputing any values."""
    wb = load_workbook(output_excel_path)
    if sheet_name not in wb.sheetnames:
        print(f"[WARNING] dashboard.py: '{sheet_name}' sheet not found -- nothing to style.")
        return
    style_dashboard_sheet(wb[sheet_name])
    excel_utils.atomic_save(wb, output_excel_path)
    print(f"[SUCCESS] '{sheet_name}' sheet styled.")


# ---------------------------------------------------------------------------
# Disclaimer: this sheet reports HISTORICAL results only (or, in LIVE
# mode, results still accumulating through the session -- see the status
# badge). Nothing here is a guarantee of future performance, and win
# rate / profit factor / expectancy on a small sample should be read as
# provisional (see the note appended under P/L Overview).
# ---------------------------------------------------------------------------
