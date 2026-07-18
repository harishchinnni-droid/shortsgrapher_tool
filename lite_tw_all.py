"""
lite_tw_all.py
--------------
NEW MODULE -- part of the 3-indicator standalone pipeline (Task 52,
18-Jul-26). Harish's spec: "only if the candle is above N-Line, and above
BLUE Ribbon, then BUY CE; if candle is below N-Line and RED Ribbon then
BUY PE; else WAIT."

Does NOT modify or replace tw_all.py -- that module (and its own, different
"flip-and-hold on Hull crossover" TW ALL Recomm logic) stays exactly as-is
for the existing full pipeline (01_Master_Code.py). This is a separate
sheet-producing module for the lean 3-indicator pipeline only
(run_pipeline_lite.py / 02_Master_Code_3Indicator.py), writing to a
freshly-provisioned workbook that never contains tw_all.py's own sheet, so
there is no naming collision in practice.

Indicator math (reuses the same Hull/EHMA construction as tw_all.py --
see that module's docstring for the EHMA translation note, reproduced
identically here so both modules plot the same line the chart does):
    HULL = EHMA(close, 16)   (Ehma is the only mode the uploaded Pine
                               script actually uses -- modeSwitch is
                               hardcoded, not exposed as an input)
    MHULL = HULL (current bar)
    SHULL = HULL, 2 bars ago
    TREND = 'Bullish' if HULL > HULL[2] else 'Bearish'  (mirrors the blue/
            red ribbon coloring: switchColor ? (HULL > HULL[2] ? blue : red))
    N-Line = EMA(close, 100)  (ema100 in the Pine script, purple line)

Per-bar 'TW ALL Recomm' (per-bar, NOT hold-based -- Harish's spec is a
state condition on THIS candle, not an event/crossover):
    BUY CE if close > N-Line AND close > MHULL AND close > SHULL AND
             TREND == 'Bullish'
    BUY PE if close < N-Line AND close < MHULL AND close < SHULL AND
             TREND == 'Bearish'
    WAIT   otherwise

Timeframe: 5-minute candles, truncated per day to CUTOFF_TIME, same
convention as every other module in this pipeline.

Disclaimer: this is a new, unbacktested rule -- not financial advice, no
result here is a guarantee of future performance. Paper-trade and backtest
before risking real capital.
"""

import os
import sys
import concurrent.futures
from datetime import time as dtime

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

CODES_DIR = os.path.dirname(os.path.abspath(__file__))
if CODES_DIR not in sys.path:
    sys.path.append(CODES_DIR)
import data_ingestion
import excel_utils

# ---------------------------------------------------------------------------
# Config -- mirrors the Pine script's default inputs (same as tw_all.py)
# ---------------------------------------------------------------------------
LENGTH = 16
CUTOFF_TIME = dtime(15, 15)
INTERVAL = "5minute"
NLINE_LENGTH = 100

# ---------------------------------------------------------------------------
# Excel styling
# ---------------------------------------------------------------------------
FILL_BULLISH = PatternFill(start_color="26A69A", end_color="26A69A", fill_type="solid")
FILL_BEARISH = PatternFill(start_color="EF5350", end_color="EF5350", fill_type="solid")
FILL_LIME = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")   # BUY CE
FILL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # BUY PE
FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")   # WAIT
FONT_WHITE = Font(color="FFFFFF")
FONT_BLACK = Font(color="000000")


# ---------------------------------------------------------------------------
# Hull EHMA (identical construction to tw_all.py's _ehma -- see that
# module's docstring for why this literal translation is a double-EMA
# smooth, not textbook Hull, and is reproduced as-is to match the chart)
# ---------------------------------------------------------------------------
def _ehma(src, length):
    inner = src.ewm(span=length, adjust=False).mean()
    smooth_len = max(int(round(np.sqrt(length))), 1)
    return inner.ewm(span=smooth_len, adjust=False).mean()


# ---------------------------------------------------------------------------
# Indicator
# ---------------------------------------------------------------------------
def calculate_tw_all_lite(df, length=LENGTH, nline_length=NLINE_LENGTH):
    df = df.copy()
    hull = _ehma(df['close'], length)
    mhull = hull
    shull = hull.shift(2)

    df['HULL'] = hull
    df['MHULL'] = mhull
    df['SHULL'] = shull
    df['N-Line'] = df['close'].ewm(span=nline_length, adjust=False).mean()
    df['TREND'] = np.where(hull > hull.shift(2), 'Bullish', 'Bearish')

    close = df['close']
    buy_ce = (close > df['N-Line']) & (close > mhull) & (close > shull) & (df['TREND'] == 'Bullish')
    buy_pe = (close < df['N-Line']) & (close < mhull) & (close < shull) & (df['TREND'] == 'Bearish')

    df['TW ALL Recomm'] = np.where(buy_ce, 'BUY CE', np.where(buy_pe, 'BUY PE', 'WAIT'))
    return df


# ---------------------------------------------------------------------------
# Per-symbol processing (same brute-force time/timezone protocol as
# tw_all.py / RSI.py)
# ---------------------------------------------------------------------------
def process_symbol(symbol_data):
    symbol, df, target_date = symbol_data
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip().str.lower()

    if not all(col in df.columns for col in ['open', 'high', 'low', 'close']):
        print(f"[WARNING] {symbol}: Missing required price columns. Discarding.")
        return symbol, None

    time_col = next((col for col in ['date', 'time', 'datetime', 'timestamp'] if col in df.columns), None)
    if time_col:
        parsed = pd.to_datetime(df[time_col], errors='coerce')
    elif pd.api.types.is_datetime64_any_dtype(df.index):
        parsed = pd.Series(df.index, index=df.index)
    else:
        unnamed_cols = [c for c in df.columns if 'unnamed' in c]
        if unnamed_cols:
            parsed = pd.to_datetime(df[unnamed_cols[0]], errors='coerce')
        else:
            print(f"[WARNING] {symbol}: Failed to locate any valid timestamp data. Discarding.")
            return symbol, None

    try:
        if parsed.dt.tz is not None:
            parsed = parsed.dt.tz_localize(None)
    except (AttributeError, TypeError):
        pass

    df['_sort_dt'] = parsed
    df['time_str'] = parsed.dt.strftime('%H:%M')
    df.dropna(subset=['time_str', '_sort_dt'], inplace=True)
    if df.empty:
        print(f"[WARNING] {symbol}: No rows with a valid timestamp. Discarding.")
        return symbol, None

    df.sort_values('_sort_dt', inplace=True)
    df = df[df['_sort_dt'].dt.time <= CUTOFF_TIME]
    if df.empty:
        print(f"[WARNING] {symbol}: No candles at/before {CUTOFF_TIME.strftime('%H:%M')}. Discarding.")
        return symbol, None

    df = calculate_tw_all_lite(df)

    df = excel_utils.restrict_to_target_date(df, target_date)
    if df is None:
        return symbol, None

    return symbol, df


# ---------------------------------------------------------------------------
# Excel export -- pivoted Symbol x Time matrix
# ---------------------------------------------------------------------------
def build_matrix(data_dict, target_date, max_workers=None):
    results = {}
    with concurrent.futures.ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(process_symbol, (sym, df, target_date)): sym for sym, df in data_dict.items()}
        for future in concurrent.futures.as_completed(futures):
            sym = futures[future]
            try:
                processed_sym, processed_df = future.result()
                if processed_df is not None and 'HULL' in processed_df.columns:
                    processed_df = processed_df.dropna(subset=['HULL'])
                    if not processed_df.empty and sym.lower() != 'summary':
                        results[sym] = processed_df
            except Exception as e:
                print(f"[ERROR] TW ALL (lite) Engine Failure on symbol {sym}: {str(e)}")

    if not results:
        raise RuntimeError("TW ALL (lite): zero symbols processed successfully for target_date -- matrix build aborted.")

    all_times = set()
    for df in results.values():
        if 'time_str' in df.columns:
            all_times.update(df['time_str'].dropna().tolist())

    if not all_times:
        raise RuntimeError("TW ALL (lite): no valid timestamps extracted across all symbols -- matrix build aborted.")

    sorted_times = sorted(all_times)
    matrix_rows = [['Symbol', 'Metrics'] + sorted_times]

    for sym in sorted(results.keys()):
        df = results[sym].drop_duplicates(subset=['time_str'], keep='last')
        df_indexed = df.set_index('time_str')

        def get_metric_row(metric_name, col_name):
            row = [sym, metric_name]
            for t in sorted_times:
                if t in df_indexed.index:
                    val = df_indexed.loc[t, col_name]
                    if isinstance(val, pd.Series):
                        val = val.iloc[-1]
                    row.append(val)
                else:
                    row.append("")
            return row

        matrix_rows.append(get_metric_row('Close', 'close'))
        matrix_rows.append(get_metric_row('N-Line', 'N-Line'))
        matrix_rows.append(get_metric_row('MHULL', 'MHULL'))
        matrix_rows.append(get_metric_row('SHULL', 'SHULL'))
        matrix_rows.append(get_metric_row('TREND', 'TREND'))
        matrix_rows.append(get_metric_row('TW ALL Recomm', 'TW ALL Recomm'))

    return matrix_rows


def write_matrix_to_workbook(matrix_rows, wb):
    sheet_name = "TW ALL"
    ws = excel_utils.replace_sheet_with_matrix(wb, sheet_name, matrix_rows)

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1), start=2):
        metric_type = ws.cell(row=r_idx, column=2).value
        for cell in row[2:]:
            val = cell.value
            if val == "":
                continue
            if metric_type == 'TREND':
                if val == 'Bullish':
                    cell.fill = FILL_BULLISH
                    cell.font = FONT_WHITE
                elif val == 'Bearish':
                    cell.fill = FILL_BEARISH
                    cell.font = FONT_WHITE
            elif metric_type == 'TW ALL Recomm':
                if val == 'BUY CE':
                    cell.fill = FILL_LIME
                    cell.font = FONT_BLACK
                elif val == 'BUY PE':
                    cell.fill = FILL_RED
                    cell.font = FONT_WHITE
                elif val == 'WAIT':
                    cell.fill = FILL_GRAY
                    cell.font = FONT_BLACK

    excel_utils.autofit_columns(ws)


def write_matrix(matrix_rows, output_excel_path):
    wb = load_workbook(output_excel_path)
    write_matrix_to_workbook(matrix_rows, wb)
    excel_utils.atomic_save(wb, output_excel_path)


def parallel_compute_and_export(data_dict, target_date, output_excel_path, max_workers=None):
    matrix_rows = build_matrix(data_dict, target_date, max_workers=max_workers)
    write_matrix(matrix_rows, output_excel_path)


def run_twall_lite_step(df_ref, target_date, output_excel_path):
    """Single entry point for run_pipeline_lite.py."""
    print("[SYSTEM] Loading 5-minute historical data for TW All In One (lite)...")
    data_dict = data_ingestion.load_interval_data(df_ref, target_date, interval=INTERVAL)
    if not data_dict:
        raise RuntimeError("TW ALL (lite): no 5-minute data files found for any symbol.")
    print(f"[PROCESS] Computing TW All In One (lite) matrix for {len(data_dict)} symbol(s)...")
    parallel_compute_and_export(data_dict, target_date, output_excel_path)
    print("[SUCCESS] TW All In One (lite) matrix written to sheet 'TW ALL'.")


# ---------------------------------------------------------------------------
# Disclaimer: this module estimates trend direction from historical/live
# price data only. It is not financial advice, no result here is a
# guarantee of future performance, and this recomm rule has not been
# backtested -- paper-trade it and run it through scripts/backtester.py
# before risking real capital.
# ---------------------------------------------------------------------------
