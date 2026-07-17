"""
zerolag.py
----------
Python translation of the core, genuinely-new pieces of a Pine v6 script
Harish shared ("Zero-Lag Strong Signals [India]"): the Zero-Lag EMA trend
cloud (ZLEMA + ATR band, with hysteresis so a flip requires price to clear
the WHOLE band, not just cross a line) and a Relative Volume (RVOL) check.

Deliberately NOT ported: the Pine script's ADX regime gate, higher-
timeframe EMA trend gate, and RSI momentum gate. This pipeline already
computes all three independently (adx_di.py, htf_bias.py, RSI.py) --
porting them again here would be re-answering the same three questions
under a new name, which is exactly the "one trend opinion asked three
times" mistake final_sheet.py's own docstring already documents and
corrected once before (that's why OBV CMF replaced BRKPRO as a vote).
Only the two pieces this pipeline didn't already have are ported.

Recommendation ('ZL Recomm') -- PERSISTENT per-bar state (BUY CE while
Trend Dir==1, BUY PE while Trend Dir==-1, WAIT before the first flip),
matching every other sheet in this workbook's convention (RSI/ADX/EMA20/
etc. are all "current state, recomputed every bar", not one-shot edge
events). This differs from the source Pine script, whose longSig/shortSig
only fire on the exact bar of a flip (an edge-triggered event) -- that
distinction matters for how this gets USED, not shown, see below.

Integration -- a GATE, not a vote:
    'ZL Recomm' is NOT added to final_sheet.py's confluence vote. Same
    reasoning as ADX's own exclusion (see final_sheet.py's docstring):
    this is used as a pre-entry STRENGTH gate in order_sheet.py instead
    (alongside the existing ADX_MIN / OI-buildup / PCR-trap gates) --
    voting AND gating on the same computation would double-count it.
    The gate checks, at the signal's own pre-entry bar:
        1. Trend Dir agrees with the Final Recomm signal direction
           (price is on the correct side of the zero-lag cloud right now)
        2. RVOL at that bar clears RVOL_MIN (real volume behind the move,
           not a thin/dead candle)
    See order_sheet.py's ENABLE_ZEROLAG_GATE flag -- off by default,
    same A/B pattern as every other experimental gate in this codebase.

Known limitation (same one breakout_probability.py already documents):
    load_interval_data() only loads ONE trading day's 5-minute candles per
    symbol (~75 bars, 09:15-15:15), not a multi-day rolling window. The
    Zero-Lag EMA's default 34-length and the ATR's 14-length warm-up eats
    into a meaningful chunk of that (~16-34 bars) before either stabilizes
    -- expect 'Trend Dir'/'RVOL' to be WAIT/blank for roughly the first
    30-45 minutes of every session. This is the same constraint every
    other indicator in this pipeline already lives with, not something
    new introduced here.

Timeframe: 5-minute candles, truncated per day to CUTOFF_TIME (15:15),
same as every other module in this pipeline.
"""

import os
import sys
import concurrent.futures
from datetime import time as dtime

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

CODES_DIR = os.path.dirname(os.path.abspath(__file__))
if CODES_DIR not in sys.path:
    sys.path.append(CODES_DIR)
import data_ingestion
import excel_utils

# ---------------------------------------------------------------------------
# Config -- mirrors the Pine script's own input defaults
# ---------------------------------------------------------------------------
ZL_LEN = 34          # `zlLen` -- Zero-Lag EMA length
ATR_LEN = 14         # `atrLen`
BAND_MULT = 1.2      # `bandMult` -- cloud width, x ATR
VOL_LEN = 20         # `volLen` -- RVOL average window
RVOL_MIN = 1.5       # `rvolMin` -- used by order_sheet.py's gate, not here
CUTOFF_TIME = dtime(15, 15)
INTERVAL = "5minute"

# ---------------------------------------------------------------------------
# Excel styling (kept visually consistent with the rest of the workbook)
# ---------------------------------------------------------------------------
FILL_LIME = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")   # BUY CE
FILL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # BUY PE
FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")   # WAIT
FONT_WHITE = Font(color="FFFFFF")
FONT_BLACK = Font(color="000000")


# ---------------------------------------------------------------------------
# Indicator -- ZLEMA/ATR cloud with hysteresis + RVOL
# ---------------------------------------------------------------------------
def _wilder_rma(series: pd.Series, period: int) -> np.ndarray:
    """Pine's ta.rma (what ta.atr() is built on): first value seeded as an
    SMA of the first `period` bars, then prev + (x - prev)/period from
    there. This is a DIFFERENT recursion from adx_di.py's own
    _wilder_running_sum (that one's a coefficient-1 running sum seeded at
    0, translating a different Pine script's raw DM+/DM-/TR smoothing,
    not ta.rma) -- don't merge the two, they're not interchangeable."""
    values = series.to_numpy()
    n = len(values)
    out = np.empty(n, dtype=float)
    if n == 0:
        return out
    seed_len = min(period, n)
    seed = float(np.mean(values[:seed_len]))
    out[:seed_len] = seed
    prev = seed
    for i in range(seed_len, n):
        prev = (values[i] - prev) / period + prev
        out[i] = prev
    return out


def calculate_zerolag(df, zl_len=ZL_LEN, atr_len=ATR_LEN, band_mult=BAND_MULT, vol_len=VOL_LEN):
    """Requires 'open','high','low','close' columns (+ 'volume' if
    present, else RVOL is 0 everywhere). Returns df with 'Zero-Lag Line',
    'Trend Dir' (1/-1/0), 'RVOL', 'ZL Recomm' columns added."""
    df = df.copy()
    close_s = df['close'].astype(float)
    high_s = df['high'].astype(float)
    low_s = df['low'].astype(float)
    n = len(df)

    # --- Zero-Lag EMA: adds the source's own lagged momentum back into
    # itself before the EMA smoothing -- literal translation of Pine's
    # `zlBasis = src + (src - src[lag]); zlema = ta.ema(zlBasis, zlLen)`.
    # ta.ema has no SMA seed (unlike ta.rma below) -- pandas .ewm(...,
    # adjust=False) matches it exactly with no special seeding needed.
    lag = (zl_len - 1) // 2
    zl_basis = close_s + (close_s - close_s.shift(lag))
    zlema = zl_basis.ewm(span=zl_len, adjust=False).mean()

    # --- ATR via Wilder's RMA (ta.atr == ta.rma(ta.tr(true), atrLen)) ---
    prev_close = close_s.shift(1)
    true_range = pd.concat([
        high_s - low_s,
        (high_s - prev_close).abs(),
        (low_s - prev_close).abs(),
    ], axis=1).max(axis=1).fillna(0.0)
    atr = pd.Series(_wilder_rma(true_range, atr_len), index=df.index)

    upper = zlema + band_mult * atr
    lower = zlema - band_mult * atr

    # --- Trend direction WITH HYSTERESIS: a plain stateful loop on
    # purpose, same pattern as breakout_probability.py's cumulative
    # counters -- `var int trendDir` in Pine only updates when price
    # clears the WHOLE band, otherwise it holds its last value. Not
    # vectorizable (each bar depends on the previous bar's decision, not
    # just this bar's inputs), but cheap at intraday row counts.
    trend_dir = np.zeros(n, dtype=int)
    state = 0
    zlema_np = zlema.to_numpy()
    upper_np = upper.to_numpy()
    lower_np = lower.to_numpy()
    close_np = close_s.to_numpy()
    for i in range(n):
        if np.isnan(zlema_np[i]) or np.isnan(upper_np[i]) or np.isnan(lower_np[i]):
            trend_dir[i] = state
            continue
        if close_np[i] > upper_np[i]:
            state = 1
        elif close_np[i] < lower_np[i]:
            state = -1
        trend_dir[i] = state

    # --- Relative Volume: current volume vs its own rolling average ---
    if 'volume' in df.columns:
        vol_s = df['volume'].astype(float)
        vol_avg = vol_s.rolling(window=vol_len, min_periods=1).mean()
        rvol = np.where(vol_avg.to_numpy() > 0, (vol_s / vol_avg).to_numpy(), 0.0)
    else:
        rvol = np.zeros(n)

    df['Zero-Lag Line'] = zlema.to_numpy()
    df['Trend Dir'] = trend_dir
    df['RVOL'] = rvol
    df['ZL Recomm'] = np.where(trend_dir == 1, 'BUY CE', np.where(trend_dir == -1, 'BUY PE', 'WAIT'))

    return df


# ---------------------------------------------------------------------------
# Per-symbol processing (same brute-force time/timezone protocol as RSI.py)
# ---------------------------------------------------------------------------
def process_symbol(symbol_data):
    symbol, df, target_date = symbol_data
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip().str.lower()

    if not all(col in df.columns for col in ['open', 'high', 'low', 'close']):
        print(f"[WARNING] {symbol}: Missing required price columns. Discarding.")
        return symbol, None

    time_col = next((col for col in ['date', 'time', 'datetime', 'timestamp'] if col in df.columns), None)
    if time_col:
        parsed = pd.to_datetime(df[time_col], errors='coerce')
    elif pd.api.types.is_datetime64_any_dtype(df.index):
        parsed = pd.Series(df.index, index=df.index)
    else:
        unnamed_cols = [c for c in df.columns if 'unnamed' in c]
        if unnamed_cols:
            parsed = pd.to_datetime(df[unnamed_cols[0]], errors='coerce')
        else:
            print(f"[WARNING] {symbol}: Failed to locate any valid timestamp data. Discarding.")
            return symbol, None

    try:
        if parsed.dt.tz is not None:
            parsed = parsed.dt.tz_localize(None)
    except (AttributeError, TypeError):
        pass

    df['_sort_dt'] = parsed
    df['time_str'] = parsed.dt.strftime('%H:%M')
    df.dropna(subset=['time_str', '_sort_dt'], inplace=True)
    if df.empty:
        print(f"[WARNING] {symbol}: No rows with a valid timestamp. Discarding.")
        return symbol, None

    df.sort_values('_sort_dt', inplace=True)
    df = df[df['_sort_dt'].dt.time <= CUTOFF_TIME]
    if df.empty:
        print(f"[WARNING] {symbol}: No candles at/before {CUTOFF_TIME.strftime('%H:%M')}. Discarding.")
        return symbol, None

    df = calculate_zerolag(df)

    df = excel_utils.restrict_to_target_date(df, target_date)
    if df is None:
        return symbol, None

    return symbol, df


# ---------------------------------------------------------------------------
# Excel export -- pivoted Symbol x Time matrix (same pattern as adx_di.py)
# ---------------------------------------------------------------------------
def build_matrix(data_dict, target_date, max_workers=None):
    results = {}
    with concurrent.futures.ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(process_symbol, (sym, df, target_date)): sym for sym, df in data_dict.items()}
        for future in concurrent.futures.as_completed(futures):
            sym = futures[future]
            try:
                processed_sym, processed_df = future.result()
                if processed_df is not None and 'Zero-Lag Line' in processed_df.columns:
                    processed_df = processed_df.dropna(subset=['Zero-Lag Line'])
                    if not processed_df.empty and sym.lower() != 'summary':
                        results[sym] = processed_df
            except Exception as e:
                print(f"[ERROR] Zero-Lag Engine Failure on symbol {sym}: {str(e)}")

    if not results:
        raise RuntimeError("Zero-Lag: zero symbols processed successfully for target_date -- matrix build aborted.")

    all_times = set()
    for df in results.values():
        if 'time_str' in df.columns:
            all_times.update(df['time_str'].dropna().tolist())

    if not all_times:
        raise RuntimeError("Zero-Lag: no valid timestamps extracted across all symbols -- matrix build aborted.")

    sorted_times = sorted(all_times)
    matrix_rows = [['Symbol', 'Metrics'] + sorted_times]

    for sym in sorted(results.keys()):
        df = results[sym].drop_duplicates(subset=['time_str'], keep='last')
        df_indexed = df.set_index('time_str')

        def get_metric_row(metric_name, col_name):
            row = [sym, metric_name]
            for t in sorted_times:
                if t in df_indexed.index:
                    val = df_indexed.loc[t, col_name]
                    if isinstance(val, pd.Series):
                        val = val.iloc[-1]
                    row.append(val)
                else:
                    row.append("")
            return row

        matrix_rows.append(get_metric_row('Zero-Lag Line', 'Zero-Lag Line'))
        matrix_rows.append(get_metric_row('Trend Dir', 'Trend Dir'))
        matrix_rows.append(get_metric_row('RVOL', 'RVOL'))
        matrix_rows.append(get_metric_row('ZL Recomm', 'ZL Recomm'))

    return matrix_rows


def write_matrix_to_workbook(matrix_rows, wb):
    sheet_name = "ZLTREND"
    ws = excel_utils.replace_sheet_with_matrix(wb, sheet_name, matrix_rows)

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1), start=2):
        metric_type = ws.cell(row=r_idx, column=2).value
        for cell in row[2:]:
            val = cell.value
            if val == "":
                continue
            if metric_type == 'ZL Recomm':
                if val == 'BUY CE':
                    cell.fill = FILL_LIME
                    cell.font = FONT_BLACK
                elif val == 'BUY PE':
                    cell.fill = FILL_RED
                    cell.font = FONT_WHITE
                elif val == 'WAIT':
                    cell.fill = FILL_GRAY
                    cell.font = FONT_BLACK

    excel_utils.autofit_columns(ws)


def write_matrix(matrix_rows, output_excel_path):
    wb = load_workbook(output_excel_path)
    write_matrix_to_workbook(matrix_rows, wb)
    excel_utils.atomic_save(wb, output_excel_path)


def parallel_compute_and_export(data_dict, target_date, output_excel_path, max_workers=None):
    matrix_rows = build_matrix(data_dict, target_date, max_workers=max_workers)
    write_matrix(matrix_rows, output_excel_path)


def run_zerolag_step(df_ref, target_date, output_excel_path):
    print("[SYSTEM] Loading 5-minute historical data for Zero-Lag Trend...")
    data_dict = data_ingestion.load_interval_data(df_ref, target_date, interval=INTERVAL)
    if not data_dict:
        raise RuntimeError("Zero-Lag: no 5-minute data files found for any symbol.")
    print(f"[PROCESS] Computing Zero-Lag Trend matrix for {len(data_dict)} symbol(s)...")
    parallel_compute_and_export(data_dict, target_date, output_excel_path)
    print("[SUCCESS] Zero-Lag Trend matrix written to sheet 'ZLTREND'.")


# ---------------------------------------------------------------------------
# Disclaimer: this module derives a signal from historical/live price data
# only. It is not financial advice, and no result here is a guarantee of
# future performance -- paper-trade it and run it through
# scripts/backtester.py before risking real capital. See order_sheet.py's
# ENABLE_ZEROLAG_GATE for how (and whether) this actually gates entries.
# ---------------------------------------------------------------------------
