"""
tw_all.py
---------
Vectorized Python translation of "TW All in One" Pine Script v4 indicator
(the Hull-MA trend/crossover half of it), wired to run as Step 10 of the
daily pipeline (after ADX & DI). Output format mirrors RSI.py / SQZMOM.py's
pivoted Symbol x Time matrix.

Indicator logic (see 'TW All In One.txt'):
    HULL = a Hull-family moving average of `close`, mode fixed to "Ehma" in
    the uploaded script (modeSwitch is hardcoded, not exposed as an input).
    MHULL = HULL (current bar), SHULL = HULL two bars ago.
    Trend color flips on HULL > HULL[2] (bullish) vs HULL < HULL[2]
    (bearish) -- replicated here as TREND.
    The script's own Buy/Sell shapes fire on:
        Sell  when SHULL crosses OVER  MHULL  (Hull turning down)
        Buy   when SHULL crosses UNDER MHULL  (Hull turning up)
    replicated here as SIGNAL.

    Translation note on EHMA: as literally written in the uploaded script,
    `EHMA(_src, _length) => ema(2 * ema(_src, _length) - ema(_src, _length),
    round(sqrt(_length)))` uses the SAME `_length` in both inner ema() calls
    (unlike the HMA() function two lines above it, which correctly halves
    the length for the fast term). Algebraically 2*X - X = X, so this
    reduces to `ema(ema(close, length), round(sqrt(length)))` -- a
    double-smoothed EMA, not a true Hull MA. That's reproduced exactly as-is
    here (not "corrected") so this matches what the script actually plots
    on the chart; `_hma()` and `_thma()` are included below, correctly
    formed, in case you want to switch to genuine Hull smoothing later
    (set MODE = "Hma" to A/B it -- see the strategic review for why this
    is worth testing).

    Scope note: the original script also plots up to 14 pivot-based
    support/resistance "Target Stoploss" lines (left=33/right=21 pivots).
    Those are chart-only visual reference lines the script itself never
    reads back into the Buy/Sell decision, so they are intentionally not
    reproduced here -- only the Hull trend + crossover signal that actually
    drives the script's own alerts is.

Discrete 'TW ALL Recomm' column (NOT part of the original indicator -- a
mapping defined here per your request): BUY CE on every Buy crossover,
BUY PE on every Sell crossover, held on every subsequent bar until the next
opposite crossover (crossovers are naturally infrequent, so this is a
straightforward flip-and-hold, no extra confirmation added).
This is a heuristic, not a validated edge -- backtest it
(scripts/backtester.py) before trusting it with real capital.

Timeframe: 5-minute candles, truncated per day to CUTOFF_TIME, same
convention as RSI.py. Timezone stripped before writing to Excel for the
same reason documented there (openpyxl can't store tz-aware datetimes).

[FIX -- 13-Jul-26] process_symbol()/build_matrix() now require target_date
and restrict the OUTPUT to that day's own rows via
excel_utils.restrict_to_target_date(), called AFTER calculate_tw_all() runs
on the full multi-day history (needed for the Hull MA warmup and the
flip-and-hold state, which must not reset at target_date's own boundary).
See that function's docstring and 01_Master_Code.py's market-open
scheduling fix.
"""

import os
import sys
import tempfile
import shutil
import concurrent.futures
from datetime import time as dtime

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

CODES_DIR = os.path.dirname(os.path.abspath(__file__))
if CODES_DIR not in sys.path:
    sys.path.append(CODES_DIR)
import data_ingestion
import excel_utils

# ---------------------------------------------------------------------------
# Config -- mirrors the Pine script's default inputs
# ---------------------------------------------------------------------------
LENGTH = 16
MODE = "Ehma"            # fixed in the uploaded script (modeSwitch is hardcoded)
CUTOFF_TIME = dtime(15, 15)
INTERVAL = "5minute"

# ---------------------------------------------------------------------------
# Excel styling (kept visually consistent with RSI.py / SQZMOM.py's palette)
# ---------------------------------------------------------------------------
FILL_BULLISH = PatternFill(start_color="26A69A", end_color="26A69A", fill_type="solid")
FILL_BEARISH = PatternFill(start_color="EF5350", end_color="EF5350", fill_type="solid")
FILL_BUY = PatternFill(start_color="0018F3", end_color="0018F3", fill_type="solid")
FILL_SELL = PatternFill(start_color="EC223D", end_color="EC223D", fill_type="solid")
FILL_LIME = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")   # BUY CE
FILL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # BUY PE
FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")   # WAIT
FONT_WHITE = Font(color="FFFFFF")
FONT_BLACK = Font(color="000000")


# ---------------------------------------------------------------------------
# Hull-family moving averages
# ---------------------------------------------------------------------------
def _wma(series, length):
    weights = np.arange(1, length + 1)
    return series.rolling(length).apply(lambda x: np.dot(x, weights) / weights.sum(), raw=True)


def _hma(src, length):
    """Textbook Hull MA -- correctly formed, unused by default (see docstring)."""
    half = max(int(length / 2), 1)
    smooth_len = max(int(round(np.sqrt(length))), 1)
    return _wma(2 * _wma(src, half) - _wma(src, length), smooth_len)


def _ehma(src, length):
    """Reproduces the uploaded script's EHMA() literally: both inner ema()
    calls use the same `length`, so this reduces to a double EMA smooth,
    not real Hull-style responsiveness. See module docstring."""
    inner = src.ewm(span=length, adjust=False).mean()
    smooth_len = max(int(round(np.sqrt(length))), 1)
    return inner.ewm(span=smooth_len, adjust=False).mean()


def _thma(src, length):
    """Textbook Triangular Hull MA -- correctly formed, unused by default."""
    third = max(int(length / 3), 1)
    half = max(int(length / 2), 1)
    return _wma(_wma(src, third) * 3 - _wma(src, half) - _wma(src, length), length)


_HULL_FUNCS = {"Hma": _hma, "Ehma": _ehma, "Thma": _thma}


# ---------------------------------------------------------------------------
# Indicator
# ---------------------------------------------------------------------------
def calculate_tw_all(df, length=LENGTH, mode=MODE):
    df = df.copy()
    hull_fn = _HULL_FUNCS.get(mode, _ehma)
    hull_len = length if mode != "Thma" else max(int(length / 2), 1)
    hull = hull_fn(df['close'], hull_len)

    mhull = hull
    shull = hull.shift(2)

    df['HULL'] = hull
    df['TREND'] = np.where(hull > hull.shift(2), 'Bullish', 'Bearish')

    sell_cross = (shull.shift(1) <= mhull.shift(1)) & (shull > mhull)   # crossover(SHULL, MHULL)
    buy_cross = (shull.shift(1) >= mhull.shift(1)) & (shull < mhull)    # crossunder(SHULL, MHULL)

    signal = np.full(len(df), '', dtype=object)
    signal[sell_cross.to_numpy(dtype=bool)] = 'Sell'
    signal[buy_cross.to_numpy(dtype=bool)] = 'Buy'
    df['SIGNAL'] = signal

    df['TW ALL Recomm'] = compute_twall_recomm(
        buy_cross.to_numpy(dtype=bool), sell_cross.to_numpy(dtype=bool)
    )
    return df


def compute_twall_recomm(buy_cross, sell_cross):
    """Flip-and-hold on the Hull crossover events. Returns a numpy object
    array of 'BUY CE' / 'BUY PE' / 'WAIT'."""
    n = len(buy_cross)
    recomm = np.full(n, 'WAIT', dtype=object)
    state = None
    for i in range(n):
        if buy_cross[i]:
            state = 'BUY CE'
        elif sell_cross[i]:
            state = 'BUY PE'
        recomm[i] = state if state is not None else 'WAIT'
    return recomm


# ---------------------------------------------------------------------------
# Per-symbol processing (same brute-force time/timezone protocol as RSI.py)
# ---------------------------------------------------------------------------
def process_symbol(symbol_data):
    symbol, df, target_date = symbol_data
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip().str.lower()

    if not all(col in df.columns for col in ['open', 'high', 'low', 'close']):
        print(f"[WARNING] {symbol}: Missing required price columns. Discarding.")
        return symbol, None

    time_col = next((col for col in ['date', 'time', 'datetime', 'timestamp'] if col in df.columns), None)
    if time_col:
        parsed = pd.to_datetime(df[time_col], errors='coerce')
    elif pd.api.types.is_datetime64_any_dtype(df.index):
        parsed = pd.Series(df.index, index=df.index)
    else:
        unnamed_cols = [c for c in df.columns if 'unnamed' in c]
        if unnamed_cols:
            parsed = pd.to_datetime(df[unnamed_cols[0]], errors='coerce')
        else:
            print(f"[WARNING] {symbol}: Failed to locate any valid timestamp data. Discarding.")
            return symbol, None

    try:
        if parsed.dt.tz is not None:
            parsed = parsed.dt.tz_localize(None)
    except (AttributeError, TypeError):
        pass

    df['_sort_dt'] = parsed
    df['time_str'] = parsed.dt.strftime('%H:%M')
    df.dropna(subset=['time_str', '_sort_dt'], inplace=True)
    if df.empty:
        print(f"[WARNING] {symbol}: No rows with a valid timestamp. Discarding.")
        return symbol, None

    df.sort_values('_sort_dt', inplace=True)
    df = df[df['_sort_dt'].dt.time <= CUTOFF_TIME]
    if df.empty:
        print(f"[WARNING] {symbol}: No candles at/before {CUTOFF_TIME.strftime('%H:%M')}. Discarding.")
        return symbol, None

    df = calculate_tw_all(df)

    # [FIX] Date-scope the OUTPUT to target_date only -- see module
    # docstring and excel_utils.restrict_to_target_date()'s docstring.
    df = excel_utils.restrict_to_target_date(df, target_date)
    if df is None:
        return symbol, None

    return symbol, df


# ---------------------------------------------------------------------------
# Excel export -- pivoted Symbol x Time matrix, merged into the existing
# workbook (same pattern as RSI.py / SQZMOM.py: load_workbook, replace only
# the 'TW ALL' sheet, leave every other sheet untouched).
# ---------------------------------------------------------------------------
def build_matrix(data_dict, target_date, max_workers=None):
    """Executes multi-core calculation and builds the pivoted Matrix
    (no Excel I/O -- safe to run concurrently with other indicators'
    build_matrix() calls since it never touches the shared workbook).

    [CHANGED] Now requires target_date -- see process_symbol()."""
    results = {}
    with concurrent.futures.ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(process_symbol, (sym, df, target_date)): sym for sym, df in data_dict.items()}
        for future in concurrent.futures.as_completed(futures):
            sym = futures[future]
            try:
                processed_sym, processed_df = future.result()
                if processed_df is not None and 'HULL' in processed_df.columns:
                    processed_df = processed_df.dropna(subset=['HULL'])
                    if not processed_df.empty and sym.lower() != 'summary':
                        results[sym] = processed_df
            except Exception as e:
                print(f"[ERROR] TW ALL Engine Failure on symbol {sym}: {str(e)}")

    if not results:
        raise RuntimeError("TW ALL: zero symbols processed successfully for target_date -- matrix build aborted.")

    all_times = set()
    for df in results.values():
        if 'time_str' in df.columns:
            all_times.update(df['time_str'].dropna().tolist())

    if not all_times:
        raise RuntimeError("TW ALL: no valid timestamps extracted across all symbols -- matrix build aborted.")

    sorted_times = sorted(all_times)
    matrix_rows = [['Symbol', 'Metrics'] + sorted_times]

    for sym in sorted(results.keys()):
        df = results[sym].drop_duplicates(subset=['time_str'], keep='last')
        df_indexed = df.set_index('time_str')

        def get_metric_row(metric_name, col_name):
            row = [sym, metric_name]
            for t in sorted_times:
                if t in df_indexed.index:
                    val = df_indexed.loc[t, col_name]
                    if isinstance(val, pd.Series):
                        val = val.iloc[-1]
                    row.append(val)
                else:
                    row.append("")
            return row

        matrix_rows.append(get_metric_row('Open', 'open'))
        matrix_rows.append(get_metric_row('Close', 'close'))
        matrix_rows.append(get_metric_row('HULL', 'HULL'))
        matrix_rows.append(get_metric_row('TREND', 'TREND'))
        matrix_rows.append(get_metric_row('SIGNAL', 'SIGNAL'))
        matrix_rows.append(get_metric_row('TW ALL Recomm', 'TW ALL Recomm'))

    return matrix_rows


def write_matrix_to_workbook(matrix_rows, wb):
    """Applies the 'TW ALL' sheet into an ALREADY-OPEN workbook (wb) --
    no file load/save here. See excel_utils.replace_sheet_with_matrix()'s
    docstring for why an orchestrator should batch all indicators
    through this instead of write_matrix()'s own load/save."""
    sheet_name = "TW ALL"
    ws = excel_utils.replace_sheet_with_matrix(wb, sheet_name, matrix_rows)

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1), start=2):
        metric_type = ws.cell(row=r_idx, column=2).value
        for cell in row[2:]:
            val = cell.value
            if val == "":
                continue
            if metric_type == 'TREND':
                if val == 'Bullish':
                    cell.fill = FILL_BULLISH
                    cell.font = FONT_WHITE
                elif val == 'Bearish':
                    cell.fill = FILL_BEARISH
                    cell.font = FONT_WHITE
            elif metric_type == 'SIGNAL':
                if val == 'Buy':
                    cell.fill = FILL_BUY
                    cell.font = FONT_WHITE
                elif val == 'Sell':
                    cell.fill = FILL_SELL
                    cell.font = FONT_WHITE
            elif metric_type == 'TW ALL Recomm':
                if val == 'BUY CE':
                    cell.fill = FILL_LIME
                    cell.font = FONT_BLACK
                elif val == 'BUY PE':
                    cell.fill = FILL_RED
                    cell.font = FONT_WHITE
                elif val == 'WAIT':
                    cell.fill = FILL_GRAY
                    cell.font = FONT_BLACK

    excel_utils.autofit_columns(ws)


def write_matrix(matrix_rows, output_excel_path):
    """Standalone/back-compat entry point: load -> apply -> save. Prefer
    write_matrix_to_workbook() when writing several indicators in the
    same run (see 01_Master_Code.py)."""
    wb = load_workbook(output_excel_path)
    write_matrix_to_workbook(matrix_rows, wb)
    excel_utils.atomic_save(wb, output_excel_path)


def parallel_compute_and_export(data_dict, target_date, output_excel_path, max_workers=None):
    """Backward-compatible fused entry point (build + write in one call).
    [CHANGED] Now requires target_date -- see build_matrix()."""
    matrix_rows = build_matrix(data_dict, target_date, max_workers=max_workers)
    write_matrix(matrix_rows, output_excel_path)


def run_twall_step(df_ref, target_date, output_excel_path):
    """Single entry point for run_pipeline.py's Step 10 (standalone use)."""
    print("[SYSTEM] Loading 5-minute historical data for TW All In One...")
    data_dict = data_ingestion.load_interval_data(df_ref, target_date, interval=INTERVAL)
    if not data_dict:
        raise RuntimeError("TW ALL: no 5-minute data files found for any symbol.")
    print(f"[PROCESS] Computing TW All In One matrix for {len(data_dict)} symbol(s)...")
    parallel_compute_and_export(data_dict, target_date, output_excel_path)
    print("[SUCCESS] TW All In One matrix written to sheet 'TW ALL'.")


# ---------------------------------------------------------------------------
# Disclaimer: this module estimates trend direction from historical/live
# price data only. It is not financial advice, no result here is a
# guarantee of future performance, and the 'TW ALL Recomm' rule has not
# been backtested in this conversation -- paper-trade it and run it through
# scripts/backtester.py before risking real capital.
# ---------------------------------------------------------------------------
