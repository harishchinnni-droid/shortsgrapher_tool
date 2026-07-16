"""
ema20.py
--------
Vectorized Python translation of the 'EMA 20' sheet from the Google Colab
reference (the `e_rec` block inside `_compute_indicators_for_symbol`).
Output format mirrors RSI.py / SQZMOM.py's pivoted Symbol x Time matrix.

Indicator logic:
    EMA_20 = 20-period EMA of close.
    'EMA 20 Recomm' (per-bar, NOT hold-based -- see note below):
        BUY CE if this bar's open AND close are both above EMA_20.
        BUY PE if this bar's open AND close are both below EMA_20.
        WAIT otherwise.
    HTF gate (applied to the Recomm above, imported directly from
    htf_bias.compute_htf_bias() so the two sheets can never disagree):
        A BUY CE is dropped to WAIT if either the 15M bias or the Daily
        bias reads DOWN. A BUY PE is dropped to WAIT if either reads UP.
        This mirrors the source's veto exactly.

Translation note -- per-bar, not flip-and-hold:
    Every OTHER Recomm column in this pipeline (SQZMOM, RSI, BRKPRO, ADX,
    TW ALL) is a stateful "arm, fire, then hold until reversal"
    signal. This one is not: the source recomputes 'EMA 20 Recomm' fresh
    on every single bar directly from that bar's own open/close vs EMA_20
    (plus the HTF gate), with no memory of previous bars. That is a
    deliberate, literal translation of the source's own logic, not an
    oversight -- flip-and-hold would be a different indicator.

Timeframe: 5-minute candles, truncated per day to CUTOFF_TIME.

[FIX -- 13-Jul-26] Two changes in this pass:
1. process_symbol()/build_matrix() now require target_date and restrict
   the OUTPUT to that day's own rows via
   excel_utils.restrict_to_target_date(), called AFTER calculate_ema20()
   (and its internal htf_bias.compute_htf_bias() call) runs on the full
   multi-day history. Previously the time_str-dedup-keep-last pivot
   trick silently fell back to a prior trading day's EMA/HTF-gated
   Recomm before target_date's own candles existed. See
   excel_utils.restrict_to_target_date()'s docstring and
   01_Master_Code.py's market-open scheduling fix.
2. write_matrix() now goes through the same
   excel_utils.replace_sheet_with_matrix() + write_matrix_to_workbook()
   split every other indicator module uses, instead of hand-rolling its
   own load/del-sheet/create-sheet logic. Purely a consistency cleanup --
   behavior for a standalone call is unchanged -- but it means this
   module can now also be batched through 01_Master_Code.py's
   single load -> write-all -> save pass like the others.
"""

import os
import sys
import concurrent.futures
from datetime import time as dtime

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

CODES_DIR = os.path.dirname(os.path.abspath(__file__))
if CODES_DIR not in sys.path:
    sys.path.append(CODES_DIR)
import excel_utils
import htf_bias  # compute_htf_bias() -- shared gate logic, see module docstring

CUTOFF_TIME = dtime(15, 15)
INTERVAL = "5minute"

# ---------------------------------------------------------------------------
# Excel styling (kept visually consistent with the other sheets)
# ---------------------------------------------------------------------------
FILL_LIME = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")   # BUY CE
FILL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # BUY PE
FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")   # WAIT
FONT_WHITE = Font(color="FFFFFF")
FONT_BLACK = Font(color="000000")


# ---------------------------------------------------------------------------
# Indicator
# ---------------------------------------------------------------------------
def calculate_ema20(df):
    """`df` must satisfy the same preconditions as
    htf_bias.compute_htf_bias() (sorted ascending by '_sort_dt', lowercase
    OHLC columns) -- this function calls that one directly for the gate."""
    df = df.copy()
    df['EMA_20'] = df['close'].ewm(span=20, adjust=False).mean()

    df = htf_bias.compute_htf_bias(df)  # adds BIAS_15M / BIAS_DAILY / ADX_VAL

    open_ = df['open'].to_numpy()
    close = df['close'].to_numpy()
    ema20 = df['EMA_20'].to_numpy()
    bias_15m = df['BIAS_15M'].to_numpy()
    bias_daily = df['BIAS_DAILY'].to_numpy()

    df['EMA 20 Recomm'] = compute_ema20_recomm(open_, close, ema20, bias_15m, bias_daily)
    return df


def compute_ema20_recomm(open_vals, close_vals, ema20_vals, bias_15m_vals, bias_daily_vals):
    """Direct per-bar rule + HTF veto (see module docstring -- no
    hold-over across bars). Fully vectorized. Returns a numpy object array
    of 'BUY CE' / 'BUY PE' / 'WAIT'."""
    raw_ce = (open_vals > ema20_vals) & (close_vals > ema20_vals)
    raw_pe = (open_vals < ema20_vals) & (close_vals < ema20_vals)

    vetoed_ce = raw_ce & ((bias_15m_vals == 'DOWN') | (bias_daily_vals == 'DOWN'))
    vetoed_pe = raw_pe & ((bias_15m_vals == 'UP') | (bias_daily_vals == 'UP'))

    recomm = np.full(len(open_vals), 'WAIT', dtype=object)
    recomm[raw_ce & ~vetoed_ce] = 'BUY CE'
    recomm[raw_pe & ~vetoed_pe] = 'BUY PE'
    return recomm


# ---------------------------------------------------------------------------
# Per-symbol processing (same brute-force time/timezone protocol as RSI.py)
# ---------------------------------------------------------------------------
def process_symbol(symbol_data):
    symbol, df, target_date = symbol_data
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip().str.lower()

    if not all(col in df.columns for col in ['open', 'high', 'low', 'close']):
        print(f"[WARNING] {symbol}: Missing required price columns. Discarding.")
        return symbol, None

    time_col = next((col for col in ['date', 'time', 'datetime', 'timestamp'] if col in df.columns), None)
    if time_col:
        parsed = pd.to_datetime(df[time_col], errors='coerce')
    elif pd.api.types.is_datetime64_any_dtype(df.index):
        parsed = pd.Series(df.index, index=df.index)
    else:
        unnamed_cols = [c for c in df.columns if 'unnamed' in c]
        if unnamed_cols:
            parsed = pd.to_datetime(df[unnamed_cols[0]], errors='coerce')
        else:
            print(f"[WARNING] {symbol}: Failed to locate any valid timestamp data. Discarding.")
            return symbol, None

    try:
        if parsed.dt.tz is not None:
            parsed = parsed.dt.tz_localize(None)
    except (AttributeError, TypeError):
        pass

    df['_sort_dt'] = parsed
    df['time_str'] = parsed.dt.strftime('%H:%M')
    df.dropna(subset=['time_str', '_sort_dt'], inplace=True)
    if df.empty:
        print(f"[WARNING] {symbol}: No rows with a valid timestamp. Discarding.")
        return symbol, None

    df.sort_values('_sort_dt', inplace=True)
    df = df[df['_sort_dt'].dt.time <= CUTOFF_TIME]
    if df.empty:
        print(f"[WARNING] {symbol}: No candles at/before {CUTOFF_TIME.strftime('%H:%M')}. Discarding.")
        return symbol, None

    df = calculate_ema20(df)

    # [FIX] Date-scope the OUTPUT to target_date only -- see module
    # docstring and excel_utils.restrict_to_target_date()'s docstring.
    df = excel_utils.restrict_to_target_date(df, target_date)
    if df is None:
        return symbol, None

    return symbol, df


# ---------------------------------------------------------------------------
# Excel export -- pivoted Symbol x Time matrix, merged into the existing
# workbook (same pattern as RSI.py / SQZMOM.py: load_workbook, replace only
# the 'EMA 20' sheet, leave every other sheet untouched).
# ---------------------------------------------------------------------------
def build_matrix(data_dict, target_date, max_workers=None):
    """Executes multi-core calculation and builds the pivoted Matrix
    (no Excel I/O -- safe to run concurrently with other indicators'
    build_matrix() calls since it never touches the shared workbook).

    [CHANGED] Now requires target_date -- see process_symbol()."""
    results = {}
    with concurrent.futures.ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(process_symbol, (sym, df, target_date)): sym for sym, df in data_dict.items()}
        for future in concurrent.futures.as_completed(futures):
            sym = futures[future]
            try:
                processed_sym, processed_df = future.result()
                if processed_df is not None and 'EMA 20 Recomm' in processed_df.columns:
                    if not processed_df.empty and sym.lower() != 'summary':
                        results[sym] = processed_df
            except Exception as e:
                print(f"[ERROR] EMA 20 Engine Failure on symbol {sym}: {str(e)}")

    if not results:
        raise RuntimeError("EMA 20: zero symbols processed successfully for target_date -- matrix build aborted.")

    all_times = set()
    for df in results.values():
        if 'time_str' in df.columns:
            all_times.update(df['time_str'].dropna().tolist())

    if not all_times:
        raise RuntimeError("EMA 20: no valid timestamps extracted across all symbols -- matrix build aborted.")

    sorted_times = sorted(all_times)
    matrix_rows = [['Symbol', 'Metrics'] + sorted_times]

    for sym in sorted(results.keys()):
        df = results[sym].drop_duplicates(subset=['time_str'], keep='last')
        df_indexed = df.set_index('time_str')

        def get_metric_row(metric_name, col_name):
            row = [sym, metric_name]
            for t in sorted_times:
                if t in df_indexed.index:
                    val = df_indexed.loc[t, col_name]
                    if isinstance(val, pd.Series):
                        val = val.iloc[-1]
                    row.append(val)
                else:
                    row.append("")
            return row

        matrix_rows.append(get_metric_row('Open', 'open'))
        matrix_rows.append(get_metric_row('Close', 'close'))
        matrix_rows.append(get_metric_row('EMA 20', 'EMA_20'))
        matrix_rows.append(get_metric_row('EMA 20 Recomm', 'EMA 20 Recomm'))

    return matrix_rows


def write_matrix_to_workbook(matrix_rows, wb):
    """[CHANGED] Applies the 'EMA 20' sheet into an ALREADY-OPEN workbook
    (wb) -- no file load/save here, same split every other indicator
    module uses (see excel_utils.replace_sheet_with_matrix()'s docstring
    for why an orchestrator should batch all indicators through this
    instead of write_matrix()'s own load/save)."""
    sheet_name = "EMA 20"
    ws = excel_utils.replace_sheet_with_matrix(wb, sheet_name, matrix_rows)
    ws.freeze_panes = "C2"

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1), start=2):
        metric_type = ws.cell(row=r_idx, column=2).value
        for cell in row[2:]:
            val = cell.value
            if val == "":
                continue
            if metric_type == 'EMA 20 Recomm':
                if val == 'BUY CE':
                    cell.fill = FILL_LIME
                    cell.font = FONT_BLACK
                elif val == 'BUY PE':
                    cell.fill = FILL_RED
                    cell.font = FONT_WHITE
                elif val == 'WAIT':
                    cell.fill = FILL_GRAY
                    cell.font = FONT_BLACK

    excel_utils.autofit_columns(ws)


def write_matrix(matrix_rows, output_excel_path):
    """Standalone/back-compat entry point: load -> apply -> save. Prefer
    write_matrix_to_workbook() when writing several indicators in the
    same run (see 01_Master_Code.py)."""
    wb = load_workbook(output_excel_path)
    write_matrix_to_workbook(matrix_rows, wb)
    excel_utils.atomic_save(wb, output_excel_path)


def parallel_compute_and_export(data_dict, target_date, output_excel_path, max_workers=None):
    """Backward-compatible fused entry point (build + write in one call).
    [CHANGED] Now requires target_date -- see build_matrix()."""
    matrix_rows = build_matrix(data_dict, target_date, max_workers=max_workers)
    write_matrix(matrix_rows, output_excel_path)


def run_ema20_step(df_ref, target_date, output_excel_path):
    """Single entry point for run_pipeline.py-style standalone use."""
    import data_ingestion
    print("[SYSTEM] Loading 5-minute historical data for EMA 20...")
    data_dict = data_ingestion.load_interval_data(df_ref, target_date, interval=INTERVAL)
    if not data_dict:
        raise RuntimeError("EMA 20: no 5-minute data files found for any symbol.")
    print(f"[PROCESS] Computing EMA 20 matrix for {len(data_dict)} symbol(s)...")
    parallel_compute_and_export(data_dict, target_date, output_excel_path)
    print("[SUCCESS] EMA 20 matrix written to sheet 'EMA 20'.")


# ---------------------------------------------------------------------------
# Disclaimer: this module estimates a directional bias from price data
# only. It is not financial advice, no result here is a guarantee of
# future performance, and the 'EMA 20 Recomm' rule has not been
# backtested in this conversation -- paper-trade it and run it through
# scripts/backtester.py before risking real capital.
# ---------------------------------------------------------------------------
