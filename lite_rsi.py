"""
lite_rsi.py
-----------
NEW MODULE -- part of the 3-indicator standalone pipeline (Task 52,
18-Jul-26). Harish's spec: "RSI movement should be compared with previous
candles, if direction is moving up then BUY CE, if moving down then BUY
PE, if it reached above 75 for Overbought and below 25 for Oversold" --
overbought/oversold OVERRIDES the up/down read to WAIT (confirmed
explicitly), rather than being just an informational label.

Does NOT modify or replace RSI.py -- that module (and its own, different
EMA25/50/100 + RSI-band recomm logic) stays exactly as-is for the existing
full pipeline (01_Master_Code.py). This is a separate sheet-producing
module for the lean 3-indicator pipeline only.

Indicator math: standard Wilder RSI(14), same formula as RSI.py's _rsi().

Per-bar 'RSI Recomm' (per-bar, recomputed fresh every bar):
    WAIT    if RSI > 75 (Overbought) or RSI < 25 (Oversold) -- overrides
            the up/down read below, takes priority
    BUY CE  if RSI > previous candle's RSI (rising)
    BUY PE  if RSI < previous candle's RSI (falling)
    WAIT    otherwise (RSI unchanged from previous candle, or NaN warmup)

Timeframe: 5-minute candles, truncated per day to CUTOFF_TIME, same
convention as every other module in this pipeline.

Disclaimer: this is a new, unbacktested rule -- not financial advice, no
result here is a guarantee of future performance. Paper-trade and backtest
before risking real capital.
"""

import os
import sys
import concurrent.futures
from datetime import time as dtime

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

CODES_DIR = os.path.dirname(os.path.abspath(__file__))
if CODES_DIR not in sys.path:
    sys.path.append(CODES_DIR)
import data_ingestion
import excel_utils

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
RSI_PERIOD = 14
CUTOFF_TIME = dtime(15, 15)
INTERVAL = "5minute"
RSI_OVERBOUGHT = 75.0
RSI_OVERSOLD = 25.0

# ---------------------------------------------------------------------------
# Excel styling
# ---------------------------------------------------------------------------
FILL_LIME = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")   # BUY CE
FILL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # BUY PE
FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")   # WAIT
FONT_WHITE = Font(color="FFFFFF")
FONT_BLACK = Font(color="000000")


# ---------------------------------------------------------------------------
# Indicator -- same Wilder RSI formula as RSI.py's _rsi()
# ---------------------------------------------------------------------------
def _rsi(close, period=RSI_PERIOD):
    delta = close.diff()
    gain = delta.where(delta > 0, 0.0).ewm(alpha=1.0 / period, adjust=False).mean()
    loss = (-delta.where(delta < 0, 0.0)).ewm(alpha=1.0 / period, adjust=False).mean()
    rs = gain / loss
    return 100 - (100 / (1 + rs))


def calculate_rsi_lite(df, period=RSI_PERIOD):
    df = df.copy()
    df['RSI'] = _rsi(df['close'], period)
    rsi = df['RSI']
    rsi_prev = rsi.shift(1)

    overbought = rsi > RSI_OVERBOUGHT
    oversold = rsi < RSI_OVERSOLD
    rising = rsi > rsi_prev
    falling = rsi < rsi_prev

    buy_ce = (~overbought) & (~oversold) & rising
    buy_pe = (~overbought) & (~oversold) & falling

    df['RSI Recomm'] = np.where(buy_ce, 'BUY CE', np.where(buy_pe, 'BUY PE', 'WAIT'))
    return df


# ---------------------------------------------------------------------------
# Per-symbol processing (same brute-force time/timezone protocol as RSI.py)
# ---------------------------------------------------------------------------
def process_symbol(symbol_data):
    symbol, df, target_date = symbol_data
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip().str.lower()

    if 'close' not in df.columns:
        print(f"[WARNING] {symbol}: Missing required 'close' column. Discarding.")
        return symbol, None

    time_col = next((col for col in ['date', 'time', 'datetime', 'timestamp'] if col in df.columns), None)
    if time_col:
        parsed = pd.to_datetime(df[time_col], errors='coerce')
    elif pd.api.types.is_datetime64_any_dtype(df.index):
        parsed = pd.Series(df.index, index=df.index)
    else:
        unnamed_cols = [c for c in df.columns if 'unnamed' in c]
        if unnamed_cols:
            parsed = pd.to_datetime(df[unnamed_cols[0]], errors='coerce')
        else:
            print(f"[WARNING] {symbol}: Failed to locate any valid timestamp data. Discarding.")
            return symbol, None

    try:
        if parsed.dt.tz is not None:
            parsed = parsed.dt.tz_localize(None)
    except (AttributeError, TypeError):
        pass

    df['_sort_dt'] = parsed
    df['time_str'] = parsed.dt.strftime('%H:%M')
    df.dropna(subset=['time_str', '_sort_dt'], inplace=True)
    if df.empty:
        print(f"[WARNING] {symbol}: No rows with a valid timestamp. Discarding.")
        return symbol, None

    df.sort_values('_sort_dt', inplace=True)
    df = df[df['_sort_dt'].dt.time <= CUTOFF_TIME]
    if df.empty:
        print(f"[WARNING] {symbol}: No candles at/before {CUTOFF_TIME.strftime('%H:%M')}. Discarding.")
        return symbol, None

    df = calculate_rsi_lite(df)

    df = excel_utils.restrict_to_target_date(df, target_date)
    if df is None:
        return symbol, None

    return symbol, df


# ---------------------------------------------------------------------------
# Excel export -- pivoted Symbol x Time matrix
# ---------------------------------------------------------------------------
def build_matrix(data_dict, target_date, max_workers=None):
    results = {}
    with concurrent.futures.ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(process_symbol, (sym, df, target_date)): sym for sym, df in data_dict.items()}
        for future in concurrent.futures.as_completed(futures):
            sym = futures[future]
            try:
                processed_sym, processed_df = future.result()
                if processed_df is not None and 'RSI' in processed_df.columns:
                    processed_df = processed_df.dropna(subset=['RSI'])
                    if not processed_df.empty and sym.lower() != 'summary':
                        results[sym] = processed_df
            except Exception as e:
                print(f"[ERROR] RSI (lite) Engine Failure on symbol {sym}: {str(e)}")

    if not results:
        raise RuntimeError("RSI (lite): zero symbols processed successfully for target_date -- matrix build aborted.")

    all_times = set()
    for df in results.values():
        if 'time_str' in df.columns:
            all_times.update(df['time_str'].dropna().tolist())

    if not all_times:
        raise RuntimeError("RSI (lite): no valid timestamps extracted across all symbols -- matrix build aborted.")

    sorted_times = sorted(all_times)
    matrix_rows = [['Symbol', 'Metrics'] + sorted_times]

    for sym in sorted(results.keys()):
        df = results[sym].drop_duplicates(subset=['time_str'], keep='last')
        df_indexed = df.set_index('time_str')

        def get_metric_row(metric_name, col_name):
            row = [sym, metric_name]
            for t in sorted_times:
                if t in df_indexed.index:
                    val = df_indexed.loc[t, col_name]
                    if isinstance(val, pd.Series):
                        val = val.iloc[-1]
                    row.append(val)
                else:
                    row.append("")
            return row

        matrix_rows.append(get_metric_row('Close', 'close'))
        matrix_rows.append(get_metric_row('RSI', 'RSI'))
        matrix_rows.append(get_metric_row('RSI Recomm', 'RSI Recomm'))

    return matrix_rows


def write_matrix_to_workbook(matrix_rows, wb):
    sheet_name = "RSI"
    ws = excel_utils.replace_sheet_with_matrix(wb, sheet_name, matrix_rows)

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1), start=2):
        metric_type = ws.cell(row=r_idx, column=2).value
        for cell in row[2:]:
            val = cell.value
            if val == "":
                continue
            if metric_type == 'RSI Recomm':
                if val == 'BUY CE':
                    cell.fill = FILL_LIME
                    cell.font = FONT_BLACK
                elif val == 'BUY PE':
                    cell.fill = FILL_RED
                    cell.font = FONT_WHITE
                elif val == 'WAIT':
                    cell.fill = FILL_GRAY
                    cell.font = FONT_BLACK

    excel_utils.autofit_columns(ws)


def write_matrix(matrix_rows, output_excel_path):
    wb = load_workbook(output_excel_path)
    write_matrix_to_workbook(matrix_rows, wb)
    excel_utils.atomic_save(wb, output_excel_path)


def parallel_compute_and_export(data_dict, target_date, output_excel_path, max_workers=None):
    matrix_rows = build_matrix(data_dict, target_date, max_workers=max_workers)
    write_matrix(matrix_rows, output_excel_path)


def run_rsi_lite_step(df_ref, target_date, output_excel_path):
    """Single entry point for run_pipeline_lite.py."""
    print("[SYSTEM] Loading 5-minute historical data for RSI (lite)...")
    data_dict = data_ingestion.load_interval_data(df_ref, target_date, interval=INTERVAL)
    if not data_dict:
        raise RuntimeError("RSI (lite): no 5-minute data files found for any symbol.")
    print(f"[PROCESS] Computing RSI (lite) matrix for {len(data_dict)} symbol(s)...")
    parallel_compute_and_export(data_dict, target_date, output_excel_path)
    print("[SUCCESS] RSI (lite) matrix written to sheet 'RSI'.")


# ---------------------------------------------------------------------------
# Disclaimer: this module derives a signal from historical/live price data
# only. It is not financial advice, and no result here is a guarantee of
# future performance -- paper-trade it and run it through
# scripts/backtester.py before risking real capital.
# ---------------------------------------------------------------------------
