"""
obv_cmf.py
----------
Volume-confirmation indicator: On-Balance Volume (OBV) + Chaikin Money
Flow (CMF), combined into a single 'OBV CMF Recomm' vote. Output format
mirrors vwap.py / ema20.py's pivoted Symbol x Time matrix.

Why this exists (08-Jul-26 pipeline audit): none of the original 8
confluence voters (SQZMOM, RSI, BRKPRO, ADX, TW ALL, Supertrend,
EMA 20, VWAP) use trading volume as a directional signal -- volume only
showed up in order_sheet.py as a post-hoc liquidity gate (skip if
option volume == 0). That's a real blind spot: none of them can tell a
price move backed by genuine participation apart from one drifting on
thin volume. Meanwhile BRKPRO -- the weakest of the original 8 on a
measured single-day check (49.95% forward hit rate, ~7,500 direction
flips across the day, i.e. no better than noise) -- was occupying a vote
without contributing real information.

INTEGRATION [CONFIRMED WIRED -- 13-Jul-26]: this module is now actually
registered in 01_Master_Code.py's INDICATORS list (so its own 'OBV CMF'
sheet gets computed and written every cycle) AND in final_sheet.py's
BREAKOUT_SHEETS, REPLACING BRKPRO's vote in the confluence check -- not
stacked on top of it as a 9th unanimous voter. BRKPRO's own sheet is
still computed and visible in the workbook (kept in
final_sheet.py's INDICATOR_SHEETS for passthrough display), it just no
longer counts toward Final Recomm. (A prior version of this docstring
described this integration before it had actually been made in
01_Master_Code.py / final_sheet.py -- that mismatch is fixed as of this
pass; both files now match what's written here.)

Indicator logic:
    OBV  : session-anchored (resets every calendar day, same convention
           as vwap.py) cumulative running volume -- adds volume on an
           up-close, subtracts it on a down-close, unchanged on a flat
           close.
    OBV_EMA : EMA(OBV, span=OBV_EMA_LEN) -- OBV's own short trend line.
           Comparing OBV to its own EMA (rather than reading OBV's raw
           bar-to-bar tick) is the noise filter here, the same way
           SQZMOM/RSI use multi-bar smoothing rather than a raw 1-bar
           read.
    CMF  : rolling Chaikin Money Flow over CMF_PERIOD bars --
           sum(Money Flow Volume) / sum(volume), where Money Flow
           Volume = ((close-low)-(high-close))/(high-low) * volume.
           Oscillates roughly -1..+1 around a zero line; positive means
           closes are skewing toward the bar's high (accumulation),
           negative toward the low (distribution).

'OBV CMF Recomm' (per-bar, NOT hold-based, same convention as
vwap.py/ema20.py -- no memory of previous bars):
    BUY CE requires BOTH: OBV > OBV_EMA (volume itself is trending up)
        AND CMF > +CMF_THRESH (money flow is meaningfully positive).
    BUY PE requires BOTH: OBV < OBV_EMA AND CMF < -CMF_THRESH.
    WAIT otherwise (including whenever OBV_EMA/CMF aren't computable
    yet, e.g. still inside the CMF_PERIOD warmup window).
    Requiring OBV and CMF to agree is deliberate: they're two
    different volume calculations (a cumulative running total vs. a
    bounded rolling oscillator built off the high/low/close range), so
    both agreeing is a real, largely independent confirmation rather
    than two views of the same number.

Timeframe / conventions: identical to vwap.py -- 5-minute candles,
CUTOFF_TIME truncation, session reset (OBV) computed BEFORE that
truncation so later bars' cumulative OBV isn't missing earlier
same-day volume. CMF is a rolling window, not session-reset -- it
naturally ages out stale data on its own, unlike OBV's running total.

[FIX -- 13-Jul-26] process_symbol()/build_matrix() now require
target_date and restrict the OUTPUT to that day's own rows via
excel_utils.restrict_to_target_date(), called AFTER calculate_obv_cmf()
runs on the full multi-day history (needed for OBV's session-reset
cumsum and CMF's rolling warmup). See that function's docstring and
01_Master_Code.py's market-open scheduling fix -- this is what stops the
'OBV CMF' sheet from silently using a prior trading day's OBV/CMF
before target_date's own candles exist, which matters even more now
that this sheet actually has a vote in the confluence check.

This is a heuristic, not a validated edge -- it has not been
backtested in this conversation across many days. Paper-trade it and
run it through scripts/backtester.py before trusting it (alone or in
the confluence vote) with real capital.
"""

import os
import sys
import concurrent.futures
from datetime import time as dtime

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

CODES_DIR = os.path.dirname(os.path.abspath(__file__))
if CODES_DIR not in sys.path:
    sys.path.append(CODES_DIR)
import excel_utils

CUTOFF_TIME = dtime(15, 15)
INTERVAL = "5minute"

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
OBV_EMA_LEN = 10     # smoothing length for OBV's own trend line
CMF_PERIOD = 20      # standard Chaikin Money Flow lookback, in bars here
CMF_THRESH = 0.05    # small buffer around the zero line. CMF crosses 0
                      # constantly in a flat/choppy tape -- a bare > 0 / < 0
                      # test would itself become a BRKPRO-style noise
                      # source, so this requires a small but real tilt
                      # before it's allowed to vote. Tune per-instrument if
                      # you find it too strict/loose once you have more
                      # days of data to check against.

# ---------------------------------------------------------------------------
# Excel styling (kept visually consistent with the other sheets)
# ---------------------------------------------------------------------------
FILL_LIME = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")   # BUY CE
FILL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # BUY PE
FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")   # WAIT
FONT_WHITE = Font(color="FFFFFF")
FONT_BLACK = Font(color="000000")


# ---------------------------------------------------------------------------
# Indicator
# ---------------------------------------------------------------------------
def calculate_obv_cmf(df, obv_ema_len=OBV_EMA_LEN, cmf_period=CMF_PERIOD, cmf_thresh=CMF_THRESH):
    """`df` must be sorted ascending by '_sort_dt' with lowercase
    'open'/'high'/'low'/'close'/'volume' columns -- the daily OBV reset
    groups on '_sort_dt'.dt.date, same convention as vwap.py."""
    df = df.copy()
    date_only = df['_sort_dt'].dt.date

    close = df['close']
    high = df['high']
    low = df['low']
    volume = df['volume']

    # --- OBV, reset every session ---
    direction = np.sign(close.diff().fillna(0))
    signed_volume = direction * volume
    df['OBV'] = signed_volume.groupby(date_only).cumsum()
    df['OBV_EMA'] = df['OBV'].ewm(span=obv_ema_len, adjust=False).mean()

    # --- Chaikin Money Flow, rolling window (a rolling window naturally
    # ages out stale data on its own -- unlike OBV's cumulative total,
    # it doesn't need an explicit per-day reset) ---
    hl_range = (high - low).replace(0, np.nan)
    mf_multiplier = ((close - low) - (high - close)) / hl_range
    mf_volume = mf_multiplier * volume
    df['CMF'] = (
        mf_volume.rolling(cmf_period).sum() / volume.rolling(cmf_period).sum().replace(0, np.nan)
    )

    df['OBV CMF Recomm'] = compute_obv_cmf_recomm(
        df['OBV'].to_numpy(), df['OBV_EMA'].to_numpy(), df['CMF'].to_numpy(), cmf_thresh
    )
    return df


def compute_obv_cmf_recomm(obv_vals, obv_ema_vals, cmf_vals, cmf_thresh=CMF_THRESH):
    """Direct per-bar rule (see module docstring -- no hold-over across
    bars, same convention as vwap.py/ema20.py). Fully vectorized.
    Returns a numpy object array of 'BUY CE' / 'BUY PE' / 'WAIT'."""
    n = len(obv_vals)
    recomm = np.full(n, 'WAIT', dtype=object)

    valid = ~np.isnan(obv_ema_vals) & ~np.isnan(cmf_vals)

    obv_up = obv_vals > obv_ema_vals
    obv_down = obv_vals < obv_ema_vals
    cmf_pos = cmf_vals > cmf_thresh
    cmf_neg = cmf_vals < -cmf_thresh

    recomm[valid & obv_up & cmf_pos] = 'BUY CE'
    recomm[valid & obv_down & cmf_neg] = 'BUY PE'
    return recomm


# ---------------------------------------------------------------------------
# Per-symbol processing (same brute-force time/timezone protocol as vwap.py)
# ---------------------------------------------------------------------------
def process_symbol(symbol_data):
    symbol, df, target_date = symbol_data
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip().str.lower()

    if not all(col in df.columns for col in ['open', 'high', 'low', 'close', 'volume']):
        print(f"[WARNING] {symbol}: Missing required price/volume columns. Discarding.")
        return symbol, None

    time_col = next((col for col in ['date', 'time', 'datetime', 'timestamp'] if col in df.columns), None)
    if time_col:
        parsed = pd.to_datetime(df[time_col], errors='coerce')
    elif pd.api.types.is_datetime64_any_dtype(df.index):
        parsed = pd.Series(df.index, index=df.index)
    else:
        unnamed_cols = [c for c in df.columns if 'unnamed' in c]
        if unnamed_cols:
            parsed = pd.to_datetime(df[unnamed_cols[0]], errors='coerce')
        else:
            print(f"[WARNING] {symbol}: Failed to locate any valid timestamp data. Discarding.")
            return symbol, None

    try:
        if parsed.dt.tz is not None:
            parsed = parsed.dt.tz_localize(None)
    except (AttributeError, TypeError):
        pass

    df['_sort_dt'] = parsed
    df['time_str'] = parsed.dt.strftime('%H:%M')
    df.dropna(subset=['time_str', '_sort_dt'], inplace=True)
    if df.empty:
        print(f"[WARNING] {symbol}: No rows with a valid timestamp. Discarding.")
        return symbol, None

    df.sort_values('_sort_dt', inplace=True)
    df = df[df['_sort_dt'].dt.time <= CUTOFF_TIME]
    if df.empty:
        print(f"[WARNING] {symbol}: No candles at/before {CUTOFF_TIME.strftime('%H:%M')}. Discarding.")
        return symbol, None

    df = calculate_obv_cmf(df)

    # [FIX] Date-scope the OUTPUT to target_date only -- see module
    # docstring and excel_utils.restrict_to_target_date()'s docstring.
    df = excel_utils.restrict_to_target_date(df, target_date)
    if df is None:
        return symbol, None

    return symbol, df


# ---------------------------------------------------------------------------
# Excel export -- pivoted Symbol x Time matrix, merged into the existing
# workbook (same pattern as vwap.py: load_workbook, replace only the
# 'OBV CMF' sheet, leave every other sheet untouched).
# ---------------------------------------------------------------------------
def build_matrix(data_dict, target_date, max_workers=None):
    """Executes multi-core calculation and builds the pivoted Matrix
    (no Excel I/O -- safe to run concurrently with other indicators'
    build_matrix() calls since it never touches the shared workbook).

    [CHANGED] Now requires target_date -- see process_symbol()."""
    results = {}
    with concurrent.futures.ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(process_symbol, (sym, df, target_date)): sym for sym, df in data_dict.items()}
        for future in concurrent.futures.as_completed(futures):
            sym = futures[future]
            try:
                processed_sym, processed_df = future.result()
                if processed_df is not None and 'OBV CMF Recomm' in processed_df.columns:
                    if not processed_df.empty and sym.lower() != 'summary':
                        results[sym] = processed_df
            except Exception as e:
                print(f"[ERROR] OBV CMF Engine Failure on symbol {sym}: {str(e)}")

    if not results:
        raise RuntimeError("OBV CMF: zero symbols processed successfully for target_date -- matrix build aborted.")

    all_times = set()
    for df in results.values():
        if 'time_str' in df.columns:
            all_times.update(df['time_str'].dropna().tolist())

    if not all_times:
        raise RuntimeError("OBV CMF: no valid timestamps extracted across all symbols -- matrix build aborted.")

    sorted_times = sorted(all_times)
    matrix_rows = [['Symbol', 'Metrics'] + sorted_times]

    for sym in sorted(results.keys()):
        df = results[sym].drop_duplicates(subset=['time_str'], keep='last')
        df_indexed = df.set_index('time_str')

        def get_metric_row(metric_name, col_name):
            row = [sym, metric_name]
            for t in sorted_times:
                if t in df_indexed.index:
                    val = df_indexed.loc[t, col_name]
                    if isinstance(val, pd.Series):
                        val = val.iloc[-1]
                    row.append(val)
                else:
                    row.append("")
            return row

        matrix_rows.append(get_metric_row('Open', 'open'))
        matrix_rows.append(get_metric_row('Close', 'close'))
        matrix_rows.append(get_metric_row('Volume', 'volume'))
        matrix_rows.append(get_metric_row('OBV', 'OBV'))
        matrix_rows.append(get_metric_row('OBV_EMA', 'OBV_EMA'))
        matrix_rows.append(get_metric_row('CMF', 'CMF'))
        matrix_rows.append(get_metric_row('OBV CMF Recomm', 'OBV CMF Recomm'))

    return matrix_rows


def write_matrix_to_workbook(matrix_rows, wb):
    """Applies the 'OBV CMF' sheet into an ALREADY-OPEN workbook (wb) --
    no file load/save here. See excel_utils.replace_sheet_with_matrix()'s
    docstring for why an orchestrator should batch all indicators
    through this instead of write_matrix()'s own load/save."""
    sheet_name = "OBV CMF"
    ws = excel_utils.replace_sheet_with_matrix(wb, sheet_name, matrix_rows)

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1), start=2):
        metric_type = ws.cell(row=r_idx, column=2).value
        for cell in row[2:]:
            val = cell.value
            if val == "":
                continue
            if metric_type == 'OBV CMF Recomm':
                if val == 'BUY CE':
                    cell.fill = FILL_LIME
                    cell.font = FONT_BLACK
                elif val == 'BUY PE':
                    cell.fill = FILL_RED
                    cell.font = FONT_WHITE
                elif val == 'WAIT':
                    cell.fill = FILL_GRAY
                    cell.font = FONT_BLACK

    excel_utils.autofit_columns(ws)


def write_matrix(matrix_rows, output_excel_path):
    """Standalone/back-compat entry point: load -> apply -> save. Prefer
    write_matrix_to_workbook() when writing several indicators in the
    same run (see 01_Master_Code.py)."""
    wb = load_workbook(output_excel_path)
    write_matrix_to_workbook(matrix_rows, wb)
    excel_utils.atomic_save(wb, output_excel_path)


def parallel_compute_and_export(data_dict, target_date, output_excel_path, max_workers=None):
    """Backward-compatible fused entry point (build + write in one call).
    [CHANGED] Now requires target_date -- see build_matrix()."""
    matrix_rows = build_matrix(data_dict, target_date, max_workers=max_workers)
    write_matrix(matrix_rows, output_excel_path)


def run_obv_cmf_step(df_ref, target_date, output_excel_path):
    """Single entry point for run_pipeline.py-style standalone use."""
    import data_ingestion
    print("[SYSTEM] Loading 5-minute historical data for OBV + CMF...")
    data_dict = data_ingestion.load_interval_data(df_ref, target_date, interval=INTERVAL)
    if not data_dict:
        raise RuntimeError("OBV CMF: no 5-minute data files found for any symbol.")
    print(f"[PROCESS] Computing OBV + CMF matrix for {len(data_dict)} symbol(s)...")
    parallel_compute_and_export(data_dict, target_date, output_excel_path)
    print("[SUCCESS] OBV + CMF matrix written to sheet 'OBV CMF'.")


# ---------------------------------------------------------------------------
# Disclaimer: this module estimates a directional bias from price/volume
# data only. It is not financial advice, no result here is a guarantee of
# future performance, and the 'OBV CMF Recomm' rule has not been
# backtested in this conversation -- paper-trade it and run it through
# scripts/backtester.py before risking real capital.
# ---------------------------------------------------------------------------
