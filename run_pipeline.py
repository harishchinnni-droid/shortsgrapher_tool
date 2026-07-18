import os
import sys
import traceback
import concurrent.futures
import pandas as pd
from datetime import datetime
from ist_clock import now_ist

# Force Python to recognize your custom codes directory
# [CHANGED -- cloud/Colab portability] derives from this file's own
# location (same convention final_sheet.py/order_sheet.py etc. already
# use) instead of hardcoding the Windows desktop path -- this constant is
# "where is this script," not "where is my data," so it doesn't need
# file_mgmt.BASE_DIR/ALGO_BASE_DIR at all, just needs to stop assuming F:\.
CODES_DIR = os.path.dirname(os.path.abspath(__file__))
if CODES_DIR not in sys.path:
    sys.path.append(CODES_DIR)

# Strict modular imports
# NOTE: excel_utils.py, breakout_probability.py, adx_di.py, tw_all.py,
# htf_bias.py, ema20.py, vwap.py and obv_cmf.py must all be saved in the
# same '05 Codes' directory as the rest of the pipeline for these imports
# to resolve.
try:
    import broker_auth
    import calendar_mgmt
    import file_mgmt
    import token_mgmt
    import data_ingestion
    import excel_utils
    import SQZMOM
    import RSI
    import breakout_probability
    import adx_di
    import tw_all
    # [ADDED] final_sheet.py's confluence vote needs all 3 of these --
    # they were missing from this list entirely, which meant the TREND
    # category only ever saw TW ALL's vote (1 of 3, needs 2) and the
    # VOLUME category (OBV CMF, the sole voter in its category) never
    # voted at all -- Final Recomm was structurally starved of votes it
    # was supposed to have, not just missing a nice-to-have indicator.
    import htf_bias
    import ema20
    import vwap
    import obv_cmf
    import zerolag
    # [ADDED -- 18-Jul-26, Task 48] Fully written (LuxAlgo SuperTrend AI
    # translation, see supertrend_ai.py) but never actually wired in --
    # its build_matrix() predated this codebase's target_date interface
    # change and would have crashed immediately if added here as-is. Now
    # fixed and added below, same as every other trend indicator.
    import supertrend_ai
    # [ADDED -- 18-Jul-26, Task 50] LuxAlgo "Support and Resistance Levels
    # with Breaks" -- see support_resistance.py's docstring. Reference
    # Pine Script was already sitting in 04_ProjectFiles; this module was
    # newly built, no interface issue to fix (unlike SuperTrend AI above).
    import support_resistance
except ImportError as e:
    print(f"[CRITICAL ERROR] Failed to import pipeline modules: {e}")
    print("Ensure all scripts are saved in the '05 Codes' directory with correct filenames.")
    sys.exit(1)

# Interval used for the intraday indicator matrices. All indicators read
# the same 5-minute CSVs that Step 5 downloads, so the data is loaded ONCE
# (Step 6) and handed to every indicator's build_matrix(), instead of each
# module re-reading the same files off disk once per module.
INDICATOR_INTERVAL = '5minute'

# (display label, module, sheet name written to the workbook) -- every
# module here exposes the same interface:
#   build_matrix(data_dict, target_date, max_workers=None) -> matrix_rows
#   write_matrix(matrix_rows, output_excel_path)
# which is what makes running them concurrently safe: the CPU-heavy part
# touches no shared state, and only the Excel write (fast: append rows +
# save) needs to happen one at a time against the single workbook file.
# HTF Bias is listed FIRST for no ordering reason that matters here --
# ema20.py calls htf_bias.compute_htf_bias() directly as a function call on
# its OWN copy of the data (not by reading the 'HTF Bias' sheet back), so
# there is no cross-module read dependency between these build_matrix()
# calls; they're all safe to run in the same parallel batch below.
INDICATORS = [
    ("HTF Bias",               htf_bias,             "HTF Bias"),
    ("SQZMOM",                 SQZMOM,               "SQZMOM"),
    ("RSI Multi Length",       RSI,                  "RSI"),
    ("Breakout Probability",   breakout_probability,  "BRKPRO"),
    ("ADX & DI",               adx_di,               "ADX"),
    ("TW All In One",          tw_all,               "TW ALL"),
    ("EMA 20",                 ema20,                "EMA 20"),
    ("VWAP",                   vwap,                 "VWAP"),
    ("OBV CMF",                obv_cmf,              "OBV CMF"),
    # [ADDED] Zero-Lag EMA/ATR trend cloud + RVOL -- see zerolag.py's
    # docstring. NOT a final_sheet.py vote (same reasoning as ADX's own
    # exclusion there); consumed as a pre-entry GATE in order_sheet.py
    # instead (ENABLE_ZEROLAG_GATE). Still listed here so its own sheet
    # gets computed/written like every other indicator.
    ("Zero-Lag Trend",         zerolag,              "ZLTREND"),
    # [ADDED -- 18-Jul-26, Task 48] LuxAlgo SuperTrend AI (Clustering) --
    # see supertrend_ai.py's docstring. Same treatment as Zero-Lag: NOT a
    # final_sheet.py vote (TREND category already has its 3 -- EMA20/
    # VWAP/TW ALL -- adding a 4th changes the "≥2/3" math and risks
    # "one trend opinion asked three times" again); consumed instead as
    # an off-by-default pre-entry GATE in order_sheet.py
    # (ENABLE_SUPERTREND_GATE) so it can be A/B tested on its own,
    # independent of the still-pending Zero-Lag/PCR re-test.
    ("SuperTrend AI",          supertrend_ai,        "Supertrend"),
    # [ADDED -- 18-Jul-26, Task 50] LuxAlgo Support/Resistance with
    # volume-confirmed breaks -- see support_resistance.py's docstring.
    # NOT a final_sheet.py vote (structural/positional signal, a
    # different dimension from TREND/MOMENTUM/VOLUME); consumed as an
    # off-by-default pre-entry GATE in order_sheet.py (ENABLE_SR_GATE).
    ("Support/Resistance",     support_resistance,   "SUPRES"),
]


def run_pipeline_for_date(smart_api, kite_api, target_date, mode):
    """Steps 3-8 only -- everything that has to happen ONCE PER TRADING
    DATE (file provisioning, token mapping, historical data, indicator
    matrices). Broker login (old Step 1) and mode/date selection (old
    Step 2) are deliberately NOT here -- 01_Master_Code.py does those
    ONCE and then calls this function in a loop, once per date, for a
    BACKTEST date range (see calendar_mgmt.get_run_config()) without
    re-logging in or re-prompting for every single date.

    Raises on failure instead of sys.exit(1) -- old run_pipeline() killed
    the whole process on a single bad date, which is wrong once a caller
    is looping over a 10-day range: one bad/missing date shouldn't take
    the other 9 down with it. Callers that loop over dates should catch
    and continue; run_pipeline() below (the standalone single-date entry
    point) still exits on failure, matching its original behavior.

    Returns (final_excel_path, df_ref).
    """
    # STEP 3: File Provisioning
    new_filename = file_mgmt.provision_daily_trade_file(target_date)
    # STEP 4: Token Synchronization & Mapping
    df_ref = token_mgmt.update_instrument_tokens(new_filename, kite_api, target_date=target_date)
    # STEP 5: Historical Data Ingestion
    try:
        # cap_to_now: LIVE must never pull/accept candles past the current
        # wall-clock time; BACKTEST is a past, fully-closed date, so the
        # whole day is legitimately wanted in one pass.
        is_live = (mode == calendar_mgmt.LIVE)

        # If today's full backfill is already on disk (e.g. pipeline was
        # re-run/restarted), skip the 90-day re-download entirely and only
        # pull the incremental delta -- saves both time and API calls.
        # [FIX] For LIVE, also verify that file isn't stale/future-dated
        # (e.g. left over from an earlier BACKTEST pass against today's
        # date) before trusting it -- otherwise a full day's worth of
        # candles can get silently adopted as "today's live data" well
        # before the market has actually reached that time.
        already_exists = data_ingestion.historical_data_exists(df_ref, target_date)
        stale = False
        if is_live and already_exists:
            first_sym = str(df_ref.dropna(subset=['Zerodha_Token']).iloc[0]['Symbol / StrikePrice']).strip().upper()
            date_str = target_date.strftime('%d-%b-%y')
            check_path = os.path.join(data_ingestion.HIST_DIR, f"{first_sym}_5minute_{date_str}.csv")
            stale = data_ingestion.file_has_future_candles(check_path, now_ist())

        if already_exists and not stale:
            print("[SYSTEM] Historical data already available for today -- skipping full backfill.")
            data_ingestion.update_incremental_data(df_ref, target_date, kite_api)
        else:
            if stale:
                print("[WARNING] Existing data on disk extends beyond the current time -- "
                      "treating as stale/leftover, not live. Purging and rebuilding.")
                data_ingestion.purge_interval_data(df_ref, target_date, interval='5minute')
            data_ingestion.download_historical_data(df_ref, target_date, kite_api, cap_to_now=is_live)
        print("[SUCCESS] Historical Data Download Complete. Pipeline fully synchronized.")
    except Exception as e:
        raise RuntimeError(f"Step 5 (Historical Ingestion) failed for {target_date.strftime('%d-%b-%y')}: {e}") from e
    print("-" * 60)

    # [CHANGED -- cloud/Colab portability] new_filename is already an
    # absolute path from file_mgmt.provision_daily_trade_file() (itself
    # joined onto file_mgmt.BASE_DIR), so this join was always a harmless
    # no-op on an absolute second argument -- but it silently hardcoded a
    # SECOND, independent copy of the Windows path that would have been
    # wrong the moment BASE_DIR pointed anywhere else. Use BASE_DIR
    # directly so there's only ever one place this pipeline's root is
    # defined.
    final_excel_path = os.path.join(file_mgmt.BASE_DIR, new_filename)

    # STEP 6: Load the shared 5-minute dataset once, then compute all five
    # indicator matrices IN PARALLEL. Each indicator's build_matrix() does
    # no Excel I/O, so there's no shared-file race between them here --
    # only the CPU-bound pandas/numpy work runs concurrently.
    try:
        print(f"[PROCESS] Loading {INDICATOR_INTERVAL} historical data (shared across all indicators)...")
        raw_data_dict = data_ingestion.load_interval_data(df_ref, target_date, interval=INDICATOR_INTERVAL)
        if not raw_data_dict:
            raise RuntimeError(f"No {INDICATOR_INTERVAL} historical data files found for any symbol.")
    except Exception as e:
        raise RuntimeError(f"Step 6 (Historical Data Load) failed for {target_date.strftime('%d-%b-%y')}: {e}") from e

    num_indicators = len(INDICATORS)
    cpu_count = os.cpu_count() or 4
    # Cap each indicator's own internal ProcessPoolExecutor so five pools
    # running at once don't oversubscribe the machine five-fold. Rounding
    # down to 1 in the worst case just means that indicator's symbols are
    # processed one at a time -- it still runs concurrently with the other
    # five indicators, which is where the actual wall-clock win comes from.
    workers_per_indicator = max(1, cpu_count // num_indicators)

    # [ADDED -- Colab/low-core portability] The line above already caps
    # each indicator's OWN process pool, but nothing capped how many of
    # those pools got created AT ONCE -- the ThreadPoolExecutor below used
    # to always launch all num_indicators (9) simultaneously regardless of
    # cpu_count, so a 9-vs-1-worker split just meant 9 separate
    # ProcessPoolExecutor pools (9 OS processes) all fighting over
    # however many logical CPUs actually exist. Harmless oversubscription
    # on a desktop with plenty of cores; on a 2-vCPU Colab VM (also a tiny
    # /dev/shm, which multiprocessing's own semaphores rely on -- the same
    # constraint that broke headless Chrome earlier) this caused one
    # indicator with genuinely trivial work (Breakout Probability's
    # single-day per-symbol loop) to sit starved for CPU time for
    # several minutes instead of milliseconds. Capping concurrent
    # indicator-level threads to cpu_count means only that many process
    # pools ever exist at once; the rest queue in the ThreadPoolExecutor
    # and start as soon as a slot frees, so Colab still gets real
    # (just narrower) parallelism instead of 9-way contention on 2 cores.
    max_concurrent_indicators = min(num_indicators, max(1, cpu_count))

    print(f"[PROCESS] Launching {num_indicators} indicator engines "
          f"({max_concurrent_indicators} running at a time, "
          f"{workers_per_indicator} worker process(es) each, {cpu_count} logical CPUs detected)...")

    matrix_results = {}   # label -> matrix_rows
    failures = {}          # label -> exception

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_concurrent_indicators) as executor:
        # [FIX] Every module's build_matrix() was updated (13-Jul-26 pass) to
        # require target_date as its 2nd positional arg -- see
        # excel_utils.restrict_to_target_date(). This call site was never
        # updated to match, so workers_per_indicator was silently landing in
        # the target_date slot (and max_workers was silently always None).
        # That's an int being fed into pd.Timestamp(target_date) inside every
        # indicator module -- would raise or silently misbehave, not just be
        # slow.
        future_to_label = {
            executor.submit(module.build_matrix, raw_data_dict, target_date, max_workers=workers_per_indicator): label
            for label, module, sheet_name in INDICATORS
        }
        for future in concurrent.futures.as_completed(future_to_label):
            label = future_to_label[future]
            try:
                matrix_results[label] = future.result()
                print(f"  -> DONE: {label} matrix computed ({len(matrix_results[label]) - 1} rows).")
            except Exception as exc:
                failures[label] = exc
                print(f"  [ERROR] {label} failed to compute: {exc}")

    print("-" * 60)

    # STEP 7: Write every successfully-computed matrix into the shared
    # workbook ONE AT A TIME, in the main thread. This is the part that
    # must stay serialized -- each write does load_workbook -> replace its
    # sheet -> save, and two concurrent writers would silently clobber each
    # other's changes (last save wins). The expensive work already happened
    # in Step 6, so this loop is just fast row-appends + saves.
    print("[PROCESS] Writing indicator matrices to workbook (sequential to avoid file races)...")
    for label, module, sheet_name in INDICATORS:
        if label not in matrix_results:
            continue  # already failed at the build stage above
        try:
            module.write_matrix(matrix_results[label], final_excel_path)
            print(f"  -> WRITTEN: '{sheet_name}' sheet updated.")
        except Exception as exc:
            failures[label] = exc
            print(f"  [ERROR] Failed writing '{sheet_name}' sheet: {exc}")

    print("-" * 60)

    # STEP 8: Auto-fit column widths across every sheet in the workbook,
    # including 'Reference' (each indicator's write_matrix() already
    # autofits its own sheet as it's written; this final pass just also
    # covers Reference and re-confirms the rest). Treated as non-fatal --
    # a cosmetic formatting failure shouldn't take down an otherwise
    # successful data run.
    try:
        print("[PROCESS] Auto-fitting column widths across all sheets...")
        from openpyxl import load_workbook
        wb = load_workbook(final_excel_path)
        excel_utils.autofit_all_sheets(wb)
        excel_utils.atomic_save(wb, final_excel_path)
        print("[SUCCESS] Column widths auto-fitted.")
    except Exception as e:
        print(f"[WARNING] Auto-fit pass failed (non-fatal): {e}")

    print("=" * 60)
    if failures:
        print(f"[WARNING] {len(failures)} of {num_indicators} indicator(s) failed:")
        for label, exc in failures.items():
            print(f"   - {label}: {exc}")
        print("   Successful indicators were still written to the workbook above.")
        print("=" * 60)
        # [CHANGED] Don't sys.exit(1) here -- a caller looping over a
        # BACKTEST date range needs to keep going for the other dates even
        # if one indicator failed on this one. The partial workbook (every
        # OTHER indicator's sheet, still written above) is still useful,
        # and final_sheet.py already degrades gracefully (treats a missing
        # sheet's vote as WAIT, not a crash) -- see its docstring. Raise so
        # the caller can log/count this date as partially-failed rather
        # than silently declaring success.
        raise RuntimeError(
            f"{len(failures)} of {num_indicators} indicator(s) failed for "
            f"{target_date.strftime('%d-%b-%y')}: {list(failures.keys())}"
        )

    print("   STATUS: READY.   ")
    print("=" * 60)
    return final_excel_path, df_ref


def run_pipeline():
    """Standalone single-date entry point -- broker login + one interactive
    date prompt + run_pipeline_for_date(). This is what running
    `python run_pipeline.py` directly still does, unchanged from before.
    For a BACKTEST date RANGE or a recurring LIVE session, use
    01_Master_Code.py instead -- it logs in once and calls
    run_pipeline_for_date() itself, once per trading date."""
    print("=" * 60)
    print("         ALGORITHMIC TRADING INITIALIZATION PIPELINE         ")
    print("=" * 60)
    # STEP 1: Broker Authentication
    try:
        smart_api = broker_auth.initialize_angel_one()
        kite_api = broker_auth.initialize_zerodha()
    except Exception as e:
        print(f"\n[FATAL] Step 1 Failed: Authentication Breakdown.\n{traceback.format_exc()}")
        sys.exit(1)
    # STEP 2: Execution Mode & Calendar Validation
    try:
        target_date, mode = calendar_mgmt.get_execution_date()
        print(f"[SYSTEM] Mode: {mode} | Date: {target_date.strftime('%d-%b-%y')}")
    except Exception as e:
        print(f"\n[FATAL] Step 2 Failed: Date Selection Error.\n{traceback.format_exc()}")
        sys.exit(1)

    try:
        run_pipeline_for_date(smart_api, kite_api, target_date, mode)
        print("   HANDING OVER TO LIVE TRACKING ENGINE.   ")
        print("=" * 60)
    except Exception as e:
        print(f"\n[FATAL] Pipeline failed for {target_date.strftime('%d-%b-%y')}: {e}")
        sys.exit(1)


if __name__ == "__main__":
    run_pipeline()