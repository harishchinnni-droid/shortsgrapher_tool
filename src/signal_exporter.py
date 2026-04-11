"""
Signal Excel Exporter Module
Exports trading signals to color-coded Excel files using openpyxl.
"""

import os
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.logging_config import LoggingManager

logger = LoggingManager.get_logger()


class SignalExcelExporter:
    """Exports signals to beautifully formatted Excel files with color coding."""
    
    def __init__(self, storage_dir: str = "./local_trading_data"):
        """
        Initialize Excel exporter.
        
        Args:
            storage_dir: Directory to store Excel files
        """
        self.storage_dir = storage_dir
        self.excel_dir = os.path.join(storage_dir, "excel_reports")
        os.makedirs(self.excel_dir, exist_ok=True)
        logger.info(f"Excel Exporter initialized at: {self.excel_dir}")
    
    def export_signals(self, signals: List[Dict[str, Any]], date_str: str = None) -> str:
        """
        Export signals to an Excel file with color coding.
        
        Args:
            signals: List of signal dictionaries
            date_str: Date string (optional, if None uses current datetime)
            
        Returns:
            str: Path to created Excel file
        """
        try:
            # Generate filename with current datetime
            if date_str is None:
                date_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            
            filename = f"{date_str}_Signals.xlsx"
            filepath = os.path.join(self.excel_dir, filename)
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Trading Signals"
            
            # Define color scheme
            colors = {
                "buy": "C6EFCE",      # Light green
                "sell": "FFC7CE",     # Light red
                "strong_buy": "00B050",  # Dark green
                "strong_sell": "FF0000",  # Dark red
                "neutral": "FFF2CC",  # Light yellow
                "header": "366092",   # Dark blue
                "text_white": "FFFFFF",
                "text_black": "000000"
            }
            
            # Define borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Create headers
            headers = [
                "Date", "Symbol", "Signal Type", "Signal Strength",
                "RSI", "VWAP", "ADX", "Squeeze Momentum",
                "EMA", "OI Dynamics", "Breakout Score",
                "Entry Price", "Stop Loss", "Take Profit"
            ]
            
            # Write header row
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = PatternFill(start_color=colors["header"], end_color=colors["header"], fill_type="solid")
                cell.font = Font(color=colors["text_white"], bold=True, size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            
            # Write data rows
            for row_idx, signal in enumerate(signals, 2):
                # Determine signal color
                signal_type = signal.get("signal_type", "").lower()
                if signal_type == "buy":
                    cell_color = colors["buy"]
                    text_color = colors["text_black"]
                elif signal_type == "sell":
                    cell_color = colors["sell"]
                    text_color = colors["text_black"]
                elif signal_type == "strong_buy":
                    cell_color = colors["strong_buy"]
                    text_color = colors["text_white"]
                elif signal_type == "strong_sell":
                    cell_color = colors["strong_sell"]
                    text_color = colors["text_white"]
                else:
                    cell_color = colors["neutral"]
                    text_color = colors["text_black"]
                
                # Write signal data
                data = [
                    signal.get("signal_date", ""),
                    signal.get("symbol", ""),
                    signal.get("signal_type", ""),
                    signal.get("signal_strength", ""),
                    self._format_value(signal.get("rsi_value")),
                    self._format_value(signal.get("vwap_value")),
                    self._format_value(signal.get("adx_value")),
                    self._format_value(signal.get("squeeze_momentum")),
                    self._format_value(signal.get("ema_value")),
                    self._format_value(signal.get("oi_dynamics")),
                    self._format_value(signal.get("breakout_score")),
                    self._format_value(signal.get("entry_price")),
                    self._format_value(signal.get("stop_loss")),
                    self._format_value(signal.get("take_profit"))
                ]
                
                # Apply colors and formatting to each cell
                for col_idx, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type="solid")
                    cell.font = Font(color=text_color, size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
            
            # Auto-adjust column widths
            column_widths = [12, 12, 15, 15, 10, 12, 10, 18, 10, 14, 15, 12, 12, 12]
            for col_idx, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            
            # Freeze header row
            ws.freeze_panes = "A2"
            
            # Save workbook
            wb.save(filepath)
            logger.info(f"✓ Excel file created: {filepath}")
            return filepath
        
        except Exception as e:
            logger.error(f"Failed to export signals to Excel: {e}")
            raise
    
    @staticmethod
    def _format_value(value: Any) -> str:
        """
        Format numeric values for display.
        
        Args:
            value: Value to format
            
        Returns:
            Formatted string or empty string
        """
        if value is None:
            return ""
        if isinstance(value, float):
            return f"{value:.2f}"
        return str(value)
    
    def get_export_directory(self) -> str:
        """Get the Excel export directory."""
        return self.excel_dir
