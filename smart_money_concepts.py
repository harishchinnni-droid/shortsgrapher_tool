"""
smart_money_concepts.py
------------------------
First version of the ICT/Smart Money Concepts signal Harish's training
material asked about (order blocks, liquidity sweeps, CHoCH, breaker
blocks -- see the 18-Jul-26 image review). Built on top of smc_lib.py
(vendored from joshyattridge/smart-money-concepts, MIT license -- see
that file's own header) rather than translating Pine Script by hand,
since a well-maintained, MIT-licensed Python implementation already
exists and this is a fundamentally different signal paradigm from every
other module in this pipeline (structure/order-flow, not an oscillator).

SCOPE -- deliberately narrow for a first version, not the full 12-pattern
catalog from the training material:
    Only Order Blocks (smc.ob()) are wired in here, on top of
    smc.swing_highs_lows() (a prerequisite input smc.ob() itself needs).
    Liquidity sweeps, Fair Value Gaps, CHoCH/BOS, and breaker blocks are
    all present in smc_lib.py (dormant, unused) for a later pass -- see
    that file's own docstring. This mirrors the review's own framing:
    Smart Money Concepts was flagged as "the biggest single ask... treated
    as one bucket, not six" and deserving its own scoped decision, not a
    quiet all-at-once addition alongside everything else.

WHAT AN ORDER BLOCK MEANS HERE:
    A bullish order block is (per ICT theory, matching the training
    images) the last down-close candle before a strong bullish move --
    "where institutions likely have resting buy orders." Price often
    returns to that zone once before continuing in the original
    direction; THAT RETURN (the "retest") is the entry, not the block's
    formation itself. smc.ob()'s own 'MitigatedIndex' marks exactly that
    retest bar -- see calculate_smc()'s docstring for how this becomes
    'SMC Recomm'.

Integration -- a GATE, not a vote:
    Same reasoning as every other experimental signal in this codebase:
    order-block retest is a structural/positional signal, not a
    restatement of TREND/MOMENTUM/VOLUME. Consumed as an off-by-default
    pre-entry GATE in order_sheet.py (ENABLE_SMC_GATE).

Timeframe: 5-minute candles, truncated per day to CUTOFF_TIME (15:15),
same convention as every other module in this pipeline. SWING_LENGTH is
intentionally much smaller than smc_lib's own default (50, built for
daily/swing charts) -- 5-min intraday data needs a much shorter lookback
to find swing points at all within a single session's ~75 bars.
"""

import os
import sys
import concurrent.futures
from datetime import time as dtime

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

CODES_DIR = os.path.dirname(os.path.abspath(__file__))
if CODES_DIR not in sys.path:
    sys.path.append(CODES_DIR)
import data_ingestion
import excel_utils
from smc_lib import smc

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
# [NOTE] smc_lib's swing_length is HALVED internally then doubled back
# (see swing_highs_lows()'s own `swing_length *= 2`) -- passing 5 here
# means a 10-bar-total centered window (~50 min at 5-min candles), a
# reasonable intraday scale. smc_lib's own default (50) is sized for
# daily/swing charts and would need ~500 bars of context, far more than
# a single session's ~75. Untested threshold -- needs its own
# sensitivity check once this gate has real backtest data behind it.
SWING_LENGTH = 5
CUTOFF_TIME = dtime(15, 15)
INTERVAL = "5minute"

# ---------------------------------------------------------------------------
# Excel styling (kept visually consistent with the rest of the workbook)
# ---------------------------------------------------------------------------
FILL_LIME = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")   # BUY CE
FILL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # BUY PE
FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")   # WAIT
FONT_WHITE = Font(color="FFFFFF")
FONT_BLACK = Font(color="000000")


# ---------------------------------------------------------------------------
# Indicator -- order blocks + retest detection
# ---------------------------------------------------------------------------
def calculate_smc(df, swing_length=SWING_LENGTH, max_zone_age=60):
    """Requires 'open','high','low','close','volume' columns. Returns df
    with 'SMC Recomm', 'SMC Zone Age', 'Last OB Dir', 'OB Top', 'OB Bottom'
    columns added.

    [DESIGN NOTE] smc.ob() itself also returns a 'MitigatedIndex' field,
    but that specifically marks where price's LOW fully broke BELOW a
    bullish block's bottom (or high above a bearish block's top) -- i.e.
    a genuine structural BREAK of the zone, which the library then either
    keeps as "touched" or fully erases if price later closes back through
    the opposite side. That is a different (and murkier, library-internal)
    concept from the classic ICT "clean retest" this module actually
    wants: price returning to trade WITHIN the block's [bottom, top]
    range (without fully breaking through) and reacting in the original
    direction -- see the training material's own order-block diagrams.
    Rather than lean on MitigatedIndex's specific semantics, this
    function tracks each order block's zone itself (using smc.ob()'s
    Top/Bottom/direction, which ARE exactly what the training images
    show) and defines the retest/entry condition explicitly and
    transparently below.

    RETEST RULE: a currently-tracked bullish zone fires 'BUY CE' the
    first bar where price's LOW dips to/into [bottom, top] (a genuine
    touch of the zone, not just an approach) WITHOUT closing below
    bottom (a full close through would be a real break, not a retest),
    AND that same bar closes bullish (close > open) -- a candle actually
    reacting off the zone, not just wicking through it. Bearish is the
    mirror. Each zone fires at most once (matches every other edge-
    triggered signal in this codebase); a zone not retested within
    max_zone_age bars is dropped as stale.

    smc_lib's positional-index logic (searchsorted, close_index-1, etc.)
    requires a plain 0..n-1 RangeIndex -- df is reset_index()'d before
    calling into it."""
    df = df.reset_index(drop=True)
    n = len(df)

    smc_input = df[['open', 'high', 'low', 'close', 'volume']]
    swing_hl = smc.swing_highs_lows(smc_input, swing_length=swing_length)
    ob_df = smc.ob(smc_input, swing_hl, close_mitigation=False)

    ob_dir = ob_df['OB'].to_numpy()      # 1 bullish / -1 bearish / NaN, at the block's ORIGIN bar
    ob_top = ob_df['Top'].to_numpy()
    ob_bottom = ob_df['Bottom'].to_numpy()

    open_np = df['open'].to_numpy(dtype=float)
    close_np = df['close'].to_numpy(dtype=float)
    high_np = df['high'].to_numpy(dtype=float)
    low_np = df['low'].to_numpy(dtype=float)

    smc_recomm = np.full(n, 'WAIT', dtype=object)
    ob_top_active = np.full(n, np.nan)
    ob_bottom_active = np.full(n, np.nan)

    # active_zones: list of dicts {dir, top, bottom, origin} for blocks
    # not yet retested (fired) or invalidated (fully broken/stale).
    active_zones = []
    for i in range(n):
        # New order block originates at this bar -- start tracking it.
        if not np.isnan(ob_dir[i]):
            active_zones.append({
                'dir': int(ob_dir[i]), 'top': ob_top[i], 'bottom': ob_bottom[i], 'origin': i,
            })

        still_active = []
        for zone in active_zones:
            if i <= zone['origin']:
                still_active.append(zone)
                continue
            age = i - zone['origin']
            if age > max_zone_age:
                continue  # stale -- drop silently, no signal

            if zone['dir'] == 1:
                # Full break: a CLOSE below bottom invalidates the zone outright.
                if close_np[i] < zone['bottom']:
                    continue
                touched = low_np[i] <= zone['top'] and low_np[i] >= zone['bottom']
                bullish_reaction = close_np[i] > open_np[i]
                if touched and bullish_reaction:
                    smc_recomm[i] = 'BUY CE'
                    ob_top_active[i] = zone['top']
                    ob_bottom_active[i] = zone['bottom']
                    continue  # fired -- drop from tracking (edge-triggered, once per zone)
                still_active.append(zone)
            else:
                if close_np[i] > zone['top']:
                    continue
                touched = high_np[i] >= zone['bottom'] and high_np[i] <= zone['top']
                bearish_reaction = close_np[i] < open_np[i]
                if touched and bearish_reaction:
                    smc_recomm[i] = 'BUY PE'
                    ob_top_active[i] = zone['top']
                    ob_bottom_active[i] = zone['bottom']
                    continue
                still_active.append(zone)
        active_zones = still_active

    # --- Zone Age + Last OB Dir: same persistent-state pairing as
    # support_resistance.py's Break Age / Last Break Dir, for the same
    # reason -- a pre-entry bar rarely lands exactly on the retest bar
    # itself, so a gate needs "how long ago, which direction" rather than
    # the edge-triggered 'SMC Recomm' alone. ---
    zone_age = np.zeros(n, dtype=int)
    last_ob_dir = np.zeros(n, dtype=int)
    age = -1
    direction_state = 0
    is_event = smc_recomm != 'WAIT'
    for i in range(n):
        if is_event[i]:
            age = 0
            direction_state = 1 if smc_recomm[i] == 'BUY CE' else -1
        elif age >= 0:
            age += 1
        zone_age[i] = age
        last_ob_dir[i] = direction_state

    df['SMC Recomm'] = smc_recomm
    df['SMC Zone Age'] = zone_age
    df['Last OB Dir'] = last_ob_dir
    df['OB Top'] = ob_top_active
    df['OB Bottom'] = ob_bottom_active

    return df


# ---------------------------------------------------------------------------
# Per-symbol processing (same brute-force time/timezone protocol as RSI.py)
# ---------------------------------------------------------------------------
def process_symbol(symbol_data):
    symbol, df, target_date = symbol_data
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip().str.lower()

    if not all(col in df.columns for col in ['open', 'high', 'low', 'close']):
        print(f"[WARNING] {symbol}: Missing required price columns. Discarding.")
        return symbol, None
    if 'volume' not in df.columns:
        df['volume'] = 0.0  # smc_lib requires the column to exist; OB volume-strength scoring degrades gracefully to 0

    time_col = next((col for col in ['date', 'time', 'datetime', 'timestamp'] if col in df.columns), None)
    if time_col:
        parsed = pd.to_datetime(df[time_col], errors='coerce')
    elif pd.api.types.is_datetime64_any_dtype(df.index):
        parsed = pd.Series(df.index, index=df.index)
    else:
        unnamed_cols = [c for c in df.columns if 'unnamed' in c]
        if unnamed_cols:
            parsed = pd.to_datetime(df[unnamed_cols[0]], errors='coerce')
        else:
            print(f"[WARNING] {symbol}: Failed to locate any valid timestamp data. Discarding.")
            return symbol, None

    try:
        if parsed.dt.tz is not None:
            parsed = parsed.dt.tz_localize(None)
    except (AttributeError, TypeError):
        pass

    df['_sort_dt'] = parsed
    df['time_str'] = parsed.dt.strftime('%H:%M')
    df.dropna(subset=['time_str', '_sort_dt'], inplace=True)
    if df.empty:
        print(f"[WARNING] {symbol}: No rows with a valid timestamp. Discarding.")
        return symbol, None

    df.sort_values('_sort_dt', inplace=True)
    df = df[df['_sort_dt'].dt.time <= CUTOFF_TIME]
    if df.empty:
        print(f"[WARNING] {symbol}: No candles at/before {CUTOFF_TIME.strftime('%H:%M')}. Discarding.")
        return symbol, None

    try:
        df = calculate_smc(df)
    except Exception as e:
        print(f"[WARNING] {symbol}: SMC calculation failed ({e}). Discarding.")
        return symbol, None

    df = excel_utils.restrict_to_target_date(df, target_date)
    if df is None:
        return symbol, None

    return symbol, df


# ---------------------------------------------------------------------------
# Excel export -- pivoted Symbol x Time matrix (same pattern as zerolag.py)
# ---------------------------------------------------------------------------
def build_matrix(data_dict, target_date, max_workers=None):
    results = {}
    with concurrent.futures.ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(process_symbol, (sym, df, target_date)): sym for sym, df in data_dict.items()}
        for future in concurrent.futures.as_completed(futures):
            sym = futures[future]
            try:
                processed_sym, processed_df = future.result()
                if processed_df is not None and 'SMC Recomm' in processed_df.columns:
                    if not processed_df.empty and sym.lower() != 'summary':
                        results[sym] = processed_df
            except Exception as e:
                print(f"[ERROR] Smart Money Concepts Engine Failure on symbol {sym}: {str(e)}")

    if not results:
        raise RuntimeError("Smart Money Concepts: zero symbols processed successfully for target_date -- matrix build aborted.")

    all_times = set()
    for df in results.values():
        if 'time_str' in df.columns:
            all_times.update(df['time_str'].dropna().tolist())

    if not all_times:
        raise RuntimeError("Smart Money Concepts: no valid timestamps extracted across all symbols -- matrix build aborted.")

    sorted_times = sorted(all_times)
    matrix_rows = [['Symbol', 'Metrics'] + sorted_times]

    for sym in sorted(results.keys()):
        df = results[sym].drop_duplicates(subset=['time_str'], keep='last')
        df_indexed = df.set_index('time_str')

        def get_metric_row(metric_name, col_name):
            row = [sym, metric_name]
            for t in sorted_times:
                if t in df_indexed.index:
                    val = df_indexed.loc[t, col_name]
                    if isinstance(val, pd.Series):
                        val = val.iloc[-1]
                    row.append(val)
                else:
                    row.append("")
            return row

        matrix_rows.append(get_metric_row('SMC Recomm', 'SMC Recomm'))
        matrix_rows.append(get_metric_row('SMC Zone Age', 'SMC Zone Age'))
        matrix_rows.append(get_metric_row('Last OB Dir', 'Last OB Dir'))
        matrix_rows.append(get_metric_row('OB Top', 'OB Top'))
        matrix_rows.append(get_metric_row('OB Bottom', 'OB Bottom'))

    return matrix_rows


def write_matrix_to_workbook(matrix_rows, wb):
    sheet_name = "SMC"
    ws = excel_utils.replace_sheet_with_matrix(wb, sheet_name, matrix_rows)

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1), start=2):
        metric_type = ws.cell(row=r_idx, column=2).value
        for cell in row[2:]:
            val = cell.value
            if val == "":
                continue
            if metric_type == 'SMC Recomm':
                if val == 'BUY CE':
                    cell.fill = FILL_LIME
                    cell.font = FONT_BLACK
                elif val == 'BUY PE':
                    cell.fill = FILL_RED
                    cell.font = FONT_WHITE
                elif val == 'WAIT':
                    cell.fill = FILL_GRAY
                    cell.font = FONT_BLACK

    excel_utils.autofit_columns(ws)


def write_matrix(matrix_rows, output_excel_path):
    wb = load_workbook(output_excel_path)
    write_matrix_to_workbook(matrix_rows, wb)
    excel_utils.atomic_save(wb, output_excel_path)


def parallel_compute_and_export(data_dict, target_date, output_excel_path, max_workers=None):
    matrix_rows = build_matrix(data_dict, target_date, max_workers=max_workers)
    write_matrix(matrix_rows, output_excel_path)


def run_smc_step(df_ref, target_date, output_excel_path):
    print("[SYSTEM] Loading 5-minute historical data for Smart Money Concepts...")
    data_dict = data_ingestion.load_interval_data(df_ref, target_date, interval=INTERVAL)
    if not data_dict:
        raise RuntimeError("Smart Money Concepts: no 5-minute data files found for any symbol.")
    print(f"[PROCESS] Computing Smart Money Concepts matrix for {len(data_dict)} symbol(s)...")
    parallel_compute_and_export(data_dict, target_date, output_excel_path)
    print("[SUCCESS] Smart Money Concepts matrix written to sheet 'SMC'.")


# ---------------------------------------------------------------------------
# Disclaimer: this module derives order-block zones and retest events from
# historical/live price and volume data only. It is not financial advice,
# and no result here is a guarantee of future performance -- paper-trade it
# and run it through scripts/backtester.py before risking real capital.
# See order_sheet.py's ENABLE_SMC_GATE for how (and whether) this actually
# gates entries.
# ---------------------------------------------------------------------------
