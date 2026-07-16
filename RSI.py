"""
RSI.py
------
Python translation of the RSI logic used in the master Colab pipeline's
`_compute_indicators_for_symbol()` (the `r_rsi` block: standard Wilder
RSI(14) plus an EMA 25/50/100 trend-confirmation filter), wired to run as
Step 7 of the daily pipeline (after SQZMOM). Output format mirrors
SQZMOM.py's pivoted Symbol x Time matrix so both sheets look and behave
the same way in the workbook.

Indicator logic (mirrors the Colab script's `r_rsi` block exactly):
    - RSI = standard Wilder RSI(14) on 'close' (ewm(alpha=1/14) smoothed
      gains/losses -- identical formula to the source).
    - EMA_25 / EMA_50 / EMA_100 = EMA(close, span=25/50/100).
    - Per-bar 'RSI Recomm' (NOT hold-based -- recomputed fresh every bar,
      same convention as ema20.py/vwap.py):
          BUY CE if EMA_50 > EMA_100 AND EMA_25 > EMA_50 AND 50 < RSI <= 70
          BUY PE if EMA_50 < EMA_100 AND EMA_25 < EMA_50 AND 30 <= RSI < 50
          WAIT   otherwise

Timeframe: 5-minute candles, truncated per day to CUTOFF_TIME (default
15:15, i.e. the 15:15-15:20 candle is the last one included).

Timezone note: Kite's historical_data() returns tz-aware (Asia/Kolkata)
timestamps. openpyxl/Excel cannot store a tz-aware datetime cell -- writing
one raises "TypeError: Excel does not support timezones in datetimes."
This module never writes a raw datetime into a cell (only a 'HH:MM' string
column header, like SQZMOM.py does), and additionally strips tz info right
after parsing as a defensive measure -- see process_symbol().

[FIX -- 13-Jul-26] process_symbol()/build_matrix() now require target_date
and restrict the OUTPUT to that day's own rows via
excel_utils.restrict_to_target_date(), called after the indicator has been
computed on the full multi-day history (needed for the EMA warmup).
Previously the Symbol x Time pivot deduped purely on time-of-day with no
date component, so a time slot target_date hadn't reached yet silently
fell back to the most recent PRIOR trading day's value -- which is what
let the workbook show a full day of RSI numbers even before the market
had opened. See excel_utils.restrict_to_target_date()'s docstring and
01_Master_Code.py's market-open scheduling fix, which the two are
designed to work together with.
"""

import os
import sys
import concurrent.futures
from datetime import time as dtime

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

CODES_DIR = os.path.dirname(os.path.abspath(__file__))
if CODES_DIR not in sys.path:
    sys.path.append(CODES_DIR)
import data_ingestion
import excel_utils

# ---------------------------------------------------------------------------
# Config -- mirrors the Colab script's RSI/EMA inputs
# ---------------------------------------------------------------------------
RSI_PERIOD = 14                       # Wilder RSI length, same as the Colab source
CUTOFF_TIME = dtime(15, 15)          # truncate each day's candles here
INTERVAL = "5minute"                  # must match data_ingestion.py's interval naming

# ---------------------------------------------------------------------------
# Excel styling (kept visually consistent with SQZMOM.py's palette)
# ---------------------------------------------------------------------------
FILL_LIME = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")   # BUY CE
FILL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # BUY PE
FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")   # WAIT
FONT_WHITE = Font(color="FFFFFF")
FONT_BLACK = Font(color="000000")


# ---------------------------------------------------------------------------
# Indicator -- exact translation of the Colab script's `r_rsi` block
# ---------------------------------------------------------------------------
def _rsi(close, period=RSI_PERIOD):
    """Standard Wilder RSI, same formula as the Colab source:
    gain/loss smoothed with ewm(alpha=1/period, adjust=False), then
    RSI = 100 - 100/(1+rs). No neutral-fill override -- if avg_loss is
    ever exactly 0, rs -> inf and RSI -> 100, matching the source's
    (unguarded) behavior rather than silently overriding it."""
    delta = close.diff()
    gain = delta.where(delta > 0, 0.0).ewm(alpha=1.0 / period, adjust=False).mean()
    loss = (-delta.where(delta < 0, 0.0)).ewm(alpha=1.0 / period, adjust=False).mean()
    rs = gain / loss
    return 100 - (100 / (1 + rs))


def calculate_rsi_ema(df, period=RSI_PERIOD):
    """Per-bar RSI + EMA25/50/100 trend-confirmation signal -- the exact
    logic of the Colab script's `r_rsi` block. Recomputed fresh on every
    bar (no hold state, no memory of previous bars), same convention as
    ema20.py/vwap.py:

        BUY CE if EMA_50 > EMA_100 AND EMA_25 > EMA_50 AND 50 < RSI <= 70
        BUY PE if EMA_50 < EMA_100 AND EMA_25 < EMA_50 AND 30 <= RSI < 50
        WAIT   otherwise
    """
    df = df.copy()

    df['EMA_25'] = df['close'].ewm(span=25, adjust=False).mean()
    df['EMA_50'] = df['close'].ewm(span=50, adjust=False).mean()
    df['EMA_100'] = df['close'].ewm(span=100, adjust=False).mean()
    df['RSI'] = _rsi(df['close'], period)

    buy_ce = (df['EMA_50'] > df['EMA_100']) & (df['EMA_25'] > df['EMA_50']) & (df['RSI'] > 50) & (df['RSI'] <= 70)
    buy_pe = (df['EMA_50'] < df['EMA_100']) & (df['EMA_25'] < df['EMA_50']) & (df['RSI'] >= 30) & (df['RSI'] < 50)

    df['RSI Recomm'] = np.where(buy_ce, 'BUY CE', np.where(buy_pe, 'BUY PE', 'WAIT'))

    return df


# ---------------------------------------------------------------------------
# Per-symbol processing
# ---------------------------------------------------------------------------
def process_symbol(symbol_data):
    """Normalizes columns, extracts a 'HH:MM' time_str (brute-force, same
    protocol as SQZMOM.py), strips timezone info, truncates to CUTOFF_TIME,
    runs the indicator on the full multi-day history, then restricts the
    returned rows to target_date only. Returns (symbol, df) or
    (symbol, None)."""
    symbol, df, target_date = symbol_data
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip().str.lower()

    if 'close' not in df.columns:
        print(f"[WARNING] {symbol}: Missing required 'close' column. Discarding.")
        return symbol, None

    time_col = next((col for col in ['date', 'time', 'datetime', 'timestamp'] if col in df.columns), None)
    if time_col:
        parsed = pd.to_datetime(df[time_col], errors='coerce')
    elif pd.api.types.is_datetime64_any_dtype(df.index):
        parsed = pd.Series(df.index, index=df.index)
    else:
        unnamed_cols = [c for c in df.columns if 'unnamed' in c]
        if unnamed_cols:
            parsed = pd.to_datetime(df[unnamed_cols[0]], errors='coerce')
        else:
            print(f"[WARNING] {symbol}: Failed to locate any valid timestamp data. Discarding.")
            return symbol, None

    # Strip tz label if present (Kite returns Asia/Kolkata tz-aware
    # timestamps). This keeps the same local wall-clock time; it just
    # removes the offset tag that openpyxl can't write to a cell.
    try:
        if parsed.dt.tz is not None:
            parsed = parsed.dt.tz_localize(None)
    except (AttributeError, TypeError):
        pass

    df['_sort_dt'] = parsed
    df['time_str'] = parsed.dt.strftime('%H:%M')
    df.dropna(subset=['time_str', '_sort_dt'], inplace=True)
    if df.empty:
        print(f"[WARNING] {symbol}: No rows with a valid timestamp. Discarding.")
        return symbol, None

    df.sort_values('_sort_dt', inplace=True)
    df = df[df['_sort_dt'].dt.time <= CUTOFF_TIME]
    if df.empty:
        print(f"[WARNING] {symbol}: No candles at/before {CUTOFF_TIME.strftime('%H:%M')}. Discarding.")
        return symbol, None

    df = calculate_rsi_ema(df)

    # [FIX] Date-scope the OUTPUT to target_date only -- see module
    # docstring and excel_utils.restrict_to_target_date()'s docstring.
    df = excel_utils.restrict_to_target_date(df, target_date)
    if df is None:
        return symbol, None

    return symbol, df


# ---------------------------------------------------------------------------
# Excel export -- pivoted Symbol x Time matrix, merged into the existing
# workbook (same pattern as SQZMOM.py: load_workbook, replace only the
# 'RSI' sheet, leave every other sheet untouched).
# ---------------------------------------------------------------------------
def build_matrix(data_dict, target_date, max_workers=None):
    """Executes multi-core calculation and builds the pivoted Matrix
    (no Excel I/O -- safe to run concurrently with other indicators'
    build_matrix() calls since it never touches the shared workbook).

    [CHANGED] Now requires target_date -- see process_symbol()."""
    results = {}
    with concurrent.futures.ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(process_symbol, (sym, df, target_date)): sym for sym, df in data_dict.items()}
        for future in concurrent.futures.as_completed(futures):
            sym = futures[future]
            try:
                processed_sym, processed_df = future.result()
                if processed_df is not None and 'RSI' in processed_df.columns:
                    processed_df = processed_df.dropna(subset=['RSI'])
                    if not processed_df.empty and sym.lower() != 'summary':
                        results[sym] = processed_df
            except Exception as e:
                print(f"[ERROR] RSI Engine Failure on symbol {sym}: {str(e)}")

    if not results:
        # [NOTE] Now also fires if called before target_date's first
        # candle has closed -- see SQZMOM.py's build_matrix() docstring
        # for why that's the intended (loud-failure-over-stale-data)
        # behavior.
        raise RuntimeError("RSI: zero symbols processed successfully for target_date -- matrix build aborted.")

    all_times = set()
    for df in results.values():
        if 'time_str' in df.columns:
            all_times.update(df['time_str'].dropna().tolist())

    if not all_times:
        raise RuntimeError("RSI: no valid timestamps extracted across all symbols -- matrix build aborted.")

    sorted_times = sorted(all_times)

    matrix_rows = [['Symbol', 'Metrics'] + sorted_times]

    for sym in sorted(results.keys()):
        df = results[sym].drop_duplicates(subset=['time_str'], keep='last')
        df_indexed = df.set_index('time_str')

        def get_metric_row(metric_name, col_name):
            row = [sym, metric_name]
            for t in sorted_times:
                if t in df_indexed.index:
                    val = df_indexed.loc[t, col_name]
                    if isinstance(val, pd.Series):
                        val = val.iloc[-1]
                    row.append(val)
                else:
                    row.append("")
            return row

        if 'open' in df_indexed.columns:
            matrix_rows.append(get_metric_row('Open', 'open'))
        matrix_rows.append(get_metric_row('Close', 'close'))
        matrix_rows.append(get_metric_row('EMA 25', 'EMA_25'))
        matrix_rows.append(get_metric_row('EMA 50', 'EMA_50'))
        matrix_rows.append(get_metric_row('EMA 100', 'EMA_100'))
        matrix_rows.append(get_metric_row('RSI', 'RSI'))
        matrix_rows.append(get_metric_row('RSI Recomm', 'RSI Recomm'))

    return matrix_rows


def write_matrix_to_workbook(matrix_rows, wb):
    """Applies the 'RSI' sheet into an ALREADY-OPEN workbook (wb) -- no
    file load/save here. See excel_utils.replace_sheet_with_matrix()'s
    docstring for why an orchestrator should batch all indicators
    through this instead of write_matrix()'s own load/save."""
    sheet_name = "RSI"
    ws = excel_utils.replace_sheet_with_matrix(wb, sheet_name, matrix_rows)

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1), start=2):
        metric_type = ws.cell(row=r_idx, column=2).value
        for cell in row[2:]:
            val = cell.value
            if val == "":
                continue
            if metric_type == 'RSI Recomm':
                if val == 'BUY CE':
                    cell.fill = FILL_LIME
                    cell.font = FONT_BLACK
                elif val == 'BUY PE':
                    cell.fill = FILL_RED
                    cell.font = FONT_WHITE
                elif val == 'WAIT':
                    cell.fill = FILL_GRAY
                    cell.font = FONT_BLACK

    excel_utils.autofit_columns(ws)


def write_matrix(matrix_rows, output_excel_path):
    """Standalone/back-compat entry point: load -> apply -> save. Prefer
    write_matrix_to_workbook() when writing several indicators in the
    same run (see 01_Master_Code.py)."""
    wb = load_workbook(output_excel_path)
    write_matrix_to_workbook(matrix_rows, wb)
    excel_utils.atomic_save(wb, output_excel_path)


def parallel_compute_and_export(data_dict, target_date, output_excel_path, max_workers=None):
    """Backward-compatible fused entry point (build + write in one call).
    [CHANGED] Now requires target_date -- see build_matrix()."""
    matrix_rows = build_matrix(data_dict, target_date, max_workers=max_workers)
    write_matrix(matrix_rows, output_excel_path)


def run_rsi_step(df_ref, target_date, output_excel_path):
    """Single entry point for run_pipeline.py's Step 7 (standalone use)."""
    print("[SYSTEM] Loading 5-minute historical data for RSI...")
    data_dict = data_ingestion.load_interval_data(df_ref, target_date, interval=INTERVAL)
    if not data_dict:
        raise RuntimeError("RSI: no 5-minute data files found for any symbol.")
    print(f"[PROCESS] Computing RSI + EMA25/50/100 matrix for {len(data_dict)} symbol(s)...")
    parallel_compute_and_export(data_dict, target_date, output_excel_path)
    print("[SUCCESS] RSI matrix written to sheet 'RSI'.")


# ---------------------------------------------------------------------------
# Disclaimer: this module derives a signal from historical/live price data
# only. It is not financial advice, and no result here is a guarantee of
# future performance -- paper-trade it and run it through
# scripts/backtester.py before risking real capital.
# ---------------------------------------------------------------------------
