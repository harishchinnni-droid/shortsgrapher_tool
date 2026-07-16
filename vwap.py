"""
vwap.py
-------
Vectorized Python translation of the 'VWAP' sheet from the Google Colab
reference (the `v_rec` block inside `_compute_indicators_for_symbol`).
Output format mirrors RSI.py / SQZMOM.py's pivoted Symbol x Time matrix.

Indicator logic:
    Session-anchored (resets every calendar day) Volume-Weighted Average
    Price:
        TP  = (high + low + close) / 3
        VWAP = cumsum(TP * volume) / cumsum(volume), cumulative sums reset
               at the start of each calendar day (groupby date).
    'VWAP Recomm' (per-bar, NOT hold-based -- see note below):
        BUY CE if close > VWAP.
        BUY PE if close < VWAP.
        WAIT if VWAP isn't computable yet (e.g. zero cumulative volume) or
        close == VWAP exactly.

Translation note -- per-bar, not flip-and-hold:
    Same as ema20.py: the source recomputes 'VWAP Recomm' fresh on every
    bar directly from that bar's close vs VWAP, with no memory of previous
    bars -- unlike SQZMOM/RSI/BRKPRO/ADX/TW ALL's flip-and-hold
    state machines. This is a literal translation of the source, not an
    oversight.

Timeframe: 5-minute candles, truncated per day to CUTOFF_TIME. The daily
VWAP reset itself is computed BEFORE that CUTOFF_TIME truncation (on the
full multi-day series, grouped by each row's own calendar date), so the
CUTOFF_TIME filter only ever drops each day's post-15:15 candles -- it
never removes the earlier same-day candles a later bar's cumulative VWAP
depends on.

[FIX -- 13-Jul-26] Two changes in this pass:
1. process_symbol()/build_matrix() now require target_date and restrict
   the OUTPUT to that day's own rows via
   excel_utils.restrict_to_target_date(), called AFTER calculate_vwap()
   runs on the full multi-day session-reset series. Previously the
   time_str-dedup-keep-last pivot trick silently fell back to a prior
   trading day's VWAP/Recomm before target_date's own candles existed.
   See excel_utils.restrict_to_target_date()'s docstring and
   01_Master_Code.py's market-open scheduling fix.
2. write_matrix() now goes through the same
   excel_utils.replace_sheet_with_matrix() + write_matrix_to_workbook()
   split every other indicator module uses, instead of hand-rolling its
   own load/del-sheet/create-sheet logic.
"""

import os
import sys
import concurrent.futures
from datetime import time as dtime

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

CODES_DIR = os.path.dirname(os.path.abspath(__file__))
if CODES_DIR not in sys.path:
    sys.path.append(CODES_DIR)
import excel_utils

CUTOFF_TIME = dtime(15, 15)
INTERVAL = "5minute"

# ---------------------------------------------------------------------------
# Excel styling (kept visually consistent with the other sheets)
# ---------------------------------------------------------------------------
FILL_LIME = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")   # BUY CE
FILL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # BUY PE
FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")   # WAIT
FONT_WHITE = Font(color="FFFFFF")
FONT_BLACK = Font(color="000000")


# ---------------------------------------------------------------------------
# Indicator
# ---------------------------------------------------------------------------
def calculate_vwap(df):
    """`df` must be sorted ascending by '_sort_dt' with lowercase
    'open'/'high'/'low'/'close'/'volume' columns -- the daily reset groups
    on '_sort_dt'.dt.date."""
    df = df.copy()
    date_only = df['_sort_dt'].dt.date

    tp = (df['high'] + df['low'] + df['close']) / 3.0
    tp_v = tp * df['volume']

    cum_tp_v = tp_v.groupby(date_only).cumsum()
    cum_vol = df['volume'].groupby(date_only).cumsum()

    df['VWAP'] = cum_tp_v / cum_vol.replace(0, np.nan)
    df['VWAP Recomm'] = compute_vwap_recomm(df['close'].to_numpy(), df['VWAP'].to_numpy())
    return df


def compute_vwap_recomm(close_vals, vwap_vals):
    """Direct per-bar close-vs-VWAP rule (see module docstring -- no
    hold-over across bars). Fully vectorized. Returns a numpy object array
    of 'BUY CE' / 'BUY PE' / 'WAIT'."""
    n = len(close_vals)
    recomm = np.full(n, 'WAIT', dtype=object)
    has_vwap = ~np.isnan(vwap_vals)
    recomm[has_vwap & (close_vals > vwap_vals)] = 'BUY CE'
    recomm[has_vwap & (close_vals < vwap_vals)] = 'BUY PE'
    return recomm


# ---------------------------------------------------------------------------
# Per-symbol processing (same brute-force time/timezone protocol as RSI.py)
# ---------------------------------------------------------------------------
def process_symbol(symbol_data):
    symbol, df, target_date = symbol_data
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip().str.lower()

    if not all(col in df.columns for col in ['open', 'high', 'low', 'close', 'volume']):
        print(f"[WARNING] {symbol}: Missing required price/volume columns. Discarding.")
        return symbol, None

    time_col = next((col for col in ['date', 'time', 'datetime', 'timestamp'] if col in df.columns), None)
    if time_col:
        parsed = pd.to_datetime(df[time_col], errors='coerce')
    elif pd.api.types.is_datetime64_any_dtype(df.index):
        parsed = pd.Series(df.index, index=df.index)
    else:
        unnamed_cols = [c for c in df.columns if 'unnamed' in c]
        if unnamed_cols:
            parsed = pd.to_datetime(df[unnamed_cols[0]], errors='coerce')
        else:
            print(f"[WARNING] {symbol}: Failed to locate any valid timestamp data. Discarding.")
            return symbol, None

    try:
        if parsed.dt.tz is not None:
            parsed = parsed.dt.tz_localize(None)
    except (AttributeError, TypeError):
        pass

    df['_sort_dt'] = parsed
    df['time_str'] = parsed.dt.strftime('%H:%M')
    df.dropna(subset=['time_str', '_sort_dt'], inplace=True)
    if df.empty:
        print(f"[WARNING] {symbol}: No rows with a valid timestamp. Discarding.")
        return symbol, None

    df.sort_values('_sort_dt', inplace=True)
    df = df[df['_sort_dt'].dt.time <= CUTOFF_TIME]
    if df.empty:
        print(f"[WARNING] {symbol}: No candles at/before {CUTOFF_TIME.strftime('%H:%M')}. Discarding.")
        return symbol, None

    df = calculate_vwap(df)

    # [FIX] Date-scope the OUTPUT to target_date only -- see module
    # docstring and excel_utils.restrict_to_target_date()'s docstring.
    df = excel_utils.restrict_to_target_date(df, target_date)
    if df is None:
        return symbol, None

    return symbol, df


# ---------------------------------------------------------------------------
# Excel export -- pivoted Symbol x Time matrix, merged into the existing
# workbook (same pattern as RSI.py / SQZMOM.py: load_workbook, replace only
# the 'VWAP' sheet, leave every other sheet untouched).
# ---------------------------------------------------------------------------
def build_matrix(data_dict, target_date, max_workers=None):
    """Executes multi-core calculation and builds the pivoted Matrix
    (no Excel I/O -- safe to run concurrently with other indicators'
    build_matrix() calls since it never touches the shared workbook).

    [CHANGED] Now requires target_date -- see process_symbol()."""
    results = {}
    with concurrent.futures.ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(process_symbol, (sym, df, target_date)): sym for sym, df in data_dict.items()}
        for future in concurrent.futures.as_completed(futures):
            sym = futures[future]
            try:
                processed_sym, processed_df = future.result()
                if processed_df is not None and 'VWAP Recomm' in processed_df.columns:
                    if not processed_df.empty and sym.lower() != 'summary':
                        results[sym] = processed_df
            except Exception as e:
                print(f"[ERROR] VWAP Engine Failure on symbol {sym}: {str(e)}")

    if not results:
        raise RuntimeError("VWAP: zero symbols processed successfully for target_date -- matrix build aborted.")

    all_times = set()
    for df in results.values():
        if 'time_str' in df.columns:
            all_times.update(df['time_str'].dropna().tolist())

    if not all_times:
        raise RuntimeError("VWAP: no valid timestamps extracted across all symbols -- matrix build aborted.")

    sorted_times = sorted(all_times)
    matrix_rows = [['Symbol', 'Metrics'] + sorted_times]

    for sym in sorted(results.keys()):
        df = results[sym].drop_duplicates(subset=['time_str'], keep='last')
        df_indexed = df.set_index('time_str')

        def get_metric_row(metric_name, col_name):
            row = [sym, metric_name]
            for t in sorted_times:
                if t in df_indexed.index:
                    val = df_indexed.loc[t, col_name]
                    if isinstance(val, pd.Series):
                        val = val.iloc[-1]
                    row.append(val)
                else:
                    row.append("")
            return row

        matrix_rows.append(get_metric_row('Open', 'open'))
        matrix_rows.append(get_metric_row('Close', 'close'))
        matrix_rows.append(get_metric_row('Volume', 'volume'))
        matrix_rows.append(get_metric_row('VWAP', 'VWAP'))
        matrix_rows.append(get_metric_row('VWAP Recomm', 'VWAP Recomm'))

    return matrix_rows


def write_matrix_to_workbook(matrix_rows, wb):
    """[CHANGED] Applies the 'VWAP' sheet into an ALREADY-OPEN workbook
    (wb) -- no file load/save here, same split every other indicator
    module uses."""
    sheet_name = "VWAP"
    ws = excel_utils.replace_sheet_with_matrix(wb, sheet_name, matrix_rows)
    ws.freeze_panes = "C2"

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1), start=2):
        metric_type = ws.cell(row=r_idx, column=2).value
        for cell in row[2:]:
            val = cell.value
            if val == "":
                continue
            if metric_type == 'VWAP Recomm':
                if val == 'BUY CE':
                    cell.fill = FILL_LIME
                    cell.font = FONT_BLACK
                elif val == 'BUY PE':
                    cell.fill = FILL_RED
                    cell.font = FONT_WHITE
                elif val == 'WAIT':
                    cell.fill = FILL_GRAY
                    cell.font = FONT_BLACK

    excel_utils.autofit_columns(ws)


def write_matrix(matrix_rows, output_excel_path):
    """Standalone/back-compat entry point: load -> apply -> save. Prefer
    write_matrix_to_workbook() when writing several indicators in the
    same run (see 01_Master_Code.py)."""
    wb = load_workbook(output_excel_path)
    write_matrix_to_workbook(matrix_rows, wb)
    excel_utils.atomic_save(wb, output_excel_path)


def parallel_compute_and_export(data_dict, target_date, output_excel_path, max_workers=None):
    """Backward-compatible fused entry point (build + write in one call).
    [CHANGED] Now requires target_date -- see build_matrix()."""
    matrix_rows = build_matrix(data_dict, target_date, max_workers=max_workers)
    write_matrix(matrix_rows, output_excel_path)


def run_vwap_step(df_ref, target_date, output_excel_path):
    """Single entry point for run_pipeline.py-style standalone use."""
    import data_ingestion
    print("[SYSTEM] Loading 5-minute historical data for VWAP...")
    data_dict = data_ingestion.load_interval_data(df_ref, target_date, interval=INTERVAL)
    if not data_dict:
        raise RuntimeError("VWAP: no 5-minute data files found for any symbol.")
    print(f"[PROCESS] Computing VWAP matrix for {len(data_dict)} symbol(s)...")
    parallel_compute_and_export(data_dict, target_date, output_excel_path)
    print("[SUCCESS] VWAP matrix written to sheet 'VWAP'.")


# ---------------------------------------------------------------------------
# Disclaimer: this module estimates a directional bias from price/volume
# data only. It is not financial advice, no result here is a guarantee of
# future performance, and the 'VWAP Recomm' rule has not been backtested
# in this conversation -- paper-trade it and run it through
# scripts/backtester.py before risking real capital.
# ---------------------------------------------------------------------------
