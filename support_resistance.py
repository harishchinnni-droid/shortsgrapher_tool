"""
support_resistance.py
----------------------
Python translation of LuxAlgo's "Support and Resistance Levels with
Breaks" Pine Script v4 indicator (see '04_ProjectFiles/Support &
Resistance_LuxAlgo.txt'), wired to run as part of the daily indicator
batch. Output format mirrors zerolag.py / supertrend_ai.py's pivoted
Symbol x Time matrix.

Indicator logic:
    A pivot high/low (leftBars=15, rightBars=15 by default) marks a
    swing point once `rightBars` bars have closed AFTER it -- i.e. a
    pivot at bar i is only KNOWABLE starting at bar i+rightBars, and this
    module additionally delays it one more bar to match the Pine
    source's own `[1]` shift (`fixnan(pivothigh(left,right)[1])`). That
    confirmed level then holds (forward-fills) as the active
    support/resistance line until the next pivot replaces it. This delay
    is deliberate, not a bug -- see 'NO-LOOKAHEAD' below.

    A "break" fires when close crosses the active support/resistance
    level AND a 5/10-period volume EMA oscillator exceeds
    VOLUME_THRESHOLD (20 by default) -- i.e. a real increase in
    participation behind the move, not just a random tick through an old
    level. A crossing WITHOUT that volume confirmation, or one that looks
    more like a rejection wick than a clean break (see WICK_FLAG), does
    NOT count as a break.

Integration -- a GATE, not a vote:
    'SR Recomm' is NOT added to final_sheet.py's confluence vote --
    support/resistance is a structural/positional signal (is price
    breaking a real level with volume), a genuinely different dimension
    from the TREND/MOMENTUM/VOLUME categories already voted there, not a
    restatement of any of them. Consumed instead as a pre-entry GATE in
    order_sheet.py (ENABLE_SR_GATE), off by default, same A/B pattern as
    every other experimental gate in this codebase.

    Unlike Zero-Lag's Trend Dir (a PERSISTENT regime state), a support/
    resistance break is inherently a momentary EVENT -- price crossing a
    level is a one-time thing, not an ongoing condition -- so 'SR Recomm'
    is edge-triggered (BUY CE/BUY PE only on the exact break bar, WAIT
    otherwise), with a separate 'SR Break Age' counter (0 on the break
    bar, incrementing after) so a gate can require FRESHNESS the same
    way ENABLE_ZEROLAG_FRESHNESS does for the Zero-Lag cloud -- this
    mirrors the lesson from that flag: a stale, long-past event
    shouldn't count the same as one that just happened.

NO-LOOKAHEAD: the pivot at bar i is only revealed starting at bar
i+rightBars+1 (see calculate_support_resistance()'s shift/ffill/shift
sequence, which reproduces Pine's `fixnan(pivothigh(...)[1])` exactly).
A pre-entry check at bar t can therefore only ever see support/
resistance levels that were ALREADY confirmed strictly before t -- never
a pivot the market hasn't finished forming yet.

Timeframe: 5-minute candles, truncated per day to CUTOFF_TIME (15:15),
same convention as every other module in this pipeline.
"""

import os
import sys
import concurrent.futures
from datetime import time as dtime

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

CODES_DIR = os.path.dirname(os.path.abspath(__file__))
if CODES_DIR not in sys.path:
    sys.path.append(CODES_DIR)
import data_ingestion
import excel_utils

# ---------------------------------------------------------------------------
# Config -- mirrors the Pine script's own input defaults
# ---------------------------------------------------------------------------
LEFT_BARS = 15         # `leftBars`
RIGHT_BARS = 15        # `rightBars`
VOLUME_THRESH = 20.0   # `volumeThresh` -- % oscillator threshold for a "confirmed" break
VOL_EMA_SHORT = 5      # `short = ema(volume, 5)`
VOL_EMA_LONG = 10      # `long = ema(volume, 10)`
CUTOFF_TIME = dtime(15, 15)
INTERVAL = "5minute"

# ---------------------------------------------------------------------------
# Excel styling (kept visually consistent with the rest of the workbook)
# ---------------------------------------------------------------------------
FILL_LIME = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")   # BUY CE
FILL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # BUY PE
FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")   # WAIT
FONT_WHITE = Font(color="FFFFFF")
FONT_BLACK = Font(color="000000")


# ---------------------------------------------------------------------------
# Indicator -- pivot-based S/R with volume-confirmed breaks
# ---------------------------------------------------------------------------
def _pivot_series(values: np.ndarray, left: int, right: int, find_max: bool) -> np.ndarray:
    """Raw pivot high/low at its TRUE bar index (ta.pivothigh/pivotlow's
    own definition: strictly the max/min of the `left+right+1`-bar window
    centered on this bar, first occurrence wins on a tie). NaN everywhere
    else. This is NOT yet lookahead-safe on its own -- see
    calculate_support_resistance() for the shift/ffill sequence that
    reproduces Pine's `fixnan(pivothigh(...)[1])` reveal timing."""
    n = len(values)
    out = np.full(n, np.nan)
    for i in range(left, n - right):
        window = values[i - left: i + right + 1]
        extreme = window.max() if find_max else window.min()
        if values[i] != extreme:
            continue
        # first occurrence of the extreme in the window must be THIS bar
        # (ta.pivothigh/pivotlow's own tie-break -- an equal high/low
        # earlier in the window wins, this bar isn't the pivot)
        idx_of_extreme = np.argmax(window) if find_max else np.argmin(window)
        if idx_of_extreme == left:
            out[i] = values[i]
    return out


def calculate_support_resistance(df, left=LEFT_BARS, right=RIGHT_BARS,
                                  volume_thresh=VOLUME_THRESH):
    """Requires 'open','high','low','close' columns (+ 'volume' if
    present, else the volume oscillator is 0 everywhere and no break is
    ever volume-confirmed). Returns df with 'Resistance', 'Support',
    'Vol Osc', 'SR Recomm', 'SR Break Age', 'Wick Flag' columns added."""
    df = df.copy()
    open_s = df['open'].astype(float)
    high_s = df['high'].astype(float)
    low_s = df['low'].astype(float)
    close_s = df['close'].astype(float)
    n = len(df)

    raw_pivot_high = _pivot_series(high_s.to_numpy(), left, right, find_max=True)
    raw_pivot_low = _pivot_series(low_s.to_numpy(), left, right, find_max=False)

    # --- Reveal timing: fixnan(pivothigh(left,right)[1]) ---
    # A pivot detected at true index i is only KNOWABLE once `right` bars
    # have closed after it (bar i+right), matching Pine's own pivothigh()
    # semantics when evaluated bar-by-bar; the source's extra `[1]` delays
    # it one bar further. shift(right) lands each raw pivot value on the
    # bar it FIRST becomes visible; ffill() holds it as "the active
    # level" until superseded; the final shift(1) reproduces the `[1]`.
    resistance = pd.Series(raw_pivot_high, index=df.index).shift(right).ffill().shift(1)
    support = pd.Series(raw_pivot_low, index=df.index).shift(right).ffill().shift(1)

    # --- Volume oscillator: short/long EMA of volume, % difference ---
    if 'volume' in df.columns:
        vol_s = df['volume'].astype(float)
        vol_short = vol_s.ewm(span=VOL_EMA_SHORT, adjust=False).mean()
        vol_long = vol_s.ewm(span=VOL_EMA_LONG, adjust=False).mean()
        vol_osc = np.where(vol_long.to_numpy() != 0,
                            100.0 * (vol_short - vol_long) / vol_long, 0.0)
    else:
        vol_osc = np.zeros(n)
    vol_osc = pd.Series(vol_osc, index=df.index)

    # --- Crossunder/crossover of close through the active level ---
    prev_close = close_s.shift(1)
    crossunder_support = (close_s < support) & (prev_close >= support.shift(1))
    crossover_resistance = (close_s > resistance) & (prev_close <= resistance.shift(1))

    # --- Wick classification (see module docstring -- Pine's own
    # "Bull Wick" / "Bear Wick" labels, mutually exclusive with a real
    # "Break"): a crossover that looks more like a rejection wick than a
    # clean break of the level.
    bear_wick = crossunder_support & ((open_s - close_s) < (high_s - open_s))
    bull_wick = crossover_resistance & ((open_s - low_s) > (close_s - open_s))

    vol_confirmed = vol_osc > volume_thresh
    support_break = crossunder_support & ~bear_wick & vol_confirmed
    resistance_break = crossover_resistance & ~bull_wick & vol_confirmed

    sr_recomm = np.where(resistance_break, 'BUY CE', np.where(support_break, 'BUY PE', 'WAIT'))

    # --- Break Age + Last Break Dir: bars since the last confirmed break
    # (either direction), 0 on the break bar itself, plus WHICH direction
    # that break was (persists until the next break, like zerolag.py's
    # Trend Dir). 'SR Recomm' above is edge-triggered (WAIT except on the
    # exact break bar) -- a pre-entry signal at t1 rarely lands exactly
    # ON a break bar, so a gate needs these two PERSISTENT columns (which
    # direction, how many bars ago) rather than 'SR Recomm' itself to
    # check "was there a recent break agreeing with this signal." Mirrors
    # zerolag.py's Trend Dir + Flip Age pairing.
    is_support_break = support_break.to_numpy()
    is_resistance_break = resistance_break.to_numpy()
    is_break = is_support_break | is_resistance_break
    break_age = np.zeros(n, dtype=int)
    last_break_dir = np.zeros(n, dtype=int)  # 0 = none yet, 1 = resistance/bullish, -1 = support/bearish
    age = -1  # -1 until the first break of the day, so age is never falsely 0 before one exists
    direction = 0
    for i in range(n):
        if is_break[i]:
            age = 0
            direction = 1 if is_resistance_break[i] else -1
        elif age >= 0:
            age += 1
        break_age[i] = age
        last_break_dir[i] = direction

    df['Resistance'] = resistance.to_numpy()
    df['Support'] = support.to_numpy()
    df['Vol Osc'] = vol_osc.to_numpy()
    df['SR Recomm'] = sr_recomm
    df['SR Break Age'] = break_age
    df['Last Break Dir'] = last_break_dir
    df['Wick Flag'] = np.where(bull_wick.to_numpy(), 'Bull Wick',
                                np.where(bear_wick.to_numpy(), 'Bear Wick', ''))

    return df


# ---------------------------------------------------------------------------
# Per-symbol processing (same brute-force time/timezone protocol as RSI.py)
# ---------------------------------------------------------------------------
def process_symbol(symbol_data):
    symbol, df, target_date = symbol_data
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip().str.lower()

    if not all(col in df.columns for col in ['open', 'high', 'low', 'close']):
        print(f"[WARNING] {symbol}: Missing required price columns. Discarding.")
        return symbol, None

    time_col = next((col for col in ['date', 'time', 'datetime', 'timestamp'] if col in df.columns), None)
    if time_col:
        parsed = pd.to_datetime(df[time_col], errors='coerce')
    elif pd.api.types.is_datetime64_any_dtype(df.index):
        parsed = pd.Series(df.index, index=df.index)
    else:
        unnamed_cols = [c for c in df.columns if 'unnamed' in c]
        if unnamed_cols:
            parsed = pd.to_datetime(df[unnamed_cols[0]], errors='coerce')
        else:
            print(f"[WARNING] {symbol}: Failed to locate any valid timestamp data. Discarding.")
            return symbol, None

    try:
        if parsed.dt.tz is not None:
            parsed = parsed.dt.tz_localize(None)
    except (AttributeError, TypeError):
        pass

    df['_sort_dt'] = parsed
    df['time_str'] = parsed.dt.strftime('%H:%M')
    df.dropna(subset=['time_str', '_sort_dt'], inplace=True)
    if df.empty:
        print(f"[WARNING] {symbol}: No rows with a valid timestamp. Discarding.")
        return symbol, None

    df.sort_values('_sort_dt', inplace=True)
    df = df[df['_sort_dt'].dt.time <= CUTOFF_TIME]
    if df.empty:
        print(f"[WARNING] {symbol}: No candles at/before {CUTOFF_TIME.strftime('%H:%M')}. Discarding.")
        return symbol, None

    df = calculate_support_resistance(df)

    df = excel_utils.restrict_to_target_date(df, target_date)
    if df is None:
        return symbol, None

    return symbol, df


# ---------------------------------------------------------------------------
# Excel export -- pivoted Symbol x Time matrix (same pattern as zerolag.py)
# ---------------------------------------------------------------------------
def build_matrix(data_dict, target_date, max_workers=None):
    results = {}
    with concurrent.futures.ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(process_symbol, (sym, df, target_date)): sym for sym, df in data_dict.items()}
        for future in concurrent.futures.as_completed(futures):
            sym = futures[future]
            try:
                processed_sym, processed_df = future.result()
                if processed_df is not None and 'SR Recomm' in processed_df.columns:
                    if not processed_df.empty and sym.lower() != 'summary':
                        results[sym] = processed_df
            except Exception as e:
                print(f"[ERROR] Support/Resistance Engine Failure on symbol {sym}: {str(e)}")

    if not results:
        raise RuntimeError("Support/Resistance: zero symbols processed successfully for target_date -- matrix build aborted.")

    all_times = set()
    for df in results.values():
        if 'time_str' in df.columns:
            all_times.update(df['time_str'].dropna().tolist())

    if not all_times:
        raise RuntimeError("Support/Resistance: no valid timestamps extracted across all symbols -- matrix build aborted.")

    sorted_times = sorted(all_times)
    matrix_rows = [['Symbol', 'Metrics'] + sorted_times]

    for sym in sorted(results.keys()):
        df = results[sym].drop_duplicates(subset=['time_str'], keep='last')
        df_indexed = df.set_index('time_str')

        def get_metric_row(metric_name, col_name):
            row = [sym, metric_name]
            for t in sorted_times:
                if t in df_indexed.index:
                    val = df_indexed.loc[t, col_name]
                    if isinstance(val, pd.Series):
                        val = val.iloc[-1]
                    row.append(val)
                else:
                    row.append("")
            return row

        matrix_rows.append(get_metric_row('Resistance', 'Resistance'))
        matrix_rows.append(get_metric_row('Support', 'Support'))
        matrix_rows.append(get_metric_row('Vol Osc', 'Vol Osc'))
        matrix_rows.append(get_metric_row('SR Recomm', 'SR Recomm'))
        matrix_rows.append(get_metric_row('SR Break Age', 'SR Break Age'))
        matrix_rows.append(get_metric_row('Last Break Dir', 'Last Break Dir'))
        matrix_rows.append(get_metric_row('Wick Flag', 'Wick Flag'))

    return matrix_rows


def write_matrix_to_workbook(matrix_rows, wb):
    sheet_name = "SUPRES"
    ws = excel_utils.replace_sheet_with_matrix(wb, sheet_name, matrix_rows)

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1), start=2):
        metric_type = ws.cell(row=r_idx, column=2).value
        for cell in row[2:]:
            val = cell.value
            if val == "":
                continue
            if metric_type == 'SR Recomm':
                if val == 'BUY CE':
                    cell.fill = FILL_LIME
                    cell.font = FONT_BLACK
                elif val == 'BUY PE':
                    cell.fill = FILL_RED
                    cell.font = FONT_WHITE
                elif val == 'WAIT':
                    cell.fill = FILL_GRAY
                    cell.font = FONT_BLACK

    excel_utils.autofit_columns(ws)


def write_matrix(matrix_rows, output_excel_path):
    wb = load_workbook(output_excel_path)
    write_matrix_to_workbook(matrix_rows, wb)
    excel_utils.atomic_save(wb, output_excel_path)


def parallel_compute_and_export(data_dict, target_date, output_excel_path, max_workers=None):
    matrix_rows = build_matrix(data_dict, target_date, max_workers=max_workers)
    write_matrix(matrix_rows, output_excel_path)


def run_support_resistance_step(df_ref, target_date, output_excel_path):
    print("[SYSTEM] Loading 5-minute historical data for Support/Resistance...")
    data_dict = data_ingestion.load_interval_data(df_ref, target_date, interval=INTERVAL)
    if not data_dict:
        raise RuntimeError("Support/Resistance: no 5-minute data files found for any symbol.")
    print(f"[PROCESS] Computing Support/Resistance matrix for {len(data_dict)} symbol(s)...")
    parallel_compute_and_export(data_dict, target_date, output_excel_path)
    print("[SUCCESS] Support/Resistance matrix written to sheet 'SUPRES'.")


# ---------------------------------------------------------------------------
# Disclaimer: this module derives levels/breaks from historical/live price
# and volume data only. It is not financial advice, and no result here is
# a guarantee of future performance -- paper-trade it and run it through
# scripts/backtester.py before risking real capital. See order_sheet.py's
# ENABLE_SR_GATE for how (and whether) this actually gates entries.
# ---------------------------------------------------------------------------
