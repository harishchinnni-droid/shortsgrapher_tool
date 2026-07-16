"""
breakout_probability.py
------------------------
Python translation of '04_ProjectFiles/Breakout_Probability.txt' (Pine v5,
"Breakout Probability (Expo)" by Zeiierman), wired into the pipeline the
same way as RSI.py/SQZMOM.py. Rebuilt from the Pine source only -- the
original Python version was lost (source file gone, only a stale
__pycache__/breakout_probability.cpython-312.pyc leftover), so this is a
fresh port, not a recovered file. Flag anything that looks off against
whatever the prior version actually did.

Scope of this translation -- level 0 only:
    The Pine script draws up to 5 percentage-step levels (0%, 1%, 2%, 3%,
    4% by default) purely for charting extra lines/labels. The only level
    that actually feeds its BULLISH/BEARISH bias/alert output is level 0
    (x = 0, i.e. "did price make any new high/low vs the previous bar at
    all"). This module implements level 0 only -- that's the tradeable
    signal; levels 1-4 in the source are visual-only and have no bearing
    on the bias this pipeline cares about.

Indicator logic (exact translation of the Pine source's running counters,
which are declared `var` -- i.e. accumulate for the life of the chart,
never reset):
    Every bar, based on the PREVIOUS bar's candle color:
        green = close[-1] > open[-1]   (previous bar was bullish)
        red   = close[-1] < open[-1]   (previous bar was bearish)
    and this bar's level-0 breakout test:
        hh = high >= high[-1]   (made a new high vs previous bar)
        ll = low  <= low[-1]    (made a new low vs previous bar)

    Four running counters, accumulated across the whole loaded history
    (not reset daily -- matches the Pine `var` semantics):
        ghh, gll -- how many green-context bars hit a new high / new low
        rhh, rll -- how many red-context bars hit a new high / new low
    and two running totals: gtotal (green-context bars seen), rtotal
    (red-context bars seen). Each percentage below only updates on the
    bar its own event fires, otherwise it carries forward its last value
    (this is a direct translation of the Pine script reading stale
    matrix cells when a branch didn't run that bar):
        Green HH% = ghh / gtotal * 100  (last updated when green & hh)
        Green LL% = gll / gtotal * 100  (last updated when green & ll)
        Red HH%   = rhh / rtotal * 100  (last updated when red & hh)
        Red LL%   = rll / rtotal * 100  (last updated when red & ll)

    Bias this bar (exact translation of the source's alert-text ternary,
    which is a plain green/not-green branch -- a flat bar where
    close == open falls into the "red" branch by the same Pine logic):
        if green: BULLISH if Green HH% >= Green LL% else BEARISH
        else:     BULLISH if Red HH%  >= Red LL%  else BEARISH

Recommendation ('BRKPRO Recomm') -- direct mapping of the bias above onto
this pipeline's BUY CE / BUY PE / WAIT vocabulary (BULLISH -> BUY CE,
BEARISH -> BUY PE), WAIT only while gtotal/rtotal haven't seen a single
green or red bar yet (i.e. before any percentage exists to compare).

Known limitation vs the live Pine indicator: on a TradingView chart these
counters accumulate since the symbol's entire listing history. This
pipeline only has whatever rolling window data_ingestion has downloaded
(commonly ~90 days) -- so the percentages here reflect breakout
probability over that window, not "all-time." Flag this to Harish if the
lost original handled the window differently.

Per-bar bias, but the underlying counters are genuinely cumulative/
stateful across the whole loaded history -- this is NOT a stateless
per-bar recompute like RSI.py/ema20.py, and NOT vectorizable the way
those are. Computed with a plain per-symbol loop (cheap at intraday row
counts, and ProcessPoolExecutor already parallelizes across symbols).

Timeframe: 5-minute candles, truncated per day to CUTOFF_TIME (15:15),
same as every other module in this pipeline.
"""

import os
import sys
import concurrent.futures
from datetime import time as dtime

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

CODES_DIR = os.path.dirname(os.path.abspath(__file__))
if CODES_DIR not in sys.path:
    sys.path.append(CODES_DIR)
import data_ingestion
import excel_utils

CUTOFF_TIME = dtime(15, 15)
INTERVAL = "5minute"

FILL_LIME = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")   # BUY CE
FILL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # BUY PE
FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")   # WAIT
FONT_WHITE = Font(color="FFFFFF")
FONT_BLACK = Font(color="000000")


# ---------------------------------------------------------------------------
# Indicator -- level-0 cumulative counters, literal translation of the
# Pine source's Score()/bias logic (see module docstring)
# ---------------------------------------------------------------------------
def calculate_breakout_probability(df):
    """Requires 'open','high','low','close' columns, sorted ascending by
    time. Returns df with Green HH%/Green LL%/Red HH%/Red LL%/BRKPRO
    Recomm columns added. Cumulative counters run over every row passed
    in (the full multi-day history), so call this BEFORE restricting the
    output to target_date, same as every other module in this pipeline."""
    df = df.copy()
    open_ = df['open'].to_numpy()
    high = df['high'].to_numpy()
    low = df['low'].to_numpy()
    close = df['close'].to_numpy()
    n = len(df)

    green_hh_pct = np.full(n, np.nan)
    green_ll_pct = np.full(n, np.nan)
    red_hh_pct = np.full(n, np.nan)
    red_ll_pct = np.full(n, np.nan)
    recomm = np.full(n, 'WAIT', dtype=object)

    gtotal = rtotal = ghh = gll = rhh = rll = 0
    a1 = b1 = a2 = b2 = None  # last-known Green HH% / Green LL% / Red HH% / Red LL%

    for i in range(1, n):
        prev_green = close[i - 1] > open_[i - 1]
        prev_red = close[i - 1] < open_[i - 1]
        hh = high[i] >= high[i - 1]
        ll = low[i] <= low[i - 1]

        if prev_green:
            gtotal += 1
        if prev_red:
            rtotal += 1

        if prev_green and hh:
            ghh += 1
            a1 = ghh / gtotal * 100
        if prev_green and ll:
            gll += 1
            b1 = gll / gtotal * 100
        if prev_red and hh:
            rhh += 1
            a2 = rhh / rtotal * 100
        if prev_red and ll:
            rll += 1
            b2 = rll / rtotal * 100

        green_hh_pct[i] = a1 if a1 is not None else np.nan
        green_ll_pct[i] = b1 if b1 is not None else np.nan
        red_hh_pct[i] = a2 if a2 is not None else np.nan
        red_ll_pct[i] = b2 if b2 is not None else np.nan

        # Exact translation of the Pine source's binary green/not-green
        # ternary -- a flat bar (close==open) falls into the "red" branch,
        # same as the original.
        if prev_green:
            if a1 is not None and b1 is not None:
                recomm[i] = 'BUY CE' if a1 >= b1 else 'BUY PE'
        else:
            if a2 is not None and b2 is not None:
                recomm[i] = 'BUY CE' if a2 >= b2 else 'BUY PE'

    df['Green HH%'] = green_hh_pct
    df['Green LL%'] = green_ll_pct
    df['Red HH%'] = red_hh_pct
    df['Red LL%'] = red_ll_pct
    df['BRKPRO Recomm'] = recomm
    return df


# ---------------------------------------------------------------------------
# Per-symbol processing (same brute-force time/timezone protocol as RSI.py)
# ---------------------------------------------------------------------------
def process_symbol(symbol_data):
    symbol, df, target_date = symbol_data
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip().str.lower()

    if not all(col in df.columns for col in ['open', 'high', 'low', 'close']):
        print(f"[WARNING] {symbol}: Missing required price columns. Discarding.")
        return symbol, None

    time_col = next((col for col in ['date', 'time', 'datetime', 'timestamp'] if col in df.columns), None)
    if time_col:
        parsed = pd.to_datetime(df[time_col], errors='coerce')
    elif pd.api.types.is_datetime64_any_dtype(df.index):
        parsed = pd.Series(df.index, index=df.index)
    else:
        unnamed_cols = [c for c in df.columns if 'unnamed' in c]
        if unnamed_cols:
            parsed = pd.to_datetime(df[unnamed_cols[0]], errors='coerce')
        else:
            print(f"[WARNING] {symbol}: Failed to locate any valid timestamp data. Discarding.")
            return symbol, None

    try:
        if parsed.dt.tz is not None:
            parsed = parsed.dt.tz_localize(None)
    except (AttributeError, TypeError):
        pass

    df['_sort_dt'] = parsed
    df['time_str'] = parsed.dt.strftime('%H:%M')
    df.dropna(subset=['time_str', '_sort_dt'], inplace=True)
    if df.empty:
        print(f"[WARNING] {symbol}: No rows with a valid timestamp. Discarding.")
        return symbol, None

    df.sort_values('_sort_dt', inplace=True)
    df = df[df['_sort_dt'].dt.time <= CUTOFF_TIME]
    if df.empty:
        print(f"[WARNING] {symbol}: No candles at/before {CUTOFF_TIME.strftime('%H:%M')}. Discarding.")
        return symbol, None

    # Counters are cumulative across the full multi-day history -- compute
    # BEFORE restricting to target_date, same ordering as RSI.py's EMA
    # warmup / ema20.py's HTF gate.
    df = calculate_breakout_probability(df)

    df = excel_utils.restrict_to_target_date(df, target_date)
    if df is None:
        return symbol, None

    return symbol, df


# ---------------------------------------------------------------------------
# Excel export -- pivoted Symbol x Time matrix (same pattern as RSI.py)
# ---------------------------------------------------------------------------
def build_matrix(data_dict, target_date, max_workers=None):
    results = {}
    with concurrent.futures.ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(process_symbol, (sym, df, target_date)): sym for sym, df in data_dict.items()}
        for future in concurrent.futures.as_completed(futures):
            sym = futures[future]
            try:
                processed_sym, processed_df = future.result()
                if processed_df is not None and 'BRKPRO Recomm' in processed_df.columns:
                    if not processed_df.empty and sym.lower() != 'summary':
                        results[sym] = processed_df
            except Exception as e:
                print(f"[ERROR] Breakout Probability Engine Failure on symbol {sym}: {str(e)}")

    if not results:
        raise RuntimeError("Breakout Probability: zero symbols processed successfully for target_date -- matrix build aborted.")

    all_times = set()
    for df in results.values():
        if 'time_str' in df.columns:
            all_times.update(df['time_str'].dropna().tolist())

    if not all_times:
        raise RuntimeError("Breakout Probability: no valid timestamps extracted across all symbols -- matrix build aborted.")

    sorted_times = sorted(all_times)
    matrix_rows = [['Symbol', 'Metrics'] + sorted_times]

    for sym in sorted(results.keys()):
        df = results[sym].drop_duplicates(subset=['time_str'], keep='last')
        df_indexed = df.set_index('time_str')

        def get_metric_row(metric_name, col_name):
            row = [sym, metric_name]
            for t in sorted_times:
                if t in df_indexed.index:
                    val = df_indexed.loc[t, col_name]
                    if isinstance(val, pd.Series):
                        val = val.iloc[-1]
                    row.append(val)
                else:
                    row.append("")
            return row

        matrix_rows.append(get_metric_row('Green HH%', 'Green HH%'))
        matrix_rows.append(get_metric_row('Green LL%', 'Green LL%'))
        matrix_rows.append(get_metric_row('Red HH%', 'Red HH%'))
        matrix_rows.append(get_metric_row('Red LL%', 'Red LL%'))
        matrix_rows.append(get_metric_row('BRKPRO Recomm', 'BRKPRO Recomm'))

    return matrix_rows


def write_matrix_to_workbook(matrix_rows, wb):
    sheet_name = "BRKPRO"
    ws = excel_utils.replace_sheet_with_matrix(wb, sheet_name, matrix_rows)

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1), start=2):
        metric_type = ws.cell(row=r_idx, column=2).value
        for cell in row[2:]:
            val = cell.value
            if val == "":
                continue
            if metric_type == 'BRKPRO Recomm':
                if val == 'BUY CE':
                    cell.fill = FILL_LIME
                    cell.font = FONT_BLACK
                elif val == 'BUY PE':
                    cell.fill = FILL_RED
                    cell.font = FONT_WHITE
                elif val == 'WAIT':
                    cell.fill = FILL_GRAY
                    cell.font = FONT_BLACK

    excel_utils.autofit_columns(ws)


def write_matrix(matrix_rows, output_excel_path):
    wb = load_workbook(output_excel_path)
    write_matrix_to_workbook(matrix_rows, wb)
    excel_utils.atomic_save(wb, output_excel_path)


def parallel_compute_and_export(data_dict, target_date, output_excel_path, max_workers=None):
    matrix_rows = build_matrix(data_dict, target_date, max_workers=max_workers)
    write_matrix(matrix_rows, output_excel_path)


def run_breakout_probability_step(df_ref, target_date, output_excel_path):
    print("[SYSTEM] Loading 5-minute historical data for Breakout Probability...")
    data_dict = data_ingestion.load_interval_data(df_ref, target_date, interval=INTERVAL)
    if not data_dict:
        raise RuntimeError("Breakout Probability: no 5-minute data files found for any symbol.")
    print(f"[PROCESS] Computing Breakout Probability matrix for {len(data_dict)} symbol(s)...")
    parallel_compute_and_export(data_dict, target_date, output_excel_path)
    print("[SUCCESS] Breakout Probability matrix written to sheet 'BRKPRO'.")


# ---------------------------------------------------------------------------
# Disclaimer: this module derives a signal from historical/live price data
# only. It is not financial advice, and no result here is a guarantee of
# future performance -- paper-trade it and run it through
# scripts/backtester.py before risking real capital.
# ---------------------------------------------------------------------------
