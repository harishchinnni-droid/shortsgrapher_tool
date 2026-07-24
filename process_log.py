"""
process_log.py
---------------
[ADDED -- Task 73, 22-Jul-26, Harish's request] Writes a 'Process Log'
sheet to the output workbook, timing how long the DATA/INDICATOR half of
the pipeline takes to run each cycle -- from historical data ingestion
(file provisioning + token sync are included too, but are normally fast/
cached and not the dominant cost) through computing every indicator
matrix and writing the indicator sheets + the Final (confluence) sheet.

Deliberately EXCLUDES order_sheet.py entirely -- both its gate-chain
entry-decision evaluation and its own Orders/Rejected/Dashboard sheet
writes, and update_open_positions_live()'s SL/TSL polling -- per Harish's
explicit boundary ("not placing orders from order sheet and TSL"). The
reasoning: the timed portion is what's actually bound by network speed
(Kite historical_data() calls) and CPU/RAM (indicator computation running
concurrently across every symbol's 5-min candle history) -- exactly what
a hardware/internet upgrade decision should be based on. order_sheet.py's
gate-chain runs over a much smaller, already-computed dataset (one row
per symbol from the Final sheet) and isn't the resource-bound part of
this pipeline the same way, so timing it would dilute the signal Harish
is actually after.

LIVE: one row APPENDED every cycle (~every 5 min), so a whole day's
session builds a time-series -- useful for spotting gradual degradation
(e.g. slowing down as more of the day's history accumulates) or a single
cycle spiking (a network blip, a rate-limit backoff).

BACKTEST: only ONE row, since a backtest date runs its data/indicator
pipeline exactly once (there's no 5-minute cycling) -- re-running the same
date overwrites that one row rather than piling up duplicates.

Best-effort throughout: any failure here is logged and swallowed, never
raised -- this is instrumentation, not a correctness-critical pipeline
step, and must never be the reason a cycle/backtest date fails.
"""
import calendar_mgmt
import excel_utils
from openpyxl import load_workbook

PROCESS_LOG_SHEET = 'Process Log'
PROCESS_LOG_HEADERS = ['Mode', 'Date', 'Cycle #', 'Time', 'Duration (sec)']


def log_cycle_timing(output_excel_path, mode, target_date, cycle_num, duration_seconds, timestamp_str):
    """Appends (LIVE) or writes/overwrites a single row (BACKTEST) to the
    'Process Log' sheet in output_excel_path.

    mode: calendar_mgmt.LIVE or calendar_mgmt.BACKTEST.
    cycle_num: LIVE's 1-based cycle counter for the day; ignored (left
        blank) for BACKTEST, which has no cycle concept.
    duration_seconds: elapsed wall-clock time for the timed span --
        caller's responsibility to measure only historical data ingestion
        through the Final/confluence sheet write, per this module's
        docstring above.
    timestamp_str: 'HH:MM:SS' IST at the moment this cycle/run finished.
    """
    try:
        wb = load_workbook(output_excel_path)
    except Exception as e:
        print(f"[WARNING] Process Log: could not open workbook ({e}) -- skipping this timing entry.")
        return

    date_str = target_date.strftime('%d-%b-%y')
    row = [mode, date_str, cycle_num if mode == calendar_mgmt.LIVE else "", timestamp_str, round(duration_seconds, 2)]

    try:
        if PROCESS_LOG_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(PROCESS_LOG_SHEET)
            ws.append(PROCESS_LOG_HEADERS)
        else:
            ws = wb[PROCESS_LOG_SHEET]

        if mode == calendar_mgmt.BACKTEST:
            # Single row only -- overwrite row 2 (the one data row) if a
            # backtest for this same date has already logged once,
            # instead of appending a duplicate every re-run.
            if ws.max_row >= 2:
                for col_idx, val in enumerate(row, start=1):
                    ws.cell(row=2, column=col_idx, value=val)
            else:
                ws.append(row)
        else:
            ws.append(row)

        excel_utils.autofit_columns(ws)
        excel_utils.atomic_save(wb, output_excel_path)
        print(f"[SYSTEM] Process Log: {mode} data/indicator pipeline took {duration_seconds:.2f}s "
              f"(file provisioning + token sync + historical data + indicators + Final sheet).")
    except Exception as e:
        print(f"[WARNING] Process Log: failed to write timing entry ({e}) -- non-fatal, continuing.")
