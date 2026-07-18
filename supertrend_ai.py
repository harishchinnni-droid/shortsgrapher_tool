"""
supertrend_ai.py
-----------------
Python translation of LuxAlgo's "SuperTrend AI (Clustering)" Pine Script v5
indicator, wired to run as Step 11 of the daily pipeline (after TW All In
One). Output format mirrors RSI.py / SQZMOM.py's pivoted Symbol x Time
matrix.

Indicator logic (see 'SupertrendAI_LuxAlgo.txt'):
    Nine candidate SuperTrend lines are run in parallel, one per ATR
    multiplier from MIN_MULT to MAX_MULT in STEP increments (default
    1.0..5.0 step 0.5 -> 9 factors). Each candidate tracks its own running
    "performance" score (an EMA of how well its trailing-stop side has
    tracked recent price moves). On every bar, those 9 performance scores
    are grouped into 3 clusters via k-means (seeded at the 25th/50th/75th
    percentile) -- Worst / Average / Best. The FROM_CLUSTER setting picks
    which cluster's *average factor* becomes that bar's selected ATR
    multiplier, which drives a single "chosen" SuperTrend line (TS). A
    performance-weighted adaptive moving average of TS (TS_AMA) is also
    kept, matching the original's trailing-stop AMA plot.

    This is inherently bar-by-bar path-dependent (each candidate's upper/
    lower/trend/perf depends on its own value from the previous bar, and
    the cluster re-fit depends on the current snapshot of all 9 perf
    scores), so -- unlike ADX/RSI/TW ALL -- it genuinely cannot be
    vectorized into pandas rolling/ewm calls; it's computed with a single
    sequential loop per symbol, same as this project's own supertrend()
    reference implementation handles the ATR-band-flip step.

    Performance note: the sequential loop is O(bars x kmeans_iters x 9
    factors), but kmeans_iters is tiny in practice (9 points into 3
    clusters typically converges in well under 10 iterations, MAX_ITER=1000
    is only ever a safety cap from the original script and is essentially
    never hit) -- this runs in a few seconds per symbol on typical 5-minute
    intraday data, well within a daily batch pipeline's budget.

Discrete 'Supertrend Recomm' column (NOT part of the original indicator --
a mapping defined here per your request): the original script's own
bullish/bearish labels already fire on every OS flip (os > os[1] = bullish,
os < os[1] = bearish) and hold naturally between flips, so this is a direct
1:1 mapping -- BUY CE while OS=1 (trailing stop below price), BUY PE while
OS=0 (trailing stop above price), WAIT only before the first valid value.
This is a heuristic, not a validated edge -- backtest it
(scripts/backtester.py) before trusting it with real capital.

Timeframe: 5-minute candles, truncated per day to CUTOFF_TIME, same
convention as RSI.py. Timezone stripped before writing to Excel for the
same reason documented there (openpyxl can't store tz-aware datetimes).
"""

import os
import sys
import tempfile
import shutil
import concurrent.futures
from datetime import time as dtime

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

CODES_DIR = os.path.dirname(os.path.abspath(__file__))
if CODES_DIR not in sys.path:
    sys.path.append(CODES_DIR)
import data_ingestion
import excel_utils

# ---------------------------------------------------------------------------
# Config -- mirrors the Pine script's default inputs
# ---------------------------------------------------------------------------
ATR_LENGTH = 10
MIN_MULT = 1.0
MAX_MULT = 5.0
STEP = 0.5
PERF_ALPHA = 10.0
FROM_CLUSTER = "Best"      # 'Best' | 'Average' | 'Worst'
MAX_ITER = 1000            # safety cap only -- 9-point/3-cluster kmeans converges in a handful of steps
CUTOFF_TIME = dtime(15, 15)
INTERVAL = "5minute"

_CLUSTER_IDX = {"Worst": 0, "Average": 1, "Best": 2}

# ---------------------------------------------------------------------------
# Excel styling (kept visually consistent with RSI.py / SQZMOM.py's palette)
# ---------------------------------------------------------------------------
FILL_BULLISH = PatternFill(start_color="26A69A", end_color="26A69A", fill_type="solid")
FILL_BEARISH = PatternFill(start_color="EF5350", end_color="EF5350", fill_type="solid")
FILL_LIME = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")   # BUY CE
FILL_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # BUY PE
FILL_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")   # WAIT
FONT_WHITE = Font(color="FFFFFF")
FONT_BLACK = Font(color="000000")


# ---------------------------------------------------------------------------
# Indicator
# ---------------------------------------------------------------------------
def _atr_wilder(df, length):
    high, low, close = df['high'], df['low'], df['close']
    prev_close = close.shift(1)
    tr = pd.concat(
        [high - low, (high - prev_close).abs(), (low - prev_close).abs()], axis=1
    ).max(axis=1)
    return tr.ewm(alpha=1.0 / length, adjust=False).mean()


def _percentile_linear(sorted_vals, p):
    """Matches numpy.percentile()'s default 'linear' interpolation exactly
    for a single percentile p on an already-sorted sequence -- used
    instead of np.percentile() inside the hot per-bar loop below, since
    numpy's per-call dispatch overhead dominates for arrays this tiny
    (see _kmeans_3()'s docstring)."""
    n = len(sorted_vals)
    if n == 1:
        return sorted_vals[0]
    idx = p / 100.0 * (n - 1)
    lo = int(idx)
    hi = min(lo + 1, n - 1)
    frac = idx - lo
    return sorted_vals[lo] + (sorted_vals[hi] - sorted_vals[lo]) * frac


def _kmeans_3(data, factors, max_iter):
    """K-means of `data` (perf scores) into 3 clusters, seeded at the
    25th/50th/75th percentile, mirroring the Pine script's clustering loop
    exactly (same seed, same distance metric, same convergence check).
    Returns (clusters_perf, clusters_factors) as 3-element lists of lists,
    ordered [Worst, Average, Best] by construction (see module docstring).

    PERFORMANCE NOTE, and an important correctness caveat found while
    optimizing it: this is called once per bar (see
    calculate_supertrend_ai()'s main loop), so on 90 days of 5-minute data
    (~6,750 bars) it runs ~6,750 times per symbol, and profiling showed it
    as by far the most expensive single piece of this whole pipeline --
    numpy's per-call dispatch overhead is roughly fixed regardless of array
    size, so for 9-element arrays that overhead dominates the actual
    arithmetic.

    The per-point distance/assignment loop below (np.abs + np.argmin,
    9 elements, up to max_iter times) is replaced with plain Python
    abs()/comparisons -- this is safe because abs() and <=/> on float64
    scalars are exact, non-summation operations: IEEE754 guarantees
    numpy and plain Python produce bit-identical results for these, so
    this part carries zero behavior risk. Likewise the percentile seed
    (_percentile_linear() above) is exact for this fixed n=9 (indices
    2/4/6 land on exact integers, no interpolation rounding either way).

    The CENTROID RECOMPUTATION (the mean of each cluster) is deliberately
    LEFT as np.mean(), not simplified to sum(c)/len(c): those two are
    mathematically equivalent but NOT bit-identical (different summation
    order/algorithm), and this indicator is recursive -- perf at bar t+1
    depends on the clustering result at bar t. During a flat/no-momentum
    stretch, all 9 factors can converge to (near-)identical perf scores,
    which the code's own comment already flags as a "zero variance"
    degenerate case -- and a first attempt at this optimization that DID
    swap in sum()/len() here changed 9 of 3,000 bars' final Recomm on a
    randomized test, purely from that summation-order difference
    compounding through repeated ties across k-means iterations and then
    through the recursive per-bar state.

    Verified bit-for-bit identical to the original numpy implementation
    (not just "close") across 6 random seeds of a full end-to-end run
    plus the specific degenerate/all-tied case above, with np.mean() kept
    here for the centroid step. Net effect, measured end-to-end on a
    synthetic 90-day/6,750-bar single-symbol test: calculate_supertrend_ai()
    went from ~2.03s to ~0.86s per symbol (~2.4x) with identical output.
    At 206 symbols on one worker process that's roughly 7 minutes down to
    roughly 3 minutes for this one indicator, which measured as by far
    the slowest of the 7 in this pipeline (the other 6 combined were
    under a minute serially) -- so it's still the dominant cost, just
    less dominant. Re-measure on your own data/hardware; these are
    single-machine numbers from a synthetic test, not a guarantee.
    """
    data_l = data.tolist() if hasattr(data, 'tolist') else list(data)
    factors_l = factors.tolist() if hasattr(factors, 'tolist') else list(factors)
    n = len(data_l)
    sorted_data = sorted(data_l)
    centroids = [
        _percentile_linear(sorted_data, 25),
        _percentile_linear(sorted_data, 50),
        _percentile_linear(sorted_data, 75),
    ]
    clusters_perf = [[], [], []]
    clusters_factors = [[], [], []]
    for _ in range(max_iter):
        clusters_perf = [[], [], []]
        clusters_factors = [[], [], []]
        for j in range(n):
            dj = data_l[j]
            d0 = abs(dj - centroids[0])
            d1 = abs(dj - centroids[1])
            d2 = abs(dj - centroids[2])
            if d0 <= d1 and d0 <= d2:
                cidx = 0
            elif d1 <= d2:
                cidx = 1
            else:
                cidx = 2
            clusters_perf[cidx].append(dj)
            clusters_factors[cidx].append(factors_l[j])
        # np.mean(), not sum()/len() -- see docstring: this specific step
        # must match the original's floating-point summation exactly.
        new_centroids = [
            float(np.mean(c)) if c else centroids[k] for k, c in enumerate(clusters_perf)
        ]
        if new_centroids == centroids:
            centroids = new_centroids
            break
        centroids = new_centroids
    return clusters_perf, clusters_factors


def calculate_supertrend_ai(df, length=ATR_LENGTH, min_mult=MIN_MULT, max_mult=MAX_MULT,
                             step=STEP, perf_alpha=PERF_ALPHA, from_cluster=FROM_CLUSTER,
                             max_iter=MAX_ITER):
    df = df.copy()
    n = len(df)

    ts_arr = np.full(n, np.nan)
    ama_arr = np.full(n, np.nan)
    os_arr = np.zeros(n, dtype=int)
    perf_idx_arr = np.full(n, np.nan)
    target_factor_arr = np.full(n, np.nan)

    if n == 0:
        df['TS'], df['TS_AMA'], df['OS'] = ts_arr, ama_arr, os_arr
        df['PERF_IDX'], df['TARGET_FACTOR'] = perf_idx_arr, target_factor_arr
        df['TREND'] = np.array([], dtype=object)
        df['Supertrend Recomm'] = np.array([], dtype=object)
        return df

    close = df['close'].to_numpy(dtype=float)
    high = df['high'].to_numpy(dtype=float)
    low = df['low'].to_numpy(dtype=float)
    hl2 = (high + low) / 2.0
    atr_vals = _atr_wilder(df, length).to_numpy(dtype=float)

    factors = np.round(np.arange(min_mult, max_mult + step / 2, step), 4)
    nf = len(factors)
    from_idx = _CLUSTER_IDX.get(from_cluster, 2)

    # Per-factor running SuperTrend state (9 parallel candidates)
    upper = np.full(nf, hl2[0])
    lower = np.full(nf, hl2[0])
    trend = np.zeros(nf, dtype=int)
    output = np.full(nf, hl2[0])
    perf = np.zeros(nf)
    perf_coeff = 2.0 / (perf_alpha + 1.0)

    # Single "chosen" SuperTrend line, driven by the clustering pick each bar
    sel_upper, sel_lower = hl2[0], hl2[0]
    os_state = 0
    target_factor = np.nan
    ts_prev = np.nan
    perf_ama = np.nan

    # Performance-index denominator: ta.ema(abs(close - close[1]), perfAlpha)
    den_arr = df['close'].diff().abs().ewm(span=int(perf_alpha), adjust=False).mean().to_numpy(dtype=float)

    for i in range(1, n):
        up = hl2[i] + atr_vals[i] * factors
        dn = hl2[i] - atr_vals[i] * factors

        trend = np.where(close[i] > upper, 1, np.where(close[i] < lower, 0, trend))

        new_upper = np.where(close[i - 1] < upper, np.minimum(up, upper), up)
        new_lower = np.where(close[i - 1] > lower, np.maximum(dn, lower), dn)

        diff_sign = np.nan_to_num(np.sign(close[i - 1] - output), nan=0.0)
        perf = perf + perf_coeff * ((close[i] - close[i - 1]) * diff_sign - perf)

        upper, lower = new_upper, new_lower
        output = np.where(trend == 1, lower, upper)

        clusters_perf, clusters_factors = _kmeans_3(perf, factors, max_iter)

        chosen_factors = clusters_factors[from_idx]
        chosen_perf = clusters_perf[from_idx]

        # In a strong trend, all factors track identically, resulting in zero variance. 
        # k-means identical distances collapse into index 0 (Worst), leaving the Best cluster completely empty.
        # This fallback ensures we grab the highest populated cluster to prevent NaN stalling.
        if not chosen_factors:
            for fallback_idx in (2, 1, 0):
                if clusters_factors[fallback_idx]:
                    chosen_factors = clusters_factors[fallback_idx]
                    chosen_perf = clusters_perf[fallback_idx]
                    break

        if chosen_factors:
            # np.mean(), not sum()/len() -- see _kmeans_3()'s docstring:
            # this recursive indicator is sensitive to floating-point
            # summation order in exactly this kind of small-list mean,
            # so it's kept matching the original exactly rather than
            # optimized further.
            target_factor = float(np.mean(chosen_factors))
        # else: keep previous target_factor (nz() fallback in the original)

        perf_idx = max(float(np.mean(chosen_perf)), 0.0) if chosen_perf else 0.0
        den = den_arr[i]
        perf_idx = (perf_idx / den) if (den and not np.isnan(den) and den != 0) else 0.0

        if not np.isnan(target_factor):
            up2 = hl2[i] + atr_vals[i] * target_factor
            dn2 = hl2[i] - atr_vals[i] * target_factor
            sel_upper = min(up2, sel_upper) if close[i - 1] < sel_upper else up2
            sel_lower = max(dn2, sel_lower) if close[i - 1] > sel_lower else dn2
            if close[i] > sel_upper:
                os_state = 1
            elif close[i] < sel_lower:
                os_state = 0
            ts = sel_lower if os_state == 1 else sel_upper
        else:
            ts = np.nan

        if not np.isnan(ts):
            if np.isnan(ts_prev):
                perf_ama = ts
            else:
                perf_ama = ts if np.isnan(perf_ama) else perf_ama + perf_idx * (ts - perf_ama)

        ts_prev = ts
        ts_arr[i] = ts
        ama_arr[i] = perf_ama
        os_arr[i] = os_state
        perf_idx_arr[i] = perf_idx
        target_factor_arr[i] = target_factor

    df['TS'] = ts_arr
    df['TS_AMA'] = ama_arr
    df['OS'] = os_arr
    df['TREND'] = np.where(df['OS'] == 1, 'Bullish', 'Bearish')
    df['PERF_IDX'] = perf_idx_arr
    df['TARGET_FACTOR'] = target_factor_arr
    df['Supertrend Recomm'] = compute_supertrend_recomm(df['OS'].to_numpy(), df['TS'].to_numpy())
    return df


def compute_supertrend_recomm(os_vals, ts_vals):
    """Direct OS-flip mapping with hold (see module docstring). Returns a
    numpy object array of 'BUY CE' / 'BUY PE' / 'WAIT'."""
    n = len(os_vals)
    recomm = np.full(n, 'WAIT', dtype=object)
    state = None
    for i in range(n):
        if np.isnan(ts_vals[i]):
            recomm[i] = 'WAIT'
            continue
        if state is None or (i > 0 and not np.isnan(ts_vals[i - 1]) and os_vals[i] != os_vals[i - 1]):
            state = 'BUY CE' if os_vals[i] == 1 else 'BUY PE'
        recomm[i] = state
    return recomm


# ---------------------------------------------------------------------------
# Per-symbol processing (same brute-force time/timezone protocol as RSI.py)
# ---------------------------------------------------------------------------
def process_symbol(symbol_data):
    symbol, df, target_date = symbol_data
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip().str.lower()

    if not all(col in df.columns for col in ['open', 'high', 'low', 'close']):
        print(f"[WARNING] {symbol}: Missing required price columns. Discarding.")
        return symbol, None

    time_col = next((col for col in ['date', 'time', 'datetime', 'timestamp'] if col in df.columns), None)
    if time_col:
        parsed = pd.to_datetime(df[time_col], errors='coerce')
    elif pd.api.types.is_datetime64_any_dtype(df.index):
        parsed = pd.Series(df.index, index=df.index)
    else:
        unnamed_cols = [c for c in df.columns if 'unnamed' in c]
        if unnamed_cols:
            parsed = pd.to_datetime(df[unnamed_cols[0]], errors='coerce')
        else:
            print(f"[WARNING] {symbol}: Failed to locate any valid timestamp data. Discarding.")
            return symbol, None

    try:
        if parsed.dt.tz is not None:
            parsed = parsed.dt.tz_localize(None)
    except (AttributeError, TypeError):
        pass

    df['_sort_dt'] = parsed
    df['time_str'] = parsed.dt.strftime('%H:%M')
    df.dropna(subset=['time_str', '_sort_dt'], inplace=True)
    if df.empty:
        print(f"[WARNING] {symbol}: No rows with a valid timestamp. Discarding.")
        return symbol, None

    df.sort_values('_sort_dt', inplace=True)
    df = df[df['_sort_dt'].dt.time <= CUTOFF_TIME]
    if df.empty:
        print(f"[WARNING] {symbol}: No candles at/before {CUTOFF_TIME.strftime('%H:%M')}. Discarding.")
        return symbol, None

    df = calculate_supertrend_ai(df)

    # [ADDED -- 18-Jul-26, Task 48] This module predates the codebase-wide
    # target_date interface change (see run_pipeline.py's own comment on
    # this) and never truncated its output to the target date -- every
    # OTHER indicator restricts its matrix to target_date's own bars
    # right here (same call, same spot) so the sheet ends up with one
    # day's worth of time columns, not the full ~90-day backfill window
    # every symbol's CSV actually contains. Without this, wiring this
    # module in would have produced a sheet with months of columns
    # instead of one trading day's ~75. This is very likely WHY it was
    # never actually wired into run_pipeline.py despite being fully
    # written -- confirmed via git history: it's untouched since the
    # initial commit, not a deliberate later removal.
    df = excel_utils.restrict_to_target_date(df, target_date)
    if df is None:
        return symbol, None

    return symbol, df


# ---------------------------------------------------------------------------
# Excel export -- pivoted Symbol x Time matrix, merged into the existing
# workbook (same pattern as RSI.py / SQZMOM.py: load_workbook, replace only
# the 'Supertrend' sheet, leave every other sheet untouched).
# ---------------------------------------------------------------------------
def build_matrix(data_dict, target_date, max_workers=None):
    """Executes multi-core calculation and builds the pivoted Matrix
    (no Excel I/O -- safe to run concurrently with other indicators'
    build_matrix() calls since it never touches the shared workbook).

    [CHANGED -- Task 48] Added target_date, matching the signature every
    other indicator module's build_matrix() already uses (run_pipeline.py
    calls all of them identically: module.build_matrix(data_dict,
    target_date, max_workers=...)). This module's build_matrix() didn't
    accept it before, which is exactly what made adding this to
    run_pipeline.py's INDICATORS list an immediate TypeError -- not a
    deliberate exclusion."""
    results = {}
    with concurrent.futures.ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(process_symbol, (sym, df, target_date)): sym for sym, df in data_dict.items()}
        for future in concurrent.futures.as_completed(futures):
            sym = futures[future]
            try:
                processed_sym, processed_df = future.result()
                if processed_df is not None and 'TS' in processed_df.columns:
                    processed_df = processed_df.dropna(subset=['TS'])
                    if not processed_df.empty and sym.lower() != 'summary':
                        results[sym] = processed_df
            except Exception as e:
                print(f"[ERROR] Supertrend Engine Failure on symbol {sym}: {str(e)}")

    if not results:
        raise RuntimeError("Supertrend: zero symbols processed successfully -- matrix build aborted.")

    all_times = set()
    for df in results.values():
        if 'time_str' in df.columns:
            all_times.update(df['time_str'].dropna().tolist())

    if not all_times:
        raise RuntimeError("Supertrend: no valid timestamps extracted across all symbols -- matrix build aborted.")

    sorted_times = sorted(all_times)
    matrix_rows = [['Symbol', 'Metrics'] + sorted_times]

    for sym in sorted(results.keys()):
        df = results[sym].drop_duplicates(subset=['time_str'], keep='last')
        df_indexed = df.set_index('time_str')

        def get_metric_row(metric_name, col_name):
            row = [sym, metric_name]
            for t in sorted_times:
                if t in df_indexed.index:
                    val = df_indexed.loc[t, col_name]
                    if isinstance(val, pd.Series):
                        val = val.iloc[-1]
                    row.append(val)
                else:
                    row.append("")
            return row

        matrix_rows.append(get_metric_row('Open', 'open'))
        matrix_rows.append(get_metric_row('Close', 'close'))
        matrix_rows.append(get_metric_row('TS', 'TS'))
        matrix_rows.append(get_metric_row('TREND', 'TREND'))
        matrix_rows.append(get_metric_row('PERF_IDX', 'PERF_IDX'))
        matrix_rows.append(get_metric_row('TARGET_FACTOR', 'TARGET_FACTOR'))
        matrix_rows.append(get_metric_row('Supertrend Recomm', 'Supertrend Recomm'))

    return matrix_rows


def write_matrix_to_workbook(matrix_rows, wb):
    """Applies the 'Supertrend' sheet into an ALREADY-OPEN workbook (wb)
    -- no file load/save here. See excel_utils.replace_sheet_with_matrix()'s
    docstring for why an orchestrator should batch all indicators
    through this instead of write_matrix()'s own load/save."""
    sheet_name = "Supertrend"
    ws = excel_utils.replace_sheet_with_matrix(wb, sheet_name, matrix_rows)

    for r_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1), start=2):
        metric_type = ws.cell(row=r_idx, column=2).value
        for cell in row[2:]:
            val = cell.value
            if val == "":
                continue
            if metric_type == 'TREND':
                if val == 'Bullish':
                    cell.fill = FILL_BULLISH
                    cell.font = FONT_WHITE
                elif val == 'Bearish':
                    cell.fill = FILL_BEARISH
                    cell.font = FONT_WHITE
            elif metric_type == 'Supertrend Recomm':
                if val == 'BUY CE':
                    cell.fill = FILL_LIME
                    cell.font = FONT_BLACK
                elif val == 'BUY PE':
                    cell.fill = FILL_RED
                    cell.font = FONT_WHITE
                elif val == 'WAIT':
                    cell.fill = FILL_GRAY
                    cell.font = FONT_BLACK

    excel_utils.autofit_columns(ws)


def write_matrix(matrix_rows, output_excel_path):
    """Standalone/back-compat entry point: load -> apply -> save. Prefer
    write_matrix_to_workbook() when writing several indicators in the
    same run (see 01_Master_Code.py)."""
    wb = load_workbook(output_excel_path)
    write_matrix_to_workbook(matrix_rows, wb)
    excel_utils.atomic_save(wb, output_excel_path)



def parallel_compute_and_export(data_dict, target_date, output_excel_path, max_workers=None):
    """Backward-compatible fused entry point (build + write in one call)."""
    matrix_rows = build_matrix(data_dict, target_date, max_workers=max_workers)
    write_matrix(matrix_rows, output_excel_path)


def run_supertrend_step(df_ref, target_date, output_excel_path):
    """Single entry point for run_pipeline.py's Step 11 (standalone use)."""
    print("[SYSTEM] Loading 5-minute historical data for SuperTrend AI...")
    data_dict = data_ingestion.load_interval_data(df_ref, target_date, interval=INTERVAL)
    if not data_dict:
        raise RuntimeError("Supertrend: no 5-minute data files found for any symbol.")
    print(f"[PROCESS] Computing SuperTrend AI matrix for {len(data_dict)} symbol(s)...")
    parallel_compute_and_export(data_dict, target_date, output_excel_path)
    print("[SUCCESS] SuperTrend AI matrix written to sheet 'Supertrend'.")


# ---------------------------------------------------------------------------
# Disclaimer: this module estimates trend direction from historical/live
# price data only. It is not financial advice, no result here is a
# guarantee of future performance, and the 'Supertrend Recomm' rule has not
# been backtested in this conversation -- paper-trade it and run it through
# scripts/backtester.py before risking real capital.
# ---------------------------------------------------------------------------
